# ACL 重复检查任务
#
# 技术栈:openpyxl、ipaddress、正则表达式、图论算法、CiscoBase（V11新增）
# 目标:分析ACL规则间的覆盖关系，识别可回收的重复策略
#
# 处理逻辑:
# 解析ACL规则:使用CiscoBase统一解析（V11优化），支持NX-OS CIDR格式和IOS-XE wildcard/host混合格式
# ACL定界:使用CiscoBase.find_acl_blocks_in_column统一处理（V11优化）
# 覆盖关系检测:分析规则A是否覆盖规则B（动作相同、协议匹配、端口匹配、源/目的网段包含关系）
# 图论算法:使用连通分量算法对覆盖规则进行分组
# 可视化标注:每组使用不同颜色标记，最大规则（不被覆盖的）标红
# 统计报告:生成详细的覆盖组统计和可回收策略数量
#
# 输入文件:LOG/DeviceBackupTask/{日期}-关键设备配置备份输出EXCEL基础任务.xlsx（V10新结构：从ACL/SourceACL迁移）
# 输出文件:LOG/ACLDupCheckTask/{日期}-大段覆盖包含明细ACL检查.xlsx（V10新结构：从ACL/ACLDupCheckTask迁移）
#
# 输出:统计覆盖组数量、保留规则数、可回收规则数，生成带颜色标注的Excel报告和Report工作表
#
# 统计指标说明:
# - 覆盖组数量:具有覆盖关系的规则分组数（每组包含相互覆盖的规则）
# - 保留规则数:每组中最大规则数量（红色标记，建议保留，覆盖其他规则）
# - 可回收规则数:每组中被覆盖的规则数量（默认字体，可删除，被其他规则覆盖）
# - 参与分析规则数:所有参与覆盖分析的规则总数（不包括无覆盖关系的规则）
#
# 输出优化:
# - 移除LOG和REPORT中的JSON格式元数据，只保留清晰的中文描述
#
# 配置说明:
# - 输入文件:自动使用当天日期的DeviceBackupTask输出文件（从settings.log_dir读取LOG目录）
# - 输出目录:LOG/ACLDupCheckTask/（从settings.log_dir读取LOG目录，V10新结构：从ACL/ACLDupCheckTask迁移）
# - 汇总输出控制:通过Config.yaml的ACLDupCheckTask.enable_summary_output开关控制汇总统计输出，默认关闭（V10新增）

# 导入标准库
import os
import re
import socket
from dataclasses import dataclass
from datetime import datetime
from ipaddress import IPv4Address, IPv4Network
from typing import Dict, Iterable, List, Optional, Set, Tuple

# 导入第三方库
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# 导入本地应用
from .TaskBase import (
    BaseTask, Level, CONFIG, get_today_str, format_datetime,
    ensure_output_dir, build_log_path, build_output_path,
    load_excel_workbook, create_excel_workbook, save_excel_workbook
)
from .CiscoBase import (
    ACLRule, parse_acl, service_to_port, ip_and_wildcard_to_network,
    host_to_network, cidr_to_network, find_acl_blocks_in_column,
    extract_acl_rules_from_column, is_acl_rule,
    is_cat1_device, is_cat2_device, analyze_first_row_for_cat1_cat2
)

# 解析辅助函数已迁移到CiscoBase
# 使用: from .CiscoBase import service_to_port, ip_and_wildcard_to_network, host_to_network, cidr_to_network

# 正则表达式（支持NX-OS/IOS-XE）
NXOS_RE = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE>\d+\.\d+\.\d+\.\d+\/\d+)\s+
    (?P<DESTINATION>\d+\.\d+\.\d+\.\d+\/\d+)
    (?:\s+eq\s+(?P<PORT>\S+))?
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# NXOS格式：序号 permit 协议 源地址 eq 端口 目标地址
NXOS_SRC_PORT_RE = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE>\d+\.\d+\.\d+\.\d+\/\d+)\s+
    eq\s+(?P<PORT>\S+)\s+
    (?P<DESTINATION>\d+\.\d+\.\d+\.\d+\/\d+)
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

IOSXE_WC_RE = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<SOURCE_WILDCARD>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+(?P<PORT_A>\S+))?
    \s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<DESTINATION_WILDCARD>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+(?P<PORT_B>\S+))?
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

IOSXE_HOST_RE = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    host\s+(?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+(?P<PORT_A>\S+))?
    \s+host\s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+(?P<PORT_B>\S+))?
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

IOSXE_MIX_RE = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?:
        host\s+(?P<SOURCE_IP_HOST>\d+\.\d+\.\d+\.\d+) |
        (?P<SOURCE_IP_WILDCARD>\d+\.\d+\.\d+\.\d+)\s+
        (?P<SOURCE_WILDCARD_WILDCARD>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+(?P<PORT_A>\S+))?
    \s+
    (?:
        host\s+(?P<DESTINATION_IP_HOST>\d+\.\d+\.\d+\.\d+) |
        (?P<DESTINATION_IP_WILDCARD>\d+\.\d+\.\d+\.\d+)\s+
        (?P<DESTINATION_WILDCARD_WILDCARD>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+(?P<PORT_B>\S+))?
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# ACLRule 和 parse_acl 已从 CiscoBase 导入，删除本地重复实现

# 覆盖关系逻辑
# 检查协议A是否覆盖协议B
def proto_covers(PROTO_A: str, PROTO_B: str) -> bool:
    """检查协议A是否覆盖协议B

    Args:
        PROTO_A: 协议A
        PROTO_B: 协议B

    Returns:
        如果PROTO_A为"ip"或与PROTO_B相同则返回True，否则返回False
    """
    return PROTO_A.lower() == "ip" or PROTO_A.lower() == PROTO_B.lower()

# 检查端口A是否覆盖端口B
def port_covers(PORT_A: Optional[int], PORT_B: Optional[int]) -> bool:
    """检查端口A是否覆盖端口B

    Args:
        PORT_A: 端口A（None表示任意端口）
        PORT_B: 端口B（None表示任意端口）

    Returns:
        如果A为None或A等于B则返回True，否则返回False
    """
    # A None => 任意端口；B None 且 A 有端口 => 不覆盖；都有端口需相等
    if PORT_A is None:
        return True
    if PORT_B is None:
        return False
    return PORT_A == PORT_B

# 检查源端口A是否覆盖源端口B
def src_port_covers(SRC_PORT_A: Optional[int], SRC_PORT_B: Optional[int]) -> bool:
    """检查源端口A是否覆盖源端口B

    Args:
        SRC_PORT_A: 源端口A（None表示任意端口）
        SRC_PORT_B: 源端口B（None表示任意端口）

    Returns:
        如果A为None或A等于B则返回True，否则返回False
    """
    if SRC_PORT_A is None:
        return True
    if SRC_PORT_B is None:
        return False
    return SRC_PORT_A == SRC_PORT_B

# 检查目标端口A是否覆盖目标端口B
def dst_port_covers(DST_PORT_A: Optional[int], DST_PORT_B: Optional[int]) -> bool:
    """检查目标端口A是否覆盖目标端口B

    Args:
        DST_PORT_A: 目标端口A（None表示任意端口）
        DST_PORT_B: 目标端口B（None表示任意端口）

    Returns:
        如果A为None或A等于B则返回True，否则返回False
    """
    if DST_PORT_A is None:
        return True  # 任意端口覆盖特定端口
    if DST_PORT_B is None:
        return False  # 特定端口不覆盖任意端口
    return DST_PORT_A == DST_PORT_B

# 检查规则A是否覆盖规则B
def rule_covers(RULE_A: ACLRule, RULE_B: ACLRule) -> bool:
    """检查规则A是否覆盖规则B

    Args:
        RULE_A: ACL规则A
        RULE_B: ACL规则B

    Returns:
        如果规则A覆盖规则B则返回True，否则返回False
    """
    if RULE_A.action != RULE_B.action:
        return False
    if not proto_covers(RULE_A.proto, RULE_B.proto):
        return False


    # 对于NXOS格式，使用port字段；对于IOS-XE格式，使用src_port和dst_port字段
    if RULE_A.style == "NXOS" and RULE_B.style == "NXOS":
        # NXOS格式：检查源端口和目标端口
        if not src_port_covers(RULE_A.src_port, RULE_B.src_port):
            return False
        if not dst_port_covers(RULE_A.dst_port, RULE_B.dst_port):
            return False
    else:
        # IOS-XE格式：检查源端口和目标端口
        # 如果src_port和dst_port都为None，则使用port字段
        if (RULE_A.src_port is None and RULE_A.dst_port is None and
                RULE_B.src_port is None and RULE_B.dst_port is None):
            # 都使用port字段
            if not port_covers(RULE_A.port, RULE_B.port):
                return False
        else:
            # 使用src_port和dst_port字段
            # 但是要确保端口信息的一致性
            if not src_port_covers(RULE_A.src_port, RULE_B.src_port):
                return False
            if not dst_port_covers(RULE_A.dst_port, RULE_B.dst_port):
                return False
            # 如果两个规则都有port字段，也要检查port字段
            if RULE_A.port is not None and RULE_B.port is not None:
                if not port_covers(RULE_A.port, RULE_B.port):
                    return False


    if not RULE_B.src.subnet_of(RULE_A.src):
        return False
    if not RULE_B.dst.subnet_of(RULE_A.dst):
        return False
    return True

# 图论算法辅助函数
# 使用图论算法计算连通分量
def connected_components(
        NODES: Iterable[int], UNDIRECTED_EDGES: List[Tuple[int, int]]
) -> List[Set[int]]:
    """计算无向图的连通分量

    Args:
        NODES: 节点集合
        UNDIRECTED_EDGES: 无向边列表

    Returns:
        连通分量列表，每个元素是一个节点集合
    """
    GRAPH: Dict[int, Set[int]] = {NODE: set() for NODE in NODES}
    for NODE_U, NODE_V in UNDIRECTED_EDGES:
        GRAPH.setdefault(NODE_U, set()).add(NODE_V)
        GRAPH.setdefault(NODE_V, set()).add(NODE_U)
    VISITED_NODES: Set[int] = set()
    COMPONENTS: List[Set[int]] = []
    for NODE in NODES:
        if NODE in VISITED_NODES:
            continue
        STACK = [NODE]
        COMPONENT = set()
        while STACK:
            CURRENT_NODE = STACK.pop()
            if CURRENT_NODE in VISITED_NODES:
                continue
            VISITED_NODES.add(CURRENT_NODE)
            COMPONENT.add(CURRENT_NODE)
            STACK.extend(GRAPH.get(CURRENT_NODE, []))
        COMPONENTS.append(COMPONENT)
    return COMPONENTS

# 严格连通分量算法：只考虑直接覆盖关系，避免间接连接
# 严格连通分量算法：只将直接有覆盖关系的规则归为一组，避免通过中间规则间接连接不相关的规则
def _strict_connected_components(
        NODES: Iterable[int], UNDIRECTED_EDGES: List[Tuple[int, int]]
) -> List[Set[int]]:
    if not UNDIRECTED_EDGES:
        return []


    # 构建邻接表
    GRAPH: Dict[int, Set[int]] = {NODE: set() for NODE in NODES}
    for NODE_U, NODE_V in UNDIRECTED_EDGES:
        GRAPH.setdefault(NODE_U, set()).add(NODE_V)
        GRAPH.setdefault(NODE_V, set()).add(NODE_U)


    VISITED_NODES: Set[int] = set()
    COMPONENTS: List[Set[int]] = []


    for NODE in NODES:
        if NODE in VISITED_NODES:
            continue


        # 使用BFS找到所有直接连接的节点
        QUEUE = [NODE]
        COMPONENT = set()


        while QUEUE:
            CURRENT_NODE = QUEUE.pop(0)
            if CURRENT_NODE in VISITED_NODES:
                continue
            VISITED_NODES.add(CURRENT_NODE)
            COMPONENT.add(CURRENT_NODE)


            # 只添加直接相邻的节点
            for NEIGHBOR in GRAPH.get(CURRENT_NODE, []):
                if NEIGHBOR not in VISITED_NODES:
                    QUEUE.append(NEIGHBOR)


        if COMPONENT:
            COMPONENTS.append(COMPONENT)


    return COMPONENTS

# 智能连通分量算法：避免通过被多个规则覆盖的中间节点创建不合理的间接连接
# 智能连通分量算法：避免通过被多个规则覆盖的中间节点创建不合理的间接连接，
# 如果某个节点被多个规则覆盖，则将其从连通图中移除，避免不合理的间接连接
def _smart_connected_components(
    NODES: Iterable[int], UNDIRECTED_EDGES: List[Tuple[int, int]],
    DIRECTED_EDGES: List[Tuple[int, int]]
) -> List[Set[int]]:
    if not UNDIRECTED_EDGES:
        return []


    # 统计每个节点的入度（被多少个规则覆盖）
    INDEGREE = {}
    for NODE in NODES:
        INDEGREE[NODE] = 0


    for FROM_NODE, TO_NODE in DIRECTED_EDGES:
        if TO_NODE in INDEGREE:
            INDEGREE[TO_NODE] += 1


    # 找出被多个规则覆盖的节点（入度 > 1）
    MULTI_COVERED_NODES = {NODE for NODE, DEGREE in INDEGREE.items() if DEGREE > 1}


    # 为被多个规则覆盖的节点找到最靠前的覆盖者
    NODE_TO_GROUP = {}
    for NODE in MULTI_COVERED_NODES:
        # 找到所有覆盖这个节点的规则
        COVERERS = []
        for FROM_NODE, TO_NODE in DIRECTED_EDGES:
            if TO_NODE == NODE:
                COVERERS.append(FROM_NODE)
        # 选择行号最小的覆盖者
        if COVERERS:
            NODE_TO_GROUP[NODE] = min(COVERERS)


    # 构建邻接表，将被多个规则覆盖的节点连接到最靠前的覆盖者
    GRAPH: Dict[int, Set[int]] = {NODE: set() for NODE in NODES}
    for NODE_U, NODE_V in UNDIRECTED_EDGES:
        # 如果节点V被多个规则覆盖，连接到其最靠前的覆盖者
        if NODE_V in MULTI_COVERED_NODES and NODE_V in NODE_TO_GROUP:
            TARGET_GROUP = NODE_TO_GROUP[NODE_V]
            if NODE_U == TARGET_GROUP:
                GRAPH.setdefault(NODE_U, set()).add(NODE_V)
                GRAPH.setdefault(NODE_V, set()).add(NODE_U)
        # 如果节点U被多个规则覆盖，连接到其最靠前的覆盖者
        elif NODE_U in MULTI_COVERED_NODES and NODE_U in NODE_TO_GROUP:
            TARGET_GROUP = NODE_TO_GROUP[NODE_U]
            if NODE_V == TARGET_GROUP:
                GRAPH.setdefault(NODE_U, set()).add(NODE_V)
                GRAPH.setdefault(NODE_V, set()).add(NODE_U)
        # 正常连接
        else:
            GRAPH.setdefault(NODE_U, set()).add(NODE_V)
            GRAPH.setdefault(NODE_V, set()).add(NODE_U)


    VISITED_NODES: Set[int] = set()
    COMPONENTS: List[Set[int]] = []


    for NODE in NODES:
        if NODE in VISITED_NODES:
            continue


        # 使用BFS找到所有连接的节点
        QUEUE = [NODE]
        COMPONENT = set()


        while QUEUE:
            CURRENT_NODE = QUEUE.pop(0)
            if CURRENT_NODE in VISITED_NODES:
                continue
            VISITED_NODES.add(CURRENT_NODE)
            COMPONENT.add(CURRENT_NODE)


            # 添加相邻的节点
            for NEIGHBOR in GRAPH.get(CURRENT_NODE, []):
                if NEIGHBOR not in VISITED_NODES:
                    QUEUE.append(NEIGHBOR)


        if COMPONENT:
            COMPONENTS.append(COMPONENT)


    return COMPONENTS


# 基础颜色调色板 - 精选20种高对比度颜色
FILL_COLORS = [
    "FFFDE68A",  # amber-300
    "FFA7F3D0",  # teal-200

    "FFFCA5A5",  # red-300
    "FF93C5FD",  # blue-300
    "FFA5B4FC",  # violet-300
    "FF86EFAC",  # green-300
    "FFFBCFE8",  # pink-200
    "FFE9D5FF",  # purple-200
    "FF67E8F9",  # cyan-300
    "FFFDE1AF",  # custom light
    "FFDBEAFE",  # blue-100
    "FFDCFCE7",  # green-100
    "FFFEE2E2",  # red-100
    "FFFEFCE8",  # yellow-100
    "FFF3E8FF",  # purple-100
    "FFFFF7ED",  # orange-100
    "FFF9FAFB",  # gray-100
    "FFECFEFF",  # cyan-100
    "FFFDF2F8",  # pink-100
    "FFEEF2FF",  # indigo-100
]

# 根据索引生成填充颜色
def fill_for_index(idx: int) -> PatternFill:
    """根据索引生成填充颜色

    Args:
        idx: 索引值

    Returns:
        PatternFill对象，颜色从FILL_COLORS列表中循环选择
    """
    color = FILL_COLORS[idx % len(FILL_COLORS)]
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

# 红字样式（仅颜色，如需加粗可设 bold=True）
RED_FONT = Font(color="FFFF0000")

# 判断文本是否为ACL规则
def _is_acl_rule(text: str) -> bool:
    text = text.strip().lower()
    # 排除证书、配置等非ACL数据
    if any(keyword in text for keyword in [
        'certificate', 'crypto', 'quit', 'exit', 'config', 'interface',
        'router', 'version', 'hostname', 'enable', 'password', 'username',
        'line', 'service', 'logging', 'ntp', 'snmp', 'tacacs', 'radius'
    ]):
        return False


    # 只处理包含permit/deny的规则
    return 'permit' in text or 'deny' in text


# 处理单个ACL块内的覆盖关系
def process_acl_block(worksheet, col, start_row, end_row):
    """处理ACL块，采集所有规则并检测覆盖关系

    Args:
        worksheet: 工作表对象
        col: 列号
        start_row: 起始行
        end_row: 结束行

    Returns:
        元组：(规则字典, 覆盖边列表, 保留规则集合)
    """
    # 采集ACL块内的所有规则
    col_rules: Dict[int, ACLRule] = {}
    for ROW in range(start_row, end_row + 1):
        cell_value = worksheet.cell(row=ROW, column=col).value
        if cell_value is None:
            continue
        cleaned_text = str(cell_value).strip()
        if not cleaned_text:
            continue


        # 只处理真正的ACL规则，忽略证书等数据
        if not _is_acl_rule(cleaned_text):
            continue


        # 解析ACL规则
        parsed_rule, parse_error = parse_acl(cleaned_text)
        if parsed_rule:
            col_rules[ROW] = parsed_rule

    # 检查是否有足够的规则
    if len(col_rules) < 2:
        return 0, 0, 0, 0  # groups, keep, recycle, total_in_groups

    # 覆盖关系（同列内）- 只考虑直接覆盖关系，避免间接连接
    directed_edges: List[Tuple[int, int]] = []    # A->B 表示 A 覆盖 B
    undirected_edges: List[Tuple[int, int]] = []  # 分组用
    rows = sorted(col_rules.keys())


    for rule_i in range(len(rows)):
        for rule_j in range(rule_i + 1, len(rows)):
            row_i, row_j = rows[rule_i], rows[rule_j]
            rule_a, rule_b = col_rules[row_i], col_rules[row_j]
            a_covers_b = rule_covers(rule_a, rule_b)
            b_covers_a = rule_covers(rule_b, rule_a)
            # 只有当两个规则之间有直接覆盖关系时才连接
            if a_covers_b or b_covers_a:
                undirected_edges.append((row_i, row_j))
                if a_covers_b:
                    directed_edges.append((row_i, row_j))
                if b_covers_a:
                    directed_edges.append((row_j, row_i))

    if not undirected_edges:
        return 0, 0, 0, 0  # 当前列没有覆盖关系

    # 仅取参与覆盖关系的节点
    nodes_in_edges: Set[int] = set()
    for node_u, node_v in undirected_edges:
        nodes_in_edges.add(node_u)
        nodes_in_edges.add(node_v)

    # 分量划分（组）- 使用智能分组算法
    # 被多个规则覆盖的节点归属到最靠前的组，避免不合理的间接连接
    comps = _smart_connected_components(sorted(nodes_in_edges), undirected_edges, directed_edges)

    groups_count = 0
    keep_count = 0
    recycle_count = 0
    total_in_groups = 0

    # 每个分量着色 & 最大规则标红 & 统计
    for idx, comp in enumerate(comps):
        groups_count += 1
        fill = fill_for_index(idx)

        # 计算入度（只看分量内部的有向边）
        indeg: Dict[int, int] = {row_number: 0 for row_number in comp}
        for edge_u, edge_v in directed_edges:
            if edge_u in comp and edge_v in comp:
                indeg[edge_v] += 1

        # 最大规则：每个组只有一条，选择入度为0且行号最小的规则
        # 如果同时被多个规则覆盖，归属到靠前的组（行号小的）
        zero_indegree_nodes = [row_number for row_number, degree in indeg.items() if degree == 0]
        if zero_indegree_nodes:
            # 选择行号最小的作为最大规则
            maxima = {min(zero_indegree_nodes)}
        else:
            # 如果没有入度为0的节点，选择行号最小的
            maxima = {min(comp)}
        group_total = len(comp)
        group_keep = len(maxima)
        group_recycle = group_total - group_keep

        total_in_groups += group_total
        keep_count += group_keep
        recycle_count += group_recycle

        # 上底色 - 应用到当前列
        for ROW_NUMBER in comp:
            worksheet.cell(row=ROW_NUMBER, column=col).fill = fill

        # 标红字体 - 应用到当前列的最大规则
        for ROW_NUMBER in maxima:
            cell = worksheet.cell(row=ROW_NUMBER, column=col)
            current_font = cell.font or Font()
            # 强制使用宋体字体
            cell.font = Font(
                name="宋体",
                sz=current_font.sz,

                b=current_font.b,

                i=current_font.i,
                underline=current_font.u,

                strike=current_font.strike,

                color="FFFF0000"
            )

    return groups_count, keep_count, recycle_count, total_in_groups

# 主处理流程
# 处理Excel文件，执行ACL覆盖检查（优化版本）
def process_file(input_path: str, output_path: str) -> Dict[str, Dict[str, int]]:
    """处理Excel文件，执行ACL覆盖检查并生成报告

    Args:
        input_path: 输入文件路径
        output_path: 输出文件路径

    Returns:
        每个工作表的统计信息字典
    """
    input_workbook = load_excel_workbook(input_path)
    # sheet -> {groups, keep, recycle, total_in_groups}
    per_sheet_stats: Dict[str, Dict[str, int]] = {}

    for worksheet in input_workbook.worksheets:
        sheet_name = worksheet.title

        groups_count = 0
        keep_count = 0       # 红色（保留）
        recycle_count = 0    # 默认字体（可回收）
        total_in_groups = 0  # 参与覆盖组的规则总数

        # 1. 分析第一行，识别cat1/cat2/cat6列
        cat1_cols, cat2_cols, cat6_cols = analyze_first_row_for_cat1_cat2(worksheet)
        target_cols = cat1_cols + cat2_cols


        # 2. 对每个目标列处理ACL块
        for COLUMN in target_cols:
            acl_blocks = find_acl_blocks_in_column(worksheet, COLUMN)


            for start_row, end_row in acl_blocks:
                # 3. 处理ACL块内的覆盖关系
                block_groups, block_keep, block_recycle, block_total = (
                    process_acl_block(worksheet, COLUMN, start_row, end_row)
                )
                groups_count += block_groups
                keep_count += block_keep
                recycle_count += block_recycle
                total_in_groups += block_total

        per_sheet_stats[sheet_name] = {
            "groups": groups_count,
            "keep": keep_count,
            "recycle": recycle_count,
            "total_in_groups": total_in_groups,
        }

    # 删除 Report 工作表（如果存在）
    if "Report" in input_workbook.sheetnames:
        del input_workbook["Report"]

    input_workbook.save(output_path)
    return per_sheet_stats

# ACL重复检查任务类：分析ACL规则覆盖关系并生成可视化报告
# ACL重复检查任务：分析ACL规则间的覆盖关系，识别可回收的重复策略，使用图论算法进行分组，最大规则标红显示，生成带颜色标注的Excel报告和统计汇总
class ACLDupCheckTask(BaseTask):
    """ACL重复检查任务

    检查大段覆盖包含明细ACL，识别冗余的ACL规则
    """

    # 初始化ACL重复检查任务：设置固定配置参数
    def __init__(self):
        super().__init__("大段覆盖包含明细ACL检查")
        # 固定配置参数
        today = get_today_str()
        # V10新结构：从 LOG/DeviceBackupTask/ 读取（ACL/SourceACL已迁移）
        self.INPUT_PATH = build_log_path("DeviceBackupTask", f"{today}-关键设备配置备份输出EXCEL基础任务.xlsx")
        # V10新结构：直接输出到 LOG/ACLDupCheckTask/
        self.OUTPUT_DIR = build_log_path("ACLDupCheckTask")
        self.NAME = "大段覆盖包含明细ACL检查"

    # 返回要处理的Sheet列表
    def items(self):
        """返回要处理的工作表列表

        Returns:
            工作表名称列表，排除Report工作表
        """
        if not os.path.exists(self.INPUT_PATH):
            return []
        try:
            input_workbook = load_excel_workbook(self.INPUT_PATH)
            sheet_names = [
                worksheet.title for worksheet in input_workbook.worksheets
                if worksheet.title != 'Report'
            ]
            input_workbook.close()
            return sheet_names
        except Exception:
            return []

    # 处理单个Sheet：执行ACL覆盖检查并生成报告
    def run_single(self, sheet_name: str):
        """处理单个工作表

        执行ACL覆盖检查并生成报告

        Args:
            sheet_name: 工作表名称
        """
        try:
            # 确保输出目录存在
            ensure_output_dir(self.OUTPUT_DIR)

            # 生成输出文件名
            today = get_today_str()
            output_filename = f"{today}-大段覆盖包含明细ACL检查.xlsx"
            output_file_path = build_output_path(self.OUTPUT_DIR, output_filename)

            # 使用原始 acl_dup.py 的处理逻辑
            per_sheet_stats = process_file(self.INPUT_PATH, output_file_path)

            if sheet_name in per_sheet_stats:
                stats = per_sheet_stats[sheet_name]
                self.add_result(
                    Level.OK,
                    f"Sheet {sheet_name} 处理完成：可回收组数={stats['groups']}，"
                    f"保留规则={stats['keep']}，可回收规则={stats['recycle']}"
                )
            else:
                self.add_result(Level.WARN, f"Sheet {sheet_name} 无有效 ACL 规则")

        except Exception as EXCEPTION:
            self.add_result(Level.ERROR, f"处理 Sheet {sheet_name} 失败: {EXCEPTION}")

    # 重写run方法：处理所有Sheet并生成最终报告
    def run(self) -> None:
        """执行ACL重复检查任务

        处理所有工作表，执行ACL覆盖检查并生成报告
        """
        task_items = list(self.items())
        if not task_items:
            self.add_result(Level.ERROR, "未找到可处理的 Sheet")
            return

        # 使用父类的进度条处理
        from tqdm import tqdm

        from .TaskBase import BAR_FORMAT, SHOW_PROGRESS

        progress = tqdm(
            total=len(task_items),
            desc=self.NAME,
            position=0,
            leave=True,
            dynamic_ncols=True,
            bar_format=BAR_FORMAT,
        ) if SHOW_PROGRESS else None

        try:
            # 收集所有 Sheet 的统计信息
            all_sheet_stats = {}

            for sheet_name in task_items:
                try:
                    # 处理单个 Sheet（这会调用 run_single）
                    self.run_single(sheet_name)

                    # 从结果中提取统计信息
                    for result in self.RESULTS:
                        if (result.meta and
                            result.meta.get("sheet_name") == sheet_name and
                            "stats" in result.meta):
                            all_sheet_stats[sheet_name] = result.meta["stats"]
                            break

                except Exception as EXCEPTION:
                    self.add_result(Level.ERROR, f"Sheet {sheet_name} 运行异常: {EXCEPTION!r}")

                if progress:
                    progress.update(1)

            # 生成最终汇总报告
            if all_sheet_stats:
                self._generate_final_report(all_sheet_stats)

        finally:
            if progress:
                progress.close()

    # 生成最终汇总报告：创建Report工作表并保存Excel文件
    def _generate_final_report(self, all_sheet_stats: Dict[str, Dict[str, int]]) -> None:
        try:
            # 生成输出文件名
            today = get_today_str()
            output_filename = f"{today}-大段覆盖包含明细ACL检查.xlsx"
            output_file_path = build_output_path(self.OUTPUT_DIR, output_filename)

            # 创建最终报告 - 使用已经处理过的Excel文件
            output_workbook = load_excel_workbook(output_file_path)

            # 删除 Report 工作表（如果存在）
            if "Report" in output_workbook.sheetnames:
                del output_workbook["Report"]

            save_excel_workbook(output_workbook, output_file_path)

            # 汇总统计（通过Config.yaml的enable_summary_output开关控制，仅输出到LOG文件）
            try:
                from .TaskBase import require_keys
                require_keys(CONFIG, ["ACLDupCheckTask"], "root")
                enable_summary = CONFIG["ACLDupCheckTask"].get("enable_summary_output", False)
            except Exception:
                enable_summary = False


            if enable_summary:
                total_groups = sum(stats.get("groups", 0) for stats in all_sheet_stats.values())
                total_keep = sum(stats.get("keep", 0) for stats in all_sheet_stats.values())
                total_recycle = sum(stats.get("recycle", 0) for stats in all_sheet_stats.values())
                total_in_groups_all = sum(
                    stats.get("total_in_groups", 0)
                    for stats in all_sheet_stats.values()
                )
                self.add_result(
                    Level.OK,
                    f"ACL覆盖检查全部完成：处理{len(all_sheet_stats)}个站点，发现{total_groups}个覆盖组，"
                    f"建议保留{total_keep}条规则（红色标记），可回收{total_recycle}条重复规则，"
                    f"共{total_in_groups_all}条规则参与覆盖分析"
                )

        except Exception as EXCEPTION:
            self.add_result(Level.ERROR, f"生成最终报告失败: {EXCEPTION}")
