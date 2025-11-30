# 通用基础类和工具函数
#
# 技术栈:Python、paramiko、yaml、tqdm、openpyxl、base64、dataclasses、enum
# 目标:提供所有巡检任务的通用基类和工具函数，实现代码复用和统一的任务执行框架
#
# 主要功能:
# - 基础任务类:BaseTask基类，提供通用的任务执行框架（items、run_single、run方法），支持进度条显示和结果记录
# - 结果模型:Level枚举（OK/WARN/CRIT/ERROR）和Result数据类，统一巡检结果的表示和存储
# - 密码加密:提供密码加密/解密功能（XOR算法），支持enc:前缀的加密密码，自动同步加密密钥
# - SSH工具:提供SSH连接创建和命令执行功能（create_ssh_connection、ssh_exec）
# - 配置管理:读取YAML配置文件（Config.yaml），提供配置验证功能（require_keys）
# - 进度条:根据配置动态生成进度条格式，支持百分比、进度条、计数、时间等显示选项
# - 日期处理:提供日期字符串格式化工具（get_today_str、format_datetime）
# - 文件路径:提供LOG目录路径构建和输出目录创建工具（build_log_path、build_output_path、ensure_output_dir）
# - Excel操作:提供Excel工作簿加载、创建、保存工具（load_excel_workbook、create_excel_workbook、save_excel_workbook）
# - 工具函数:提供站点名提取（extract_site_from_filename、extract_site_from_device）、Excel工作表名称安全化（safe_sheet_name）、百分比分级（grade_percent）等工具函数
#
# 使用说明:
# - 所有巡检任务应继承BaseTask基类，实现items()和run_single()方法
# - 使用add_result()方法记录巡检结果，结果会自动分级（OK/WARN/CRIT/ERROR）
# - 使用CONFIG全局变量访问配置，使用require_keys()验证配置完整性
# - 使用decrypt_password()解密配置中的加密密码
# - 使用工具函数处理日期、文件路径、Excel操作等通用需求
#
# 配置依赖:
# - YAML/Config.yaml:主配置文件，包含settings、timeouts、network、progress等配置节
# - 支持通过settings.suppress_ok_logs控制是否抑制OK级别日志
# - 支持通过settings.show_progress控制是否显示进度条

# 导入标准库
import base64
import os
from dataclasses import dataclass
from datetime import datetime
from enum import Enum
from typing import Any, Dict, Iterable, Optional, Tuple

# 导入第三方库
import paramiko
import yaml
from tqdm import tqdm

# 读取配置
with open("YAML/Config.yaml", "r", encoding="utf-8") as fileHandle:
    CONFIG = yaml.safe_load(fileHandle)

_DEFAULT_SYNC_TOKEN = "inspection-v9-default-2024"

# 使用密钥解密密码：XOR解密算法，将enc:前缀的加密密码解密为明文
def _decrypt_with_key(encrypted_password: str, key: bytes) -> str:
    if not encrypted_password.startswith("enc:"):
        return encrypted_password
    try:
        ENCRYPTED_BYTES = base64.b64decode(encrypted_password[4:])
        DECRYPTED_BYTES = bytearray()
        for I, BYTE in enumerate(ENCRYPTED_BYTES):
            DECRYPTED_BYTES.append(BYTE ^ key[I % len(key)])
        return DECRYPTED_BYTES.decode('utf-8')
    except Exception:
        return None

# 使用密钥加密密码：XOR加密算法，将明文密码加密并添加enc:前缀
def _encrypt_with_key(plain_password: str, key: bytes) -> str:
    PLAIN_BYTES = plain_password.encode('utf-8')
    ENCRYPTED_BYTES = bytearray()
    for I, BYTE in enumerate(PLAIN_BYTES):
        ENCRYPTED_BYTES.append(BYTE ^ key[I % len(key)])
    ENCRYPTED_B64 = base64.b64encode(bytes(ENCRYPTED_BYTES)).decode('ascii')
    return f"enc:{ENCRYPTED_B64}"

# 重新加密所有密码：递归遍历配置字典，使用新密钥重新加密所有密码字段
def _re_encrypt_all_passwords(old_key: str, new_key: str, config_dict: dict, path: str = ""):
    UPDATED = False
    if isinstance(config_dict, dict):
        for KEY, VALUE in config_dict.items():
            CURRENT_PATH = f"{path}.{KEY}" if path else KEY
            if KEY == "password" and isinstance(VALUE, str) and VALUE.startswith("enc:"):
                OLD_KEY_BYTES = old_key.encode()
                PLAIN = _decrypt_with_key(VALUE, OLD_KEY_BYTES)
                if PLAIN:
                    NEW_KEY_BYTES = new_key.encode()
                    NEW_ENCRYPTED = _encrypt_with_key(PLAIN, NEW_KEY_BYTES)
                    config_dict[KEY] = NEW_ENCRYPTED
                    UPDATED = True
            elif isinstance(VALUE, dict):
                if _re_encrypt_all_passwords(old_key, new_key, VALUE, CURRENT_PATH):
                    UPDATED = True
            elif isinstance(VALUE, list):
                for I, ITEM in enumerate(VALUE):
                    if isinstance(ITEM, dict):
                        if _re_encrypt_all_passwords(
                                old_key, new_key, ITEM, f"{CURRENT_PATH}[{I}]"
                        ):
                            UPDATED = True
    return UPDATED

# 同步加密密钥到配置：检查并更新配置中的加密密钥，必要时重新加密所有密码
def _sync_encrypt_key_to_config():
    try:
        SETTINGS = CONFIG.get("settings", {})
        OLD_CONFIG_KEY = SETTINGS.get("config_version")
        NEW_CONFIG_KEY = _DEFAULT_SYNC_TOKEN


        if OLD_CONFIG_KEY and OLD_CONFIG_KEY != NEW_CONFIG_KEY:
            NEED_RE_ENCRYPT = _re_encrypt_all_passwords(OLD_CONFIG_KEY, NEW_CONFIG_KEY, CONFIG)
            if NEED_RE_ENCRYPT:
                if "config_version" in SETTINGS:
                    del CONFIG["settings"]["config_version"]
                with open("YAML/Config.yaml", "w", encoding="utf-8") as FILE_HANDLE:
                    yaml.dump(
                        CONFIG, FILE_HANDLE, allow_unicode=True,
                        default_flow_style=False, sort_keys=False
                    )
        elif "config_version" in SETTINGS:
            del CONFIG["settings"]["config_version"]
            with open("YAML/Config.yaml", "w", encoding="utf-8") as FILE_HANDLE:
                yaml.dump(
                    CONFIG, FILE_HANDLE, allow_unicode=True,
                    default_flow_style=False, sort_keys=False
                )
    except Exception:
        pass

_sync_encrypt_key_to_config()
with open("YAML/Config.yaml", "r", encoding="utf-8") as FILE_HANDLE:
    CONFIG = yaml.safe_load(FILE_HANDLE)

_ENCRYPT_KEY = os.environ.get("ENCRYPT_KEY", _DEFAULT_SYNC_TOKEN).encode()

# 加密密码：使用默认密钥加密明文密码，返回enc:前缀的加密字符串
def encrypt_password(plain_password: str) -> str:
    """加密密码


    Args:
        plain_password: 明文密码


    Returns:
        str: 加密后的密码（带enc:前缀）
    """
    PLAIN_BYTES = plain_password.encode('utf-8')
    KEY_BYTES = _ENCRYPT_KEY
    ENCRYPTED_BYTES = bytearray()
    for I, BYTE in enumerate(PLAIN_BYTES):
        ENCRYPTED_BYTES.append(BYTE ^ KEY_BYTES[I % len(KEY_BYTES)])
    ENCRYPTED_B64 = base64.b64encode(bytes(ENCRYPTED_BYTES)).decode('ascii')
    return f"enc:{ENCRYPTED_B64}"

# 解密密码：使用默认密钥解密enc:前缀的加密密码，返回明文密码
def decrypt_password(encrypted_password: str) -> str:
    """解密密码


    Args:
        encrypted_password: 加密的密码（带enc:前缀）


    Returns:
        str: 明文密码


    Raises:
        ValueError: 如果解密失败
    """
    if not encrypted_password.startswith("enc:"):
        return encrypted_password
    try:
        ENCRYPTED_BYTES = base64.b64decode(encrypted_password[4:])
        KEY_BYTES = _ENCRYPT_KEY
        DECRYPTED_BYTES = bytearray()
        for I, BYTE in enumerate(ENCRYPTED_BYTES):
            DECRYPTED_BYTES.append(BYTE ^ KEY_BYTES[I % len(KEY_BYTES)])
        return DECRYPTED_BYTES.decode('utf-8')
    except Exception as ERROR:
        raise ValueError(f"处理失败: {ERROR}")

# 从配置读取常量
# 从YAML配置中读取超时设置（移除硬编码默认值，只从配置文件读取）
DEFAULT_SSH_TIMEOUT = CONFIG["timeouts"]["ssh"]
DEFAULT_HTTP_TIMEOUT = CONFIG["timeouts"]["http"]
DEFAULT_PAGE_GOTO_TIMEOUT = CONFIG["timeouts"]["page_goto"]
DEFAULT_SELECTOR_TIMEOUT = CONFIG["timeouts"]["selector"]

# 从YAML配置中读取网络设置（移除硬编码默认值，只从配置文件读取）
WEB_MAX_WORKERS = CONFIG["network"]["max_workers"]
WEB_DELAY_RANGE = tuple(CONFIG["network"]["delay_range"])
BLOCK_RES_TYPES = set(CONFIG["network"]["block_resources"])

# 根据配置参数动态生成进度条格式：从YAML配置构建自定义进度条显示格式
def _build_progress_format():
    PROGRESS_CONFIG = CONFIG["progress"]

    # 获取配置参数（移除硬编码默认值，只从配置文件读取）
    SHOW_PERCENTAGE = PROGRESS_CONFIG["show_percentage"]
    SHOW_BAR = PROGRESS_CONFIG["show_bar"]
    SHOW_COUNT = PROGRESS_CONFIG["show_count"]
    SHOW_ELAPSED = PROGRESS_CONFIG["show_elapsed"]
    SHOW_REMAINING = PROGRESS_CONFIG["show_remaining"]
    BAR_LENGTH = PROGRESS_CONFIG["bar_length"]
    PREFIX = PROGRESS_CONFIG["prefix"]
    SUFFIX = PROGRESS_CONFIG["suffix"]

    # 构建格式字符串
    PARTS = []

    if SHOW_PERCENTAGE:
        PARTS.append("{percentage:3.0f}%")

    if SHOW_BAR:
        PARTS.append(f"|{{bar:{BAR_LENGTH}}}|")

    if SHOW_COUNT:
        PARTS.append("{n:>3d}/{total:>3d}")

    if SHOW_ELAPSED or SHOW_REMAINING:
        TIME_PARTS = []
        ELAPSED_LABEL = PROGRESS_CONFIG["elapsed_label"]
        REMAINING_LABEL = PROGRESS_CONFIG["remaining_label"]
        if SHOW_ELAPSED:
            TIME_PARTS.append(f"{ELAPSED_LABEL}:{{elapsed}}")
        if SHOW_REMAINING:
            TIME_PARTS.append(f"{REMAINING_LABEL}:{{remaining}}")
        PARTS.append(f"[{'  '.join(TIME_PARTS)}]")

    # 组合最终格式
    FORMAT_STR = " ".join(PARTS)
    if PREFIX:
        FORMAT_STR = f"{PREFIX} {FORMAT_STR}"
    if SUFFIX:
        FORMAT_STR = f"{FORMAT_STR} {SUFFIX}"

    return FORMAT_STR

BAR_FORMAT = _build_progress_format()

# 检查配置键是否存在：验证配置字典中是否包含必需的键
def require_keys(d: Dict[str, Any], keys: Iterable[str], ctx: str) -> None:
    """检查配置字典中是否包含必需的键

    Args:
        d: 配置字典
        keys: 必需的键列表
        ctx: 上下文名称（用于错误信息）

    Raises:
        ValueError: 如果缺少必需的键
    """
    for KEY in keys:
        if KEY not in d:
            raise ValueError(f"配置缺失: {ctx}.{KEY}")

require_keys(CONFIG, ["settings", "MirrorFortiGateTask", "OxidizedTask", "ESServer"], "root")

SHOW_PROGRESS = bool(CONFIG["settings"]["show_progress"])

# 分级结果模型
# 告警级别枚举：定义巡检结果的严重程度等级
class Level(Enum):
    """告警级别枚举


    定义巡检结果的严重程度等级
    """
    OK = "OK"
    WARN = "WARN"
    CRIT = "CRIT"
    ERROR = "ERROR"

# 巡检结果数据类：存储单个巡检任务的结果信息
@dataclass
class Result:
    """巡检结果数据类


    存储单个巡检任务的结果信息
    """
    level: str
    message: str
    meta: Optional[dict] = None

# 通用工具
# 百分比分级函数：根据阈值将数值映射到告警级别
def grade_percent(value: Optional[float], warn: int, crit: int) -> Level:
    """百分比分级函数


    根据阈值将数值映射到告警级别


    Args:
        value: 要分级的数值
        warn: 警告阈值
        crit: 严重阈值


    Returns:
        Level: 告警级别
    """
    if value is None:
        return Level.ERROR
    if value >= crit:
        return Level.CRIT
    if value >= warn:
        return Level.WARN
    return Level.OK

# 从文件名提取站点名：使用正则表达式从文件名中提取站点标识
def extract_site_from_filename(filename: str, pattern: str = r"^\d{8}-([^-]+)-") -> str:
    """从文件名提取站点名


    Args:
        filename: 文件名
        pattern: 正则表达式模式


    Returns:
        str: 站点名
    """
    import re
    BASE = os.path.basename(filename)
    MATCH = re.match(pattern, BASE)
    return MATCH.group(1) if MATCH else "DEFAULT"

# 从设备名提取站点名：从设备名称中提取站点标识
def extract_site_from_device(device_name: str) -> str:
    """从设备名提取站点名


    Args:
        device_name: 设备名称


    Returns:
        str: 站点名，如果无法提取则返回None
    """
    import re
    # 支持 HX 和 P 开头的站点
    MATCH = re.match(r'^(HX\d+|P\d+)', device_name)
    return MATCH.group(1) if MATCH else None

# 创建SSH连接：建立到远程主机的SSH连接
def create_ssh_connection(
    host: str, port: int, username: str, password: str,
    timeout: int = DEFAULT_SSH_TIMEOUT
) -> paramiko.SSHClient:
    """创建SSH连接


    Args:
        host: 主机地址
        port: 端口号
        username: 用户名
        password: 密码
        timeout: 超时时间（秒）


    Returns:
        paramiko.SSHClient: SSH客户端对象
    """
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(host, port=port, username=username, password=password, timeout=timeout)
    return ssh

# 生成安全的Excel工作表名称：确保工作表名称符合Excel规范
def safe_sheet_name(name: str) -> str:
    """生成安全的Excel工作表名称


    Args:
        name: 原始名称


    Returns:
        str: 安全的Excel工作表名称（最多31个字符）
    """
    import re
    safe_name = re.sub(r'[:\\/*?\[\]]', "_", name)[:31]
    return safe_name or "Sheet1"

# 执行SSH命令：通过SSH连接执行远程命令并返回结果
def ssh_exec(
    ssh: paramiko.SSHClient, cmd: str,
    timeout: int = DEFAULT_SSH_TIMEOUT, label: str = ""
) -> Tuple[int, str, str]:
    """执行SSH命令


    Args:
        ssh: SSH客户端对象
        cmd: 要执行的命令
        timeout: 超时时间（秒）
        label: 命令标签（用于错误信息）


    Returns:
        Tuple[int, str, str]: (退出码, 标准输出, 标准错误)


    Raises:
        RuntimeError: 如果命令执行失败
    """
    try:
        stdin, stdout, stderr = ssh.exec_command(cmd, timeout=timeout)
        stdout_text = stdout.read().decode("utf-8", "ignore")
        stderr_text = stderr.read().decode("utf-8", "ignore")
        exit_code = stdout.channel.recv_exit_status()
        return exit_code, stdout_text, stderr_text
    except Exception as error:
        raise RuntimeError(f"SSH 命令执行失败 {label or cmd}: {error}")

# 任务框架
# 基础任务类：所有巡检任务的基类，提供通用的任务执行框架
class BaseTask:
    """基础任务类


    所有巡检任务的基类，提供通用的任务执行框架
    """
    # 初始化基础任务：设置任务名称和结果列表
    def __init__(self, name: str):
        self.NAME = name
        self.RESULTS: list[Result] = []
        # 从配置读取是否压降OK级别日志（默认False，不抑制）
        try:
            self.SUPPRESS_OK_LOGS: bool = bool(
                CONFIG.get("settings", {}).get("suppress_ok_logs", False)
            )
        except Exception:
            self.SUPPRESS_OK_LOGS = False

    # 获取任务项目列表：子类必须实现，返回要处理的项目列表
    def items(self):
        """获取任务项目列表


        子类必须实现，返回要处理的项目列表


        Returns:
            list: 要处理的项目列表
        """
        raise NotImplementedError

    # 处理单个项目：子类必须实现，处理单个任务项目
    def run_single(self, item):
        """处理单个项目


        子类必须实现，处理单个任务项目


        Args:
            item: 要处理的项目
        """
        raise NotImplementedError

    # 添加结果记录：将巡检结果添加到结果列表
    def add_result(self, level: Level, message: str, meta: Optional[dict] = None) -> None:
        """添加结果记录


        Args:
            level: 告警级别
            message: 结果消息
            meta: 附加元数据
        """
        # 可选抑制OK级别的非关键日志
        if level == Level.OK and getattr(self, "SUPPRESS_OK_LOGS", False):
            return
        entry = Result(level=level.value, message=message, meta=meta)
        self.RESULTS.append(entry)

    # 执行任务：遍历所有项目并执行巡检，显示进度条
    def run(self) -> None:
        """执行任务


        遍历所有项目并执行巡检，显示进度条
        """
        task_items = list(self.items())
        progress = tqdm(
            total=len(task_items),
            desc=self.NAME,
            position=0,
            leave=True,
            dynamic_ncols=True,
            bar_format=BAR_FORMAT,
        ) if SHOW_PROGRESS else None

        try:
            for single_item in task_items:
                try:
                    self.run_single(single_item)
                except Exception as error:
                    self.add_result(Level.ERROR, f"{single_item} 运行异常: {error!r}")
                if progress:
                    progress.update(1)
        finally:
            if progress:
                progress.close()

# ============================================================================
# 日期处理工具函数
# ============================================================================

# 获取今天的日期字符串（YYYYMMDD格式）
def get_today_str() -> str:
    """获取今天的日期字符串

    Returns:
        str: 今天的日期字符串（YYYYMMDD格式），例如"20241129"
    """
    return datetime.now().strftime("%Y%m%d")

# 格式化日期时间
def format_datetime(dt: datetime, fmt: str = "%Y%m%d") -> str:
    """格式化日期时间

    Args:
        dt: 日期时间对象
        fmt: 格式字符串，默认为"%Y%m%d"

    Returns:
        str: 格式化后的日期时间字符串
    """
    return dt.strftime(fmt)

# ============================================================================
# 文件路径处理工具函数
# ============================================================================

# 确保输出目录存在
def ensure_output_dir(output_dir: str) -> None:
    """确保输出目录存在

    如果目录不存在则创建，已存在则不报错

    Args:
        output_dir: 输出目录路径
    """
    os.makedirs(output_dir, exist_ok=True)

# 构建LOG目录下的路径
def build_log_path(*parts: str) -> str:
    """构建LOG目录下的路径

    Args:
        *parts: 路径部分

    Returns:
        str: 完整的路径
    """
    return os.path.join("LOG", *parts)

# 构建输出文件路径
def build_output_path(output_dir: str, filename: str) -> str:
    """构建输出文件路径

    自动确保输出目录存在

    Args:
        output_dir: 输出目录
        filename: 文件名

    Returns:
        str: 完整的输出文件路径
    """
    ensure_output_dir(output_dir)
    return os.path.join(output_dir, filename)

# ============================================================================
# Excel操作工具函数
# ============================================================================

# 加载Excel工作簿
def load_excel_workbook(file_path: str):
    """加载Excel工作簿

    Args:
        file_path: Excel文件路径

    Returns:
        Workbook: openpyxl Workbook对象

    Raises:
        FileNotFoundError: 如果文件不存在
        PermissionError: 如果文件被占用
    """
    from openpyxl import load_workbook
    return load_workbook(file_path)

# 创建新的Excel工作簿
def create_excel_workbook():
    """创建新的Excel工作簿

    自动移除默认的"Sheet"工作表

    Returns:
        Workbook: openpyxl Workbook对象
    """
    from openpyxl import Workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    return wb

# 保存Excel工作簿
def save_excel_workbook(workbook, file_path: str) -> None:
    """保存Excel工作簿

    自动确保输出目录存在

    Args:
        workbook: openpyxl Workbook对象
        file_path: 保存路径
    """
    output_dir = os.path.dirname(file_path)
    if output_dir:
        ensure_output_dir(output_dir)
    workbook.save(file_path)
