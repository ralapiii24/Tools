# N9K&LINKAS ACL交叉检查任务
#
# 技术栈:openpyxl、ipaddress、正则表达式、网络地址计算、图论算法、Excel样式处理、CiscoBase（V11新增）
# 目标:实现跨Sheet的ACL规则匹配检查，识别不同Sheet之间的ACL规则匹配关系，使用多颜色标记系统区分不同类型的匹配规则
#
# 处理逻辑:
# - 输入文件:自动使用当天日期的DeviceBackupTask输出文件（从settings.log_dir读取LOG目录）
# - 设备分类:自动识别CS-N9K（N9K核心交换机）、LINK-AS（LINKAS接入交换机）、OOB-DS（OOB-DS交换机）设备
# - 规则解析:使用CiscoBase.parse_acl统一解析（V11优化），支持NX-OS CIDR格式和IOS-XE wildcard/host混合格式，解析源/目的网段、协议、端口等信息
# - ACL定界:使用CiscoBase.find_acl_blocks_in_column和extract_acl_rules_from_column统一处理（V11优化）
# - 跨Sheet匹配:实现跨平台步骤1-10的复杂匹配逻辑，包括完全匹配、覆盖匹配、包含匹配、平台外覆盖、特殊规则检查等
# - 双向检查:跨平台步骤1-7支持双向检查（Sheet A -> Sheet B和Sheet B -> Sheet A方向）
# - 颜色标记:使用多颜色标记系统（绿色、蓝色、橙色、紫色、深黄色）区分不同类型的匹配规则
# - 颜色优先级:绿色和蓝色具有最高优先级，不会被后续步骤覆盖；橙色、紫色、深黄色按优先级顺序标记
# - 自动标记:同一Sheet内相同LINK-AS规则自动标记相同颜色
#
# 输出文件:LOG/ACLCrossCheckTask/{日期}-跨平台N9K&LINKAS&OOB ACL交叉检查.xlsx（从settings.log_dir读取LOG目录）
# 输出内容:
# - 每个Sheet保留原始配置内容
# - 匹配规则使用不同颜色标记（绿色、蓝色、橙色、紫色、深黄色）
# - 颜色标记说明在染色逻辑文档中详细说明
#
# 13个步骤说明:
# - 基础步骤1:识别设备列（分析第一行，识别CS-N9K/LINK-AS/OOB-DS列）- 自动识别N9K核心交换机、LINKAS接入交换机、OOB-DS交换机设备列
# - 基础步骤2:提取ACL配置（按照定界find_acl_blocks_in_column截取ACL配置）- 从源Excel中提取ACL配置块，清除样式准备后续染色
# - 基础步骤3:删除同平台策略（对CS-N9K和OOB-DS设备进行多余规则检测）- 仅对预定义的Sheet进行多余规则检测，删除同平台策略
#   在删除之前，先提取同平台 ACL 到新的 Excel 文件中（从settings.log_dir读取LOG目录，LOG/ACLCrossCheckTask/{日期}-同平台N9K ACL检查.xlsx），每个平台一个 sheet
#   仅收集 CS-N9K（N9K核心交换机）和 OOB-DS（OOB-DS交换机）设备的同平台 ACL，不收集 LINK-AS（LINKAS接入交换机）设备
# - 同平台步骤1:同Sheet同列CS-N9K/OOB-DS反向完全匹配检查（标绿色）- 处理同平台N9K ACL检查.xlsx文件（从settings.log_dir读取LOG目录）
#   - 忽略规则前面的序号和末尾的 log 关键字
#   - 仅在同一个Sheet、同一列内对比，不跨Sheet、不跨列
#   - 对 CS-N9K 和 OOB-DS 设备的 ACL 规则进行不同 ip access-list 块之间的反向完全匹配检查
# - 同平台步骤2:同Sheet同列目的地址为special_network_map的规则检查（标橙色）
#   - 仅在同一个Sheet、同一列内对比，不跨Sheet、不跨列
#   - 对 CS-N9K 和 OOB-DS 设备的 ACL 规则中，目的地址落在本Sheet对应的special_network_map内的规则标记为橙色（不覆盖已有绿色）
# - 同平台步骤3:同Sheet内CS-N9K和OOB-DS反向完全匹配检查（标蓝色）
#   - 仅在同一个Sheet内对比，不跨Sheet
#   - 对 CS-N9K 列与 OOB-DS 列之间，源/目的地址和端口互为反向完全匹配的规则对标记为蓝色（不覆盖已有绿色、橙色）
# - 跨平台步骤1:CS-N9K完全匹配检查（标绿色）
#   - 跨Sheet匹配：CS-N9K vs LINK-AS匹配相同（两个Sheet）
#   - 同Sheet内匹配：CS-N9K vs CS-N9K反向匹配 + LINK-AS vs LINK-AS反向匹配
# - 跨平台步骤2:CS-N9K覆盖匹配检查（标绿色）
#   - 跨Sheet匹配：CS-N9K vs LINK-AS覆盖（两个Sheet）
#   - 同Sheet内匹配：CS-N9K vs CS-N9K反向匹配 + LINK-AS vs LINK-AS反向匹配
# - 跨平台步骤3:LINK-AS覆盖CS-N9K匹配检查（标绿色）
#   - 跨Sheet匹配：LINK-AS vs CS-N9K覆盖（两个Sheet）
#   - 同Sheet内匹配：CS-N9K vs CS-N9K反向匹配 + LINK-AS vs LINK-AS反向匹配
# - 跨平台步骤4:OOB-DS完全匹配检查（标蓝色）
#   - 跨Sheet匹配：OOB-DS vs LINK-AS匹配相同（两个Sheet）
#   - 同Sheet内匹配：OOB-DS vs OOB-DS反向匹配相同 + LINK-AS vs LINK-AS反向匹配相同
# - 跨平台步骤5:OOB-DS覆盖匹配检查（标蓝色）
#   - 跨Sheet匹配：OOB-DS vs LINK-AS覆盖（两个Sheet）
#   - 同Sheet内匹配：OOB-DS vs OOB-DS反向匹配相同 + LINK-AS vs LINK-AS反向匹配相同
# - 跨平台步骤6:LINK-AS覆盖匹配OOB-DS检查（标蓝色）
#   - 跨Sheet匹配：LINK-AS vs OOB-DS覆盖（两个Sheet）
#   - 同Sheet内匹配：OOB-DS vs OOB-DS反向匹配相同 + LINK-AS vs LINK-AS反向匹配相同
# - 跨平台步骤7:OOB-DS-CS-N9K包含匹配（标蓝色）
#   - 匹配条件：A.OOB-DS覆盖A.LINK-AS + A.OOB-DS与B.CS-N9K双向包含/反向匹配 + A.LINK-AS与B.LINK-AS反向匹配 + B.CS-N9K覆盖B.LINK-AS
# - 跨平台步骤8:平台外覆盖检查（标橙色）
#   - 匹配条件：CS-N9K覆盖LINK-AS规则，源为本平台，目的为非platform_network_map地址
# - 跨平台步骤9:特殊规则检查（标紫色）
#   - 匹配条件：CS-N9K源地址为本平台地址，目的地址为特殊规则定义IP段地址
#   - 附加条件：同平台CS-N9K覆盖LINK-AS，跨平台LINK-AS反向检查通过
# - 跨平台步骤10:特殊地址段检查（标深黄色）
#   - 匹配条件：LINK-AS规则源地址和目的地址都在任何Sheet的special_network_map中
#
# 性能优化:
# - 使用预构建索引避免重复计算（LINK-AS反向匹配索引、OOB-DS与CS-N9K包含关系索引、CS-N9K覆盖LINK-AS索引）
# - 优化查找效率（将列表转换为set使用O(1)查找替代O(n)的any()函数）
# - 限制检查范围（组合数超过50000时限制检查前50000个）
# - 文件保存优化（默认只在最后保存一次）
#
# 配置说明:
# - 输入文件:自动使用当天日期的DeviceBackupTask输出文件（从settings.log_dir读取LOG目录）
# - 输出目录:LOG/ACLCrossCheckTask/（从settings.log_dir读取LOG目录）
# - 配置依赖:platform_network_map（平台网段映射，使用公共配置settings.platform_network_map，兼容ACLCrossCheckTask.platform_network_map）、special_network_map（特殊网段映射，从ACLCrossCheckTask.special_network_map读取）
# - 详细染色逻辑说明:参考TASK/ACLCrossCheckTask_染色逻辑说明.md文档

# 导入标准库
import os
import re
import socket
from dataclasses import dataclass
from datetime import datetime
from ipaddress import IPv4Address, IPv4Network
from typing import Dict, List, Optional, Tuple

# 导入第三方库
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

# 导入本地应用
from .TaskBase import (
    BaseTask, Level, CONFIG, get_today_str, format_datetime,
    ensure_output_dir, build_log_path, build_output_path,
    load_excel_workbook, create_excel_workbook, save_excel_workbook
)
from .CiscoBase import (
    ACLRule, parse_acl, service_to_port,
    ip_and_wildcard_to_network, host_to_network, cidr_to_network,
    find_acl_blocks_in_column, extract_acl_rules_from_column, is_acl_rule,
    get_device_classification_rules, is_cat1_device, is_cat2_device, is_cat6_device,
    extract_device_number, analyze_first_row_for_cat1_cat2
)

# ACL规则判断功能已迁移到CiscoBase
# 使用: from .CiscoBase import is_acl_rule

# 设备分类相关函数已迁移到CiscoBase
# 使用: from .CiscoBase import get_device_classification_rules, is_cat1_device, is_cat2_device, is_cat6_device, extract_device_number, analyze_first_row_for_cat1_cat2

# 以下函数定义已删除，改用CiscoBase中的版本
# 所有设备分类相关函数已迁移到CiscoBase，直接使用导入的函数即可

# ACL定界功能已迁移到CiscoBase
# 使用: from .CiscoBase import find_acl_blocks_in_column, extract_acl_rules_from_column

# 规则覆盖判断函数（从ACLDupCheckTask复制）
# 检查协议A是否覆盖协议B：支持三种情况：1.TCP覆盖TCP 2.IP覆盖IP 3.IP覆盖TCP/UDP
def proto_covers(PROTO_A: str, PROTO_B: str) -> bool:
    """检查协议A是否覆盖协议B

    支持三种情况：1.TCP覆盖TCP 2.IP覆盖IP 3.IP覆盖TCP/UDP
    如果A是IP协议，可以覆盖任何协议（IP、TCP、UDP等）
    如果A和B是相同协议，可以覆盖

    Args:
        PROTO_A: 协议A
        PROTO_B: 协议B

    Returns:
        如果A是IP协议或A与B相同则返回True，否则返回False
    """
    # 如果A是IP协议，可以覆盖任何协议（IP、TCP、UDP等）- 对应情况2和情况3
    # 如果A和B是相同协议，可以覆盖 - 对应情况1和情况2
    return PROTO_A.lower() == "ip" or PROTO_A.lower() == PROTO_B.lower()

# 检查端口A是否覆盖端口B：支持多个端口的情况，任意端口可以覆盖特定端口，
# 多个端口可以覆盖单个端口
def port_covers(
    PORT_A: Optional[int], PORT_B: Optional[int],
    RULE_A: Optional[ACLRule] = None, RULE_B: Optional[ACLRule] = None,
    port_type: str = "dst"
) -> bool:
    """检查端口A是否覆盖端口B

    支持多个端口的情况，任意端口可以覆盖特定端口，多个端口可以覆盖单个端口

    Args:
        PORT_A: 端口A（None表示任意端口）
        PORT_B: 端口B（None表示任意端口）
        RULE_A: ACL规则A（可选，用于检查多个端口）
        RULE_B: ACL规则B（可选，用于检查多个端口）
        port_type: 端口类型（"dst"或"src"）

    Returns:
        如果A覆盖B则返回True，否则返回False
    """
    # A None => 任意端口；B None 且 A 有端口 => 不覆盖；都有端口需相等
    if PORT_A is None:
        return True
    if PORT_B is None:
        # 检查RULE_B是否有多个端口（使用ports字段）
        if RULE_B and RULE_B.ports and len(RULE_B.ports) > 0:
            # RULE_B有多个端口，PORT_A必须也是None（任意端口）或者RULE_A有相同的多个端口才能覆盖
            if PORT_A is None:
                return True  # 任意端口可以覆盖多个端口
            # 如果RULE_A也有多个端口，需要检查端口集合是否完全相等
            if RULE_A and RULE_A.ports and len(RULE_A.ports) > 1:
                return RULE_A.ports == RULE_B.ports
            # PORT_A是单个端口，不能覆盖多个端口
            return False
        return False


    # 如果RULE_A有多个端口，可以覆盖单个端口的规则（如果单端口在多端口集合中）
    if RULE_A and RULE_A.ports and len(RULE_A.ports) > 1:
        # RULE_A有多个端口，需要检查RULE_B的端口集合
        if not RULE_B or not RULE_B.ports or len(RULE_B.ports) == 0:
            # RULE_B没有ports字段或为空，使用port
            # 如果RULE_B的port在RULE_A的ports集合中，可以覆盖
            if PORT_B is not None and PORT_B in RULE_A.ports:
                return True
            return False
        elif len(RULE_B.ports) == 1:
            # RULE_B只有一个端口，如果这个端口在RULE_A的ports集合中，可以覆盖
            single_port_b = next(iter(RULE_B.ports))
            if single_port_b in RULE_A.ports:
                return True
            return False
        elif RULE_A.ports != RULE_B.ports:
            # RULE_A和RULE_B的端口集合不完全相等
            # 如果RULE_B的所有端口都在RULE_A的ports集合中，可以覆盖
            if RULE_B.ports.issubset(RULE_A.ports):
                return True
            return False
        else:
            # 端口集合完全相等，可以覆盖
            return True


    # 如果RULE_B有多个端口，RULE_A必须也有相同的多个端口才能覆盖（或者RULE_A是None表示任意端口）
    if RULE_B and RULE_B.ports and len(RULE_B.ports) > 1:
        # RULE_B有多个端口，单个端口不能覆盖多个端口
        # 只有当RULE_A也是None（任意端口）或者RULE_A有相同的多个端口时才能覆盖
        if PORT_A is None:
            return True  # 任意端口可以覆盖多个端口
        # 如果RULE_A也有多个端口，需要检查端口集合是否完全相等
        if RULE_A and RULE_A.ports and len(RULE_A.ports) > 1:
            return RULE_A.ports == RULE_B.ports
        # RULE_A是单个端口，不能覆盖多个端口
        return False


    return PORT_A == PORT_B

# 检查源端口A是否覆盖源端口B：支持多个端口的情况，任意端口可以覆盖特定端口
def source_port_covers(
    SRC_PORT_A: Optional[int], SRC_PORT_B: Optional[int],
    RULE_A: Optional[ACLRule] = None, RULE_B: Optional[ACLRule] = None
) -> bool:
    """检查源端口A是否覆盖源端口B

    支持多个端口的情况

    Args:
        SRC_PORT_A: 源端口A（None表示任意端口）
        SRC_PORT_B: 源端口B（None表示任意端口）
        RULE_A: ACL规则A（可选，用于检查多个端口）
        RULE_B: ACL规则B（可选，用于检查多个端口）

    Returns:
        如果A覆盖B则返回True，否则返回False
    """
    if SRC_PORT_A is None:
        return True
    if SRC_PORT_B is None:
        # 检查RULE_B是否有多个源端口（使用ports字段）
        if RULE_B and RULE_B.ports and len(RULE_B.ports) > 0:
            if len(RULE_B.ports) == 1:
                # RULE_B只有一个端口，允许SRC_PORT_A等于这个端口来覆盖
                single_port = next(iter(RULE_B.ports))
                return SRC_PORT_A == single_port
            # RULE_B有多个端口，SRC_PORT_A必须也是None（任意端口）或者RULE_A有相同的多个端口才能覆盖
            if SRC_PORT_A is None:
                return True  # 任意端口可以覆盖多个端口
            # 如果RULE_A也有多个端口，需要检查端口集合是否完全相等
            if RULE_A and RULE_A.ports and len(RULE_A.ports) > 1:
                return RULE_A.ports == RULE_B.ports
            # SRC_PORT_A是单个端口，不能覆盖多个端口
            return False
        return False


    # 如果RULE_A有多个源端口，可以覆盖单个端口的规则（如果单端口在多端口集合中）
    if RULE_A and RULE_A.ports and len(RULE_A.ports) > 1:
        # RULE_A有多个端口，需要检查RULE_B的端口集合
        if not RULE_B or not RULE_B.ports or len(RULE_B.ports) == 0:
            # RULE_B没有ports字段或为空，使用src_port
            # 如果RULE_B的src_port在RULE_A的ports集合中，可以覆盖
            if SRC_PORT_B is not None and SRC_PORT_B in RULE_A.ports:
                return True
            return False
        elif len(RULE_B.ports) == 1:
            # RULE_B只有一个端口，如果这个端口在RULE_A的ports集合中，可以覆盖
            single_port_b = next(iter(RULE_B.ports))
            if single_port_b in RULE_A.ports:
                return True
            return False
        elif RULE_A.ports != RULE_B.ports:
            # RULE_A和RULE_B的端口集合不完全相等
            # 如果RULE_B的所有端口都在RULE_A的ports集合中，可以覆盖
            if RULE_B.ports.issubset(RULE_A.ports):
                return True
            return False
        else:
            # 端口集合完全相等，可以覆盖
            return True


    # 如果RULE_B有多个源端口，RULE_A必须也有相同的多个源端口才能覆盖（或者RULE_A是None表示任意端口）
    if RULE_B and RULE_B.ports and len(RULE_B.ports) > 1:
        # RULE_B有多个端口，单个端口不能覆盖多个端口
        # 只有当RULE_A也是None（任意端口）或者RULE_A有相同的多个端口时才能覆盖
        if SRC_PORT_A is None:
            return True  # 任意端口可以覆盖多个端口
        # 如果RULE_A也有多个端口，需要检查端口集合是否完全相等
        if RULE_A and RULE_A.ports and len(RULE_A.ports) > 1:
            return RULE_A.ports == RULE_B.ports
        # RULE_A是单个端口，不能覆盖多个端口
        return False


    return SRC_PORT_A == SRC_PORT_B

# 检查目标端口A是否覆盖目标端口B：支持多个端口的情况，任意端口可以覆盖特定端口
def destination_port_covers(
    DST_PORT_A: Optional[int], DST_PORT_B: Optional[int],
    RULE_A: Optional[ACLRule] = None, RULE_B: Optional[ACLRule] = None
) -> bool:
    """检查目标端口A是否覆盖目标端口B

    支持多个端口的情况

    Args:
        DST_PORT_A: 目标端口A（None表示任意端口）
        DST_PORT_B: 目标端口B（None表示任意端口）
        RULE_A: ACL规则A（可选，用于检查多个端口）
        RULE_B: ACL规则B（可选，用于检查多个端口）

    Returns:
        如果A覆盖B则返回True，否则返回False
    """
    if DST_PORT_A is None:
        return True  # 任意端口覆盖特定端口
    if DST_PORT_B is None:
        # 检查RULE_B是否有多个目标端口（使用ports字段）
        if RULE_B and RULE_B.ports and len(RULE_B.ports) > 0:
            # RULE_B有多个端口，DST_PORT_A必须也是None（任意端口）或者RULE_A有相同的多个端口才能覆盖
            if DST_PORT_A is None:
                return True  # 任意端口可以覆盖多个端口
            # 如果RULE_A也有多个端口，需要检查端口集合是否完全相等
            if RULE_A and RULE_A.ports and len(RULE_A.ports) > 1:
                return RULE_A.ports == RULE_B.ports
            # DST_PORT_A是单个端口，不能覆盖多个端口
            return False
        return False  # 特定端口不覆盖任意端口


    # 如果RULE_A有多个目标端口，可以覆盖单个端口的规则（如果单端口在多端口集合中）
    if RULE_A and RULE_A.ports and len(RULE_A.ports) > 1:
        # RULE_A有多个端口，需要检查RULE_B的端口集合
        if not RULE_B or not RULE_B.ports or len(RULE_B.ports) == 0:
            # RULE_B没有ports字段或为空，使用dst_port
            # 如果RULE_B的dst_port在RULE_A的ports集合中，可以覆盖
            if DST_PORT_B is not None and DST_PORT_B in RULE_A.ports:
                return True
            return False
        elif len(RULE_B.ports) == 1:
            # RULE_B只有一个端口，如果这个端口在RULE_A的ports集合中，可以覆盖
            single_port_b = next(iter(RULE_B.ports))
            if single_port_b in RULE_A.ports:
                return True
            return False
        elif RULE_A.ports != RULE_B.ports:
            # RULE_A和RULE_B的端口集合不完全相等
            # 如果RULE_B的所有端口都在RULE_A的ports集合中，可以覆盖
            if RULE_B.ports.issubset(RULE_A.ports):
                return True
            return False
        else:
            # 端口集合完全相等，可以覆盖
            return True


    # 如果RULE_B有多个目标端口，RULE_A必须也有相同的多个目标端口才能覆盖（或者RULE_A是None表示任意端口）
    if RULE_B and RULE_B.ports and len(RULE_B.ports) > 1:
        # RULE_B有多个端口，单个端口不能覆盖多个端口
        # 只有当RULE_A也是None（任意端口）或者RULE_A有相同的多个端口时才能覆盖
        if DST_PORT_A is None:
            return True  # 任意端口可以覆盖多个端口
        # 如果RULE_A也有多个端口，需要检查端口集合是否完全相等
        if RULE_A and RULE_A.ports and len(RULE_A.ports) > 1:
            return RULE_A.ports == RULE_B.ports
        # RULE_A是单个端口，不能覆盖多个端口
        return False


    return DST_PORT_A == DST_PORT_B

# 检查端口是否反向匹配：用于rule_reverse_matches函数，检查端口A和端口B是否相等
# （任意端口可匹配任何端口，特定端口必须相等）
def _port_reverse_matches(
    PORT_A: Optional[int], PORT_B: Optional[int],
    RULE_A: Optional[ACLRule] = None, RULE_B: Optional[ACLRule] = None
) -> bool:
    # 如果PORT_A是None（任意端口），可以匹配任何PORT_B
    if PORT_A is None:
        return True
    # 如果PORT_B是None（任意端口），可以匹配任何PORT_A
    if PORT_B is None:
        return True


    # 两者都是特定端口，需要检查是否相等
    # 如果RULE_A有多个端口（ports集合），检查PORT_B是否在集合中
    if RULE_A and RULE_A.ports and len(RULE_A.ports) > 0:
        if PORT_B in RULE_A.ports:
            # 如果RULE_B也有多个端口，需要检查端口集合是否完全相等
            if RULE_B and RULE_B.ports and len(RULE_B.ports) > 0:
                return RULE_A.ports == RULE_B.ports
            return True
        return False


    # 如果RULE_B有多个端口（ports集合），检查PORT_A是否在集合中
    if RULE_B and RULE_B.ports and len(RULE_B.ports) > 0:
        if PORT_A in RULE_B.ports:
            return True
        return False


    # 都没有ports集合，直接比较端口值
    return PORT_A == PORT_B

# 检查规则A是否覆盖规则B：支持三种覆盖情况：
# 1.TCP端口覆盖TCP端口（相同协议，需端口匹配）
# 2.IP覆盖范围大的IP覆盖范围小的IP（相同协议，需网络范围匹配）
# 3.IP覆盖TCP端口（IP协议可覆盖TCP/UDP，跳过端口检查，需网络范围匹配）
def rule_covers(RULE_A: ACLRule, RULE_B: ACLRule) -> bool:
    """检查规则A是否覆盖规则B

    支持三种情况：
    1. TCP覆盖TCP（范围大的TCP覆盖范围小的TCP）
    2. IP覆盖IP（范围大的IP覆盖范围小的IP）
    3. IP覆盖TCP（IP协议覆盖所有协议和所有端口）

    Args:
        RULE_A: ACL规则A
        RULE_B: ACL规则B

    Returns:
        如果规则A覆盖规则B则返回True，否则返回False
    """
    if RULE_A.action != RULE_B.action:
        return False
    if not proto_covers(RULE_A.proto, RULE_B.proto):
        return False


    # 情况2和情况3：如果RULE_A是IP协议，可以覆盖RULE_B的任何协议和端口（只要网络匹配）
    # 情况2：IP覆盖IP（范围大的IP覆盖范围小的IP）
    # 情况3：IP覆盖TCP（IP协议覆盖所有协议（IP、TCP、UDP等）和所有端口，跳过端口检查）
    if RULE_A.proto.lower() == "ip":
        # IP协议覆盖所有端口，跳过端口检查
        pass
    else:
        # 情况1：RULE_A是TCP/UDP协议，需要检查端口匹配
        # 情况1：TCP端口覆盖TCP端口（需端口匹配）、UDP端口覆盖UDP端口（需端口匹配）
        # cat1是NXOS，cat2是IOS-XE
        # NXOS格式：使用src_port和dst_port字段
        # IOS-XE格式：可能使用port字段或src_port/dst_port字段
        if RULE_A.style == "NXOS" and RULE_B.style == "IOS-XE":
            # 对于IOS-XE，如果src_port和dst_port都为None，则使用port字段
            if RULE_B.src_port is None and RULE_B.dst_port is None:
                # IOS-XE使用port字段
                if not port_covers(RULE_A.port, RULE_B.port, RULE_A, RULE_B):
                    return False
            else:
                # IOS-XE使用src_port和dst_port字段
                if not source_port_covers(RULE_A.src_port, RULE_B.src_port, RULE_A, RULE_B):
                    return False
                if not destination_port_covers(RULE_A.dst_port, RULE_B.dst_port, RULE_A, RULE_B):
                    return False
        else:
            # 其他情况：使用src_port和dst_port字段，或者两个规则都是IOS-XE且都使用port字段
            # 如果两个规则都是IOS-XE且都使用port字段（src_port和dst_port都为None），需要特殊处理
            if (RULE_A.style == "IOS-XE" and RULE_B.style == "IOS-XE" and
                    RULE_A.src_port is None and RULE_A.dst_port is None and
                    RULE_B.src_port is None and RULE_B.dst_port is None):
                # 两个规则都使用port字段
                if not port_covers(RULE_A.port, RULE_B.port, RULE_A, RULE_B):
                    return False
            else:
                # 使用src_port和dst_port字段
                if not source_port_covers(RULE_A.src_port, RULE_B.src_port, RULE_A, RULE_B):
                    return False
                if not destination_port_covers(RULE_A.dst_port, RULE_B.dst_port, RULE_A, RULE_B):
                    return False


            # 特殊检查1：如果RULE_B包含端口范围或多端口，且RULE_A是单个端口，则不应该覆盖
            # 例如：RULE_A的dst_port=22，RULE_B包含多个端口（22和22222）
            # 单个端口不能覆盖端口范围或多端口
            if RULE_A.dst_port is not None and RULE_B.dst_port is None:
                # RULE_A有特定目的端口，RULE_B的目的端口是None（可能是端口范围或多端口）
                # 优先使用ports字段检查
                if RULE_B.ports and len(RULE_B.ports) > 1:
                    # RULE_B包含多个端口，单个端口不能覆盖多个端口
                    return False
                # 如果没有ports字段，检查原始文本是否包含"range"关键字（向后兼容）
                if "range" in RULE_B.raw.lower():
                    return False


            # 同样检查源端口的情况
            if RULE_A.src_port is not None and RULE_B.src_port is None:
                # RULE_A有特定源端口，RULE_B的源端口是None（可能是端口范围或多端口）
                # 优先使用ports字段检查
                if RULE_B.ports and len(RULE_B.ports) > 1:
                    # RULE_B包含多个端口，单个端口不能覆盖多个端口
                    return False
                # 如果没有ports字段，检查原始文本是否包含"range"关键字（向后兼容）
                if "range" in RULE_B.raw.lower():
                    return False


    # 检查网络覆盖：RULE_B的源和目的网络必须是RULE_A的子网（范围大的覆盖范围小的）
    # 这确保了：RULE_A的网络范围 >= RULE_B的网络范围
    if not RULE_B.src.subnet_of(RULE_A.src):
        return False
    if not RULE_B.dst.subnet_of(RULE_A.dst):
        return False
    return True

# 检查CS-N9K规则的端口是否是LINK-AS规则端口集合的一部分：用于CS-N9K/LINK-AS覆盖匹配相关步骤，识别多个CS-N9K规则一起覆盖LINK-AS规则的所有端口
def rule_port_in_cat2_ports(RULE_A: ACLRule, RULE_B: ACLRule) -> bool:
    """检查CS-N9K规则的端口是否在LINK-AS规则的端口集合中

    Args:
        RULE_A: CS-N9K规则（ACLRule对象）
        RULE_B: LINK-AS规则（ACLRule对象）

    Returns:
        如果CS-N9K规则的端口在LINK-AS规则的端口集合中则返回True，否则返回False
    """
    if RULE_A.action != RULE_B.action:
        return False
    if not proto_covers(RULE_A.proto, RULE_B.proto):
        return False


    # 如果CS-N9K是IP协议，端口是任意端口，可以覆盖任何端口
    if RULE_A.proto.lower() == "ip":
        return True


    # 检查网络是否匹配
    if not RULE_B.src.subnet_of(RULE_A.src):
        return False
    if not RULE_B.dst.subnet_of(RULE_A.dst):
        return False


    # 检查端口：CS-N9K规则的端口是否是LINK-AS规则端口集合的一部分
    # 对于NXOS格式的CS-N9K规则，使用dst_port字段
    # 对于IOS-XE格式的LINK-AS规则，使用ports字段
    if RULE_A.style == "NXOS" and RULE_B.style == "IOS-XE":
        # 获取CS-N9K规则的目的端口
        category1_destination_port = (
            RULE_A.dst_port if RULE_A.dst_port is not None else RULE_A.port
        )


        if category1_destination_port is None:
            # CS-N9K规则是任意端口，可以覆盖任何端口
            return True


        # 使用ports字段
        if not RULE_B.ports or len(RULE_B.ports) == 0:
            # LINK-AS规则没有端口限制，CS-N9K的特定端口不能覆盖
            return False


        # 检查CS-N9K的端口是否在LINK-AS的端口集合中
        return category1_destination_port in RULE_B.ports


    return False

# 从IOS-XE规则的原始文本中提取所有端口：包括服务名转换为端口号和range端口
def _extract_ports_from_iosxe_rule(rule_raw: str) -> set:
    PORTS = set()
    RULE_LOWER = rule_raw.lower()


    # 首先检查是否有range端口
    # 匹配 "range 起始端口 结束端口"
    RANGE_PATTERN = r'\brange\s+(\d+)\s+(\d+)'
    RANGE_MATCH = re.search(RANGE_PATTERN, RULE_LOWER)


    if RANGE_MATCH:
        START_PORT_NUMBER = int(RANGE_MATCH.group(1))
        END_PORT_NUMBER = int(RANGE_MATCH.group(2))
        # 添加范围内的所有端口
        for PORT in range(START_PORT_NUMBER, END_PORT_NUMBER + 1):
            PORTS.add(PORT)
        return PORTS


    # 匹配 "eq" 后面的端口或服务名
    # 例如：eq www 443, eq 22 22222, eq domain ntp
    EQUAL_PATTERN = r'\beq\s+([\w\s]+)'
    EQUAL_MATCH = re.search(EQUAL_PATTERN, RULE_LOWER)


    if EQUAL_MATCH:
        PORTS_STRING = EQUAL_MATCH.group(1).strip()
        # 分割多个端口（用空格分隔）
        PORT_ITEMS = PORTS_STRING.split()


        for PORT_ITEM in PORT_ITEMS:
            PORT_ITEM = PORT_ITEM.strip()
            if not PORT_ITEM:
                continue


            # 尝试转换为端口号
            if PORT_ITEM.isdigit():
                PORTS.add(int(PORT_ITEM))
            else:
                # 尝试将服务名转换为端口号
                try:
                    PORT = socket.getservbyname(PORT_ITEM)
                    PORTS.add(PORT)
                except (OSError, socket.gaierror):
                    # 如果无法转换，跳过
                    pass


    return PORTS

# 检查规则A和规则B是否匹配相同：不是覆盖关系，而是相等或重叠关系（网络相等或重叠，端口相等）
def rule_matches(RULE_A: ACLRule, RULE_B: ACLRule) -> bool:
    """检查规则A是否完全匹配规则B

    要求协议兼容、端口相等、网络匹配

    Args:
        RULE_A: ACL规则A
        RULE_B: ACL规则B

    Returns:
        如果规则A完全匹配规则B则返回True，否则返回False
    """
    if RULE_A.action != RULE_B.action:
        return False


    # 检查协议是否兼容：ip协议可以匹配任何协议，相同协议可以匹配
    if (RULE_A.proto.lower() != "ip" and RULE_B.proto.lower() != "ip" and
            RULE_A.proto.lower() != RULE_B.proto.lower()):
        return False


    # 检查端口：必须相等（不是覆盖关系）
    # 首先检查ports集合：如果其中一个规则有多个端口，另一个规则也必须有多端口且集合相等
    if RULE_A.ports and len(RULE_A.ports) > 1:
        # RULE_A有多个端口，RULE_B也必须有多端口且集合相等
        if (not RULE_B.ports or len(RULE_B.ports) != len(RULE_A.ports) or
                RULE_A.ports != RULE_B.ports):
            return False
    elif RULE_B.ports and len(RULE_B.ports) > 1:
        # RULE_B有多个端口，RULE_A也必须有多端口且集合相等
        if (not RULE_A.ports or len(RULE_A.ports) != len(RULE_B.ports) or
                RULE_A.ports != RULE_B.ports):
            return False
    elif RULE_A.ports and RULE_B.ports:
        # 两个规则都有ports集合（但都是单个端口），检查是否相等
        if RULE_A.ports != RULE_B.ports:
            return False
    elif RULE_A.ports or RULE_B.ports:
        # 只有一个规则有ports集合，另一个没有，不匹配
        return False


    # 如果都没有ports集合或ports集合都是单个端口，检查port字段
    if RULE_A.style == "NXOS" and RULE_B.style == "IOS-XE":
        # NXOS vs IOS-XE
        if RULE_B.src_port is None and RULE_B.dst_port is None:
            # IOS-XE使用port字段
            if RULE_A.port != RULE_B.port:
                return False
        else:
            # IOS-XE使用src_port和dst_port字段
            if RULE_A.src_port != RULE_B.src_port or RULE_A.dst_port != RULE_B.dst_port:
                return False
    elif RULE_A.style == "IOS-XE" and RULE_B.style == "NXOS":
        # IOS-XE vs NXOS（反向情况）
        if RULE_A.src_port is None and RULE_A.dst_port is None:
            # IOS-XE使用port字段
            if RULE_A.port != RULE_B.port:
                return False
        else:
            # IOS-XE使用src_port和dst_port字段
            if RULE_A.src_port != RULE_B.src_port or RULE_A.dst_port != RULE_B.dst_port:
                return False
    elif RULE_A.style == "NXOS" and RULE_B.style == "NXOS":
        # NXOS vs NXOS：使用port字段
        if RULE_A.port != RULE_B.port:
            return False
    else:
        # 相同格式或都是IOS-XE
        if (RULE_A.style == "IOS-XE" and RULE_B.style == "IOS-XE" and
                RULE_A.src_port is None and RULE_A.dst_port is None and
                RULE_B.src_port is None and RULE_B.dst_port is None):
            # 两个规则都使用port字段
            if RULE_A.port != RULE_B.port:
                return False
        else:
            # 使用src_port和dst_port字段
            if RULE_A.src_port != RULE_B.src_port or RULE_A.dst_port != RULE_B.dst_port:
                return False


    # 检查网络匹配：网络必须相等或重叠（不是子网关系）
    source_match = (RULE_A.src == RULE_B.src) or RULE_A.src.overlaps(RULE_B.src)
    destination_match = (RULE_A.dst == RULE_B.dst) or RULE_A.dst.overlaps(RULE_B.dst)


    if not source_match or not destination_match:
        return False


    return True

# 检查规则A和规则B是否反向匹配：源目地址互换，其他条件相同（规则A的源地址=规则B的目的地址，规则A的目的地址=规则B的源地址）
def rule_reverse_matches(RULE_A: ACLRule, RULE_B: ACLRule) -> bool:
    """检查规则A是否反向匹配规则B

    规则A的目的端口应该等于规则B的源端口，规则A的源端口应该等于规则B的目的端口

    Args:
        RULE_A: ACL规则A
        RULE_B: ACL规则B

    Returns:
        如果规则A反向匹配规则B则返回True，否则返回False
    """
    if RULE_A.action != RULE_B.action:
        return False
    # 检查协议是否兼容：ip协议可以匹配任何协议，相同协议可以匹配
    if (RULE_A.proto.lower() != "ip" and RULE_B.proto.lower() != "ip" and
            RULE_A.proto.lower() != RULE_B.proto.lower()):
        return False


    # 检查端口：规则A的目的端口应该等于规则B的源端口，规则A的源端口应该等于规则B的目的端口
    # 对于NXOS格式，使用port字段；对于IOS-XE格式，使用src_port和dst_port字段，也可能使用port字段
    if RULE_A.style == "NXOS" and RULE_B.style == "IOS-XE":
        # cat1是NXOS，cat2是IOS-XE
        # 规则A的目的端口应该等于规则B的源端口（如果都有）
        # 规则A的源端口应该等于规则B的目的端口（如果都有）
        # 如果规则B使用port字段（src_port和dst_port都为None）
        if RULE_B.src_port is None and RULE_B.dst_port is None:
            # 规则B使用port字段，规则A的目的端口应该等于规则B的port
            # 注意：对于反向匹配，规则A的目的端口对应规则B的源端口，但规则B只有port字段
            # 这里我们检查规则A的目的端口是否匹配规则B的port（考虑ports集合）
            if RULE_A.dst_port is not None and RULE_B.port is not None:
                # 如果规则B有ports集合，检查规则A的目的端口是否在集合中
                if RULE_B.ports and len(RULE_B.ports) > 0:
                    if RULE_A.dst_port not in RULE_B.ports:
                        return False
                else:
                    # 如果没有ports集合，直接比较
                    if RULE_A.dst_port != RULE_B.port:
                        return False
            elif RULE_A.dst_port is None and RULE_B.port is not None:
                # 规则A的目的端口是任意，规则B的port是特定，不匹配
                return False
            elif RULE_A.dst_port is not None and RULE_B.port is None:
                # 规则A的目的端口是特定，规则B的port是任意，匹配
                pass
        else:
            # 规则B使用src_port和dst_port字段
            # 规则A的目的端口应该等于规则B的源端口（相等关系，不是覆盖关系）
            if not _port_reverse_matches(RULE_A.dst_port, RULE_B.src_port, RULE_A, RULE_B):
                return False
            # 规则A的源端口应该等于规则B的目的端口（相等关系，不是覆盖关系）
            if not _port_reverse_matches(RULE_A.src_port, RULE_B.dst_port, RULE_A, RULE_B):
                return False
    elif RULE_A.style == "IOS-XE" and RULE_B.style == "NXOS":
        # LINK-AS是IOS-XE，CS-N9K是NXOS（反向情况）
        # 规则A的目的端口应该等于规则B的源端口（如果都有）
        # 规则A的源端口应该等于规则B的目的端口（如果都有）
        # 如果规则A使用port字段（src_port和dst_port都为None）
        if RULE_A.src_port is None and RULE_A.dst_port is None:
            # 规则A使用port字段（表示目的端口），规则B使用port字段（表示源端口和目的端口）
            # 对于反向匹配：规则A的目的端口（RULE_A.port）应该等于规则B的源端口（RULE_B.port）
            # 如果规则B的port是None（任意端口），则可以匹配规则A的特定端口
            if RULE_A.port is not None and RULE_B.port is not None:
                # 两者都有特定端口，必须相等（考虑ports集合）
                # 如果规则A有ports集合，检查规则B的port是否在集合中
                if RULE_A.ports and len(RULE_A.ports) > 0:
                    if RULE_B.port not in RULE_A.ports:
                        return False
                # 如果规则B有ports集合，检查规则A的port是否在集合中
                elif RULE_B.ports and len(RULE_B.ports) > 0:
                    if RULE_A.port not in RULE_B.ports:
                        return False
                else:
                    # 都没有ports集合，直接比较
                    if RULE_A.port != RULE_B.port:
                        return False
            elif RULE_A.port is not None and RULE_B.port is None:
                # 规则A有特定端口，规则B是任意端口，匹配（规则B可以匹配规则A的端口）
                pass
            elif RULE_A.port is None and RULE_B.port is not None:
                # 规则A是任意端口，规则B是特定端口，不匹配（规则A的任意端口不能匹配规则B的特定端口）
                return False
        else:
            # 规则A使用src_port和dst_port字段
            # 规则A的目的端口应该等于规则B的源端口（相等关系，不是覆盖关系）
            if not _port_reverse_matches(RULE_A.dst_port, RULE_B.src_port, RULE_A, RULE_B):
                return False
            # 规则A的源端口应该等于规则B的目的端口（相等关系，不是覆盖关系）
            if not _port_reverse_matches(RULE_A.src_port, RULE_B.dst_port, RULE_A, RULE_B):
                return False
    else:
        # 其他情况：两个规则都是IOS-XE格式
        # 如果两个规则都使用port字段（src_port和dst_port都为None）
        if (RULE_A.src_port is None and RULE_A.dst_port is None and
                RULE_B.src_port is None and RULE_B.dst_port is None):
            # 两个规则都使用port字段
            # 对于反向匹配：规则A的目的端口（RULE_A.port）应该等于规则B的源端口（RULE_B.port）
            # 规则A的源端口（RULE_A.port）应该等于规则B的目的端口（RULE_B.port）
            # 由于IOS-XE的port字段表示目的端口，对于反向匹配，我们需要检查端口是否匹配
            if RULE_A.port is not None and RULE_B.port is not None:
                # 两者都有特定端口，必须相等（考虑ports集合）
                # 如果规则A有ports集合，检查规则B的port是否在集合中
                if RULE_A.ports and len(RULE_A.ports) > 0:
                    if RULE_B.port not in RULE_A.ports:
                        return False
                # 如果规则B有ports集合，检查规则A的port是否在集合中
                elif RULE_B.ports and len(RULE_B.ports) > 0:
                    if RULE_A.port not in RULE_B.ports:
                        return False
                else:
                    # 都没有ports集合，直接比较
                    if RULE_A.port != RULE_B.port:
                        return False
                # 特殊检查：如果规则A有多个端口，规则B也必须有多端口且集合相等
                if RULE_A.ports and len(RULE_A.ports) > 1:
                    if (not RULE_B.ports or
                            len(RULE_B.ports) != len(RULE_A.ports) or
                            RULE_A.ports != RULE_B.ports):
                        return False
                # 如果规则B有多个端口，规则A也必须有多端口且集合相等
                elif RULE_B.ports and len(RULE_B.ports) > 1:
                    if (not RULE_A.ports or
                            len(RULE_A.ports) != len(RULE_B.ports) or
                            RULE_A.ports != RULE_B.ports):
                        return False
            elif RULE_A.port is not None and RULE_B.port is None:
                # 规则A有特定端口，规则B是任意端口，匹配（规则B可以匹配规则A的端口）
                pass
            elif RULE_A.port is None and RULE_B.port is not None:
                # 规则A是任意端口，规则B是特定端口，不匹配（规则A的任意端口不能匹配规则B的特定端口）
                return False
        else:
            # 使用src_port和dst_port字段
            # 对于反向匹配，规则A的目的端口应该等于规则B的源端口（相等关系，不是覆盖关系）
            # 规则A的源端口应该等于规则B的目的端口（相等关系，不是覆盖关系）
            # 如果端口是None（任意端口），则任何特定端口都可以匹配它
            # 如果端口是特定值，则必须相等（考虑ports集合）


            # 检查规则A的目的端口是否匹配规则B的源端口
            if not _port_reverse_matches(RULE_A.dst_port, RULE_B.src_port, RULE_A, RULE_B):
                return False
            # 检查规则A的源端口是否匹配规则B的目的端口
            if not _port_reverse_matches(RULE_A.src_port, RULE_B.dst_port, RULE_A, RULE_B):
                return False


    # 检查网络反向匹配：规则A的源地址应该等于规则B的目的地址，规则A的目的地址应该等于规则B的源地址
    # 使用网络相等或重叠关系（允许网络重叠，不要求完全相等）
    # 规则A: 源地址A，目的地址B
    # 规则B: 源地址B，目的地址A
    # 检查：A.src 与 B.dst 是否匹配，A.dst 与 B.src 是否匹配
    source_match = (
        (RULE_A.src == RULE_B.dst) or RULE_A.src.overlaps(RULE_B.dst) or
        RULE_A.src.subnet_of(RULE_B.dst) or RULE_B.dst.subnet_of(RULE_A.src)
    )
    destination_match = (
        (RULE_A.dst == RULE_B.src) or RULE_A.dst.overlaps(RULE_B.src) or
        RULE_A.dst.subnet_of(RULE_B.src) or RULE_B.src.subnet_of(RULE_A.dst)
    )


    if not source_match or not destination_match:
        return False


    return True

# N9K&LINKAS ACL交叉检查任务类：从源Excel提取CS-N9K（N9K核心交换机）、LINK-AS（LINKAS接入交换机）和OOB-DS（OOB-DS交换机）的ACL配置，输出到Excel文件
class ACLCrossCheckTask(BaseTask):
    """ACL跨平台检查任务

    检查CS-N9K和LINK-AS之间的ACL规则匹配情况，包括完全匹配、覆盖匹配等
    """

    @staticmethod
    # 从配置文件加载Sheet名称到平台网段的映射（接入层同平台策略）
    def _load_platform_network_map() -> Dict[str, List[IPv4Network]]:
        # 从公共配置settings.platform_network_map加载平台网段映射：返回Sheet名称到平台网段列表的映射
        try:
            # 优先从公共配置settings.platform_network_map读取，兼容旧配置ACLCrossCheckTask.platform_network_map
            configuration_map = CONFIG.get("settings", {}).get("platform_network_map", {})
            if not configuration_map:
                # 兼容旧配置位置
                configuration_map = CONFIG.get("ACLCrossCheckTask", {}).get(
                    "platform_network_map", {}
                )
            RESULT = {}
            for SHEET_NAME, NETWORK_STRINGS in configuration_map.items():
                NETWORKS = [IPv4Network(NET_STR, strict=False) for NET_STR in NETWORK_STRINGS]
                RESULT[SHEET_NAME] = NETWORKS
            return RESULT
        except (KeyError, ValueError, TypeError, AttributeError):
            # 如果配置读取失败，返回空字典
            return {}


    @staticmethod
    # 从配置文件加载特殊地址段映射：返回Sheet名称到特殊地址段列表的映射
    def _load_special_network_map() -> Dict[str, List[IPv4Network]]:
        # 从配置文件加载特殊地址段映射：返回Sheet名称到特殊地址段列表的映射
        try:
            configuration_map = CONFIG.get("ACLCrossCheckTask", {}).get("special_network_map", {})
            RESULT = {}
            for SHEET_NAME, NETWORK_STRINGS in configuration_map.items():
                NETWORKS = [IPv4Network(NET_STR, strict=False) for NET_STR in NETWORK_STRINGS]
                RESULT[SHEET_NAME] = NETWORKS
            return RESULT
        except (KeyError, ValueError, TypeError, AttributeError):
            # 如果配置读取失败，返回空字典
            return {}


    # 预定义的Sheet名称到平台网段的映射（从配置文件加载）
    PLATFORM_NETWORK_MAP = _load_platform_network_map.__func__()


    # 预定义的Sheet名称到特殊地址段的映射（从配置文件加载）
    SPECIAL_NETWORK_MAP = _load_special_network_map.__func__()

    # 初始化ACL交叉检查任务：设置固定配置参数
    def __init__(self):
        super().__init__("N9K&LINKAS ACL交叉检查")
        # 固定配置参数
        TODAY = get_today_str()
        # 从 LOG/DeviceBackupTask/ 读取
        self.INPUT_PATH = build_log_path("DeviceBackupTask", f"{TODAY}-关键设备配置备份输出EXCEL基础任务.xlsx")


        # 输出到 LOG/ACLCrossCheckTask/
        self.OUTPUT_DIR = build_log_path("ACLCrossCheckTask")
        self.NAME = "N9K&LINKAS ACL交叉检查"

    # 返回要处理的Sheet列表
    def items(self):
        """返回要处理的工作表列表

        Returns:
            工作表名称列表，排除Report工作表
        """
        if not os.path.exists(self.INPUT_PATH):
            self.add_result(Level.ERROR, f"输入文件不存在: {self.INPUT_PATH}")
            return []
        try:
            INPUT_WORKBOOK = load_workbook(self.INPUT_PATH)
            SHEET_NAMES = [
                WORKSHEET.title for WORKSHEET in INPUT_WORKBOOK.worksheets
                if WORKSHEET.title != 'Report'
            ]
            INPUT_WORKBOOK.close()
            return SHEET_NAMES
        except (FileNotFoundError, PermissionError, KeyError, AttributeError):
            return []

    # 获取平台网段：从预定义映射中获取平台网段列表，如果Sheet不在映射中则返回None
    def _get_platform_networks(self, sheet_name: str) -> Optional[List[IPv4Network]]:
        return self.PLATFORM_NETWORK_MAP.get(sheet_name)

    # 判断网络是否在平台网段内：支持多个平台网段，检查网络是否在任一平台网段内（重叠或包含关系）
    def _network_in_platform(
            self, network: IPv4Network,
            platform_networks: List[IPv4Network]
    ) -> bool:
        if not platform_networks:
            return False
        try:
            for PLATFORM_NETWORK in platform_networks:
                # 检查网络是否与平台网段重叠，或者网络是平台网段的子网
                if network.overlaps(PLATFORM_NETWORK) or network.subnet_of(PLATFORM_NETWORK):
                    return True
            return False
        except (ValueError, TypeError, AttributeError):
            return False


    # 从ACL规则文本中提取源地址和目的地址（即使parse_acl失败）：使用正则表达式匹配CIDR格式的IP地址，返回(源地址, 目的地址)或None
    def _extract_networks_from_rule_text(
            self, rule_text: str
    ) -> Optional[Tuple[IPv4Network, IPv4Network]]:
        if not rule_text:
            return None


        # 匹配CIDR格式的IP地址：\d+\.\d+\.\d+\.\d+/\d+
        cidr_pattern = re.compile(r'\b(\d+\.\d+\.\d+\.\d+/\d+)\b')
        matches = cidr_pattern.findall(rule_text)


        if len(matches) >= 2:
            try:
                source_network = IPv4Network(matches[0], strict=False)
                destination_network = IPv4Network(matches[1], strict=False)
                return (source_network, destination_network)
            except (ValueError, AttributeError):
                return None


        return None

    # ========== 辅助方法：减少重复代码 ==========


    # 从指定列收集规则：从target_cols中收集规则，返回[(col, row, parsed_rule, raw_text), ...]
    def _collect_rules_from_cols(self, target_cols, col_mapping, rule_row_mapping):
        rules = []
        for col, device_number, device_name in target_cols:
            output_col = col_mapping[col]
            if output_col in rule_row_mapping:
                for row, (raw_text, parsed_rule) in rule_row_mapping[output_col].items():
                    rules.append((output_col, row, parsed_rule, raw_text))
        return rules


    # 设置单元格字体颜色（统一方法）：设置openpyxl Cell对象的字体颜色，color为颜色值（如"FF00FF00"表示绿色），preserve_style表示是否保留原有样式
    def _set_cell_font_color(self, cell, color, preserve_style=True):
        # 如果color是字符串，转换为Color对象
        if isinstance(color, str):
            color = Color(rgb=color)


        if preserve_style and cell.font:
            cell.font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=color
            )
        else:
            cell.font = Font(
                name="宋体",
                size=11,
                bold=False,
                italic=False,
                color=color
            )


    # 检查颜色优先级，判断是否可以标记：检查cell是否已有exclude_colors中的颜色，返回True表示可以标记，False表示应该跳过
    def _check_color_priority(self, cell, exclude_colors):
        if not exclude_colors:
            return True


        current_color = cell.font.color if cell.font and cell.font.color else None
        if not current_color:
            return True


        try:
            color_string = str(current_color).upper()
            for EXCLUDE_COLOR in exclude_colors:
                if EXCLUDE_COLOR.upper() in color_string:
                    return False
            return True
        except (AttributeError, TypeError):
            return True


    # ========== 公共辅助方法（优化提取） ==========


    # 从缓存获取规则，如果没有缓存则重新收集：rule_type为'cat1'(CS-N9K), 'cat2'(LINK-AS), 'cat6'(OOB-DS)
    def _get_rules_from_cache_or_collect(self, sheet_name, rule_type, sheet_info, rules_cache):
        if rules_cache and sheet_name in rules_cache:
            return rules_cache[sheet_name].get(rule_type, [])
        else:
            # 重新收集规则
            col_mapping = sheet_info['col_mapping']
            rule_row_mapping = sheet_info['rule_row_mapping']
            target_cols = sheet_info[f'{rule_type}_target_cols']
            return self._collect_rules_from_cols(target_cols, col_mapping, rule_row_mapping)


    # 统一的单元格标记方法：cells为[(sheet_name, col, row), ...]，color为颜色代码，exclude_colors为不覆盖的颜色列表，mark_same_cat2为是否标记相同LINK-AS规则，返回(marked_count, skipped_count, all_cells)
    def _mark_cells_with_color(
            self, cells, color, exclude_colors, output_workbook,
            sheet_info_list=None, mark_same_cat2=False
    ):
        if not cells:
            return 0, 0, []


        # 构建排除颜色的单元格集合
        exclude_sets = {}
        if exclude_colors:
            for exclude_color in exclude_colors:
                exclude_sets[exclude_color] = set()


        marked_count = 0
        skipped_count = 0
        all_cells = list(cells)


        for sheet_name, col, row in cells:
            # 检查是否已被排除颜色标记
            should_skip = False
            if exclude_colors:
                if sheet_name in output_workbook.sheetnames:
                    ws = output_workbook[sheet_name]
                    cell = ws.cell(row=row, column=col)
                    if not self._check_color_priority(cell, exclude_colors):
                        should_skip = True


            if should_skip:
                skipped_count += 1
                continue


            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                if self._check_color_priority(cell, exclude_colors):
                    self._set_cell_font_color(cell, color)
                    marked_count += 1


        # 处理相同LINK-AS规则标记
        if mark_same_cat2 and sheet_info_list:
            for sheet_info in sheet_info_list:
                additional_cells = self._mark_same_cat2_rules(
                    sheet_info, all_cells, color, exclude_colors, output_workbook
                )
                all_cells.extend(additional_cells)


        return marked_count, skipped_count, all_cells


    # 处理相同规则的LINK-AS设备列标记：对于已标记的LINK-AS规则，检查同一Sheet中其他LINK-AS设备列的相同规则并标记，返回额外标记的单元格列表
    def _mark_same_cat2_rules(
            self, sheet_info, marked_cells, color,
            exclude_colors=None, output_workbook=None
    ):
        if not output_workbook:
            return []


        additional_cells = []
        marked_cells_set = set(marked_cells)


        sheet_name = sheet_info['sheet_name']
        if sheet_name not in output_workbook.sheetnames:
            return []


        ws = output_workbook[sheet_name]
        cat2_target_cols = sheet_info['cat2_target_cols']
        col_mapping = sheet_info['col_mapping']


        # 对于每个已标记的LINK-AS规则，检查同一Sheet中其他LINK-AS设备列的相同规则
        for sheet_name_marked, col_marked, row_marked in marked_cells:
            if sheet_name_marked != sheet_name:
                continue


            # 检查这个单元格是否是LINK-AS设备列
            is_cat2_col = False
            for col, device_number, device_name in cat2_target_cols:
                output_col = col_mapping[col]
                if output_col == col_marked:
                    is_cat2_col = True
                    break


            if not is_cat2_col:
                continue


            # 获取这个单元格的值
            cell_marked = ws.cell(row=row_marked, column=col_marked)
            cell_value = cell_marked.value
            if not cell_value:
                continue


            # 检查同一Sheet中其他LINK-AS设备列（同一行）是否有相同的规则
            for col, device_number, device_name in cat2_target_cols:
                output_col = col_mapping[col]
                if output_col == col_marked:
                    continue  # 跳过自己


                # 检查同一行的单元格是否有相同的值
                cell_other = ws.cell(row=row_marked, column=output_col)
                if cell_other.value == cell_value:
                    # 如果还没有被标记，且不是更高优先级的颜色，添加到额外标记列表
                    if (sheet_name, output_col, row_marked) not in marked_cells_set:
                        # 检查颜色优先级
                        if self._check_color_priority(cell_other, exclude_colors or []):
                            additional_cells.append((sheet_name, output_col, row_marked))


        # 标记额外的单元格
        for sheet_name_add, col_add, row_add in additional_cells:
            if sheet_name_add in output_workbook.sheetnames:
                ws_add = output_workbook[sheet_name_add]
                cell_add = ws_add.cell(row=row_add, column=col_add)
                self._set_cell_font_color(cell_add, color)


        return additional_cells

    # 统一的步骤执行包装器，处理进度更新、日志记录、文件保存等
    def _execute_step(
            self, step_num, step_name, step_func, sheet_info_list,
            output_workbook, *args, progress=None, save_after_step=False,
            output_path=None, stop_at_step=None, **kwargs
    ):
        self.add_result(Level.OK, f"开始执行步骤{step_num}：{step_name}...")
        if progress:
            progress.set_description(f"{self.NAME} (步骤{step_num}: {step_name})")


        result = step_func(sheet_info_list, output_workbook, *args, **kwargs)


        if save_after_step and output_path:
            save_excel_workbook(output_workbook, output_path)


        # 根据结果类型生成完成消息
        if isinstance(result, list):
            self.add_result(Level.OK, f"步骤{step_num}执行完成，标记了 {len(result)} 个单元格")
        else:
            self.add_result(Level.OK, f"步骤{step_num}执行完成")


        if progress:
            progress.update(1)


        if stop_at_step == step_num:
            self.add_result(Level.WARN, f"已执行到步骤{step_num}，停止执行")
            return None


        return result

    # ========== 跨平台步骤1-6：跨Sheet比较和标记方法 ==========


    # 跨Sheet匹配检查的公共辅助方法（支持双向检查）：检查Sheet内部匹配和跨Sheet匹配，返回匹配的规则对列表
    def _cross_sheet_match_check_helper(
            self, sheet_info_list, match_func_internal, match_func_cross,
            rules_cache=None
    ):
        matched_pairs = []


        if len(sheet_info_list) < 2:
            return matched_pairs


        # 对每两个不同的Sheet进行比较
        for SHEET_INDEX_A, sheet_info_a in enumerate(sheet_info_list):
            for SHEET_INDEX_B, sheet_info_b in enumerate(sheet_info_list):
                if SHEET_INDEX_A >= SHEET_INDEX_B:  # 避免重复比较
                    continue


                platform_network_a = sheet_info_a['platform_network']
                platform_network_b = sheet_info_b['platform_network']


                # 只比较不同平台的Sheet
                if not (platform_network_a and platform_network_b and
                        platform_network_a != platform_network_b):
                    continue


                sheet_name_a = sheet_info_a['sheet_name']
                sheet_name_b = sheet_info_b['sheet_name']


                # 使用规则缓存（性能优化）
                if rules_cache:
                    cat1_rules_a = rules_cache[sheet_name_a]['cat1']
                    cat2_rules_a = rules_cache[sheet_name_a]['cat2']
                    cat1_rules_b = rules_cache[sheet_name_b]['cat1']
                    cat2_rules_b = rules_cache[sheet_name_b]['cat2']
                else:
                    # 兼容旧代码：如果没有缓存则重新收集
                    rule_row_mapping_a = sheet_info_a['rule_row_mapping']
                    rule_row_mapping_b = sheet_info_b['rule_row_mapping']
                    col_mapping_a = sheet_info_a['col_mapping']
                    col_mapping_b = sheet_info_b['col_mapping']
                    cat1_target_cols_a = sheet_info_a['cat1_target_cols']
                    cat2_target_cols_a = sheet_info_a['cat2_target_cols']
                    cat1_rules_a = self._collect_rules_from_cols(
                        cat1_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat2_rules_a = self._collect_rules_from_cols(
                        cat2_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat1_target_cols_b = sheet_info_b['cat1_target_cols']
                    cat2_target_cols_b = sheet_info_b['cat2_target_cols']
                    cat1_rules_b = self._collect_rules_from_cols(
                        cat1_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )
                    cat2_rules_b = self._collect_rules_from_cols(
                        cat2_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )


                # ========== 性能优化：预构建匹配索引 ==========
                # 预构建CS-N9K反向匹配索引（A -> B）
                # {(cat1_col_a, cat1_row_a): [(cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b), ...]}
                cat1_reverse_index_ab = {}
                for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                    matches = []
                    for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                        if match_func_cross(cat1_rule_a, cat1_rule_b):
                            matches.append((cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b))
                    if matches:
                        cat1_reverse_index_ab[(cat1_col_a, cat1_row_a)] = matches


                # 预构建LINK-AS反向匹配索引（A -> B）
                # {(cat2_col_a, cat2_row_a): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                cat2_reverse_index_ab = {}
                for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        if match_func_cross(cat2_rule_a, cat2_rule_b):
                            matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat2_reverse_index_ab[(cat2_col_a, cat2_row_a)] = matches


                # 预构建CS-N9K反向匹配索引（B -> A）
                # {(cat1_col_b, cat1_row_b): [(cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a), ...]}
                cat1_reverse_index_ba = {}
                for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                    matches = []
                    for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                        if match_func_cross(cat1_rule_b, cat1_rule_a):
                            matches.append((cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a))
                    if matches:
                        cat1_reverse_index_ba[(cat1_col_b, cat1_row_b)] = matches


                # 预构建LINK-AS反向匹配索引（B -> A）
                # {(cat2_col_b, cat2_row_b): [(cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a), ...]}
                cat2_reverse_index_ba = {}
                for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                    matches = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        if match_func_cross(cat2_rule_b, cat2_rule_a):
                            matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                    if matches:
                        cat2_reverse_index_ba[(cat2_col_b, cat2_row_b)] = matches


                # 检查Sheet A -> Sheet B方向（使用预构建索引）
                for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                    cat1_key_a = (cat1_col_a, cat1_row_a)
                    # 早期退出：如果cat1没有匹配的，跳过
                    if cat1_key_a not in cat1_reverse_index_ab:
                        continue


                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        # 条件1：Sheet A内部匹配
                        if not match_func_internal(cat1_rule_a, cat2_rule_a):
                            continue


                        cat2_key_a = (cat2_col_a, cat2_row_a)
                        # 早期退出：如果cat2没有匹配的，跳过
                        if cat2_key_a not in cat2_reverse_index_ab:
                            continue


                        # 条件2和3：使用预构建索引（O(1)查找）
                        found_match = False
                        for matched_cat1_b in cat1_reverse_index_ab[cat1_key_a]:
                            for matched_cat2_b in cat2_reverse_index_ab[cat2_key_a]:
                                # 条件4：Sheet B内部匹配
                                if match_func_internal(matched_cat1_b[2], matched_cat2_b[2]):
                                    matched_pairs.append((
                                        sheet_info_a['sheet_name'], cat1_col_a, cat1_row_a,
                                        cat2_col_a, cat2_row_a,
                                        sheet_info_b['sheet_name'], matched_cat1_b[0],
                                        matched_cat1_b[1], matched_cat2_b[0], matched_cat2_b[1]
                                    ))
                                    found_match = True
                                    break  # 找到一个匹配即可
                            if found_match:
                                break  # 已找到匹配，跳出内层循环


                # 检查Sheet B -> Sheet A方向（使用预构建索引）
                for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                    cat1_key_b = (cat1_col_b, cat1_row_b)
                    # 早期退出：如果cat1没有匹配的，跳过
                    if cat1_key_b not in cat1_reverse_index_ba:
                        continue


                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        # 条件1：Sheet B内部匹配
                        if not match_func_internal(cat1_rule_b, cat2_rule_b):
                            continue


                        cat2_key_b = (cat2_col_b, cat2_row_b)
                        # 早期退出：如果cat2没有匹配的，跳过
                        if cat2_key_b not in cat2_reverse_index_ba:
                            continue


                        # 条件2和3：使用预构建索引（O(1)查找）
                        found_match = False
                        for matched_cat1_a in cat1_reverse_index_ba[cat1_key_b]:
                            for matched_cat2_a in cat2_reverse_index_ba[cat2_key_b]:
                                # 条件4：Sheet A内部匹配
                                if match_func_internal(matched_cat1_a[2], matched_cat2_a[2]):
                                    matched_pairs.append((
                                        sheet_info_b['sheet_name'], cat1_col_b, cat1_row_b,
                                        cat2_col_b, cat2_row_b,
                                        sheet_info_a['sheet_name'], matched_cat1_a[0],
                                        matched_cat1_a[1], matched_cat2_a[0], matched_cat2_a[1]
                                    ))
                                    found_match = True
                                    break  # 找到一个匹配即可
                            if found_match:
                                break  # 已找到匹配，跳出内层循环


        return matched_pairs


    # 跨Sheet cat6和cat2匹配检查辅助函数
    def _cross_sheet_cat6_cat2_match_check_helper(
        self, sheet_info_list, match_func_internal, match_func_cross,
        rules_cache=None
    ):
        matched_pairs = []


        if len(sheet_info_list) >= 2:
            # 对每两个不同的Sheet进行比较
            for SHEET_INDEX_A, sheet_info_a in enumerate(sheet_info_list):
                for SHEET_INDEX_B, sheet_info_b in enumerate(sheet_info_list):
                    if SHEET_INDEX_A >= SHEET_INDEX_B:  # 避免重复比较
                        continue


                    platform_network_a = sheet_info_a['platform_network']
                    platform_network_b = sheet_info_b['platform_network']


                    # 只比较不同平台的Sheet
                    if not (platform_network_a and platform_network_b and
                        platform_network_a != platform_network_b):
                        continue


                    sheet_name_a = sheet_info_a['sheet_name']
                    sheet_name_b = sheet_info_b['sheet_name']


                    # 使用规则缓存（性能优化）
                    if rules_cache:
                        cat6_rules_a = rules_cache[sheet_name_a]['cat6']
                        cat2_rules_a = rules_cache[sheet_name_a]['cat2']
                        cat6_rules_b = rules_cache[sheet_name_b]['cat6']
                        cat2_rules_b = rules_cache[sheet_name_b]['cat2']
                    else:
                        # 兼容旧代码：如果没有缓存则重新收集
                        rule_row_mapping_a = sheet_info_a['rule_row_mapping']
                        rule_row_mapping_b = sheet_info_b['rule_row_mapping']
                        col_mapping_a = sheet_info_a['col_mapping']
                        col_mapping_b = sheet_info_b['col_mapping']
                        cat6_target_cols_a = sheet_info_a['cat6_target_cols']
                        cat2_target_cols_a = sheet_info_a['cat2_target_cols']
                        cat6_rules_a = self._collect_rules_from_cols(
                            cat6_target_cols_a, col_mapping_a, rule_row_mapping_a
                        )
                        cat2_rules_a = self._collect_rules_from_cols(
                            cat2_target_cols_a, col_mapping_a, rule_row_mapping_a
                        )
                        cat6_target_cols_b = sheet_info_b['cat6_target_cols']
                        cat2_target_cols_b = sheet_info_b['cat2_target_cols']
                        cat6_rules_b = self._collect_rules_from_cols(
                            cat6_target_cols_b, col_mapping_b, rule_row_mapping_b
                        )
                        cat2_rules_b = self._collect_rules_from_cols(
                            cat2_target_cols_b, col_mapping_b, rule_row_mapping_b
                        )


                    # ========== 性能优化：预构建匹配索引 ==========
                    # 预构建OOB-DS反向匹配索引（A -> B）
                    # {(cat6_col_a, cat6_row_a): [(cat6_col_b, cat6_row_b, cat6_rule_b, cat6_raw_b), ...]}
                    cat6_reverse_index_ab = {}
                    for cat6_col_a, cat6_row_a, cat6_rule_a, cat6_raw_a in cat6_rules_a:
                        matches = []
                        for cat6_col_b, cat6_row_b, cat6_rule_b, cat6_raw_b in cat6_rules_b:
                            if match_func_cross(cat6_rule_a, cat6_rule_b):
                                matches.append((cat6_col_b, cat6_row_b, cat6_rule_b, cat6_raw_b))
                        if matches:
                            cat6_reverse_index_ab[(cat6_col_a, cat6_row_a)] = matches


                    # 预构建LINK-AS反向匹配索引（A -> B）
                    # {(cat2_col_a, cat2_row_a): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                    cat2_reverse_index_ab = {}
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        matches = []
                        for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                            if match_func_cross(cat2_rule_a, cat2_rule_b):
                                matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                        if matches:
                            cat2_reverse_index_ab[(cat2_col_a, cat2_row_a)] = matches


                    # 预构建OOB-DS反向匹配索引（B -> A）
                    # {(cat6_col_b, cat6_row_b): [(cat6_col_a, cat6_row_a, cat6_rule_a, cat6_raw_a), ...]}
                    cat6_reverse_index_ba = {}
                    for cat6_col_b, cat6_row_b, cat6_rule_b, cat6_raw_b in cat6_rules_b:
                        matches = []
                        for cat6_col_a, cat6_row_a, cat6_rule_a, cat6_raw_a in cat6_rules_a:
                            if match_func_cross(cat6_rule_b, cat6_rule_a):
                                matches.append((cat6_col_a, cat6_row_a, cat6_rule_a, cat6_raw_a))
                        if matches:
                            cat6_reverse_index_ba[(cat6_col_b, cat6_row_b)] = matches


                    # 预构建LINK-AS反向匹配索引（B -> A）
                    # {(cat2_col_b, cat2_row_b): [(cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a), ...]}
                    cat2_reverse_index_ba = {}
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        matches = []
                        for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                            if match_func_cross(cat2_rule_b, cat2_rule_a):
                                matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                        if matches:
                            cat2_reverse_index_ba[(cat2_col_b, cat2_row_b)] = matches


                    # 检查Sheet A -> Sheet B方向（使用预构建索引）
                    for cat6_col_a, cat6_row_a, cat6_rule_a, cat6_raw_a in cat6_rules_a:
                        cat6_key_a = (cat6_col_a, cat6_row_a)
                        # 早期退出：如果cat6没有匹配的，跳过
                        if cat6_key_a not in cat6_reverse_index_ab:
                            continue


                        for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                            # 条件1：Sheet A内部匹配
                            if not match_func_internal(cat6_rule_a, cat2_rule_a):
                                continue


                            cat2_key_a = (cat2_col_a, cat2_row_a)
                            # 早期退出：如果cat2没有匹配的，跳过
                            if cat2_key_a not in cat2_reverse_index_ab:
                                continue


                            # 条件2和3：使用预构建索引（O(1)查找）
                            found_match = False
                            for matched_cat6_b in cat6_reverse_index_ab[cat6_key_a]:
                                for matched_cat2_b in cat2_reverse_index_ab[cat2_key_a]:
                                    # 条件4：Sheet B内部匹配
                                    if match_func_internal(matched_cat6_b[2], matched_cat2_b[2]):
                                        matched_pairs.append((
                                            sheet_info_a['sheet_name'], cat6_col_a, cat6_row_a,
                                            cat2_col_a, cat2_row_a,
                                            sheet_info_b['sheet_name'], matched_cat6_b[0],
                                            matched_cat6_b[1], matched_cat2_b[0], matched_cat2_b[1]
                                        ))
                                        found_match = True
                                        break  # 找到一个匹配即可
                                if found_match:
                                    break  # 已找到匹配，跳出内层循环


                    # 检查Sheet B -> Sheet A方向（使用预构建索引）
                    for cat6_col_b, cat6_row_b, cat6_rule_b, cat6_raw_b in cat6_rules_b:
                        cat6_key_b = (cat6_col_b, cat6_row_b)
                        # 早期退出：如果cat6没有匹配的，跳过
                        if cat6_key_b not in cat6_reverse_index_ba:
                            continue


                        for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                            # 条件1：Sheet B内部匹配
                            if not match_func_internal(cat6_rule_b, cat2_rule_b):
                                continue


                            cat2_key_b = (cat2_col_b, cat2_row_b)
                            # 早期退出：如果cat2没有匹配的，跳过
                            if cat2_key_b not in cat2_reverse_index_ba:
                                continue


                            # 条件2和3：使用预构建索引（O(1)查找）
                            found_match = False
                            for matched_cat6_a in cat6_reverse_index_ba[cat6_key_b]:
                                for matched_cat2_a in cat2_reverse_index_ba[cat2_key_b]:
                                    # 条件4：Sheet A内部匹配
                                    if match_func_internal(matched_cat6_a[2], matched_cat2_a[2]):
                                        matched_pairs.append((
                                            sheet_info_b['sheet_name'], cat6_col_b, cat6_row_b,
                                            cat2_col_b, cat2_row_b,
                                            sheet_info_a['sheet_name'], matched_cat6_a[0],
                                            matched_cat6_a[1], matched_cat2_a[0], matched_cat2_a[1]
                                        ))
                                        found_match = True
                                        break  # 找到一个匹配即可
                                if found_match:
                                    break  # 已找到匹配，跳出内层循环


        return matched_pairs


    # ========== 合并后的跨平台步骤1-3：cat1/cat2匹配检查（统一函数） ==========


    # 统一的cat1/cat2匹配检查函数（对应跨平台步骤1-3，内部step_num为4/5/6）：
    # - match_type为'complete'（完全匹配）、'cat1_cover'（cat1覆盖cat2）、'cat2_cover'（cat2覆盖cat1）
    # - exclude_cells为已标记的单元格列表（用于避免覆盖）
    def _match_cat1_cat2_cross_platform(self, sheet_info_list, output_workbook, step_num,

                                        match_type, exclude_cells=None, rules_cache=None):
        # 根据match_type选择内部匹配函数
        if match_type == 'complete':
            match_func_internal = rule_matches
            step_desc = "cat1完全匹配"
            match_desc = "cat1 vs cat2匹配相同（两个Sheet）"
        elif match_type == 'cat1_cover':
            match_func_internal = lambda cat1, cat2: rule_covers(cat1, cat2)
            step_desc = "cat1覆盖匹配"
            match_desc = "cat1 vs cat2覆盖（两个Sheet）"
        elif match_type == 'cat2_cover':
            match_func_internal = lambda cat1, cat2: rule_covers(cat2, cat1)
            step_desc = "cat2覆盖cat1匹配"
            match_desc = "cat2 vs cat1覆盖（两个Sheet）"
        else:
            raise ValueError(f"Invalid match_type: {match_type}")


        # 使用辅助方法进行匹配检查
        matched_pairs = self._cross_sheet_match_check_helper(
            sheet_info_list,
            match_func_internal=match_func_internal,
            match_func_cross=rule_reverse_matches,
            rules_cache=rules_cache
        )


        # 转换为单元格列表
        new_cells = []
        for sheet_name_a, cat1_col_a, cat1_row_a, cat2_col_a, cat2_row_a, \
            sheet_name_b, cat1_col_b, cat1_row_b, cat2_col_b, cat2_row_b in matched_pairs:
            new_cells.append((sheet_name_a, cat1_col_a, cat1_row_a))
            new_cells.append((sheet_name_a, cat2_col_a, cat2_row_a))
            new_cells.append((sheet_name_b, cat1_col_b, cat1_row_b))
            new_cells.append((sheet_name_b, cat2_col_b, cat2_row_b))


        cross_match_count = len(matched_pairs)


        # 构建排除颜色列表
        exclude_colors = ["FF00FF00"] if exclude_cells else None
        exclude_cell_set = set(exclude_cells) if exclude_cells else set()


        # 使用统一的标记方法
        marked_count, skipped_count, all_cells = self._mark_cells_with_color(
            new_cells, "FF00FF00", exclude_colors, output_workbook,
            sheet_info_list, mark_same_cat2=True
        )


        # 过滤掉已排除的单元格
        filtered_cells = [cell for cell in all_cells if cell not in exclude_cell_set]


        if cross_match_count > 0:
            self.add_result(Level.OK,

                f"步骤{step_num}{step_desc}检查完成：发现{cross_match_count}对匹配规则"
                f"（标绿色，四个条件全部满足：{match_desc} + cat1 vs cat1反向匹配 + cat2 vs cat2反向匹配）")


        return filtered_cells


    # ========== 保持向后兼容的独立步骤函数（调用统一函数） ==========


    # 步骤4：cat1完全匹配检查（标绿色）
    def _step_cat1_complete_match_check(self, sheet_info_list, output_workbook, rules_cache=None):
        return self._match_cat1_cat2_cross_platform(
            sheet_info_list, output_workbook, 4, 'complete', None, rules_cache
        )


    # 步骤5：cat1覆盖匹配检查（标绿色）
    def _step_cat1_cover_match_check(
            self, sheet_info_list, output_workbook, green_cells,
            rules_cache=None
    ):
        return self._match_cat1_cat2_cross_platform(
            sheet_info_list, output_workbook, 5, 'cat1_cover', green_cells, rules_cache
        )


    # 步骤6：cat2覆盖cat1匹配检查（标绿色）
    def _step_cat2_cover_cat1_match_check(
        self, sheet_info_list, output_workbook, green_cells,
        dark_green_cells, rules_cache=None
    ):
        # 合并所有已排除的单元格
        exclude_cells = (
            list(green_cells) + list(dark_green_cells)
            if dark_green_cells else green_cells
        )
        return self._match_cat1_cat2_cross_platform(
            sheet_info_list, output_workbook, 6, 'cat2_cover', exclude_cells, rules_cache
        )


    # 步骤10：跨Sheet cat1和cat2匹配检查（标绿色）：检查条件（三个条件必须全部满足）：
    # 1.cat1反向匹配 2.cat2反向匹配覆盖 3.Sheet A内部覆盖
    def _step_cross_sheet_cat1_cat2_reverse_match_check(
        self, sheet_info_list, output_workbook, green_cells, rules_cache=None
    ):
        new_green_cells = []


        if len(sheet_info_list) < 2:
            return green_cells


        for sheet_index_a, sheet_info_a in enumerate(sheet_info_list):
            for sheet_index_b, sheet_info_b in enumerate(sheet_info_list):
                if sheet_index_a >= sheet_index_b:  # 避免重复比较
                    continue


                platform_network_a = sheet_info_a['platform_network']
                platform_network_b = sheet_info_b['platform_network']


                # 只比较不同平台的Sheet
                if not (platform_network_a and platform_network_b and
                        platform_network_a != platform_network_b):
                    continue


                sheet_name_a = sheet_info_a['sheet_name']
                sheet_name_b = sheet_info_b['sheet_name']


                # 使用规则缓存（性能优化）
                if rules_cache:
                    cat1_rules_a = rules_cache[sheet_name_a]['cat1']
                    cat2_rules_a = rules_cache[sheet_name_a]['cat2']
                    cat1_rules_b = rules_cache[sheet_name_b]['cat1']
                    cat2_rules_b = rules_cache[sheet_name_b]['cat2']
                else:
                    # 兼容旧代码：如果没有缓存则重新收集
                    rule_row_mapping_a = sheet_info_a['rule_row_mapping']
                    rule_row_mapping_b = sheet_info_b['rule_row_mapping']
                    col_mapping_a = sheet_info_a['col_mapping']
                    col_mapping_b = sheet_info_b['col_mapping']
                    cat1_target_cols_a = sheet_info_a['cat1_target_cols']
                    cat2_target_cols_a = sheet_info_a['cat2_target_cols']
                    cat1_rules_a = self._collect_rules_from_cols(
                        cat1_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat2_rules_a = self._collect_rules_from_cols(
                        cat2_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat1_target_cols_b = sheet_info_b['cat1_target_cols']
                    cat2_target_cols_b = sheet_info_b['cat2_target_cols']
                    cat1_rules_b = self._collect_rules_from_cols(
                        cat1_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )
                    cat2_rules_b = self._collect_rules_from_cols(
                        cat2_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )


                # 预构建CS-N9K反向匹配索引（A -> B）
                # {(cat1_col_a, cat1_row_a): [(cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b), ...]}
                cat1_reverse_index_ab = {}
                for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                    matches = []
                    for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                        if rule_reverse_matches(cat1_rule_a, cat1_rule_b):
                            matches.append((cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b))
                    if matches:
                        cat1_reverse_index_ab[(cat1_col_a, cat1_row_a)] = matches


                # 预构建LINK-AS反向匹配索引（A -> B），同时检查覆盖关系
                # cat2_reverse_index_ab: {(cat2_col_a, cat2_row_a): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                # 其中cat2_rule_b与cat2_rule_a反向匹配，且cat2_rule_b覆盖cat2_rule_a（反向后）
                # 注意：对于反向匹配的规则，覆盖检查需要特殊处理
                cat2_reverse_cover_index_ab = {}
                for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        # 条件1：反向匹配
                        if not rule_reverse_matches(cat2_rule_a, cat2_rule_b):
                            continue


                        # 条件2：Sheet B的cat2覆盖Sheet A的cat2（反向后）
                        # 如果cat2_rule_a反向匹配cat2_rule_b，那么：
                        # cat2_rule_a反向: src=cat2_rule_a.dst, dst=cat2_rule_a.src
                        # cat2_rule_b: src=cat2_rule_b.src, dst=cat2_rule_b.dst
                        # 要检查cat2_rule_b是否覆盖cat2_rule_a（反向后），需要：
                        # - cat2_rule_b.src覆盖cat2_rule_a.dst
                        # - cat2_rule_b.dst覆盖cat2_rule_a.src
                        # - 协议和端口匹配
                        # cat2_rule_a反向的src是cat2_rule_b.src的子网
                        # cat2_rule_a反向的dst是cat2_rule_b.dst的子网
                        if (cat2_rule_a.action == cat2_rule_b.action and
                            proto_covers(cat2_rule_b.proto, cat2_rule_a.proto) and
                            cat2_rule_a.dst.subnet_of(cat2_rule_b.src) and
                            cat2_rule_a.src.subnet_of(cat2_rule_b.dst)):
                            # 检查端口（如果协议不是IP）
                            if cat2_rule_b.proto.lower() != "ip":
                                # 端口需要匹配（因为都是特定端口）
                                if cat2_rule_a.port == cat2_rule_b.port:
                                    matches.append(
                                        (cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b)
                                    )
                            else:
                                # IP协议，任意端口
                                matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat2_reverse_cover_index_ab[(cat2_col_a, cat2_row_a)] = matches


                # 检查Sheet A -> Sheet B方向
                for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                    cat1_key_a = (cat1_col_a, cat1_row_a)


                    if cat1_key_a not in cat1_reverse_index_ab:
                        continue


                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        # 条件3：Sheet A内部覆盖 - cat1覆盖cat2
                        if not rule_covers(cat1_rule_a, cat2_rule_a):
                            continue


                        cat2_key_a = (cat2_col_a, cat2_row_a)
                        if cat2_key_a not in cat2_reverse_cover_index_ab:
                            continue


                        # 找到cat1反向匹配且cat2反向匹配覆盖的组合
                        for matched_cat1_b in cat1_reverse_index_ab[cat1_key_a]:
                            for matched_cat2_b in cat2_reverse_cover_index_ab[cat2_key_a]:
                                new_green_cells.append(
                                    (sheet_name_a, cat1_col_a, cat1_row_a)
                                )
                                new_green_cells.append(
                                    (sheet_name_a, cat2_col_a, cat2_row_a)
                                )
                                new_green_cells.append(
                                    (sheet_name_b, matched_cat1_b[0], matched_cat1_b[1])
                                )
                                new_green_cells.append(
                                    (sheet_name_b, matched_cat2_b[0], matched_cat2_b[1])
                                )


                # 检查Sheet B -> Sheet A方向（双向检查）
                # 预构建CS-N9K反向匹配索引（B -> A）
                # {(cat1_col_b, cat1_row_b): [(cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a), ...]}
                cat1_reverse_index_ba = {}
                for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                    matches = []
                    for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                        if rule_reverse_matches(cat1_rule_b, cat1_rule_a):
                            matches.append((cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a))
                    if matches:
                        cat1_reverse_index_ba[(cat1_col_b, cat1_row_b)] = matches


                # 预构建LINK-AS反向匹配覆盖索引（B -> A）
                cat2_reverse_cover_index_ba = {}
                for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                    matches = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        # 条件1：反向匹配
                        if not rule_reverse_matches(cat2_rule_b, cat2_rule_a):
                            continue


                        # 条件2：Sheet A的cat2覆盖Sheet B的cat2（反向后）
                        if (cat2_rule_b.action == cat2_rule_a.action and
                            proto_covers(cat2_rule_a.proto, cat2_rule_b.proto) and
                            cat2_rule_b.dst.subnet_of(cat2_rule_a.src) and
                            cat2_rule_b.src.subnet_of(cat2_rule_a.dst)):
                            if cat2_rule_a.proto.lower() != "ip":
                                if cat2_rule_b.port == cat2_rule_a.port:
                                    matches.append(
                                        (cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a)
                                    )
                            else:
                                matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                    if matches:
                        cat2_reverse_cover_index_ba[(cat2_col_b, cat2_row_b)] = matches


                # 检查Sheet B -> Sheet A方向
                for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                    cat1_key_b = (cat1_col_b, cat1_row_b)
                    if cat1_key_b not in cat1_reverse_index_ba:
                        continue


                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        # 条件3：Sheet B内部覆盖 - cat1覆盖cat2
                        if not rule_covers(cat1_rule_b, cat2_rule_b):
                            continue


                        cat2_key_b = (cat2_col_b, cat2_row_b)
                        if cat2_key_b not in cat2_reverse_cover_index_ba:
                            continue


                        # 找到cat1反向匹配且cat2反向匹配覆盖的组合
                        for matched_cat1_a in cat1_reverse_index_ba[cat1_key_b]:
                            for matched_cat2_a in cat2_reverse_cover_index_ba[cat2_key_b]:
                                new_green_cells.append(
                                    (sheet_name_b, cat1_col_b, cat1_row_b)
                                )
                                new_green_cells.append(
                                    (sheet_name_b, cat2_col_b, cat2_row_b)
                                )
                                new_green_cells.append(
                                    (sheet_name_a, matched_cat1_a[0], matched_cat1_a[1])
                                )
                                new_green_cells.append(
                                    (sheet_name_a, matched_cat2_a[0], matched_cat2_a[1])
                                )


        # 去重
        new_green_cells = list(set(new_green_cells))


        # 统一标绿色字体（不覆盖已标绿色的单元格）
        green_cell_set = self._create_cell_set(green_cells)
        skipped_count = 0
        marked_count = 0


        for sheet_name, col, row in new_green_cells:
            if (sheet_name, col, row) in green_cell_set:
                skipped_count += 1
                continue


            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                self._set_cell_font_color(cell, "FF00FF00")  # 绿色
                marked_count += 1


        # 处理相同规则的cat2设备列标记
        for sheet_info in sheet_info_list:
            additional_cells = self._mark_same_cat2_rules(
                sheet_info, new_green_cells, "FF00FF00", None, output_workbook
            )
            new_green_cells.extend(additional_cells)


        # 合并到green_cells
        green_cells.extend(new_green_cells)
        green_cells = list(set(green_cells))


        # 每组4个单元格
        cross_match_count = (
            len([cell for cell in new_green_cells if cell not in green_cell_set]) // 4
        )
        if cross_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤8跨Sheet cat1和cat2匹配检查完成："
                f"发现{cross_match_count}对匹配规则，"
                f"标记{marked_count}个单元格"
                f"（跳过{skipped_count}个已标绿色的单元格）"
            )
        elif len(new_green_cells) > 0:
            self.add_result(
                Level.WARN,
                f"步骤8发现{len(new_green_cells)}个匹配单元格，"
                f"但全部已被其他步骤标记（跳过{skipped_count}个）"
            )


        return green_cells


    # 步骤8：平台源地址特殊目的地址cat1和cat2匹配检查（标绿色）：
    # 检查条件（三个条件必须全部满足）：
    # 1.源地址为本平台地址，目的地址为special_network_map
    # 2.Sheet A内部覆盖 3.Sheet A的cat2规则与Sheet B的cat2规则反向匹配完全相同
    def _step_platform_src_special_dst_cat1_cat2_check(
        self, sheet_info_list, output_workbook, green_cells, rules_cache=None
    ):
        # 步骤8：平台源地址特殊目的地址cat1和cat2匹配检查（标绿色）：检查条件（三个条件必须全部满足）：1.源地址为本平台地址，目的地址为special_network_map 2.Sheet A内部覆盖 3.Sheet A的cat2规则与Sheet B的cat2规则反向匹配完全相同
        new_green_cells = []


        if len(sheet_info_list) < 2:
            return green_cells


        # 收集所有special_network_map中的网络（任何Sheet的特殊网段）
        all_special_networks = []
        for SHEET_NETWORKS in self.SPECIAL_NETWORK_MAP.values():
            if SHEET_NETWORKS:
                all_special_networks.extend(SHEET_NETWORKS)


        # 对每两个不同的Sheet进行比较
        for sheet_index_a, sheet_info_a in enumerate(sheet_info_list):
            for sheet_index_b, sheet_info_b in enumerate(sheet_info_list):
                if sheet_index_a >= sheet_index_b:  # 避免重复比较
                    continue


                platform_network_a = sheet_info_a['platform_network']
                platform_network_b = sheet_info_b['platform_network']


                # 只比较不同平台的Sheet
                if not (platform_network_a and platform_network_b and
                        platform_network_a != platform_network_b):
                    continue


                sheet_name_a = sheet_info_a['sheet_name']
                sheet_name_b = sheet_info_b['sheet_name']


                # 获取完整的平台网段列表（用于检查源地址是否在本平台）
                platform_networks_a = sheet_info_a.get('platform_networks', [])
                platform_networks_b = sheet_info_b.get('platform_networks', [])


                # 使用规则缓存（性能优化）
                if rules_cache:
                    cat1_rules_a = rules_cache[sheet_name_a]['cat1']
                    cat2_rules_a = rules_cache[sheet_name_a]['cat2']
                    cat2_rules_b = rules_cache[sheet_name_b]['cat2']
                else:
                    # 兼容旧代码：如果没有缓存则重新收集
                    rule_row_mapping_a = sheet_info_a['rule_row_mapping']
                    rule_row_mapping_b = sheet_info_b['rule_row_mapping']
                    col_mapping_a = sheet_info_a['col_mapping']
                    col_mapping_b = sheet_info_b['col_mapping']
                    cat1_target_cols_a = sheet_info_a['cat1_target_cols']
                    cat2_target_cols_a = sheet_info_a['cat2_target_cols']
                    cat1_rules_a = self._collect_rules_from_cols(
                        cat1_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat2_rules_a = self._collect_rules_from_cols(
                        cat2_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat2_target_cols_b = sheet_info_b['cat2_target_cols']
                    cat2_rules_b = self._collect_rules_from_cols(
                        cat2_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )


                # 预构建Sheet A的cat1规则索引（满足条件1的cat1规则）
                # [(cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a), ...]
                valid_cat1_rules_a = []
                for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                    # 条件1：源地址为本平台地址，目的地址为special_network_map
                    source_in_platform = self._network_in_platform(
                        cat1_rule_a.src, platform_networks_a
                    )
                    if not source_in_platform:
                        continue


                    # 检查cat1目的地址是否在special_network_map中（任何Sheet的特殊网段）
                    dst_in_special = False
                    for special_net in all_special_networks:
                        if (cat1_rule_a.dst.overlaps(special_net) or
                                cat1_rule_a.dst.subnet_of(special_net)):
                            dst_in_special = True
                            break
                    if dst_in_special:
                        valid_cat1_rules_a.append(
                            (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a)
                        )


                # 预构建Sheet A内部cat1覆盖cat2的索引
                # {(cat1_col_a, cat1_row_a): [(cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a), ...]}
                cat1_cover_cat2_index_a = {}
                for cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a in valid_cat1_rules_a:
                    matches = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        if rule_covers(cat1_rule_a, cat2_rule_a):
                            matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                    if matches:
                        cat1_cover_cat2_index_a[(cat1_col_a, cat1_row_a)] = matches


                # 预构建Sheet A和Sheet B的cat2反向匹配索引
                # {(cat2_col_a, cat2_row_a): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                cat2_reverse_index_ab = {}
                for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        if rule_reverse_matches(cat2_rule_a, cat2_rule_b):
                            matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat2_reverse_index_ab[(cat2_col_a, cat2_row_a)] = matches


                # 检查Sheet A -> Sheet B方向
                for CAT1_KEY_A, CAT1_DATA in cat1_cover_cat2_index_a.items():
                    CAT1_COL_A, CAT1_ROW_A = CAT1_KEY_A
                    for CAT2_COL_A, CAT2_ROW_A, CAT2_RULE_A, CAT2_RAW_A in CAT1_DATA:
                        CAT2_KEY_A = (CAT2_COL_A, CAT2_ROW_A)
                        if CAT2_KEY_A not in cat2_reverse_index_ab:
                            continue


                        # 所有条件满足，标记单元格
                        for (CAT2_COL_B, CAT2_ROW_B, CAT2_RULE_B, CAT2_RAW_B) in (
                                cat2_reverse_index_ab[CAT2_KEY_A]
                        ):
                            new_green_cells.append((sheet_name_a, CAT1_COL_A, CAT1_ROW_A))
                            new_green_cells.append((sheet_name_a, CAT2_COL_A, CAT2_ROW_A))
                            new_green_cells.append((sheet_name_b, CAT2_COL_B, CAT2_ROW_B))


                # 检查Sheet B -> Sheet A方向（双向检查）
                # 使用规则缓存（性能优化）
                if rules_cache:
                    cat1_rules_b = rules_cache[sheet_name_b]['cat1']
                else:
                    cat1_target_cols_b = sheet_info_b['cat1_target_cols']
                    cat1_rules_b = self._collect_rules_from_cols(
                        cat1_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )


                # 预构建Sheet B的cat1规则索引（满足条件1的cat1规则）
                # [(cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b), ...]
                valid_cat1_rules_b = []
                for (cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b) in cat1_rules_b:
                    # 条件1：源地址为本平台地址，目的地址为special_network_map
                    source_in_platform = self._network_in_platform(
                        cat1_rule_b.src, platform_networks_b
                    )
                    if not source_in_platform:
                        continue


                    # 检查cat1目的地址是否在special_network_map中（任何Sheet的特殊网段）
                    dst_in_special = False
                    for special_net in all_special_networks:
                        if (cat1_rule_b.dst.overlaps(special_net) or
                                cat1_rule_b.dst.subnet_of(special_net)):
                            dst_in_special = True
                            break
                    if dst_in_special:
                        valid_cat1_rules_b.append(
                            (cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b)
                        )


                # 预构建Sheet B内部cat1覆盖cat2的索引
                # {(cat1_col_b, cat1_row_b): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                cat1_cover_cat2_index_b = {}
                for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in valid_cat1_rules_b:
                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        if rule_covers(cat1_rule_b, cat2_rule_b):
                            matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat1_cover_cat2_index_b[(cat1_col_b, cat1_row_b)] = matches


                # 预构建Sheet B和Sheet A的cat2反向匹配索引
                # {(cat2_col_b, cat2_row_b): [(cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a), ...]}
                cat2_reverse_index_ba = {}
                for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                    matches = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        if rule_reverse_matches(cat2_rule_b, cat2_rule_a):
                            matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                    if matches:
                        cat2_reverse_index_ba[(cat2_col_b, cat2_row_b)] = matches


                # 检查Sheet B -> Sheet A方向
                for CAT1_KEY_B, CAT1_DATA in cat1_cover_cat2_index_b.items():
                    CAT1_COL_B, CAT1_ROW_B = CAT1_KEY_B
                    for CAT2_COL_B, CAT2_ROW_B, CAT2_RULE_B, CAT2_RAW_B in CAT1_DATA:
                        CAT2_KEY_B = (CAT2_COL_B, CAT2_ROW_B)
                        if CAT2_KEY_B not in cat2_reverse_index_ba:
                            continue


                        # 所有条件满足，标记单元格
                        for (CAT2_COL_A, CAT2_ROW_A, CAT2_RULE_A, CAT2_RAW_A) in (
                                cat2_reverse_index_ba[CAT2_KEY_B]
                        ):
                            new_green_cells.append((sheet_name_b, CAT1_COL_B, CAT1_ROW_B))
                            new_green_cells.append((sheet_name_b, CAT2_COL_B, CAT2_ROW_B))
                            new_green_cells.append((sheet_name_a, CAT2_COL_A, CAT2_ROW_A))


        # 去重
        new_green_cells = list(set(new_green_cells))


        # 统一标绿色字体（不覆盖已标绿色的单元格）
        green_cell_set = self._create_cell_set(green_cells)
        skipped_count = 0
        marked_count = 0


        for sheet_name, col, row in new_green_cells:
            if (sheet_name, col, row) in green_cell_set:
                skipped_count += 1
                continue


            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                self._set_cell_font_color(cell, "FF00FF00")  # 绿色
                marked_count += 1


        # 处理相同规则的cat2设备列标记
        for sheet_info in sheet_info_list:
            additional_cells = self._mark_same_cat2_rules(
                sheet_info, new_green_cells, "FF00FF00", None, output_workbook
            )
            new_green_cells.extend(additional_cells)


        # 合并到green_cells
        green_cells.extend(new_green_cells)
        green_cells = list(set(green_cells))


        # 每组3个单元格（cat1_a, cat2_a, cat2_b）
        cross_match_count = len([
            CELL for CELL in new_green_cells if CELL not in green_cell_set
        ]) // 3
        if cross_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤8平台源地址特殊目的地址cat1和cat2匹配检查完成："
                f"发现{cross_match_count}对匹配规则，"
                f"标记{marked_count}个单元格"
                f"（跳过{skipped_count}个已标绿色的单元格）"
            )
        elif len(new_green_cells) > 0:
            self.add_result(
                Level.WARN,
                f"步骤8发现{len(new_green_cells)}个匹配单元格，"
                f"但全部已被其他步骤标记（跳过{skipped_count}个）"
            )


        return green_cells


    # 步骤9：cat2 IP协议覆盖cat1 TCP协议检查（标绿色）
    # 检查条件（三个条件必须全部满足）：
    # 1.源地址为本平台地址，目的地址为special_network_map
    # 2.Sheet A的cat2 IP协议覆盖Sheet A的cat1 TCP协议
    # 3.Sheet A的cat2规则与Sheet B的cat2规则反向匹配完全相同
    def _step_cat2_ip_cover_cat1_tcp_check(
            self, sheet_info_list, output_workbook, green_cells,
            rules_cache=None
    ):
        # 步骤9：cat2 IP协议覆盖cat1 TCP协议检查（标绿色）：检查条件（三个条件必须全部满足）：
        # 1.源地址为本平台地址，目的地址为special_network_map 2.Sheet A的cat2 IP协议覆盖Sheet A的cat1 TCP协议 3.Sheet A的cat2规则与Sheet B的cat2规则反向匹配完全相同
        new_green_cells = []


        if len(sheet_info_list) < 2:
            return green_cells


        # 收集所有special_network_map中的网络（任何Sheet的特殊网段）
        all_special_networks = []
        for SHEET_NETWORKS in self.SPECIAL_NETWORK_MAP.values():
            if SHEET_NETWORKS:
                all_special_networks.extend(SHEET_NETWORKS)


        # 对每两个不同的Sheet进行比较
        for sheet_index_a, sheet_info_a in enumerate(sheet_info_list):
            for sheet_index_b, sheet_info_b in enumerate(sheet_info_list):
                if sheet_index_a >= sheet_index_b:  # 避免重复比较
                    continue


                platform_network_a = sheet_info_a['platform_network']
                platform_network_b = sheet_info_b['platform_network']


                # 只比较不同平台的Sheet
                if not (platform_network_a and platform_network_b and
                        platform_network_a != platform_network_b):
                    continue


                sheet_name_a = sheet_info_a['sheet_name']
                sheet_name_b = sheet_info_b['sheet_name']


                # 获取完整的平台网段列表（用于检查源地址是否在本平台）
                platform_networks_a = sheet_info_a.get('platform_networks', [])
                platform_networks_b = sheet_info_b.get('platform_networks', [])


                # 使用规则缓存（性能优化）
                if rules_cache:
                    cat1_rules_a = rules_cache[sheet_name_a]['cat1']
                    cat2_rules_a = rules_cache[sheet_name_a]['cat2']
                    cat2_rules_b = rules_cache[sheet_name_b]['cat2']
                else:
                    # 兼容旧代码：如果没有缓存则重新收集
                    rule_row_mapping_a = sheet_info_a['rule_row_mapping']
                    rule_row_mapping_b = sheet_info_b['rule_row_mapping']
                    col_mapping_a = sheet_info_a['col_mapping']
                    col_mapping_b = sheet_info_b['col_mapping']
                    cat1_target_cols_a = sheet_info_a['cat1_target_cols']
                    cat2_target_cols_a = sheet_info_a['cat2_target_cols']
                    cat1_rules_a = self._collect_rules_from_cols(
                        cat1_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat2_rules_a = self._collect_rules_from_cols(
                        cat2_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat2_target_cols_b = sheet_info_b['cat2_target_cols']
                    cat2_rules_b = self._collect_rules_from_cols(
                        cat2_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )


                # 预构建Sheet A的cat1规则索引（满足条件1的cat1规则）
                # [(cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a), ...]
                valid_cat1_rules_a = []
                for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                    # 条件1：源地址为本平台地址，目的地址为special_network_map
                    source_in_platform = self._network_in_platform(
                        cat1_rule_a.src, platform_networks_a
                    )
                    if not source_in_platform:
                        continue


                    # 检查cat1目的地址是否在special_network_map中（任何Sheet的特殊网段）
                    dst_in_special = False
                    for special_net in all_special_networks:
                        if (cat1_rule_a.dst.overlaps(special_net) or
                                cat1_rule_a.dst.subnet_of(special_net)):
                            dst_in_special = True
                            break
                    if dst_in_special:
                        valid_cat1_rules_a.append(
                            (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a)
                        )


                # 预构建Sheet A内部cat2 IP协议覆盖cat1 TCP协议的索引
                # 条件2：Sheet A的cat2 IP协议覆盖Sheet A的cat1 TCP协议
                # {(cat1_col_a, cat1_row_a): [(cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a), ...]}
                cat2_ip_cover_cat1_tcp_index_a = {}
                for cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a in valid_cat1_rules_a:
                    # 检查cat1是否是TCP协议
                    if cat1_rule_a.proto.lower() != "tcp":
                        continue


                    matches = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        # 检查cat2是否是IP协议
                        if cat2_rule_a.proto.lower() != "ip":
                            continue


                        # 检查cat2 IP协议覆盖cat1 TCP协议
                        if rule_covers(cat2_rule_a, cat1_rule_a):
                            matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                    if matches:
                        cat2_ip_cover_cat1_tcp_index_a[(cat1_col_a, cat1_row_a)] = matches


                # 预构建Sheet A和Sheet B的cat2反向匹配索引
                # 条件3：Sheet A的cat2规则与Sheet B的cat2规则反向匹配完全相同
                # {(cat2_col_a, cat2_row_a): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                cat2_reverse_index_ab = {}
                for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        if rule_reverse_matches(cat2_rule_a, cat2_rule_b):
                            matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat2_reverse_index_ab[(cat2_col_a, cat2_row_a)] = matches


                # 检查Sheet A -> Sheet B方向
                for CAT1_KEY_A, CAT1_DATA in cat2_ip_cover_cat1_tcp_index_a.items():
                    CAT1_COL_A, CAT1_ROW_A = CAT1_KEY_A
                    for CAT2_COL_A, CAT2_ROW_A, CAT2_RULE_A, CAT2_RAW_A in CAT1_DATA:
                        CAT2_KEY_A = (CAT2_COL_A, CAT2_ROW_A)
                        if CAT2_KEY_A not in cat2_reverse_index_ab:
                            continue


                        # 所有条件满足，标记单元格
                        for (CAT2_COL_B, CAT2_ROW_B, CAT2_RULE_B, CAT2_RAW_B) in (
                                cat2_reverse_index_ab[CAT2_KEY_A]
                        ):
                            new_green_cells.append((sheet_name_a, CAT1_COL_A, CAT1_ROW_A))
                            new_green_cells.append((sheet_name_a, CAT2_COL_A, CAT2_ROW_A))
                            new_green_cells.append((sheet_name_b, CAT2_COL_B, CAT2_ROW_B))


                # 检查Sheet B -> Sheet A方向（双向检查）
                # 使用规则缓存（性能优化）
                if rules_cache:
                    cat1_rules_b = rules_cache[sheet_name_b]['cat1']
                else:
                    cat1_target_cols_b = sheet_info_b['cat1_target_cols']
                    rule_row_mapping_b = sheet_info_b['rule_row_mapping']
                    col_mapping_b = sheet_info_b['col_mapping']
                    cat1_rules_b = self._collect_rules_from_cols(
                        cat1_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )


                # 预构建Sheet B的cat1规则索引（满足条件1的cat1规则）
                # [(cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b), ...]
                valid_cat1_rules_b = []
                for (cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b) in cat1_rules_b:
                    # 条件1：源地址为本平台地址，目的地址为special_network_map
                    source_in_platform = self._network_in_platform(
                        cat1_rule_b.src, platform_networks_b
                    )
                    if not source_in_platform:
                        continue


                    # 检查cat1目的地址是否在special_network_map中（任何Sheet的特殊网段）
                    dst_in_special = False
                    for special_net in all_special_networks:
                        if (cat1_rule_b.dst.overlaps(special_net) or
                                cat1_rule_b.dst.subnet_of(special_net)):
                            dst_in_special = True
                            break
                    if dst_in_special:
                        valid_cat1_rules_b.append(
                            (cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b)
                        )


                # 预构建Sheet B内部cat2 IP协议覆盖cat1 TCP协议的索引
                # 条件2：Sheet B的cat2 IP协议覆盖Sheet B的cat1 TCP协议
                # {(cat1_col_b, cat1_row_b): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                cat2_ip_cover_cat1_tcp_index_b = {}
                for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in valid_cat1_rules_b:
                    # 检查cat1是否是TCP协议
                    if cat1_rule_b.proto.lower() != "tcp":
                        continue


                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        # 检查cat2是否是IP协议
                        if cat2_rule_b.proto.lower() != "ip":
                            continue


                        # 检查cat2 IP协议覆盖cat1 TCP协议
                        if rule_covers(cat2_rule_b, cat1_rule_b):
                            matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat2_ip_cover_cat1_tcp_index_b[(cat1_col_b, cat1_row_b)] = matches


                # 预构建Sheet B和Sheet A的cat2反向匹配索引
                # 条件3：Sheet B的cat2规则与Sheet A的cat2规则反向匹配完全相同
                # {(cat2_col_b, cat2_row_b): [(cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a), ...]}
                cat2_reverse_index_ba = {}
                for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                    matches = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        if rule_reverse_matches(cat2_rule_b, cat2_rule_a):
                            matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                    if matches:
                        cat2_reverse_index_ba[(cat2_col_b, cat2_row_b)] = matches


                # 检查Sheet B -> Sheet A方向
                for CAT1_KEY_B, CAT1_DATA in cat2_ip_cover_cat1_tcp_index_b.items():
                    CAT1_COL_B, CAT1_ROW_B = CAT1_KEY_B
                    for CAT2_COL_B, CAT2_ROW_B, CAT2_RULE_B, CAT2_RAW_B in CAT1_DATA:
                        CAT2_KEY_B = (CAT2_COL_B, CAT2_ROW_B)
                        if CAT2_KEY_B not in cat2_reverse_index_ba:
                            continue


                        # 所有条件满足，标记单元格
                        for (CAT2_COL_A, CAT2_ROW_A, CAT2_RULE_A, CAT2_RAW_A) in (
                                cat2_reverse_index_ba[CAT2_KEY_B]
                        ):
                            new_green_cells.append((sheet_name_b, CAT1_COL_B, CAT1_ROW_B))
                            new_green_cells.append((sheet_name_b, CAT2_COL_B, CAT2_ROW_B))
                            new_green_cells.append((sheet_name_a, CAT2_COL_A, CAT2_ROW_A))


        # 去重
        new_green_cells = list(set(new_green_cells))


        # 统一标绿色字体（不覆盖已标绿色的单元格）
        green_cell_set = self._create_cell_set(green_cells)
        skipped_count = 0
        marked_count = 0


        for sheet_name, col, row in new_green_cells:
            if (sheet_name, col, row) in green_cell_set:
                skipped_count += 1
                continue


            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                self._set_cell_font_color(cell, "FF00FF00")  # 绿色
                marked_count += 1


        # 处理相同规则的cat2设备列标记
        for sheet_info in sheet_info_list:
            additional_cells = self._mark_same_cat2_rules(
                sheet_info, new_green_cells, "FF00FF00", None, output_workbook
            )
            new_green_cells.extend(additional_cells)


        # 合并到green_cells
        green_cells.extend(new_green_cells)
        green_cells = list(set(green_cells))


        # 每组3个单元格（cat1_a, cat2_a, cat2_b）
        cross_match_count = len([
            CELL for CELL in new_green_cells if CELL not in green_cell_set
        ]) // 3
        if cross_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤8 cat2 IP协议覆盖cat1 TCP协议检查完成："
                f"发现{cross_match_count}对匹配规则，"
                f"标记{marked_count}个单元格"
                f"（跳过{skipped_count}个已标绿色的单元格）"
            )
        elif len(new_green_cells) > 0:
            self.add_result(
                Level.WARN,
                f"步骤8发现{len(new_green_cells)}个匹配单元格，"
                f"但全部已被其他步骤标记（跳过{skipped_count}个）"
            )


        return green_cells


    # 步骤11：cat1与cat2匹配且cat2反向匹配检查（标绿色）：
    # 检查条件（三个条件必须全部满足）：
    # 1.Sheet A的cat1目的地址为special_network_map
    # 2.Sheet A的cat2规则覆盖Sheet A的cat1规则
    # 3.Sheet A的cat2规则与Sheet B的cat2规则反向匹配
    def _step_cat1_cat2_match_cat2_reverse_check(
        self, sheet_info_list, output_workbook, green_cells, rules_cache=None
    ):
        # 步骤11：cat1与cat2匹配且cat2反向匹配检查（标绿色）：检查条件（三个条件必须全部满足）：
        # 1.Sheet A的cat1目的地址为special_network_map 2.Sheet A的cat2规则覆盖Sheet A的cat1规则 3.Sheet A的cat2规则与Sheet B的cat2规则反向匹配
        new_green_cells = []


        if len(sheet_info_list) < 2:
            return green_cells


        # 收集所有special_network_map中的网络（任何Sheet的特殊网段）
        all_special_networks = []
        for SHEET_NETWORKS in self.SPECIAL_NETWORK_MAP.values():
            if SHEET_NETWORKS:
                all_special_networks.extend(SHEET_NETWORKS)


        # 对每两个不同的Sheet进行比较
        for sheet_index_a, sheet_info_a in enumerate(sheet_info_list):
            for sheet_index_b, sheet_info_b in enumerate(sheet_info_list):
                if sheet_index_a >= sheet_index_b:  # 避免重复比较
                    continue


                platform_network_a = sheet_info_a['platform_network']
                platform_network_b = sheet_info_b['platform_network']


                # 只比较不同平台的Sheet
                if not (platform_network_a and platform_network_b and
                        platform_network_a != platform_network_b):
                    continue


                sheet_name_a = sheet_info_a['sheet_name']
                sheet_name_b = sheet_info_b['sheet_name']


                # 使用规则缓存（性能优化）
                if rules_cache:
                    cat1_rules_a = rules_cache[sheet_name_a]['cat1']
                    cat2_rules_a = rules_cache[sheet_name_a]['cat2']
                    cat2_rules_b = rules_cache[sheet_name_b]['cat2']
                else:
                    # 兼容旧代码：如果没有缓存则重新收集
                    rule_row_mapping_a = sheet_info_a['rule_row_mapping']
                    rule_row_mapping_b = sheet_info_b['rule_row_mapping']
                    col_mapping_a = sheet_info_a['col_mapping']
                    col_mapping_b = sheet_info_b['col_mapping']
                    cat1_target_cols_a = sheet_info_a['cat1_target_cols']
                    cat2_target_cols_a = sheet_info_a['cat2_target_cols']
                    cat1_rules_a = self._collect_rules_from_cols(
                        cat1_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat2_rules_a = self._collect_rules_from_cols(
                        cat2_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat2_target_cols_b = sheet_info_b['cat2_target_cols']
                    cat2_rules_b = self._collect_rules_from_cols(
                        cat2_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )


                # 预构建Sheet A的cat1规则索引（满足条件1的cat1规则：目的地址在special_network_map中）
                # [(cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a), ...]
                valid_cat1_rules_a = []
                for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                    # 条件1：cat1目的地址在special_network_map中（任何Sheet的特殊网段）
                    dst_in_special = False
                    for special_net in all_special_networks:
                        if (cat1_rule_a.dst.overlaps(special_net) or
                                cat1_rule_a.dst.subnet_of(special_net)):
                            dst_in_special = True
                            break
                    if dst_in_special:
                        valid_cat1_rules_a.append(
                            (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a)
                        )


                # 预构建Sheet A内部cat2覆盖cat1的索引
                # 条件2：Sheet A的cat2规则覆盖Sheet A的cat1规则
                # {(cat1_col_a, cat1_row_a): [(cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a), ...]}
                cat2_cover_cat1_index_a = {}
                for cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a in valid_cat1_rules_a:
                    matches = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        # 检查cat2覆盖cat1（注意参数顺序：cat2覆盖cat1）
                        if rule_covers(cat2_rule_a, cat1_rule_a):
                            matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                    if matches:
                        cat2_cover_cat1_index_a[(cat1_col_a, cat1_row_a)] = matches


                # 预构建Sheet A和Sheet B的cat2反向匹配索引
                # 条件3：Sheet A的cat2规则与Sheet B的cat2规则反向匹配
                # {(cat2_col_a, cat2_row_a): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                cat2_reverse_index_ab = {}
                for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        if rule_reverse_matches(cat2_rule_a, cat2_rule_b):
                            matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat2_reverse_index_ab[(cat2_col_a, cat2_row_a)] = matches


                # 检查Sheet A -> Sheet B方向
                for CAT1_KEY_A, CAT1_DATA in cat2_cover_cat1_index_a.items():
                    CAT1_COL_A, CAT1_ROW_A = CAT1_KEY_A
                    for CAT2_COL_A, CAT2_ROW_A, CAT2_RULE_A, CAT2_RAW_A in CAT1_DATA:
                        CAT2_KEY_A = (CAT2_COL_A, CAT2_ROW_A)
                        if CAT2_KEY_A not in cat2_reverse_index_ab:
                            continue


                        # 所有条件满足，标记单元格
                        for (CAT2_COL_B, CAT2_ROW_B, CAT2_RULE_B, CAT2_RAW_B) in (
                                cat2_reverse_index_ab[CAT2_KEY_A]
                        ):
                            new_green_cells.append((sheet_name_a, CAT1_COL_A, CAT1_ROW_A))
                            new_green_cells.append((sheet_name_a, CAT2_COL_A, CAT2_ROW_A))
                            new_green_cells.append((sheet_name_b, CAT2_COL_B, CAT2_ROW_B))


                # 检查Sheet B -> Sheet A方向（双向检查）
                # 预构建Sheet B的cat1规则索引（满足条件1的cat1规则：目的地址在special_network_map中）
                valid_cat1_rules_b = []
                if rules_cache:
                    cat1_rules_b = rules_cache[sheet_name_b]['cat1']
                else:
                    cat1_target_cols_b = sheet_info_b['cat1_target_cols']
                    cat1_rules_b = self._collect_rules_from_cols(
                        cat1_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )


                for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                    # 条件1：cat1目的地址在special_network_map中
                    dst_in_special = False
                    for special_net in all_special_networks:
                        if (cat1_rule_b.dst.overlaps(special_net) or
                                cat1_rule_b.dst.subnet_of(special_net)):
                            dst_in_special = True
                            break
                    if dst_in_special:
                        valid_cat1_rules_b.append(
                            (cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b)
                        )


                # 预构建Sheet B内部cat2覆盖cat1的索引
                cat2_cover_cat1_index_b = {}
                for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in valid_cat1_rules_b:
                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        if rule_covers(cat2_rule_b, cat1_rule_b):
                            matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat2_cover_cat1_index_b[(cat1_col_b, cat1_row_b)] = matches


                # 预构建Sheet B和Sheet A的cat2反向匹配索引
                cat2_reverse_index_ba = {}
                for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                    matches = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        if rule_reverse_matches(cat2_rule_b, cat2_rule_a):
                            matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                    if matches:
                        cat2_reverse_index_ba[(cat2_col_b, cat2_row_b)] = matches


                # 检查Sheet B -> Sheet A方向
                for CAT1_KEY_B, CAT1_DATA in cat2_cover_cat1_index_b.items():
                    CAT1_COL_B, CAT1_ROW_B = CAT1_KEY_B
                    for CAT2_COL_B, CAT2_ROW_B, CAT2_RULE_B, CAT2_RAW_B in CAT1_DATA:
                        CAT2_KEY_B = (CAT2_COL_B, CAT2_ROW_B)
                        if CAT2_KEY_B not in cat2_reverse_index_ba:
                            continue


                        # 所有条件满足，标记单元格
                        for (CAT2_COL_A, CAT2_ROW_A, CAT2_RULE_A, CAT2_RAW_A) in (
                                cat2_reverse_index_ba[CAT2_KEY_B]
                        ):
                            new_green_cells.append((sheet_name_b, CAT1_COL_B, CAT1_ROW_B))
                            new_green_cells.append((sheet_name_b, CAT2_COL_B, CAT2_ROW_B))
                            new_green_cells.append((sheet_name_a, CAT2_COL_A, CAT2_ROW_A))


        # 去重
        new_green_cells = list(set(new_green_cells))


        # 统一标绿色字体（不覆盖已标绿色的单元格）
        green_cell_set = self._create_cell_set(green_cells)
        skipped_count = 0
        marked_count = 0


        for sheet_name, col, row in new_green_cells:
            if (sheet_name, col, row) in green_cell_set:
                skipped_count += 1
                continue


            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                self._set_cell_font_color(cell, "FF00FF00")  # 绿色
                marked_count += 1


        # 处理相同规则的cat2设备列标记
        for sheet_info in sheet_info_list:
            additional_cells = self._mark_same_cat2_rules(
                sheet_info, new_green_cells, "FF00FF00", None, output_workbook
            )
            new_green_cells.extend(additional_cells)


        # 合并到green_cells
        green_cells.extend(new_green_cells)
        green_cells = list(set(green_cells))


        # 每组3个单元格
        cross_match_count = len([
            cell for cell in new_green_cells if cell not in green_cell_set
        ]) // 3
        if cross_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤11 cat1与cat2匹配且cat2反向匹配检查完成："
                f"发现{cross_match_count}对匹配规则，"
                f"标记{marked_count}个单元格"
                f"（跳过{skipped_count}个已标绿色的单元格）"
            )
        elif len(new_green_cells) > 0:
            self.add_result(
                Level.WARN,
                f"步骤11发现{len(new_green_cells)}个匹配单元格，"
                f"但全部已被其他步骤标记（跳过{skipped_count}个）"
            )


        return green_cells


    # ========== 合并后的跨平台步骤4-6：cat6/cat2匹配检查（统一函数） ==========


    # 统一的cat6/cat2匹配检查函数（对应跨平台步骤4-6，内部step_num为11/12/13）：
    # - match_type为'complete'（完全匹配）、'cat6_cover'（cat6覆盖cat2）、'cat2_cover'（cat2覆盖cat6）
    # - green_cells为已标记的绿色单元格列表（用于避免覆盖）
    def _match_cat6_cat2_cross_platform(
            self, sheet_info_list, output_workbook, step_num,
            match_type, green_cells, rules_cache=None
    ):
        # 根据match_type选择内部匹配函数
        if match_type == 'complete':
            match_func_internal = rule_matches
            step_desc = "cat6完全匹配"
            match_desc = "cat6 vs cat2匹配相同（两个Sheet）"
            cell_order = 'cat6_first'  # cat6在前
        elif match_type == 'cat6_cover':
            match_func_internal = rule_covers
            step_desc = "cat6覆盖匹配"
            match_desc = "cat6 vs cat2覆盖（两个Sheet）"
            cell_order = 'cat6_first'  # cat6在前
        elif match_type == 'cat2_cover':
            match_func_internal = lambda cat6_rule, cat2_rule: rule_covers(cat2_rule, cat6_rule)
            step_desc = "cat2覆盖cat6匹配"
            match_desc = "cat2 vs cat6覆盖（两个Sheet）"
            cell_order = 'cat2_first'  # cat2在前
        else:
            raise ValueError(f"Invalid match_type: {match_type}")


        # 使用辅助方法进行匹配检查
        matched_pairs = self._cross_sheet_cat6_cat2_match_check_helper(
            sheet_info_list,
            match_func_internal=match_func_internal,
            match_func_cross=rule_reverse_matches,
            rules_cache=rules_cache
        )


        # 转换为单元格列表（根据cell_order决定顺序）
        new_cells = []
        for sheet_name_a, cat6_col_a, cat6_row_a, cat2_col_a, cat2_row_a, \
            sheet_name_b, cat6_col_b, cat6_row_b, cat2_col_b, cat2_row_b in matched_pairs:
            if cell_order == 'cat6_first':
                new_cells.append((sheet_name_a, cat6_col_a, cat6_row_a))
                new_cells.append((sheet_name_a, cat2_col_a, cat2_row_a))
                new_cells.append((sheet_name_b, cat6_col_b, cat6_row_b))
                new_cells.append((sheet_name_b, cat2_col_b, cat2_row_b))
            else:  # cat2_first
                new_cells.append((sheet_name_a, cat2_col_a, cat2_row_a))
                new_cells.append((sheet_name_a, cat6_col_a, cat6_row_a))
                new_cells.append((sheet_name_b, cat2_col_b, cat2_row_b))
                new_cells.append((sheet_name_b, cat6_col_b, cat6_row_b))


        cross_match_count = len(matched_pairs)


        # 使用统一的标记方法（蓝色，不覆盖绿色）
        marked_count, skipped_count, all_cells = self._mark_cells_with_color(
            new_cells, "0000FF", ["FF00FF00"], output_workbook,
            sheet_info_list, mark_same_cat2=True
        )


        # 过滤掉已标记的绿色单元格
        green_cell_set = self._create_cell_set(green_cells)
        filtered_cells = [cell for cell in all_cells if cell not in green_cell_set]


        if cross_match_count > 0:
            self.add_result(Level.OK,

                f"步骤{step_num}{step_desc}检查完成：发现{cross_match_count}对匹配规则"
                f"（标蓝色，四个条件全部满足：{match_desc} + cat6 vs cat6反向匹配相同 + cat2 vs cat2反向匹配相同）")


        # 返回合并后的单元格列表
        if step_num == 11:  # 第一个步骤，返回去重后的合并列表
            return list(set(green_cells + filtered_cells)) if green_cells else filtered_cells
        else:  # 后续步骤，直接追加
            return (green_cells or []) + filtered_cells


    # ========== 保持向后兼容的独立步骤函数（调用统一函数） ==========


    # 步骤11：cat6完全匹配检查（标蓝色）
    def _step_cat6_complete_match_check(
            self, sheet_info_list, output_workbook, green_cells,
            rules_cache=None
    ):
        return self._match_cat6_cat2_cross_platform(
            sheet_info_list, output_workbook, 11, 'complete', green_cells, rules_cache
        )


    # 步骤12：cat6覆盖匹配检查（标蓝色）
    def _step_cat6_cover_match_check(
            self, sheet_info_list, output_workbook, green_cells,
            rules_cache=None
    ):
        return self._match_cat6_cat2_cross_platform(
            sheet_info_list, output_workbook, 12, 'cat6_cover', green_cells, rules_cache
        )


    # 步骤13：cat2覆盖cat6匹配检查（标蓝色）
    def _step_cat2_cover_cat6_match_check(
            self, sheet_info_list, output_workbook, green_cells,
            rules_cache=None
    ):
        return self._match_cat6_cat2_cross_platform(
            sheet_info_list, output_workbook, 13, 'cat2_cover', green_cells, rules_cache
        )


    # 步骤15：cat6-cat1包含匹配（标蓝色）
    # 检查条件（四个条件必须全部满足）：
    # 1.Sheet A的cat6覆盖cat2 2.cat6与cat1双向包含/反向匹配
    # 3.cat2反向匹配 4.Sheet B的cat1覆盖cat2
    def _step_cat6_cat1_containment_check(
            self, sheet_info_list, output_workbook, green_cells,
            rules_cache=None
    ):
        existing_green_cell_set = self._create_cell_set(green_cells)
        new_blue_cells = []
        blue_match_count = 0


        if len(sheet_info_list) < 2:
            return green_cells


        # 使用规则缓存（性能优化）
        if rules_cache:
            sheet_rules_cache = rules_cache
        else:
            # 兼容旧代码：如果没有缓存则重新收集
            sheet_rules_cache = {}
            for sheet_info in sheet_info_list:
                sheet_name = sheet_info['sheet_name']
                rule_row_mapping = sheet_info['rule_row_mapping']
                col_mapping = sheet_info['col_mapping']


                sheet_rules_cache[sheet_name] = {
                    'cat6': self._collect_rules_from_cols(
                        sheet_info['cat6_target_cols'], col_mapping, rule_row_mapping
                    ),
                    'cat2': self._collect_rules_from_cols(
                        sheet_info['cat2_target_cols'], col_mapping, rule_row_mapping
                    ),
                    'cat1': self._collect_rules_from_cols(
                        sheet_info['cat1_target_cols'], col_mapping, rule_row_mapping
                    ),
                }


        total_pairs = sum(
            1 for SHEET_INDEX_A in range(len(sheet_info_list))
            for SHEET_INDEX_B in range(len(sheet_info_list))
            if SHEET_INDEX_A != SHEET_INDEX_B
        )
        current_pair = 0


        for SHEET_INDEX_A, sheet_info_a in enumerate(sheet_info_list):
            for SHEET_INDEX_B, sheet_info_b in enumerate(sheet_info_list):
                if SHEET_INDEX_A >= SHEET_INDEX_B:  # 优化：只检查一次，避免重复
                    continue


                current_pair += 1
                if current_pair % 5 == 0 or current_pair == total_pairs:
                    self.add_result(
                        Level.OK,
                        f"步骤12进度: {current_pair}/{total_pairs} 个Sheet对已检查，"
                        f"当前发现 {blue_match_count} 对匹配规则"
                    )


                sheet_name_a = sheet_info_a['sheet_name']
                sheet_name_b = sheet_info_b['sheet_name']
                platform_network_a = sheet_info_a['platform_network']
                platform_network_b = sheet_info_b['platform_network']


                if not (platform_network_a and platform_network_b and
                        platform_network_a != platform_network_b):
                    continue


                # 从缓存获取规则（双向）
                cat6_rules_a = sheet_rules_cache[sheet_name_a]['cat6']
                cat2_rules_a = sheet_rules_cache[sheet_name_a]['cat2']
                cat1_rules_a = sheet_rules_cache[sheet_name_a]['cat1']
                cat1_rules_b = sheet_rules_cache[sheet_name_b]['cat1']
                cat2_rules_b = sheet_rules_cache[sheet_name_b]['cat2']
                cat6_rules_b = sheet_rules_cache[sheet_name_b]['cat6']


                # 限制检查范围：如果组合数太大，只检查前N个
                MAX_COMBINATIONS = 100000  # 最多检查10万个组合
                total_combinations = (
                    (len(cat6_rules_a) * len(cat2_rules_a)) +
                    (len(cat6_rules_b) * len(cat2_rules_b))
                )
                if total_combinations > MAX_COMBINATIONS:
                    self.add_result(
                        Level.WARN,
                        f"步骤12: Sheet对 {sheet_name_a} <-> {sheet_name_b} "
                        f"组合数过多({total_combinations})，限制检查前{MAX_COMBINATIONS}个"
                    )
                    # 只取前部分规则
                    max_cat6_a = min(
                        len(cat6_rules_a),
                        MAX_COMBINATIONS // 2 // max(len(cat2_rules_a), 1)
                    )
                    max_cat6_b = min(
                        len(cat6_rules_b),
                        MAX_COMBINATIONS // 2 // max(len(cat2_rules_b), 1)
                    )
                    cat6_rules_a = cat6_rules_a[:max_cat6_a]
                    cat6_rules_b = cat6_rules_b[:max_cat6_b]
                    total_combinations = (
                        (len(cat6_rules_a) * len(cat2_rules_a)) +
                        (len(cat6_rules_b) * len(cat2_rules_b))
                    )


                # 预构建所有索引（优化：避免重复计算）
                # 1. cat2反向匹配索引（双向对称，只需构建一次）
                # {(cat2_col_a, cat2_row_a): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                cat2_reverse_index = {}
                for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        if rule_reverse_matches(cat2_rule_a, cat2_rule_b):
                            matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat2_reverse_index[(cat2_col_a, cat2_row_a)] = matches


                # 2. cat6与cat1的包含关系索引（方向1：A的cat6 vs B的cat1）
                # {cat6_rule_a: [(cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b), ...]}
                cat6_cat1_index_ab = {}
                for cat6_col_a, cat6_row_a, cat6_rule_a, cat6_raw_a in cat6_rules_a:
                    matches = []
                    for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                        if (
                            rule_covers(cat6_rule_a, cat1_rule_b)
                            or rule_covers(cat1_rule_b, cat6_rule_a)
                            or rule_reverse_matches(cat6_rule_a, cat1_rule_b)
                        ):
                            matches.append((cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b))
                    if matches:
                        cat6_cat1_index_ab[(cat6_col_a, cat6_row_a)] = matches


                # 3. cat6与cat1的包含关系索引（方向2：B的cat6 vs A的cat1）
                # {cat6_rule_b: [(cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a), ...]}
                cat6_cat1_index_ba = {}
                for cat6_col_b, cat6_row_b, cat6_rule_b, cat6_raw_b in cat6_rules_b:
                    matches = []
                    for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                        if (
                            rule_covers(cat6_rule_b, cat1_rule_a)
                            or rule_covers(cat1_rule_a, cat6_rule_b)
                            or rule_reverse_matches(cat6_rule_b, cat1_rule_a)
                        ):
                            matches.append((cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a))
                    if matches:
                        cat6_cat1_index_ba[(cat6_col_b, cat6_row_b)] = matches


                # 4. cat1覆盖cat2索引（Sheet A）
                # {(cat1_col_a, cat1_row_a): [(cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a), ...]}
                cat1_cat2_index_a = {}
                for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                    matches = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        if rule_covers(cat1_rule_a, cat2_rule_a):
                            matches.append((cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a))
                    if matches:
                        cat1_cat2_index_a[(cat1_col_a, cat1_row_a)] = matches


                # 5. cat1覆盖cat2索引（Sheet B）
                # {(cat1_col_b, cat1_row_b): [(cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b), ...]}
                cat1_cat2_index_b = {}
                for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                    matches = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        if rule_covers(cat1_rule_b, cat2_rule_b):
                            matches.append((cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b))
                    if matches:
                        cat1_cat2_index_b[(cat1_col_b, cat1_row_b)] = matches


                # 使用预构建索引进行快速匹配（方向1：A的cat6覆盖A的cat2，A的cat6与B的cat1匹配，A的cat2与B的cat2反向匹配，B的cat1覆盖B的cat2）
                checked_combinations = 0
                for cat6_col_a, cat6_row_a, cat6_rule_a, cat6_raw_a in cat6_rules_a:
                    cat6_key = (cat6_col_a, cat6_row_a)
                    # 如果cat6没有匹配的cat1，跳过
                    if cat6_key not in cat6_cat1_index_ab:
                        continue
                    matched_cat1_b_list = cat6_cat1_index_ab[cat6_key]


                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        checked_combinations += 1
                        if checked_combinations % 5000 == 0:
                            self.add_result(
                                Level.OK,
                                f"步骤12进度: {sheet_name_a} <-> {sheet_name_b} (方向1): "
                                f"{checked_combinations}/{total_combinations}，"
                                f"已发现 {blue_match_count} 对"
                            )


                        # 条件1：cat6覆盖cat2
                        if not rule_covers(cat6_rule_a, cat2_rule_a):
                            continue


                        # 条件3：cat2反向匹配（使用预构建索引）
                        cat2_key = (cat2_col_a, cat2_row_a)
                        if cat2_key not in cat2_reverse_index:
                            continue
                        matched_cat2_b_list = cat2_reverse_index[cat2_key]


                        # 条件2和4：使用预构建索引快速匹配
                        found_match = False
                        for matched_cat1_b in matched_cat1_b_list:
                            cat1_key = (matched_cat1_b[0], matched_cat1_b[1])
                            if cat1_key not in cat1_cat2_index_b:
                                continue
                            # 检查cat1覆盖的cat2是否在cat2反向匹配列表中
                            cat1_covered_cat2_list = cat1_cat2_index_b[cat1_key]
                            for matched_cat2_b in matched_cat2_b_list:
                                # 检查matched_cat2_b是否在cat1覆盖的cat2列表中
                                if any(
                                    matched_cat2_b[0] == COVERED_CAT2[0] and
                                    matched_cat2_b[1] == COVERED_CAT2[1]
                                    for COVERED_CAT2 in cat1_covered_cat2_list
                                ):
                                    new_blue_cells.append((sheet_name_a, cat6_col_a, cat6_row_a))
                                    new_blue_cells.append((sheet_name_a, cat2_col_a, cat2_row_a))
                                    new_blue_cells.append(
                                        (sheet_name_b, matched_cat1_b[0], matched_cat1_b[1])
                                    )
                                    new_blue_cells.append(
                                        (sheet_name_b, matched_cat2_b[0], matched_cat2_b[1])
                                    )
                                    blue_match_count += 1
                                    found_match = True
                                    break  # 找到一个匹配即可
                            if found_match:
                                break  # 已找到匹配，跳出cat1循环
                        if found_match:
                            continue  # 继续下一个cat2规则


                # 使用预构建索引进行快速匹配（方向2：B的cat6覆盖B的cat2，B的cat6与A的cat1匹配，B的cat2与A的cat2反向匹配，A的cat1覆盖A的cat2）
                for cat6_col_b, cat6_row_b, cat6_rule_b, cat6_raw_b in cat6_rules_b:
                    cat6_key = (cat6_col_b, cat6_row_b)
                    # 如果cat6没有匹配的cat1，跳过
                    if cat6_key not in cat6_cat1_index_ba:
                        continue
                    matched_cat1_a_list = cat6_cat1_index_ba[cat6_key]


                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        checked_combinations += 1
                        if checked_combinations % 5000 == 0:
                            self.add_result(
                                Level.OK,
                                f"步骤12进度: {sheet_name_a} <-> {sheet_name_b} (方向2): "
                                f"{checked_combinations}/{total_combinations}，"
                                f"已发现 {blue_match_count} 对"
                            )


                        # 条件1：cat6覆盖cat2
                        if not rule_covers(cat6_rule_b, cat2_rule_b):
                            continue


                        # 条件3：cat2反向匹配（使用预构建索引，注意方向）
                        # 需要找到A的cat2规则，使得B的cat2与A的cat2反向匹配
                        cat2_key_b = (cat2_col_b, cat2_row_b)
                        matched_cat2_a_list = []
                        for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                            if rule_reverse_matches(cat2_rule_b, cat2_rule_a):
                                matched_cat2_a_list.append(
                                    (cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a)
                                )


                        if not matched_cat2_a_list:
                            continue


                        # 条件2和4：使用预构建索引快速匹配
                        found_match = False
                        for matched_cat1_a in matched_cat1_a_list:
                            cat1_key = (matched_cat1_a[0], matched_cat1_a[1])
                            if cat1_key not in cat1_cat2_index_a:
                                continue
                            # 检查cat1覆盖的cat2是否在cat2反向匹配列表中
                            cat1_covered_cat2_list = cat1_cat2_index_a[cat1_key]
                            for matched_cat2_a in matched_cat2_a_list:
                                # 检查matched_cat2_a是否在cat1覆盖的cat2列表中
                                if any(
                                    matched_cat2_a[0] == COVERED_CAT2[0] and
                                    matched_cat2_a[1] == COVERED_CAT2[1]
                                    for COVERED_CAT2 in cat1_covered_cat2_list
                                ):
                                    new_blue_cells.append((sheet_name_b, cat6_col_b, cat6_row_b))
                                    new_blue_cells.append((sheet_name_b, cat2_col_b, cat2_row_b))
                                    new_blue_cells.append(
                                        (sheet_name_a, matched_cat1_a[0], matched_cat1_a[1])
                                    )
                                    new_blue_cells.append(
                                        (sheet_name_a, matched_cat2_a[0], matched_cat2_a[1])
                                    )
                                    blue_match_count += 1
                                    found_match = True
                                    break  # 找到一个匹配即可
                            if found_match:
                                break  # 已找到匹配，跳出cat1循环
                        if found_match:
                            continue  # 继续下一个cat2规则


        # 标记新发现的蓝色单元格（不覆盖已标绿色和蓝色的单元格）
        for sheet_name, col, row in new_blue_cells:
            if (sheet_name, col, row) in existing_green_cell_set:
                continue
            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                if self._check_color_priority(cell, ["FF00FF00", "0000FF"]):  # 不覆盖绿色和蓝色
                    self._set_cell_font_color(cell, "0000FF")  # 蓝色


        if blue_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤12包含匹配检查完成：发现{blue_match_count}对匹配规则"
                f"（标蓝色：A.cat6覆盖A.cat2 + A.cat6与B.cat1双向包含/反向匹配 + "
                f"A.cat2与B.cat2反向匹配 + B.cat1覆盖B.cat2）"
            )


        # 返回合并后的绿色单元格列表（步骤12标蓝色，不添加到green_cells）
        return green_cells


    # 步骤7：多个cat1规则覆盖cat2规则检查（标绿色）
    # 检查条件（四个条件必须全部满足）：
    # 1.多个cat1规则一起覆盖cat2（两个Sheet） 2.cat1反向匹配 3.cat2反向匹配
    def _step_multi_cat1_cover_cat2_check(
            self, sheet_info_list, output_workbook, green_cells,
            rules_cache=None
    ):
        # 步骤7：多个cat1规则覆盖cat2规则检查（标绿色）：检查条件（四个条件必须全部满足）：1.多个cat1规则一起覆盖cat2（两个Sheet） 2.cat1反向匹配 3.cat2反向匹配
        new_green_cells = []
        green_match_count = 0


        if len(sheet_info_list) < 2:
            return green_cells if green_cells else []


        for SHEET_INDEX_A, sheet_info_a in enumerate(sheet_info_list):
            for SHEET_INDEX_B, sheet_info_b in enumerate(sheet_info_list):
                if SHEET_INDEX_A >= SHEET_INDEX_B:
                    continue


                platform_network_a = sheet_info_a['platform_network']
                platform_network_b = sheet_info_b['platform_network']


                if not (platform_network_a and platform_network_b and
                        platform_network_a != platform_network_b):
                    continue


                sheet_name_a = sheet_info_a['sheet_name']
                sheet_name_b = sheet_info_b['sheet_name']


                # 使用规则缓存（性能优化）
                if rules_cache:
                    cat1_rules_a = rules_cache[sheet_name_a]['cat1']
                    cat2_rules_a = rules_cache[sheet_name_a]['cat2']
                    cat1_rules_b = rules_cache[sheet_name_b]['cat1']
                    cat2_rules_b = rules_cache[sheet_name_b]['cat2']
                else:
                    # 兼容旧代码：如果没有缓存则重新收集
                    rule_row_mapping_a = sheet_info_a['rule_row_mapping']
                    rule_row_mapping_b = sheet_info_b['rule_row_mapping']
                    col_mapping_a = sheet_info_a['col_mapping']
                    col_mapping_b = sheet_info_b['col_mapping']
                    cat1_target_cols_a = sheet_info_a['cat1_target_cols']
                    cat2_target_cols_a = sheet_info_a['cat2_target_cols']
                    cat1_rules_a = self._collect_rules_from_cols(
                        cat1_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat2_rules_a = self._collect_rules_from_cols(
                        cat2_target_cols_a, col_mapping_a, rule_row_mapping_a
                    )
                    cat1_target_cols_b = sheet_info_b['cat1_target_cols']
                    cat2_target_cols_b = sheet_info_b['cat2_target_cols']
                    cat1_rules_b = self._collect_rules_from_cols(
                        cat1_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )
                    cat2_rules_b = self._collect_rules_from_cols(
                        cat2_target_cols_b, col_mapping_b, rule_row_mapping_b
                    )


                # 检查Sheet A -> Sheet B方向
                for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                    # 收集所有覆盖这个cat2规则的cat1规则
                    # 注意：这里需要检查cat1的目的地址是cat2目的地址的子网（反向覆盖）
                    covering_cat1_rules_a = []
                    for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                        # 检查基本条件：协议、动作、源地址覆盖
                        if cat1_rule_a.action != cat2_rule_a.action:
                            continue
                        if not proto_covers(cat1_rule_a.proto, cat2_rule_a.proto):
                            continue
                        # cat2的源地址必须是cat1的源地址的子网（或相等）
                        if (not cat2_rule_a.src.subnet_of(cat1_rule_a.src) and
                                cat2_rule_a.src != cat1_rule_a.src):
                            continue
                        # cat1的目的地址必须是cat2的目的地址的子网（反向覆盖）
                        if (not cat1_rule_a.dst.subnet_of(cat2_rule_a.dst) and
                                cat1_rule_a.dst != cat2_rule_a.dst):
                            continue
                        covering_cat1_rules_a.append(
                            (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a)
                        )


                    if not covering_cat1_rules_a:
                        continue


                    # 检查cat2规则是否有端口范围或多端口
                    cat2_ports_a = (
                        cat2_rule_a.ports
                        if cat2_rule_a.ports and len(cat2_rule_a.ports) > 0
                        else None
                    )
                    if not cat2_ports_a:
                        # 尝试从原始文本提取端口
                        cat2_ports_a = _extract_ports_from_iosxe_rule(cat2_raw_a)


                    # 收集所有覆盖cat2规则的cat1规则的端口
                    covered_ports_a = set()
                    for cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a in covering_cat1_rules_a:
                        # 对于range端口，需要从原始文本提取
                        if "range" in cat1_raw_a.lower():
                            cat1_ports = _extract_ports_from_iosxe_rule(cat1_raw_a)
                            covered_ports_a.update(cat1_ports)
                        else:
                            category1_destination_port = (
                                cat1_rule_a.dst_port if cat1_rule_a.dst_port is not None
                                else cat1_rule_a.port
                            )
                            if category1_destination_port is not None:
                                covered_ports_a.add(category1_destination_port)
                            elif cat1_rule_a.ports:
                                covered_ports_a.update(cat1_rule_a.ports)


                    # 检查是否所有cat2的端口都被cat1规则覆盖
                    if cat2_ports_a and not covered_ports_a.issuperset(cat2_ports_a):
                        continue


                    # ========== 性能优化：预构建匹配索引 ==========
                    # 预构建LINK-AS反向匹配索引（A -> B）
                    cat2_key_a = (cat2_col_a, cat2_row_a)
                    matched_cat2_b_list = []
                    for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                        if rule_reverse_matches(cat2_rule_a, cat2_rule_b):
                            matched_cat2_b_list.append(
                                (cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b)
                            )


                    if not matched_cat2_b_list:
                        continue


                    # 预构建CS-N9K反向匹配索引（A -> B）
                    # {(cat1_col_a, cat1_row_a): [(cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b), ...]}
                    cat1_reverse_index_ab = {}
                    for cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a in covering_cat1_rules_a:
                        matches = []
                        for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                            if rule_reverse_matches(cat1_rule_a, cat1_rule_b):
                                matches.append((cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b))
                        if matches:
                            cat1_reverse_index_ab[(cat1_col_a, cat1_row_a)] = matches


                    if not cat1_reverse_index_ab:
                        continue


                    # 使用预构建索引查找匹配
                    matched_cat1_rules_b = []
                    for cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a in covering_cat1_rules_a:
                        cat1_key_a = (cat1_col_a, cat1_row_a)
                        if cat1_key_a in cat1_reverse_index_ab:
                            matched_cat1_rules_b.extend(cat1_reverse_index_ab[cat1_key_a])


                    if not matched_cat1_rules_b:
                        continue


                    # 对每个匹配的cat2_b检查
                    for matched_cat2_b in matched_cat2_b_list:
                        # 条件4：Sheet B的多个cat1规则一起覆盖Sheet B的cat2规则
                        cat2_ports_b = (
                            matched_cat2_b[2].ports
                            if matched_cat2_b[2].ports and len(matched_cat2_b[2].ports) > 0
                            else None
                        )
                        if not cat2_ports_b:
                            cat2_ports_b = _extract_ports_from_iosxe_rule(
                                matched_cat2_b[3] if len(matched_cat2_b) > 3 else ""
                            )


                        covered_ports_b = set()
                        for (cat1_col_b, cat1_row_b, cat1_rule_b,
                                cat1_raw_b) in matched_cat1_rules_b:
                            # 检查cat1规则是否覆盖cat2规则（使用相同的反向覆盖逻辑）
                            cat2_rule_b = matched_cat2_b[2]
                            if (cat1_rule_b.action == cat2_rule_b.action and
                                proto_covers(cat1_rule_b.proto, cat2_rule_b.proto) and
                                (cat1_rule_b.src.subnet_of(cat2_rule_b.src) or
                                    cat1_rule_b.src == cat2_rule_b.src) and
                                (cat1_rule_b.dst.subnet_of(cat2_rule_b.dst) or
                                    cat1_rule_b.dst == cat2_rule_b.dst)
                            ):
                                # 对于range端口，需要从原始文本提取
                                if "range" in cat1_raw_b.lower():
                                    cat1_ports = _extract_ports_from_iosxe_rule(cat1_raw_b)
                                    covered_ports_b.update(cat1_ports)
                                else:
                                    category1_destination_port = (
                                        cat1_rule_b.dst_port
                                        if cat1_rule_b.dst_port is not None
                                        else cat1_rule_b.port
                                    )
                                    if category1_destination_port is not None:
                                        covered_ports_b.add(category1_destination_port)
                                    elif cat1_rule_b.ports:
                                        covered_ports_b.update(cat1_rule_b.ports)


                        if cat2_ports_b and not covered_ports_b.issuperset(cat2_ports_b):
                            continue


                        # 所有条件满足，标记所有相关规则
                        for cat1_col_a, cat1_row_a, _, _ in covering_cat1_rules_a:
                            new_green_cells.append(
                                (sheet_info_a['sheet_name'], cat1_col_a, cat1_row_a)
                            )
                        new_green_cells.append(
                            (sheet_info_a['sheet_name'], cat2_col_a, cat2_row_a)
                        )


                        for cat1_col_b, cat1_row_b, _, _ in matched_cat1_rules_b:
                            new_green_cells.append(
                                (sheet_info_b['sheet_name'], cat1_col_b, cat1_row_b)
                            )
                        new_green_cells.append(
                            (sheet_info_b['sheet_name'], matched_cat2_b[0], matched_cat2_b[1])
                        )


                        green_match_count += 1
                        break  # 找到一个匹配的cat2_b即可


                # 检查Sheet B -> Sheet A方向（双向检查）
                for cat2_col_b, cat2_row_b, cat2_rule_b, cat2_raw_b in cat2_rules_b:
                    covering_cat1_rules_b = []
                    for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in cat1_rules_b:
                        # 检查基本条件：协议、动作、源地址覆盖
                        if cat1_rule_b.action != cat2_rule_b.action:
                            continue
                        if not proto_covers(cat1_rule_b.proto, cat2_rule_b.proto):
                            continue
                        # cat2的源地址必须是cat1的源地址的子网（或相等）
                        if (not cat2_rule_b.src.subnet_of(cat1_rule_b.src) and
                                cat2_rule_b.src != cat1_rule_b.src):
                            continue
                        # cat1的目的地址必须是cat2的目的地址的子网（反向覆盖）
                        if (not cat1_rule_b.dst.subnet_of(cat2_rule_b.dst) and
                                cat1_rule_b.dst != cat2_rule_b.dst):
                            continue
                        covering_cat1_rules_b.append(
                            (cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b)
                        )


                    if not covering_cat1_rules_b:
                        continue


                    cat2_ports_b = (
                        cat2_rule_b.ports
                        if cat2_rule_b.ports and len(cat2_rule_b.ports) > 0
                        else None
                    )
                    if not cat2_ports_b:
                        cat2_ports_b = _extract_ports_from_iosxe_rule(cat2_raw_b)


                    covered_ports_b = set()
                    for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in covering_cat1_rules_b:
                        # 对于range端口，需要从原始文本提取
                        if "range" in cat1_raw_b.lower():
                            cat1_ports = _extract_ports_from_iosxe_rule(cat1_raw_b)
                            covered_ports_b.update(cat1_ports)
                        else:
                            category1_destination_port = (
                                cat1_rule_b.dst_port
                                if cat1_rule_b.dst_port is not None
                                else cat1_rule_b.port
                            )
                            if category1_destination_port is not None:
                                covered_ports_b.add(category1_destination_port)
                            elif cat1_rule_b.ports:
                                covered_ports_b.update(cat1_rule_b.ports)


                    if cat2_ports_b and not covered_ports_b.issuperset(cat2_ports_b):
                        continue


                    # ========== 性能优化：预构建匹配索引（B -> A） ==========
                    # 预构建LINK-AS反向匹配索引（B -> A）
                    cat2_key_b = (cat2_col_b, cat2_row_b)
                    matched_cat2_a_list = []
                    for cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a in cat2_rules_a:
                        if rule_reverse_matches(cat2_rule_b, cat2_rule_a):
                            matched_cat2_a_list.append(
                                (cat2_col_a, cat2_row_a, cat2_rule_a, cat2_raw_a)
                            )


                    if not matched_cat2_a_list:
                        continue


                    # 预构建CS-N9K反向匹配索引（B -> A）
                    # {(cat1_col_b, cat1_row_b): [(cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a), ...]}
                    cat1_reverse_index_ba = {}
                    for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in covering_cat1_rules_b:
                        matches = []
                        for (cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a) in cat1_rules_a:
                            if rule_reverse_matches(cat1_rule_b, cat1_rule_a):
                                matches.append((cat1_col_a, cat1_row_a, cat1_rule_a, cat1_raw_a))
                        if matches:
                            cat1_reverse_index_ba[(cat1_col_b, cat1_row_b)] = matches


                    if not cat1_reverse_index_ba:
                        continue


                    # 使用预构建索引查找匹配
                    matched_cat1_rules_a = []
                    for cat1_col_b, cat1_row_b, cat1_rule_b, cat1_raw_b in covering_cat1_rules_b:
                        cat1_key_b = (cat1_col_b, cat1_row_b)
                        if cat1_key_b in cat1_reverse_index_ba:
                            matched_cat1_rules_a.extend(cat1_reverse_index_ba[cat1_key_b])


                    if not matched_cat1_rules_a:
                        continue


                    # 对每个匹配的cat2_a检查
                    for matched_cat2_a in matched_cat2_a_list:


                        # 条件4：Sheet A的多个cat1规则一起覆盖Sheet A的cat2规则
                        cat2_ports_a = (
                            matched_cat2_a[2].ports
                            if matched_cat2_a[2].ports and len(matched_cat2_a[2].ports) > 0
                            else None
                        )
                        if not cat2_ports_a:
                            cat2_ports_a = _extract_ports_from_iosxe_rule(
                                matched_cat2_a[3] if len(matched_cat2_a) > 3 else ""
                            )


                        covered_ports_a = set()
                        for (cat1_col_a, cat1_row_a, cat1_rule_a,
                                cat1_raw_a) in matched_cat1_rules_a:
                            # 检查cat1规则是否覆盖cat2规则（使用相同的反向覆盖逻辑）
                            cat2_rule_a = matched_cat2_a[2]
                            if (cat1_rule_a.action == cat2_rule_a.action and
                                proto_covers(cat1_rule_a.proto, cat2_rule_a.proto) and
                                (cat1_rule_a.src.subnet_of(cat2_rule_a.src) or
                                    cat1_rule_a.src == cat2_rule_a.src) and
                                (cat1_rule_a.dst.subnet_of(cat2_rule_a.dst) or
                                    cat1_rule_a.dst == cat2_rule_a.dst)
                            ):
                                # 对于range端口，需要从原始文本提取
                                if "range" in cat1_raw_a.lower():
                                    cat1_ports = _extract_ports_from_iosxe_rule(cat1_raw_a)
                                    covered_ports_a.update(cat1_ports)
                                else:
                                    category1_destination_port = (
                                cat1_rule_a.dst_port if cat1_rule_a.dst_port is not None
                                else cat1_rule_a.port
                            )
                                    if category1_destination_port is not None:
                                        covered_ports_a.add(category1_destination_port)
                                    elif cat1_rule_a.ports:
                                        covered_ports_a.update(cat1_rule_a.ports)


                        if cat2_ports_a and not covered_ports_a.issuperset(cat2_ports_a):
                            continue


                        # 所有条件满足，标记所有相关规则
                        for cat1_col_b, cat1_row_b, _, _ in covering_cat1_rules_b:
                            new_green_cells.append(
                                (sheet_info_b['sheet_name'], cat1_col_b, cat1_row_b)
                            )
                        new_green_cells.append(
                            (sheet_info_b['sheet_name'], cat2_col_b, cat2_row_b)
                        )


                        for cat1_col_a, cat1_row_a, _, _ in matched_cat1_rules_a:
                            new_green_cells.append(
                                (sheet_info_a['sheet_name'], cat1_col_a, cat1_row_a)
                            )
                        new_green_cells.append(
                            (sheet_info_a['sheet_name'], matched_cat2_a[0], matched_cat2_a[1])
                        )


                        green_match_count += 1
                        break  # 找到一个匹配的cat2_a即可


        # 标记绿色单元格（不覆盖已标绿色的单元格）
        green_cell_set = self._create_cell_set(green_cells)
        for sheet_name, col, row in new_green_cells:
            # 检查是否已被绿色标记（不覆盖已标绿色的单元格）
            if (sheet_name, col, row) in green_cell_set:
                continue
            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                # 检查是否已标绿色，避免覆盖
                if self._check_color_priority(cell, ["FF00FF00"]):
                    self._set_cell_font_color(cell, "FF00FF00")  # 绿色


        # 处理相同规则的cat2设备列标记
        for sheet_info in sheet_info_list:
            additional_cells = self._mark_same_cat2_rules(
                sheet_info, new_green_cells, "FF00FF00", ["FF00FF00"], output_workbook  # 绿色
            )
            new_green_cells.extend(additional_cells)


        if green_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤7多个cat1覆盖cat2检查完成：发现{green_match_count}对匹配规则"
                f"（标绿色：多个cat1规则一起覆盖cat2规则（两个Sheet）+ "
                f"cat1反向匹配 + cat2反向匹配）"
            )


        # 返回合并后的绿色单元格列表
        return list(set(green_cells + new_green_cells))


    # 步骤16：平台外覆盖检查（标橙色）
    # 检查同Sheet内，源地址为本平台地址、目的为非platform_network_map地址的cat1覆盖cat2规则
    def _step_platform_outside_check(
            self, sheet_info_list, output_workbook, green_cells,
            rules_cache=None
    ):
        # 步骤16：平台外覆盖检查（标橙色）：检查同Sheet内，源地址为本平台地址、目的为非platform_network_map地址的cat1覆盖cat2规则
        yellow_cells = []
        yellow_match_count = 0


        all_platform_networks = []
        for SHEET_NETWORKS in self.PLATFORM_NETWORK_MAP.values():
            if SHEET_NETWORKS:
                all_platform_networks.extend(SHEET_NETWORKS)


        total_sheets = len(sheet_info_list)
        current_sheet = 0
        for sheet_info in sheet_info_list:
            current_sheet += 1
            if current_sheet % 5 == 0 or current_sheet == total_sheets:
                self.add_result(
                    Level.OK,
                    f"步骤13进度: {current_sheet}/{total_sheets} 个Sheet已检查，"
                    f"当前发现 {yellow_match_count} 对匹配规则"
                )


            sheet_name = sheet_info['sheet_name']
            platform_networks = sheet_info['platform_networks']


            if not platform_networks:
                continue


            # 使用规则缓存（性能优化）
            if rules_cache:
                cat1_rules = rules_cache[sheet_name]['cat1']
                cat2_rules = rules_cache[sheet_name]['cat2']
            else:
                # 兼容旧代码：如果没有缓存则重新收集
                rule_row_mapping = sheet_info['rule_row_mapping']
                col_mapping = sheet_info['col_mapping']
                cat1_target_cols = sheet_info['cat1_target_cols']
                cat2_target_cols = sheet_info['cat2_target_cols']
                cat1_rules = self._collect_rules_from_cols(
                    cat1_target_cols, col_mapping, rule_row_mapping
                )
                cat2_rules = self._collect_rules_from_cols(
                    cat2_target_cols, col_mapping, rule_row_mapping
                )


            total_cat2_rules = len(cat2_rules)
            checked_cat2_rules = 0
            for cat2_col, cat2_row, cat2_rule, cat2_raw in cat2_rules:
                checked_cat2_rules += 1
                if total_cat2_rules > 500 and checked_cat2_rules % 500 == 0:
                    self.add_result(
                        Level.OK,
                        f"步骤13进度: Sheet {sheet_name}: "
                        f"{checked_cat2_rules}/{total_cat2_rules} 个cat2规则已检查"
                    )
                cat2_src_in_platform = self._network_in_platform(
                    cat2_rule.src, platform_networks
                )
                cat2_dst_in_any_platform = self._network_in_platform(
                    cat2_rule.dst, all_platform_networks
                )


                if not cat2_src_in_platform or cat2_dst_in_any_platform:
                    continue


                covering_cat1_rules = []
                partial_covering_cat1_rules = []


                for (cat1_col, cat1_row, cat1_rule, cat1_raw) in cat1_rules:
                    cat1_src_in_platform = self._network_in_platform(
                        cat1_rule.src, platform_networks
                    )
                    cat1_dst_in_any_platform = self._network_in_platform(
                        cat1_rule.dst, all_platform_networks
                    )


                    if not cat1_src_in_platform or cat1_dst_in_any_platform:
                        continue


                    if rule_covers(cat1_rule, cat2_rule):
                        covering_cat1_rules.append(
                            (cat1_col, cat1_row, cat1_rule, cat1_raw)
                        )
                    elif rule_port_in_cat2_ports(cat1_rule, cat2_rule):
                        partial_covering_cat1_rules.append(
                            (cat1_col, cat1_row, cat1_rule, cat1_raw)
                        )


                marked_cat1_cells = set()
                marked_cat2_cell = False


                if covering_cat1_rules:
                    for cat1_col, cat1_row, cat1_rule, cat1_raw in covering_cat1_rules:
                        marked_cat1_cells.add((cat1_col, cat1_row))
                        yellow_cells.append((sheet_name, cat1_col, cat1_row))
                    yellow_cells.append((sheet_name, cat2_col, cat2_row))
                    marked_cat2_cell = True
                    yellow_match_count += 1


                if partial_covering_cat1_rules:
                    cat2_ports = (
                        cat2_rule.ports
                        if cat2_rule.ports and len(cat2_rule.ports) > 0
                        else None
                    )
                    if not cat2_ports:
                        cat2_ports = _extract_ports_from_iosxe_rule(cat2_raw)
                    if cat2_ports:
                        covered_ports = set()
                        for (cat1_col, cat1_row, cat1_rule, cat1_raw) in covering_cat1_rules:
                            category1_destination_port = (
                                cat1_rule.dst_port
                                if cat1_rule.dst_port is not None
                                else cat1_rule.port
                            )
                            if category1_destination_port is not None:
                                covered_ports.add(category1_destination_port)
                        for (cat1_col, cat1_row, cat1_rule,
                                cat1_raw) in partial_covering_cat1_rules:
                            category1_destination_port = (
                                cat1_rule.dst_port
                                if cat1_rule.dst_port is not None
                                else cat1_rule.port
                            )
                            if category1_destination_port is not None:
                                covered_ports.add(category1_destination_port)


                        if covered_ports >= cat2_ports:
                            for (cat1_col, cat1_row, cat1_rule,
                                cat1_raw) in partial_covering_cat1_rules:
                                if (cat1_col, cat1_row) not in marked_cat1_cells:
                                    marked_cat1_cells.add((cat1_col, cat1_row))
                                    yellow_cells.append((sheet_name, cat1_col, cat1_row))
                            if not marked_cat2_cell:
                                yellow_cells.append((sheet_name, cat2_col, cat2_row))
                                yellow_match_count += 1


        # 统一标橙色字体（不覆盖已标绿色的单元格）
        green_cell_set = self._create_cell_set(green_cells)
        for sheet_name, col, row in yellow_cells:
            if (sheet_name, col, row) in green_cell_set:
                continue


            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                if self._check_color_priority(cell, ["FF00FF00"]):
                    self._set_cell_font_color(cell, "FFA500")  # 橙色


        # 处理相同规则的cat2设备列标记
        for sheet_info in sheet_info_list:
            self._mark_same_cat2_rules(
                sheet_info, yellow_cells, "FFA500", ["FF00FF00"], output_workbook
            )


        if yellow_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤13平台外覆盖检查完成：发现{yellow_match_count}对cat1覆盖cat2规则"
                f"（源为本平台，目的为非platform_network_map地址，cat1和cat2都标橙色）"
            )


        return yellow_cells


    # 步骤17：特殊规则检查（标橙色）
    # 检查条件：1.cat1源地址为本平台地址 2.cat1目的地址为特殊规则定义IP段地址
    # 3.cat1规则覆盖cat2规则 4.跨平台cat2反向检查通过
    def _step_special_rule_check(
            self, sheet_info_list, output_workbook, green_cells, yellow_cells,
            rules_cache=None
    ):
        # 步骤17：特殊规则检查（标橙色）：检查条件：1.cat1源地址为本平台地址 2.cat1目的地址为特殊规则定义IP段地址 3.cat1规则覆盖cat2规则 4.跨平台cat2反向检查通过
        gray_cells = []
        gray_match_count = 0


        all_special_networks = []
        for SHEET_NETWORKS in self.SPECIAL_NETWORK_MAP.values():
            if SHEET_NETWORKS:
                all_special_networks.extend(SHEET_NETWORKS)


        total_sheets = len(sheet_info_list)
        current_sheet = 0
        for sheet_info in sheet_info_list:
            current_sheet += 1
            if current_sheet % 5 == 0 or current_sheet == total_sheets:
                self.add_result(
                    Level.OK,
                    f"步骤14进度: {current_sheet}/{total_sheets} 个Sheet已检查，"
                    f"当前发现 {gray_match_count} 对匹配规则"
                )


            sheet_name = sheet_info['sheet_name']
            platform_networks = sheet_info['platform_networks']


            if not platform_networks:
                continue


            # 使用规则缓存（性能优化）
            if rules_cache:
                cat1_rules = rules_cache[sheet_name]['cat1']
                cat2_rules = rules_cache[sheet_name]['cat2']
            else:
                # 兼容旧代码：如果没有缓存则重新收集
                rule_row_mapping = sheet_info['rule_row_mapping']
                col_mapping = sheet_info['col_mapping']
                cat1_target_cols = sheet_info['cat1_target_cols']
                cat2_target_cols = sheet_info['cat2_target_cols']
                cat1_rules = self._collect_rules_from_cols(
                    cat1_target_cols, col_mapping, rule_row_mapping
                )
                cat2_rules = self._collect_rules_from_cols(
                    cat2_target_cols, col_mapping, rule_row_mapping
                )


            total_cat1_rules = len(cat1_rules)
            checked_cat1_rules = 0
            for cat1_col, cat1_row, cat1_rule, cat1_raw in cat1_rules:
                checked_cat1_rules += 1
                if total_cat1_rules > 500 and checked_cat1_rules % 500 == 0:
                    self.add_result(
                        Level.OK,
                        f"步骤14进度: Sheet {sheet_name}: "
                        f"{checked_cat1_rules}/{total_cat1_rules} 个cat1规则已检查"
                    )
                cat1_src_in_platform = self._network_in_platform(
                    cat1_rule.src, platform_networks
                )
                cat1_dst_in_special = self._network_in_platform(
                    cat1_rule.dst, all_special_networks
                )


                if not cat1_src_in_platform or not cat1_dst_in_special:
                    continue


                for cat2_col, cat2_row, cat2_rule, cat2_raw in cat2_rules:
                    if not rule_covers(cat1_rule, cat2_rule):
                        continue


                    cat2_reverse_match_found = False
                    for other_sheet_info in sheet_info_list:
                        if other_sheet_info['sheet_name'] == sheet_name:
                            continue


                        other_platform_networks = other_sheet_info['platform_networks']
                        if not other_platform_networks:
                            continue


                        if platform_networks == other_platform_networks:
                            continue


                        other_rule_row_mapping = other_sheet_info['rule_row_mapping']
                        other_col_mapping = other_sheet_info['col_mapping']
                        other_cat2_target_cols = other_sheet_info['cat2_target_cols']


                        for (other_col, other_device_number,
                                other_device_name) in other_cat2_target_cols:
                            other_output_col = other_col_mapping[other_col]
                            if other_output_col in other_rule_row_mapping:
                                for OTHER_ROW, (OTHER_RAW_TEXT, OTHER_PARSED_RULE) in (
                                    other_rule_row_mapping[other_output_col].items()
                                ):
                                    if rule_reverse_matches(cat2_rule, OTHER_PARSED_RULE):
                                        cat2_reverse_match_found = True
                                        break
                                if cat2_reverse_match_found:
                                    break
                            if cat2_reverse_match_found:
                                break
                        if cat2_reverse_match_found:
                            break


                    if cat2_reverse_match_found:
                        gray_cells.append((sheet_name, cat1_col, cat1_row))
                        gray_cells.append((sheet_name, cat2_col, cat2_row))
                        gray_match_count += 1


        # 统一标橙色字体（不覆盖已标绿色、橙色的单元格）
        green_cell_set = self._create_cell_set(green_cells)
        yellow_cell_set = self._create_cell_set(yellow_cells)
        for sheet_name, col, row in gray_cells:
            if (sheet_name, col, row) in green_cell_set:
                continue
            if (sheet_name, col, row) in yellow_cell_set:
                continue


            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                if self._check_color_priority(cell, ["FF00FF00", "FFA500"]):
                    self._set_cell_font_color(cell, "FFA500")  # 橙色


        # 处理相同规则的cat2设备列标记
        for sheet_info in sheet_info_list:
            self._mark_same_cat2_rules(
                sheet_info, gray_cells, "FFA500", ["FF00FF00", "FFA500"], output_workbook
            )


        if gray_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤13特殊规则检查完成：发现{gray_match_count}对匹配规则"
                f"（cat1源地址为本平台地址，目的地址为特殊规则定义IP段地址，"
                f"同平台cat1覆盖cat2，跨平台cat2反向检查通过，标橙色）"
            )


        return gray_cells


    # 步骤18：特殊地址段检查（标红色）：
    # 检查cat2规则的源地址和目的地址是否都在任何Sheet的special_network_map中
    def _step_special_address_check(
        self, sheet_info_list, output_workbook, green_cells,
        yellow_cells, gray_cells, rules_cache=None
    ):
        # 步骤18：特殊地址段检查（标红色）：检查cat2规则的源地址和目的地址是否都在任何Sheet的special_network_map中
        yellow_light_cells = []
        yellow_light_match_count = 0


        all_special_networks = []
        for SHEET_NETWORKS in self.SPECIAL_NETWORK_MAP.values():
            if SHEET_NETWORKS:
                all_special_networks.extend(SHEET_NETWORKS)


        total_sheets = len(sheet_info_list)
        current_sheet = 0
        for sheet_info in sheet_info_list:
            current_sheet += 1
            if current_sheet % 5 == 0 or current_sheet == total_sheets:
                self.add_result(
                    Level.OK,
                    f"步骤15进度: {current_sheet}/{total_sheets} 个Sheet已检查，"
                    f"当前发现 {yellow_light_match_count} 条匹配规则"
                )


            sheet_name = sheet_info['sheet_name']


            # 使用规则缓存（性能优化）
            if rules_cache:
                cat2_rules = rules_cache[sheet_name]['cat2']
            else:
                # 兼容旧代码：如果没有缓存则重新收集
                rule_row_mapping = sheet_info['rule_row_mapping']
                col_mapping = sheet_info['col_mapping']
                cat2_target_cols = sheet_info['cat2_target_cols']
                cat2_rules = self._collect_rules_from_cols(
                    cat2_target_cols, col_mapping, rule_row_mapping
                )


            total_cat2_rules = len(cat2_rules)
            checked_cat2_rules = 0
            for cat2_col, cat2_row, cat2_rule, cat2_raw in cat2_rules:
                checked_cat2_rules += 1
                if total_cat2_rules > 1000 and checked_cat2_rules % 1000 == 0:
                    self.add_result(
                        Level.OK,
                        f"步骤15进度: Sheet {sheet_name}: "
                        f"{checked_cat2_rules}/{total_cat2_rules} 个cat2规则已检查"
                    )
                cat2_src_in_special = self._network_in_platform(
                    cat2_rule.src, all_special_networks
                )
                cat2_dst_in_special = self._network_in_platform(
                    cat2_rule.dst, all_special_networks
                )


                if cat2_src_in_special and cat2_dst_in_special:
                    yellow_light_cells.append((sheet_name, cat2_col, cat2_row))
                    yellow_light_match_count += 1


        # 统一标红色字体（不覆盖已标绿色、橙色的单元格）
        green_cell_set = self._create_cell_set(green_cells)
        yellow_cell_set = self._create_cell_set(yellow_cells)
        gray_cell_set = self._create_cell_set(gray_cells)
        for sheet_name, col, row in yellow_light_cells:
            if (sheet_name, col, row) in green_cell_set:
                continue
            if (sheet_name, col, row) in yellow_cell_set:
                continue
            if (sheet_name, col, row) in gray_cell_set:
                continue


            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                if self._check_color_priority(cell, ["FF00FF00", "FFA500"]):
                    self._set_cell_font_color(cell, "FF0000")  # 红色


        # 处理相同规则的cat2设备列标记
        for sheet_info in sheet_info_list:
            self._mark_same_cat2_rules(
                sheet_info, yellow_light_cells, "FF0000", ["FF00FF00", "FFA500"], output_workbook
            )


        if yellow_light_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤15特殊地址段检查完成：发现{yellow_light_match_count}条cat2规则"
                f"（源地址和目的地址都在任何Sheet的special_network_map中，标红色）"
            )


        return yellow_light_cells


    # 步骤19：cat2特殊源地址非平台目的地址检查（标红色）：
    # 检查cat2规则的源地址在special_network_map中，
    # 目的地址不在任何Sheet的platform_network_map中
    def _step_cat2_special_src_non_platform_dst_check(
        self, sheet_info_list, output_workbook, green_cells,
        yellow_cells, gray_cells, yellow_light_cells, rules_cache=None
    ):
        # 步骤19：cat2特殊源地址非平台目的地址检查（标红色）：检查cat2规则的源地址在special_network_map中，目的地址不在任何Sheet的platform_network_map中
        red_cells = []
        red_match_count = 0


        # 收集所有special_network_map中的网络
        all_special_networks = []
        for SHEET_NETWORKS in self.SPECIAL_NETWORK_MAP.values():
            if SHEET_NETWORKS:
                all_special_networks.extend(SHEET_NETWORKS)


        # 收集所有platform_network_map中的网络
        all_platform_networks = []
        for SHEET_NETWORKS in self.PLATFORM_NETWORK_MAP.values():
            if SHEET_NETWORKS:
                all_platform_networks.extend(SHEET_NETWORKS)


        total_sheets = len(sheet_info_list)
        current_sheet = 0
        for sheet_info in sheet_info_list:
            current_sheet += 1
            if current_sheet % 5 == 0 or current_sheet == total_sheets:
                self.add_result(
                    Level.OK,
                    f"步骤16进度: {current_sheet}/{total_sheets} 个Sheet已检查，"
                    f"当前发现 {red_match_count} 条匹配规则"
                )


            sheet_name = sheet_info['sheet_name']


            # 使用规则缓存（性能优化）
            if rules_cache:
                cat2_rules = rules_cache[sheet_name]['cat2']
            else:
                # 兼容旧代码：如果没有缓存则重新收集
                rule_row_mapping = sheet_info['rule_row_mapping']
                col_mapping = sheet_info['col_mapping']
                cat2_target_cols = sheet_info['cat2_target_cols']
                cat2_rules = self._collect_rules_from_cols(
                    cat2_target_cols, col_mapping, rule_row_mapping
                )


            total_cat2_rules = len(cat2_rules)
            checked_cat2_rules = 0
            for cat2_col, cat2_row, cat2_rule, cat2_raw in cat2_rules:
                checked_cat2_rules += 1
                if total_cat2_rules > 1000 and checked_cat2_rules % 1000 == 0:
                    self.add_result(
                        Level.OK,
                        f"步骤16进度: Sheet {sheet_name}: "
                        f"{checked_cat2_rules}/{total_cat2_rules} 个cat2规则已检查"
                    )


                # 检查源地址是否在special_network_map中
                cat2_src_in_special = self._network_in_platform(
                    cat2_rule.src, all_special_networks
                )


                # 检查目的地址是否不在任何Sheet的platform_network_map中
                cat2_dst_in_platform = self._network_in_platform(
                    cat2_rule.dst, all_platform_networks
                )


                if cat2_src_in_special and not cat2_dst_in_platform:
                    red_cells.append((sheet_name, cat2_col, cat2_row))
                    red_match_count += 1


        # 统一标红色字体（不覆盖已标绿色、橙色的单元格）
        green_cell_set = self._create_cell_set(green_cells)
        yellow_cell_set = self._create_cell_set(yellow_cells)
        gray_cell_set = self._create_cell_set(gray_cells)
        yellow_light_cell_set = self._create_cell_set(yellow_light_cells)
        for sheet_name, col, row in red_cells:
            if (sheet_name, col, row) in green_cell_set:
                continue
            if (sheet_name, col, row) in yellow_cell_set:
                continue
            if (sheet_name, col, row) in gray_cell_set:
                continue
            if (sheet_name, col, row) in yellow_light_cell_set:
                continue


            if sheet_name in output_workbook.sheetnames:
                ws = output_workbook[sheet_name]
                cell = ws.cell(row=row, column=col)
                if self._check_color_priority(cell, ["FF00FF00", "FFA500"]):
                    self._set_cell_font_color(cell, "FF0000")  # 红色


        # 处理相同规则的cat2设备列标记
        for sheet_info in sheet_info_list:
            self._mark_same_cat2_rules(
                sheet_info, red_cells, "FF0000", ["FF00FF00", "FFA500"], output_workbook
            )


        if red_match_count > 0:
            self.add_result(
                Level.OK,
                f"步骤16特殊源地址非平台目的地址检查完成：发现{red_match_count}条cat2规则"
                f"（源地址在special_network_map中，"
                f"目的地址不在任何Sheet的platform_network_map中，标红色）"
            )


        return red_cells


    # 步骤14已移除（与步骤6逻辑相似，已删除）

    # 处理单个Sheet：提取ACL配置到Excel
    def run_single(self, sheet_name: str, output_workbook: Workbook, input_worksheet):
        """处理单个工作表

        执行ACL跨平台检查，包括完全匹配、覆盖匹配等步骤

        Args:
            sheet_name: 工作表名称
            output_workbook: 输出工作簿对象
            input_worksheet: 输入工作表对象
        """
        try:
            # 创建新的工作表
            output_worksheet = output_workbook.create_sheet(title=sheet_name)

            # 步骤1：识别设备列（分析第一行，识别cat1/cat2/cat6列）
            cat1_cols, cat2_cols, cat6_cols = analyze_first_row_for_cat1_cat2(input_worksheet)

            # 选择所有cat1、cat2和cat6列（不再限制特定设备）
            cat1_target_cols = list(cat1_cols)
            cat2_target_cols = list(cat2_cols)
            cat6_target_cols = list(cat6_cols)

            # 合并所有目标列，并创建列映射（原始列号 -> 输出列号，从1开始）
            all_target_cols = cat1_target_cols + cat2_target_cols + cat6_target_cols
            # 按原始列号排序，确保输出列顺序一致
            all_target_cols.sort(key=lambda x: x[0])


            # 创建列映射：原始列号 -> 输出列号（从1开始连续）
            col_mapping = {}
            for output_col_idx, (original_col, device_number,
                                    device_name) in enumerate(
                all_target_cols, start=1
            ):
                col_mapping[original_col] = output_col_idx

            # 第一行：只复制目标列的内容（从源Excel的第一行复制）
            # 步骤2显式清除所有样式，避免复制源文件的颜色
            for ORIGINAL_COL, OUTPUT_COL in col_mapping.items():
                SOURCE_CELL = input_worksheet.cell(row=1, column=ORIGINAL_COL)
                TARGET_CELL = output_worksheet.cell(row=1, column=OUTPUT_COL)
                TARGET_CELL.value = SOURCE_CELL.value
                # 显式清除所有样式（font和fill），确保没有颜色
                TARGET_CELL.font = Font()  # 清除字体样式（默认黑色，无特殊格式）
                TARGET_CELL.fill = PatternFill()  # 清除填充样式（无填充）
                # 样式和颜色由后续步骤（跨平台步骤1-6）统一设置

            # 获取平台网段（用于cat1设备多余规则检测）
            platform_networks = self._get_platform_networks(sheet_name)


            # 步骤2：提取ACL配置（按照定界find_acl_blocks_in_column截取ACL配置）
            # 对于每个目标列，独立提取ACL配置（使用映射后的列位置）
            # 先收集所有需要写入的数据，跳过多余规则（不写入即相当于删除）
            data_to_write = {}  # {output_col: [(row, value, font, fill), ...]}
            redundant_count = {}  # {output_col: count} 统计删除的多余规则数量

            # 收集同平台 ACL 数据（在删除之前提取）
            # 结构：{platform_name: {device_name: [acl_block_lines, ...]}}
            # acl_block_lines 包含完整的 ACL 块，包括 "ip access-list VLAN108-ACL" 定义行
            same_platform_acls = {}  # {platform_name: {device_name: [acl_block_lines, ...]}}
            # 使用 sheet_name 作为平台标识
            platform_name = sheet_name


            # 预构建集合用于快速查找（避免在循环内重复创建）
            cat1_cols_set = {col for col, _, _ in cat1_target_cols}
            cat6_cols_set = {col for col, _, _ in cat6_target_cols}


            for original_col, device_number, device_name in all_target_cols:
                output_col = col_mapping[original_col]
                acl_blocks = find_acl_blocks_in_column(input_worksheet, original_col)
                current_row = 2  # 每个列从第二行开始


                # 判断是否为cat1或cat6设备（用于多余规则检测）
                is_cat1 = original_col in cat1_cols_set
                is_cat6 = original_col in cat6_cols_set


                if output_col not in data_to_write:
                    data_to_write[output_col] = []
                    redundant_count[output_col] = 0


                if len(acl_blocks) == 0:
                    self.add_result(
                        Level.WARN,
                        f"Sheet {sheet_name} 列{original_col} "
                        f"({device_name}) 未找到ACL块"
                    )
                else:
                    col_data_count = 0
                    col_redundant_count = 0


                    for start_row, end_row in acl_blocks:
                        for row_idx in range(start_row, end_row + 1):
                            source_cell = input_worksheet.cell(row=row_idx, column=original_col)
                            if source_cell.value is not None:
                                col_data_count += 1


                    # 统计ACL块信息
                    self.add_result(
                        Level.OK,
                        f"Sheet {sheet_name} 列{original_col} ({device_name}) "
                        f"找到{len(acl_blocks)}个ACL块，共{col_data_count}个非空单元格"
                    )


                for start_row, end_row in acl_blocks:
                    # 先收集完整的 ACL 块（用于同平台 ACL 提取）
                    acl_block_lines = []  # 存储当前 ACL 块的所有行（用于输出）
                    same_platform_acl_block_lines = []  # 存储同平台 ACL 块的行（只包含同平台规则）
                    has_redundant_rule = False  # 标记当前 ACL 块是否包含多余规则
                    has_same_platform_rule = False  # 标记当前 ACL 块是否包含同平台规则（源和目的都在平台网段内）
                    acl_block_header = None  # 存储 ACL 块定义行（ip access-list ...）

                    # 从start_row到end_row提取ACL配置
                    for row_idx in range(start_row, end_row + 1):
                        source_cell = input_worksheet.cell(row=row_idx, column=original_col)
                        if source_cell.value is not None:
                            cell_text = str(source_cell.value).strip()
                            # 收集 ACL 块的所有行（包括 ip access-list 定义行）
                            acl_block_lines.append(cell_text)

                            # 保存 ACL 块定义行
                            if "ip access-list" in cell_text.lower():
                                acl_block_header = cell_text
                                # 如果是同平台ACL收集，也添加定义行
                                if (is_cat1 or is_cat6) and platform_networks:
                                    same_platform_acl_block_lines.append(cell_text)

                            # 复制样式
                            source_font = source_cell.font
                            source_fill = source_cell.fill


                            # 步骤3：删除同平台策略（对cat1和cat6设备进行多余规则检测，仅对预定义的Sheet）
                            # 删除条件：
                            # 1. 如果源地址和目的地址都在平台网段内，不写入（相当于删除）
                            # 2. 源any目的any，不写入（相当于删除）
                            # 3. 源any目的掩码8的 x.x.x.x/8，不写入（相当于删除）
                            # 4. icmp any x.x.x.252/30 或 x.x.x.253/30 或 x.x.x.254/30，不写入（相当于删除）
                            is_redundant = False


                            if (is_cat1 or is_cat6) and platform_networks:
                                cell_text = str(source_cell.value).strip()
                                # 只对ACL规则行（包含permit或deny）进行多余规则检测
                                if 'permit' in cell_text.lower() or 'deny' in cell_text.lower():
                                    cell_text_lower = cell_text.lower()


                                    # 删除条件1：源any目的any
                                    # 检查是否包含"any"关键字（源和目的都是any）
                                    # 匹配格式：permit/deny 协议 any any 或 permit/deny 协议 any eq port any
                                    if 'any' in cell_text_lower:
                                        # 使用正则表达式检查源any和目的any的模式
                                        # 匹配：any后面直接跟着any，或any后面有端口等参数再跟着any
                                        any_pattern = re.compile(
                                            r'\bany\s+(?:eq\s+\S+\s+)?any\b',
                                            re.IGNORECASE
                                        )
                                        if any_pattern.search(cell_text):
                                            is_redundant = True
                                            redundant_count[output_col] += 1
                                            has_redundant_rule = True


                                    # 删除条件2：源any目的掩码8的 x.x.x.x/8
                                    if not is_redundant:
                                        # 检查源any和目的/8掩码
                                        # 匹配格式：permit/deny 协议 any ... x.x.x.x/8
                                        any_src_dst_8_pattern = re.compile(
                                            r'\bany\s+(?:eq\s+\S+\s+)?(\d+\.\d+\.\d+\.\d+)/8\b',
                                            re.IGNORECASE
                                        )
                                        if any_src_dst_8_pattern.search(cell_text):
                                            is_redundant = True
                                            redundant_count[output_col] += 1
                                            has_redundant_rule = True


                                    # 删除条件3：icmp any x.x.x.252/30 或 x.x.x.253/30 或 x.x.x.254/30
                                    if not is_redundant:
                                        # 检查icmp协议、源any、目的地址为/30掩码且最后一段是252、253或254
                                        # 匹配格式：permit/deny icmp any x.x.x.252/30 或 x.x.x.253/30 或 x.x.x.254/30
                                        # any后面可能直接是IP地址，也可能有其他参数（如log等）
                                        icmp_any_252_254_pattern = re.compile(
                                            r'\bicmp\s+any\s+'
                                            r'(\d+\.\d+\.\d+\.(?:252|253|254))/30\b',
                                            re.IGNORECASE
                                        )
                                        if icmp_any_252_254_pattern.search(cell_text):
                                            is_redundant = True
                                            redundant_count[output_col] += 1
                                            has_redundant_rule = True


                                    # 删除条件4：源地址和目的地址都在平台网段内（原有逻辑）
                                    if not is_redundant:
                                        # 首先尝试使用parse_acl解析
                                        parsed_rule, parse_error = parse_acl(cell_text)
                                        if parsed_rule:
                                            # 检查源网络和目的网络是否都在平台网段内
                                            source_in_platform = self._network_in_platform(
                                                parsed_rule.src, platform_networks
                                            )
                                            destination_in_platform = self._network_in_platform(
                                                parsed_rule.dst, platform_networks
                                            )


                                            # 如果源和目的都在平台网段内，标记为多余规则（不写入）
                                            if source_in_platform and destination_in_platform:
                                                is_redundant = True
                                                redundant_count[output_col] += 1
                                                has_same_platform_rule = True
                                                # 收集同平台规则到同平台ACL块
                                                if (is_cat1 or is_cat6) and platform_networks:
                                                    same_platform_acl_block_lines.append(cell_text)
                                        else:
                                            # parse_acl失败时，尝试从文本中直接提取IP地址
                                            # 例如：368 permit tcp 10.66.110.0/24 eq 3366 10.66.120.90/32 eq 7080
                                            networks = (
                                                self._extract_networks_from_rule_text(
                                                    cell_text
                                                )
                                            )
                                            if networks:
                                                source_network, destination_network = networks
                                                source_in_platform = (
                                                    self._network_in_platform(
                                                        source_network, platform_networks
                                                    )
                                                )
                                                destination_in_platform = (
                                                    self._network_in_platform(
                                                        destination_network,
                                                        platform_networks
                                                    )
                                                )


                                                # 如果源和目的都在平台网段内，标记为多余规则（不写入）
                                                if source_in_platform and destination_in_platform:
                                                    is_redundant = True
                                                    redundant_count[output_col] += 1
                                                    has_redundant_rule = True
                                                    has_same_platform_rule = True
                                                    # 收集同平台规则到同平台ACL块
                                                    if ((is_cat1 or is_cat6) and
                                                            platform_networks):
                                                        same_platform_acl_block_lines.append(
                                                            cell_text
                                                        )

                            # 如果不是多余规则，则添加到写入列表（包括ip access-list行、注释行等）
                            if not is_redundant:
                                data_to_write[output_col].append((
                                    current_row, source_cell.value,
                                    source_font, source_fill
                                ))
                                current_row += 1

                    # 如果当前 ACL 块包含同平台规则（源和目的都在平台网段内），且是 cat1 或 cat6 设备，则收集到同平台 ACL 数据
                    # 注意：只收集同平台规则，不收集any any规则或其他非同平台规则
                    if has_same_platform_rule and (is_cat1 or is_cat6) and platform_networks:
                        if platform_name not in same_platform_acls:
                            same_platform_acls[platform_name] = {}
                        if device_name not in same_platform_acls[platform_name]:
                            same_platform_acls[platform_name][device_name] = []
                        # 只保存包含同平台规则的ACL块（不包含any any规则和非同平台规则）
                        # same_platform_acl_block_lines 已经包含了ACL定义行和同平台规则行
                        if len(same_platform_acl_block_lines) > 1:  # 至少有定义行和一条规则
                            same_platform_acls[platform_name][device_name].append(
                                same_platform_acl_block_lines
                            )


            # 写入数据到输出工作表（跳过多余规则，相当于删除）
            # 同时记录每个规则的行号和原始文本，用于第三步比较
            rule_row_mapping = {}  # {output_col: {row: (raw_text, parsed_rule)}}


            total_items = sum(len(DATA_LIST) for DATA_LIST in data_to_write.values())
            for OUTPUT_COL, DATA_LIST in data_to_write.items():
                self.add_result(
                    Level.OK,
                    f"Sheet {sheet_name} 输出列{OUTPUT_COL} "
                    f"有{len(DATA_LIST)}条数据待写入"
                )


            if total_items == 0:
                # 如果未找到符合条件的设备，会在后面统一输出WARN，此处不再重复
                if cat1_target_cols or cat2_target_cols or cat6_target_cols:
                    # 找到了设备列但data_to_write为空，输出详细信息
                    for original_col, device_number, device_name in all_target_cols:
                        output_col = col_mapping[original_col]
                        acl_blocks_count = len(
                            find_acl_blocks_in_column(
                                input_worksheet, original_col
                            )
                        )
                        data_count = len(data_to_write.get(output_col, []))
                        if acl_blocks_count > 0:
                            self.add_result(
                                Level.WARN,
                                f"Sheet {sheet_name} 列{original_col} ({device_name}) "
                                f"找到{acl_blocks_count}个ACL块，"
                                f"但data_to_write[列{output_col}]只有{data_count}条数据"
                                f"（可能所有规则都被删除）"
                            )
                # 如果未找到设备列，不在这里输出WARN，由后面的逻辑统一处理
            else:
                self.add_result(Level.OK, f"Sheet {sheet_name} 准备写入 {total_items} 条数据到输出工作表")


            for OUTPUT_COL, DATA_LIST in data_to_write.items():
                if OUTPUT_COL not in rule_row_mapping:
                    rule_row_mapping[OUTPUT_COL] = {}


                for ROW, VALUE, FONT, FILL in DATA_LIST:
                    TARGET_CELL = output_worksheet.cell(row=ROW, column=OUTPUT_COL)
                    TARGET_CELL.value = VALUE


                    # 步骤2显式清除所有样式，确保不继承源文件的颜色
                    # 样式和颜色由后续步骤（跨平台步骤1-6）统一设置
                    TARGET_CELL.font = Font()  # 清除字体样式（默认黑色，无特殊格式）
                    TARGET_CELL.fill = PatternFill()  # 清除填充样式（无填充）


                    # 记录规则信息（用于第三步比较）
                    if VALUE and isinstance(VALUE, str):
                        CELL_TEXT = str(VALUE).strip()
                        if is_acl_rule(CELL_TEXT):
                            PARSED_RULE, PARSE_ERROR = parse_acl(CELL_TEXT)
                            if PARSED_RULE:
                                rule_row_mapping[OUTPUT_COL][ROW] = (CELL_TEXT, PARSED_RULE)


            # 复制列宽设置（只复制目标列）
            # 首先找到第一个有列宽的原始列，作为默认列宽
            default_width = None
            for ORIGINAL_COL in col_mapping.keys():
                ORIGINAL_COL_LETTER = get_column_letter(ORIGINAL_COL)
                if ORIGINAL_COL_LETTER in input_worksheet.column_dimensions:
                    COL_DIM = input_worksheet.column_dimensions[ORIGINAL_COL_LETTER]
                    if COL_DIM.width:
                        default_width = COL_DIM.width
                        break
            # 如果没有找到有列宽的列，使用默认值80.0
            if default_width is None:
                default_width = 80.0


            # 复制列宽设置：如果原始列有列宽则使用原始列宽，否则使用默认列宽
            for ORIGINAL_COL, OUTPUT_COL in col_mapping.items():
                ORIGINAL_COL_LETTER = get_column_letter(ORIGINAL_COL)
                OUTPUT_COL_LETTER = get_column_letter(OUTPUT_COL)
                WIDTH_TO_SET = default_width


                # 如果原始列有列宽设置，使用原始列宽
                if ORIGINAL_COL_LETTER in input_worksheet.column_dimensions:
                    COL_DIM = input_worksheet.column_dimensions[ORIGINAL_COL_LETTER]
                    if COL_DIM.width:
                        WIDTH_TO_SET = COL_DIM.width


                # 设置输出列的列宽
                output_worksheet.column_dimensions[OUTPUT_COL_LETTER].width = WIDTH_TO_SET

            # 统计实际写入的ACL规则数量（已删除多余规则）
            total_cat1_rules = 0
            total_cat1_redundant = 0
            for col, device_number, device_name in cat1_target_cols:
                output_col = col_mapping[col]
                if output_col in data_to_write:
                    total_cat1_rules += len(data_to_write[output_col])
                if output_col in redundant_count:
                    total_cat1_redundant += redundant_count[output_col]


            total_cat2_rules = 0
            for col, device_number, device_name in cat2_target_cols:
                output_col = col_mapping[col]
                if output_col in data_to_write:
                    total_cat2_rules += len(data_to_write[output_col])


            total_cat6_rules = 0
            total_cat6_redundant = 0
            for col, device_number, device_name in cat6_target_cols:
                output_col = col_mapping[col]
                if output_col in data_to_write:
                    total_cat6_rules += len(data_to_write[output_col])
                if output_col in redundant_count:
                    total_cat6_redundant += redundant_count[output_col]

            if cat1_target_cols and (cat2_target_cols or cat6_target_cols):
                # 格式化平台网段信息用于显示
                if platform_networks:
                    platform_str = ", ".join([str(net) for net in platform_networks])
                    platform_info = f"，平台网段: {platform_str}"
                else:
                    platform_info = ""
                # 统计删除的多余规则（cat1和cat6）
                total_redundant = total_cat1_redundant + total_cat6_redundant
                redundant_info = ""
                if total_redundant > 0:
                    parts = []
                    if total_cat1_redundant > 0:
                        parts.append(f"cat1 {total_cat1_redundant}条")
                    if total_cat6_redundant > 0:
                        parts.append(f"cat6 {total_cat6_redundant}条")
                    redundant_info = f"，删除同平台策略({', '.join(parts)})"
                cat6_info = f"，cat6 ACL规则{total_cat6_rules}条" if total_cat6_rules > 0 else ""
                self.add_result(
                    Level.OK,
                    f"Sheet {sheet_name} 处理完成（步骤1-3）："
                    f"提取cat1 ACL规则{total_cat1_rules}条，"
                    f"cat2 ACL规则{total_cat2_rules}条{cat6_info}"
                    f"{redundant_info}{platform_info}"
                )
            else:
                # 未找到符合条件的设备（需要cat1和cat2/cat6列）
                # 如果total_items也为0，说明既没有设备列也没有数据，只输出一条WARN
                if total_items == 0:
                    self.add_result(
                        Level.WARN,
                        f"Sheet {sheet_name} 处理完成："
                        f"未找到符合条件的设备（需要cat1和cat2/cat6列），"
                        f"且没有数据需要写入"
                    )
                else:
                    self.add_result(
                        Level.WARN,
                        f"Sheet {sheet_name} 处理完成："
                        f"未找到符合条件的设备（需要cat1和cat2/cat6列）"
                    )


            # 返回规则映射和平台网段信息
            # 对于跨Sheet比较，使用第一个平台网段（如果存在）
            platform_network_for_cross_check = platform_networks[0] if platform_networks else None
            return {
                'sheet_name': sheet_name,
                'platform_network': platform_network_for_cross_check,  # 用于跨Sheet比较（使用第一个网段）
                'platform_networks': platform_networks,  # 完整的平台网段列表
                'rule_row_mapping': rule_row_mapping,
                'col_mapping': col_mapping,
                'cat1_target_cols': cat1_target_cols,
                'cat2_target_cols': cat2_target_cols,
                'cat6_target_cols': cat6_target_cols,
                'output_worksheet': output_worksheet,
                'same_platform_acls': same_platform_acls  # 同平台 ACL 数据
            }

        except (ValueError, TypeError, AttributeError, KeyError, OSError, IOError) as EXCEPTION:
            import traceback
            error_msg = f"处理 Sheet {sheet_name} 失败: {EXCEPTION}"
            error_detail = traceback.format_exc()
            self.add_result(Level.ERROR, error_msg)
            self.add_result(Level.ERROR, f"详细错误信息: {error_detail}")
            return None

    # 重写run方法：处理所有Sheet并生成最终报告
    # ========== run方法拆分出的辅助方法 ==========

    def _extract_same_platform_acls_to_excel(
        self,
        all_same_platform_acls: Dict[str, Dict[str, List[List[str]]]],
        progress=None,
    ) -> None:
        """提取同平台 ACL 到新的 Excel 文件（仅收集 cat1 和 cat6 设备）并执行同平台步骤。

        Args:
            all_same_platform_acls: {platform_name: {device_name: [acl_block_lines, ...]}}
                acl_block_lines 是包含完整 ACL 块的列表，包括 "ip access-list VLAN108-ACL" 定义行
                仅包含 cat1（N9K核心交换机）和 cat6（OOB-DS交换机）设备的同平台 ACL
            progress: tqdm 进度条对象，可选，用于更新同平台步骤的进度
        """
        try:
            # 生成输出文件名
            today = get_today_str()
            output_filename = f"{today}-同平台N9K ACL检查.xlsx"
            output_path = build_output_path(self.OUTPUT_DIR, output_filename)

            # 创建新的工作簿
            same_platform_workbook = Workbook()
            if "Sheet" in same_platform_workbook.sheetnames:
                same_platform_workbook.remove(same_platform_workbook["Sheet"])

            # 用于步骤4：同Sheet同列反向完全匹配检查的数据收集
            # 结构：{(sheet_name, col): [(row, acl_name, parsed_rule), ...]}
            reverse_match_rules: Dict[
                Tuple[str, int],
                List[Tuple[int, str, ACLRule]]
            ] = {}

            # 为每个平台创建一个 sheet
            for platform_name, devices in all_same_platform_acls.items():
                # 创建平台 sheet（sheet 名称不能超过 31 个字符）
                sheet_name = (
                    platform_name[:31]
                    if len(platform_name) <= 31
                    else platform_name[:28] + "..."
                )
                ws = same_platform_workbook.create_sheet(title=sheet_name)

                # 将设备按 cat1 和 cat6 分类，cat1设备按设备编号（01/03）进一步分类
                cat1_devices_01 = {}  # {device_name: [acl_blocks, ...]} 设备编号为01的cat1设备
                cat1_devices_03 = {}  # {device_name: [acl_blocks, ...]} 设备编号为03的cat1设备
                cat6_devices = {}  # {device_name: [acl_blocks, ...]}

                for device_name, acl_blocks in devices.items():
                    if is_cat1_device(device_name):
                        device_number = extract_device_number(device_name)
                        if device_number == 1:
                            cat1_devices_01[device_name] = acl_blocks
                        elif device_number == 3:
                            cat1_devices_03[device_name] = acl_blocks
                    elif is_cat6_device(device_name):
                        cat6_devices[device_name] = acl_blocks

                # 写入 cat1 设备（01设备写入第1列，03设备写入第2列）
                # 第1列：cat1设备01
                cat1_01_row = 1
                for device_name, acl_blocks in cat1_devices_01.items():
                    # 如果不是第一行，添加空行分隔
                    if cat1_01_row > 1:
                        cat1_01_row += 1

                    # 写入设备名称（去掉 "# 设备: " 前缀，不加粗）
                    device_cell = ws.cell(row=cat1_01_row, column=1)
                    device_cell.value = device_name
                    cat1_01_row += 1

                    # 写入每个 ACL 块
                    for acl_block_lines in acl_blocks:
                        acl_name = ""
                        for line in acl_block_lines:
                            cell_text = str(line).strip()
                            cell = ws.cell(row=cat1_01_row, column=1)
                            cell.value = cell_text

                            # 记录ACL名称（ip access-list开头的行）
                            if "ip access-list" in cell_text.lower():
                                acl_name = cell_text
                            else:
                                # 同平台步骤1：仅对permit/deny规则行进行解析和收集（忽略行号和log等由parse_acl处理）
                                if (
                                    "permit" in cell_text.lower()
                                    or "deny" in cell_text.lower()
                                ):
                                    parsed_rule, parse_error = parse_acl(cell_text)
                                    if parsed_rule:
                                        key = (sheet_name, 1)
                                        if key not in reverse_match_rules:
                                            reverse_match_rules[key] = []
                                        reverse_match_rules[key].append(
                                            (cat1_01_row, acl_name, parsed_rule)
                                        )

                            cat1_01_row += 1
                        # ACL 块之间添加空行
                        cat1_01_row += 1

                # 第2列：cat1设备03（如果存在）
                cat1_03_row = 1
                for device_name, acl_blocks in cat1_devices_03.items():
                    # 如果不是第一行，添加空行分隔
                    if cat1_03_row > 1:
                        cat1_03_row += 1

                    # 写入设备名称（去掉 "# 设备: " 前缀，不加粗）
                    device_cell = ws.cell(row=cat1_03_row, column=2)
                    device_cell.value = device_name
                    cat1_03_row += 1

                    # 写入每个 ACL 块
                    for acl_block_lines in acl_blocks:
                        acl_name = ""
                        for line in acl_block_lines:
                            cell_text = str(line).strip()
                            cell = ws.cell(row=cat1_03_row, column=2)
                            cell.value = cell_text

                            # 记录ACL名称（ip access-list开头的行）
                            if "ip access-list" in cell_text.lower():
                                acl_name = cell_text
                            else:
                                # 同平台步骤1：仅对permit/deny规则行进行解析和收集（忽略行号和log等由parse_acl处理）
                                if (
                                    "permit" in cell_text.lower()
                                    or "deny" in cell_text.lower()
                                ):
                                    parsed_rule, parse_error = parse_acl(cell_text)
                                    if parsed_rule:
                                        key = (sheet_name, 2)
                                        if key not in reverse_match_rules:
                                            reverse_match_rules[key] = []
                                        reverse_match_rules[key].append(
                                            (cat1_03_row, acl_name, parsed_rule)
                                        )

                            cat1_03_row += 1
                        # ACL 块之间添加空行
                        cat1_03_row += 1

                # 写入 cat6 设备（如果第2列被03占用，cat6放在第3列；否则放在第2列）
                cat6_col = 3 if cat1_devices_03 else 2  # 如果第2列被03占用，cat6放在第3列
                cat6_row = 1
                for device_name, acl_blocks in cat6_devices.items():
                    # 如果不是第一行，添加空行分隔
                    if cat6_row > 1:
                        cat6_row += 1

                    # 写入设备名称（去掉 "# 设备: " 前缀，不加粗）
                    device_cell = ws.cell(row=cat6_row, column=cat6_col)
                    device_cell.value = device_name
                    cat6_row += 1

                    # 写入每个 ACL 块
                    for acl_block_lines in acl_blocks:
                        acl_name = ""
                        for line in acl_block_lines:
                            cell_text = str(line).strip()
                            cell = ws.cell(row=cat6_row, column=cat6_col)
                            cell.value = cell_text

                            # 记录ACL名称（ip access-list开头的行）
                            if "ip access-list" in cell_text.lower():
                                acl_name = cell_text
                            else:
                                # 同平台步骤1：仅对permit/deny规则行进行解析和收集（忽略行号和log等由parse_acl处理）
                                if (
                                    "permit" in cell_text.lower()
                                    or "deny" in cell_text.lower()
                                ):
                                    parsed_rule, parse_error = parse_acl(cell_text)
                                    if parsed_rule:
                                        key = (sheet_name, cat6_col)
                                        if key not in reverse_match_rules:
                                            reverse_match_rules[key] = []
                                        reverse_match_rules[key].append(
                                            (cat6_row, acl_name, parsed_rule)
                                        )

                            cat6_row += 1
                        # ACL 块之间添加空行
                        cat6_row += 1

                # 设置列宽（根据实际使用的列数）
                ws.column_dimensions['A'].width = 120.0
                ws.column_dimensions['B'].width = 120.0
                if cat6_col == 3:
                    ws.column_dimensions['C'].width = 120.0

            # 同平台步骤1：在同一Sheet同一列内，不同ACL块之间做反向完全匹配检查，并标记为绿色
            total_pairs = 0
            for (sheet_name, col), rules in reverse_match_rules.items():
                if len(rules) < 2:
                    continue

                ws = same_platform_workbook[sheet_name]
                matched_cells = set()
                rule_count = len(rules)

                # rules: [(row, acl_name, parsed_rule), ...]
                for i in range(rule_count):
                    row_a, acl_a, rule_a = rules[i]
                    for j in range(i + 1, rule_count):
                        row_b, acl_b, rule_b = rules[j]
                        # 只比较不同ACL块之间的规则
                        if not acl_a or not acl_b or acl_a == acl_b:
                            continue
                        # 反向完全匹配：A反向匹配B且B反向匹配A
                        if (
                            rule_reverse_matches(rule_a, rule_b)
                            and rule_reverse_matches(rule_b, rule_a)
                        ):
                            matched_cells.add((row_a, col))
                            matched_cells.add((row_b, col))
                            total_pairs += 1

                # 标记匹配单元格为绿色字体
                for row, col_idx in matched_cells:
                    cell = ws.cell(row=row, column=col_idx)
                    old_font = cell.font or Font()
                    cell.font = Font(
                        name=old_font.name,
                        size=old_font.size,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        vertAlign=old_font.vertAlign,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color="FF00FF00",
                    )

            if total_pairs > 0:
                self.add_result(
                    Level.OK,
                    f"同平台步骤1：反向完全匹配检查完成，共标记 {total_pairs} 对规则（绿色）"
                )
            else:
                self.add_result(
                    Level.OK,
                    "同平台步骤1：反向完全匹配检查完成，未发现反向完全匹配规则"
                )
            if progress:
                progress.update(1)

            # 同平台步骤2：目的地址为special_network_map标记橙色（仅同Sheet同列，不跨Sheet/列）
            total_special = 0
            for (sheet_name, col), rules in reverse_match_rules.items():
                ws = same_platform_workbook[sheet_name]
                special_networks = self.SPECIAL_NETWORK_MAP.get(sheet_name, [])
                if not special_networks:
                    continue

                for row, _acl_name, parsed_rule in rules:
                    if not parsed_rule:
                        continue

                    # 检查目的地址是否在任一特殊网段内
                    dst_in_special = False
                    for special_net in special_networks:
                        try:
                            if (
                                parsed_rule.dst.overlaps(special_net)
                                or parsed_rule.dst.subnet_of(special_net)
                            ):
                                dst_in_special = True
                                break
                        except (AttributeError, ValueError, TypeError):
                            continue

                    if not dst_in_special:
                        continue

                    cell = ws.cell(row=row, column=col)
                    # 如果已经是绿色（同平台步骤1标记），则不覆盖
                    try:
                        color_str = str(cell.font.color or "").upper()
                        if "00FF00" in color_str:
                            continue
                    except Exception:
                        pass

                    old_font = cell.font or Font()
                    cell.font = Font(
                        name=old_font.name,
                        size=old_font.size,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        vertAlign=old_font.vertAlign,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color="FFA500",  # 橙色
                    )
                    total_special += 1

            if total_special > 0:
                self.add_result(
                    Level.OK,
                    f"同平台步骤2：目的地址为special_network_map检查完成，共标记 {total_special} 条规则为橙色"
                )
            else:
                self.add_result(
                    Level.OK,
                    "同平台步骤2：目的地址为special_network_map检查完成，未发现匹配规则"
                )
            if progress:
                progress.update(1)

            # 同平台步骤3：同Sheet内cat1和cat6反向完全匹配检查（标蓝色，仅cat1列与cat6列之间）
            # cat1的01设备在第1列，03设备在第2列（如果存在）
            # cat6设备在第2列（如果03不存在）或第3列（如果03存在）
            total_cat1_cat6_pairs = 0
            for sheet_name in same_platform_workbook.sheetnames:
                ws = same_platform_workbook[sheet_name]

                # 找出cat1列和cat6列
                cat1_cols = []  # cat1列号列表（可能是1和/或2）
                cat6_col = None  # cat6列号（可能是2或3）

                # 检查所有可能的列（1, 2, 3）
                for col in [1, 2, 3]:
                    key = (sheet_name, col)
                    if key in reverse_match_rules:
                        # 判断是cat1还是cat6：通过检查该列的第一个设备名称
                        # 查找该列的第一行（设备名称行）
                        first_row_with_value = None
                        for row in range(1, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col)
                            if cell.value and isinstance(cell.value, str):
                                device_name = str(cell.value).strip()
                                if device_name and len(device_name) > 10:  # 设备名称通常较长
                                    first_row_with_value = row
                                    if is_cat1_device(device_name):
                                        if col not in cat1_cols:
                                            cat1_cols.append(col)
                                    elif is_cat6_device(device_name):
                                        if cat6_col is None:
                                            cat6_col = col
                                    break

                # 如果没有找到cat6列，跳过
                if cat6_col is None or not cat1_cols:
                    continue

                # 获取cat6规则
                cat6_rules = reverse_match_rules.get((sheet_name, cat6_col), [])
                if not cat6_rules:
                    continue

                # 对每个cat1列，与cat6列进行匹配检查
                for cat1_col in cat1_cols:
                    cat1_rules = reverse_match_rules.get((sheet_name, cat1_col), [])
                    if not cat1_rules:
                        continue

                    for row1, _acl1, rule1 in cat1_rules:
                        for row6, _acl6, rule6 in cat6_rules:
                            if (
                                rule_reverse_matches(rule1, rule6)
                                and rule_reverse_matches(rule6, rule1)
                            ):
                                # 标记cat1和cat6对应规则为蓝色，不覆盖已有绿色、橙色
                                for row, col_idx in ((row1, cat1_col), (row6, cat6_col)):
                                    cell = ws.cell(row=row, column=col_idx)
                                    try:
                                        color_str = str(cell.font.color or "").upper()
                                        # 已经是绿色(FF00FF00)或橙色(FFA500)则不覆盖
                                        if "00FF00" in color_str or "FFA500" in color_str:
                                            continue
                                    except Exception:
                                        pass
                                    old_font = cell.font or Font()
                                    cell.font = Font(
                                        name=old_font.name,
                                        size=old_font.size,
                                        bold=old_font.bold,
                                        italic=old_font.italic,
                                        vertAlign=old_font.vertAlign,
                                        underline=old_font.underline,
                                        strike=old_font.strike,
                                        color="0000FF",  # 蓝色
                                    )
                                total_cat1_cat6_pairs += 1

            if total_cat1_cat6_pairs > 0:
                self.add_result(
                    Level.OK,
                    "同平台步骤3：cat1和cat6反向完全匹配检查完成，"
                    f"共标记 {total_cat1_cat6_pairs} 对规则为蓝色"
                )
            else:
                self.add_result(
                    Level.OK,
                    "同平台步骤3：cat1和cat6反向完全匹配检查完成，未发现匹配规则"
                )
            if progress:
                progress.update(1)

            # 保存文件
            save_excel_workbook(same_platform_workbook, output_path)
            self.add_result(
                Level.OK,
                f"同平台 ACL 提取完成：已保存到 {output_path}"
            )

            # 统计信息
            total_platforms = len(all_same_platform_acls)
            total_devices = sum(len(devices) for devices in all_same_platform_acls.values())
            total_acl_blocks = sum(
                len(acl_blocks)
                for devices in all_same_platform_acls.values()
                for acl_blocks in devices.values()
            )
            self.add_result(
                Level.OK,
                f"同平台 ACL 统计：{total_platforms}个平台，"
                f"{total_devices}个设备，{total_acl_blocks}个ACL块"
            )

        except Exception as e:
            self.add_result(
                Level.ERROR,
                f"提取同平台 ACL 到 Excel 失败: {e}"
            )
            import traceback
            self.add_result(Level.ERROR, f"详细错误: {traceback.format_exc()}")

    # 处理所有Sheet，返回sheet_info_list
    def _process_all_sheets(self, task_items, input_workbook, output_workbook, progress):
        import time
        sheet_info_list = []
        total_start_time = time.time()
        # 收集所有 sheet 的同平台 ACL 数据
        all_same_platform_acls = {}  # {platform_name: {device_name: [acl_block_lines, ...]}}


        for idx, sheet_name in enumerate(task_items, 1):
            sheet_start_time = time.time()
            try:
                if sheet_name not in input_workbook.sheetnames:
                    self.add_result(
                        Level.WARN,
                        f"Sheet {sheet_name} 在输入文件中不存在，跳过"
                    )
                    if progress:
                        progress.set_description(
                            f"{self.NAME} (Sheet {idx}/{len(task_items)}: {sheet_name})"
                        )
                    continue


                input_worksheet = input_workbook[sheet_name]
                sheet_info = self.run_single(sheet_name, output_workbook, input_worksheet)

                # 收集同平台 ACL 数据
                if sheet_info and 'same_platform_acls' in sheet_info:
                    sheet_same_platform_acls = sheet_info.get('same_platform_acls', {})
                    for platform_name, devices in sheet_same_platform_acls.items():
                        if platform_name not in all_same_platform_acls:
                            all_same_platform_acls[platform_name] = {}
                        for device_name, acl_blocks in devices.items():
                            if device_name not in all_same_platform_acls[platform_name]:
                                all_same_platform_acls[platform_name][device_name] = []
                            all_same_platform_acls[platform_name][device_name].extend(acl_blocks)


                has_data = False
                if sheet_info:
                    sheet_info_list.append(sheet_info)
                    rule_row_mapping = sheet_info.get('rule_row_mapping', {})
                    total_rules = sum(len(rules) for rules in rule_row_mapping.values())
                    has_data = total_rules > 0


                sheet_elapsed = time.time() - sheet_start_time
                total_elapsed = time.time() - total_start_time
                status = "处理" if has_data else "读取（无数据）"
                self.add_result(
                    Level.OK,
                    f"Sheet {sheet_name} ({idx}/{len(task_items)}) {status}完成，"
                    f"耗时 {sheet_elapsed:.2f} 秒（累计 {total_elapsed:.2f} 秒）"
                )


            except (ValueError, TypeError, AttributeError, KeyError,
                    OSError, IOError) as EXCEPTION:
                self.add_result(Level.ERROR, f"Sheet {sheet_name} 运行异常: {EXCEPTION!r}")
                sheet_elapsed = time.time() - sheet_start_time
                total_elapsed = time.time() - total_start_time
                self.add_result(
                    Level.OK,
                    f"Sheet {sheet_name} ({idx}/{len(task_items)}) 异常，"
                    f"耗时 {sheet_elapsed:.2f} 秒（累计 {total_elapsed:.2f} 秒）"
                )

            if progress:
                progress.set_description(
                    f"{self.NAME} (Sheet {idx}/{len(task_items)}: {sheet_name})"
                )

        # 生成同平台 ACL Excel 文件（在删除之前），并执行同平台步骤1-3
        if all_same_platform_acls:
            self._extract_same_platform_acls_to_excel(all_same_platform_acls, progress)

        return sheet_info_list


    # 预构建规则缓存（性能优化）
    def _build_rules_cache(self, sheet_info_list, progress):
        self.add_result(Level.OK, "预构建规则缓存（性能优化）...")
        if progress:
            progress.set_description(f"{self.NAME} (预构建规则缓存)")


        rules_cache = {}
        for sheet_info in sheet_info_list:
            sheet_name = sheet_info['sheet_name']
            col_mapping = sheet_info['col_mapping']
            rule_row_mapping = sheet_info['rule_row_mapping']
            rules_cache[sheet_name] = {
                'cat1': self._collect_rules_from_cols(
                    sheet_info['cat1_target_cols'], col_mapping, rule_row_mapping),
                'cat2': self._collect_rules_from_cols(
                    sheet_info['cat2_target_cols'], col_mapping, rule_row_mapping),
                'cat6': self._collect_rules_from_cols(
                    sheet_info['cat6_target_cols'], col_mapping, rule_row_mapping),
            }


        self.add_result(Level.OK, f"规则缓存构建完成，共 {len(rules_cache)} 个Sheet")
        return rules_cache


    # 统一创建单元格集合
    def _create_cell_set(self, cells):
        if not cells:
            return set()
        return set(cells)


    # 统一创建排除单元格集合
    def _create_exclude_sets(self, *cell_lists):
        exclude_sets = {}
        for cells in cell_lists:
            if cells:
                for cell in cells:
                    if isinstance(cell, tuple) and len(cell) == 3:
                        sheet_name, col, row = cell
                        if sheet_name not in exclude_sets:
                            exclude_sets[sheet_name] = set()
                        exclude_sets[sheet_name].add((col, row))
        return exclude_sets


    # 执行所有步骤4-19（按“跨平台步骤1-10”概念步骤进行进度统计）
    def _execute_all_steps(
            self, sheet_info_list, output_workbook, rules_cache, progress,
            output_path
    ):
        STOP_AT_STEP = None
        SAVE_AFTER_EACH_STEP = False


        self.add_result(Level.OK, f"开始执行步骤4-19，共有 {len(sheet_info_list)} 个Sheet需要处理")
        if STOP_AT_STEP:
            self.add_result(Level.WARN, f"将执行到步骤{STOP_AT_STEP}后停止")
        if SAVE_AFTER_EACH_STEP:
            self.add_result(Level.WARN, "每个步骤后都会保存文件（可能影响性能）")


        # 跨平台步骤1：cat1完全匹配检查（标绿色）
        green_cells = self._execute_step(
            4, "cat1完全匹配检查", self._step_cat1_complete_match_check,
            sheet_info_list, output_workbook, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None
        if progress:
            progress.update(1)

        # 跨平台步骤2：cat1覆盖匹配检查（标绿色）
        dark_green_cells = self._execute_step(
            5, "cat1覆盖匹配检查", self._step_cat1_cover_match_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if dark_green_cells is None:
            return None
        if progress:
            progress.update(1)

        # 跨平台步骤3：cat2覆盖cat1匹配检查（标绿色，含多个cat1规则覆盖cat2）
        light_green_cells = self._execute_step(
            6, "cat2覆盖cat1匹配检查", self._step_cat2_cover_cat1_match_check,
            sheet_info_list, output_workbook, green_cells, dark_green_cells,
            rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if light_green_cells is None:
            return None

        all_green_cells = list(set(green_cells + dark_green_cells + light_green_cells))
        green_cells = self._execute_step(
            7, "多个cat1规则覆盖cat2规则检查", self._step_multi_cat1_cover_cat2_check,
            sheet_info_list, output_workbook, all_green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None
        if progress:
            progress.update(1)

        # 跨平台步骤4：平台源地址特殊目的地址cat1和cat2匹配检查（标绿色）
        green_cells = self._execute_step(
            8, "平台源地址特殊目的地址cat1和cat2匹配检查", self._step_platform_src_special_dst_cat1_cat2_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None
        if progress:
            progress.update(1)

        # 跨平台步骤5：cat2 IP协议覆盖cat1 TCP协议检查（标绿色）
        green_cells = self._execute_step(
            9, "cat2 IP协议覆盖cat1 TCP协议检查", self._step_cat2_ip_cover_cat1_tcp_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None
        if progress:
            progress.update(1)

        # 跨平台步骤6：跨Sheet cat1和cat2匹配 + cat2反向匹配检查（标绿色）
        green_cells = self._execute_step(
            10, "跨Sheet cat1和cat2匹配检查", self._step_cross_sheet_cat1_cat2_reverse_match_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None

        green_cells = self._execute_step(
            11, "cat1与cat2匹配且cat2反向匹配检查", self._step_cat1_cat2_match_cat2_reverse_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None
        if progress:
            progress.update(1)

        # 跨平台步骤7：cat6与cat2的完全/覆盖/反向匹配检查（标蓝色）
        green_cells = self._execute_step(
            12, "cat6完全匹配检查", self._step_cat6_complete_match_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None

        green_cells = self._execute_step(
            13, "cat6覆盖匹配检查", self._step_cat6_cover_match_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None

        green_cells = self._execute_step(
            14, "cat2覆盖匹配cat6检查", self._step_cat2_cover_cat6_match_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None
        if progress:
            progress.update(1)

        # 跨平台步骤8：cat6-cat1包含匹配检查（标蓝色）
        green_cells = self._execute_step(
            15, "cat6- cat1 包含匹配检查", self._step_cat6_cat1_containment_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if green_cells is None:
            return None
        if progress:
            progress.update(1)

        # 跨平台步骤9：平台外覆盖检查 + 特殊规则检查（标橙色）
        yellow_cells = self._execute_step(
            16, "平台外覆盖检查", self._step_platform_outside_check,
            sheet_info_list, output_workbook, green_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if yellow_cells is None:
            return None

        gray_cells = self._execute_step(
            17, "特殊规则检查", self._step_special_rule_check,
            sheet_info_list, output_workbook, green_cells, yellow_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if gray_cells is None:
            return None
        if progress:
            progress.update(1)

        # 跨平台步骤10：特殊地址段检查 + cat2特殊源地址非平台目的地址检查（标红色）
        yellow_light_cells = self._execute_step(
            18, "特殊地址段检查", self._step_special_address_check,
            sheet_info_list, output_workbook, green_cells, yellow_cells,
            gray_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if yellow_light_cells is None:
            return None

        red_cells = self._execute_step(
            19, "cat2特殊源地址非平台目的地址检查",
            self._step_cat2_special_src_non_platform_dst_check,
            sheet_info_list, output_workbook, green_cells, yellow_cells,
            gray_cells, yellow_light_cells, rules_cache=rules_cache,
            progress=None, save_after_step=SAVE_AFTER_EACH_STEP,
            output_path=output_path, stop_at_step=STOP_AT_STEP
        )
        if red_cells is None:
            return None

        # 删除不包含规则的access-list 也算入最后一个跨平台步骤的进度
        self._remove_empty_acls(sheet_info_list, progress=None)
        if progress:
            progress.update(1)


        self.add_result(Level.OK, "所有步骤执行完成")
        return True


    # 删除不包含规则的access-list
    def _remove_empty_acls(self, sheet_info_list, progress):
        self.add_result(Level.OK, "开始执行最后一步：删除不包含规则的access-list...")
        if progress:
            progress.set_description(f"{self.NAME} (最后一步: 删除空access-list)")


        total_empty_acl_all = 0
        cells_to_clear = []


        # 第一阶段：标记要清空的单元格
        for sheet_info in sheet_info_list:
            sheet_name = sheet_info['sheet_name']
            output_worksheet = sheet_info['output_worksheet']
            saved_max_row = output_worksheet.max_row


            for output_col in range(1, output_worksheet.max_column + 1):
                empty_acl_count = 0
                current_acl_row = None


                for row in range(2, saved_max_row + 1):
                    cell = output_worksheet.cell(row=row, column=output_col)
                    if cell.value and isinstance(cell.value, str):
                        cell_text = str(cell.value).strip().lower()


                        if cell_text.startswith('ip access-list '):
                            if current_acl_row is not None:
                                cells_to_clear.append((sheet_name, output_col, current_acl_row))
                                empty_acl_count += 1
                            current_acl_row = row
                        elif 'permit' in cell_text or 'deny' in cell_text:
                            current_acl_row = None


                if current_acl_row is not None:
                    cells_to_clear.append((sheet_name, output_col, current_acl_row))
                    empty_acl_count += 1


                if empty_acl_count > 0:
                    total_empty_acl_all += empty_acl_count


        # 第二阶段：清空标记的单元格
        if cells_to_clear:
            for sheet_name, output_col, row in cells_to_clear:
                for sheet_info in sheet_info_list:
                    if sheet_info['sheet_name'] == sheet_name:
                        cell = sheet_info['output_worksheet'].cell(row=row, column=output_col)
                        cell.value = None
                        cell.font = Font()
                        cell.fill = PatternFill()
                        break


        # 第三阶段：批量上移空单元格
        for sheet_info in sheet_info_list:
            sheet_name = sheet_info['sheet_name']
            output_worksheet = sheet_info['output_worksheet']


            cols_to_shift = {col for sname, col, _ in cells_to_clear if sname == sheet_name}
            cells_by_col = {}
            for sname, col, row in cells_to_clear:
                if sname == sheet_name:
                    cells_by_col.setdefault(col, []).append(row)


            for output_col in cols_to_shift:
                if output_col not in cells_by_col:
                    continue


                rows_to_shift = sorted(cells_by_col[output_col], reverse=True)
                current_max_row = output_worksheet.max_row


                for i, target_row in enumerate(rows_to_shift):
                    adjusted_row = target_row - i
                    if adjusted_row > current_max_row or adjusted_row < 2:
                        continue


                    cell = output_worksheet.cell(row=adjusted_row, column=output_col)
                    if (cell.value is not None and
                            (not isinstance(cell.value, str) or
                                cell.value.strip())):
                        continue


                    if adjusted_row < current_max_row:
                        for shift_row in range(adjusted_row, current_max_row):
                            source_cell = output_worksheet.cell(
                                row=shift_row + 1, column=output_col
                            )
                            target_cell = output_worksheet.cell(
                                row=shift_row, column=output_col
                            )
                            if source_cell.value is not None:
                                target_cell.value = source_cell.value
                                source_font = source_cell.font
                                target_cell.font = Font(
                                    name=source_font.name, size=source_font.size,
                                    bold=source_font.bold, italic=source_font.italic,
                                    vertAlign=source_font.vertAlign,
                                    underline=source_font.underline,
                                    strike=source_font.strike, color=source_font.color
                                )
                                source_fill = source_cell.fill
                                target_cell.fill = PatternFill(
                                    fill_type=source_fill.fill_type,
                                    start_color=source_fill.start_color,
                                    end_color=source_fill.end_color
                                )
                            else:
                                target_cell.value = None
                                target_cell.font = Font()
                                target_cell.fill = PatternFill()


                    if current_max_row >= 2:
                        last_cell = output_worksheet.cell(row=current_max_row, column=output_col)
                        last_cell.value = None
                        last_cell.font = Font()
                        last_cell.fill = PatternFill()


                    current_max_row -= 1


        if total_empty_acl_all > 0:
            self.add_result(Level.OK, f"最后一步执行完成：共删除{total_empty_acl_all}个不包含规则的access-list")
        else:
            self.add_result(Level.OK, "最后一步执行完成：未发现不包含规则的access-list")


        if progress:
            progress.update(1)


    # 主执行方法
    def run(self) -> None:
        """执行ACL跨平台检查任务

        处理所有工作表，执行ACL跨平台检查并生成报告
        """
        task_items = list(self.items())
        if not task_items:
            self.add_result(Level.ERROR, "未找到可处理的 Sheet")
            return

        # 确保输出目录存在
        ensure_output_dir(self.OUTPUT_DIR)

        # 生成输出文件名
        today = get_today_str()
        output_filename = f"{today}-跨平台N9K&LINKAS&OOB ACL交叉检查.xlsx"
        output_path = build_output_path(self.OUTPUT_DIR, output_filename)

        # 打开输入Excel文件
        try:
            input_workbook = load_excel_workbook(self.INPUT_PATH)
        except (FileNotFoundError, PermissionError, IOError, OSError, ValueError) as EXCEPTION:
            self.add_result(Level.ERROR, f"无法打开输入文件 {self.INPUT_PATH}: {EXCEPTION}")
            return

        # 创建输出Excel工作簿
        output_workbook = create_excel_workbook()

        # 使用父类的进度条处理
        from tqdm import tqdm

        from .TaskBase import BAR_FORMAT, SHOW_PROGRESS

        # total_steps 说明（完全按“概念步骤”统计）：
        # - 基础步骤1-3：3 步（所有Sheet的基础处理完成后一次性+3）
        # - 同平台步骤1-3：3 步（在提取同平台ACL时分别+1）
        # - 跨平台步骤1-10：10 步（内部多个细分子步骤合并为10个概念步骤）
        total_steps = 3 + 3 + 10
        progress = tqdm(
            total=total_steps,
            desc=self.NAME,
            position=0,
            leave=True,
            dynamic_ncols=True,
            bar_format=BAR_FORMAT,
        ) if SHOW_PROGRESS else None

        try:
            # 处理所有Sheet（基础步骤1-3）
            sheet_info_list = self._process_all_sheets(
                task_items, input_workbook, output_workbook, progress
            )
            if progress:
                progress.update(3)  # 基础步骤1-3完成

            # 如果当前模式仅执行同平台步骤，可以在此处直接返回
            # return

            # 预构建规则缓存
            rules_cache = self._build_rules_cache(sheet_info_list, progress=None)

            # 执行跨平台步骤1-10（内部按概念步骤更新进度）
            result = self._execute_all_steps(
                sheet_info_list, output_workbook, rules_cache, progress,
                output_path
            )
            if result is None:
                return

        finally:
            # 先保存输出Excel文件（在关闭进度条之前，避免进度条错误影响保存）
            try:
                save_excel_workbook(output_workbook, output_path)
                self.add_result(Level.OK, f"输出文件已保存: {output_path}")
            except (PermissionError, IOError, OSError, ValueError) as EXCEPTION:
                self.add_result(Level.ERROR, f"保存输出文件失败: {EXCEPTION}")


            # 关闭进度条（可能出错，但不影响文件保存）
            if progress:
                try:
                    progress.close()
                except (AttributeError, RuntimeError):
                    pass


            # 关闭工作簿
            try:
                input_workbook.close()
            except (AttributeError, IOError, OSError):
                pass
            try:
                output_workbook.close()
            except (AttributeError, IOError, OSError):
                pass

