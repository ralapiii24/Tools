# Cisco ACL 解析基础模块（V11新增）
#
# 技术栈:Python, 正则表达式, ipaddress, openpyxl, 网络地址计算
# 目标:统一ACL解析功能，提供通用的ACL规则解析和Excel处理函数，供多个ACL任务共用
#
# 主要功能:
# - parse_acl: 统一ACL规则解析函数，支持NX-OS CIDR格式和IOS-XE wildcard/host混合格式，解析源/目的网段、协议、端口等信息
# - find_acl_blocks_in_column: 在Excel列中查找ACL块，支持登录ACL结束标记检测（大小写不敏感，匹配包含vty和ip的ip access-list行）
# - extract_acl_rules_from_column: 从Excel列中提取ACL规则列表
# - is_acl_rule: 判断是否为ACL规则行
# - 服务端口解析: service_to_port函数，支持服务名和端口号转换
#
# 使用场景:
# - ACLCrossCheckTask: 使用parse_acl、find_acl_blocks_in_column、extract_acl_rules_from_column
# - ACLArpCheckTask: 使用find_acl_blocks_in_column
# - ACLDupCheckTask: 使用find_acl_blocks_in_column
#

# 导入标准库
import re
import socket
from dataclasses import dataclass
from ipaddress import IPv4Address, IPv4Network
from typing import Dict, List, Optional, Set, Tuple

# 导入第三方库
try:
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    WORKSHEET_TYPE = None  # 类型提示用

# 导入本地应用
# (无本地应用依赖)

# 解析辅助函数
# 仅数字或系统services数据库解析；失败返回None（视为任意端口）
def service_to_port(SERVICE_TEXT: str) -> Optional[int]:
    """将服务名或端口号字符串转换为端口号

    Args:
        SERVICE_TEXT: 服务名（如 "http"）或端口号字符串（如 "80"）

    Returns:
        端口号，转换失败返回None
    """
    if SERVICE_TEXT is None:
        return None
    SERVICE_TEXT = SERVICE_TEXT.strip().lower()
    if not SERVICE_TEXT:
        return None
    if SERVICE_TEXT.isdigit():
        PORT_NUMBER = int(SERVICE_TEXT)
        return PORT_NUMBER if 0 <= PORT_NUMBER <= 65535 else None
    try:
        return socket.getservbyname(SERVICE_TEXT)
    except (OSError, socket.gaierror):
        return None

# IOS-XE：ip + wildcard转网络（假设通配位连续）
def ip_and_wildcard_to_network(IP_TEXT: str, WILDCARD_TEXT: str) -> Optional[IPv4Network]:
    """将IP地址和通配符转换为IPv4Network对象

    Args:
        IP_TEXT: IP地址字符串
        WILDCARD_TEXT: 通配符字符串

    Returns:
        IPv4Network对象，转换失败返回None
    """
    try:
        IP_ADDRESS_VALUE = int(IPv4Address(IP_TEXT))
        WILDCARD_ADDRESS_VALUE = int(IPv4Address(WILDCARD_TEXT))
        NETMASK_INTEGER_VALUE = (~WILDCARD_ADDRESS_VALUE) & 0xFFFFFFFF
        PREFIX_LENGTH = 32 - bin(WILDCARD_ADDRESS_VALUE).count("1")
        NETWORK_INTEGER_VALUE = IP_ADDRESS_VALUE & NETMASK_INTEGER_VALUE
        return IPv4Network((IPv4Address(NETWORK_INTEGER_VALUE), PREFIX_LENGTH), strict=False)
    except (ValueError, TypeError):
        return None

# 将IP地址转换为/32网络
def host_to_network(IP_STRING: str) -> Optional[IPv4Network]:
    """将IP地址字符串转换为/32网络

    Args:
        IP_STRING: IP地址字符串

    Returns:
        IPv4Network对象，转换失败返回None
    """
    try:
        return IPv4Network(f"{IP_STRING}/32", strict=False)
    except (ValueError, TypeError):
        return None

# 将CIDR字符串转换为网络对象
def cidr_to_network(CIDR_STRING: str) -> Optional[IPv4Network]:
    """将CIDR字符串转换为IPv4Network对象

    Args:
        CIDR_STRING: CIDR格式字符串（如 "192.168.1.0/24"）

    Returns:
        IPv4Network对象，转换失败返回None
    """
    try:
        return IPv4Network(CIDR_STRING, strict=False)
    except (ValueError, TypeError):
        return None

# 将'any'关键字转换为0.0.0.0/0网络
def any_to_network() -> IPv4Network:
    """将'any'关键字转换为0.0.0.0/0网络

    Returns:
        IPv4Network对象，表示任意网络（0.0.0.0/0）
    """
    return IPv4Network("0.0.0.0/0", strict=False)

# 从端口字符串中提取所有端口：支持多个端口（用空格分隔），例如"22 22222"或"domain ntp"，返回端口号集合
def _extract_all_ports(port_str: str) -> Set[int]:
    # 从端口字符串中提取所有端口号
    if not port_str:
        return set()


    PORTS = set()
    PORT_ITEM_LIST = port_str.strip().split()


    for PORT_ITEM_ELEMENT in PORT_ITEM_LIST:
        PORT_ITEM_ELEMENT = PORT_ITEM_ELEMENT.strip()
        if not PORT_ITEM_ELEMENT:
            continue


        # 尝试转换为端口号
        if PORT_ITEM_ELEMENT.isdigit():
            PORTS.add(int(PORT_ITEM_ELEMENT))
        else:
            # 尝试将服务名转换为端口号
            try:
                PORT = service_to_port(PORT_ITEM_ELEMENT)
                if PORT is not None:
                    PORTS.add(PORT)
            except (ValueError, TypeError, OSError):
                pass


    return PORTS

# ============================================================================
# 正则表达式定义 - 支持所有Cisco ACL格式
# ============================================================================

# NXOS格式 - 标准格式（目标端口）
# 示例: 10 permit tcp 10.10.0.0/16 10.20.0.0/16 eq 80
# 注意: 使用宽松匹配 (?:\s+\S+)* 以兼容log-input、time-range等关键字（与ACLCrossCheckTask保持一致）
NXOS_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE>\d+\.\d+\.\d+\.\d+/\d+)\s+
    (?P<DESTINATION>\d+\.\d+\.\d+\.\d+/\d+)
    (?:\s+eq\s+(?P<PORT>\S+))?
    (?:\s+\S+)*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# NXOS格式 - 源端口格式
# 示例: 162 permit tcp 10.10.100.31/32 eq 55888 10.10.108.63/32
NXOS_SOURCE_PORT_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE>\d+\.\d+\.\d+\.\d+/\d+)\s+
    eq\s+(?P<PORT>\S+)\s+
    (?P<DESTINATION>\d+\.\d+\.\d+\.\d+/\d+)
    (?:\s+eq\s+(?P<DST_PORT>\S+))?
    (?:\s+log)?
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# NXOS格式 - 端口范围格式（两种格式）
# 格式1: 源地址 目标地址 range 起始端口 结束端口
# 格式2: 源地址 range 起始端口 结束端口 目标地址
# 示例: 182 permit tcp 10.10.106.40/32 range 8001 8002 10.10.62.32/31
NXOS_RANGE_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE>\d+\.\d+\.\d+\.\d+/\d+)\s+
    (?:
        (?P<DESTINATION_BEFORE>\d+\.\d+\.\d+\.\d+/\d+)\s+range\s+
        (?P<PORT_START>\d+)\s+(?P<PORT_END>\d+) |
        range\s+(?P<PORT_START2>\d+)\s+(?P<PORT_END2>\d+)\s+
        (?P<DESTINATION_AFTER>\d+\.\d+\.\d+\.\d+/\d+)
    )
    (?:\s+log)?
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# IOS-XE格式 - Wildcard格式（标准）
# 示例: 10 permit tcp 10.10.0.0 0.0.255.255 10.20.0.0 0.0.255.255 eq 80
IOSXE_WILDCARD_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<SOURCE_WILDCARD>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+(?P<PORT_A>\S+))?
    \s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<DESTINATION_WILDCARD>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+(?P<PORT_B>\S+))?
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# IOS-XE格式 - Host格式
# 示例: 10 permit tcp host 10.10.0.1 host 10.20.0.1 eq 80
IOSXE_HOST_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    host\s+(?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+(?P<PORT_A>\S+))?
    \s+host\s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+(?P<PORT_B>\S+))?
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# IOS-XE格式 - 混合格式（host和wildcard混合）
# 示例: 10 permit tcp host 10.10.0.1 10.20.0.0 0.0.255.255 eq 80
IOSXE_MIX_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?:
        host\s+(?P<SOURCE_IP_HOST>\d+\.\d+\.\d+\.\d+) |
        (?P<SOURCE_IP_WILDCARD>\d+\.\d+\.\d+\.\d+)\s+
        (?P<SOURCE_WILDCARD_WILDCARD>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+(?P<PORT_A>[\w\s]+))?
    \s+
    (?:
        host\s+(?P<DESTINATION_IP_HOST>\d+\.\d+\.\d+\.\d+) |
        (?P<DESTINATION_IP_WILDCARD>\d+\.\d+\.\d+\.\d+)\s+
        (?P<DESTINATION_WILDCARD_WILDCARD>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+(?P<PORT_B>[\w\s]+))?
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# IOS-XE格式 - host <ip> eq <port1> <port2> ... <dst_ip> <wildcard>
# 示例: 4190 permit udp host 10.65.16.53 eq domain ntp 10.70.130.0 0.0.0.255 log
IOSXE_HOST_MULTI_EQ_WILDCARD_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    host\s+(?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)
    \s+eq\s+(?P<PORT_MULTI>[\w\s]+)
    \s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<DESTINATION_WILDCARD>\d+\.\d+\.\d+\.\d+)
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# IOS-XE格式 - 支持range和多个eq端口（源IP+通配符，目的host）
# 示例: 290 permit tcp 10.70.130.0 0.0.0.31 range 7180 8088 host 10.65.63.55 log
# 示例: 300 permit tcp 10.70.130.0 0.0.0.31 eq 8888 9000 9010 9020 9030 9083 9870 10000 host 10.65.63.55 log
IOSXE_RANGE_MULTI_EQ_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<SOURCE_WILDCARD>\d+\.\d+\.\d+\.\d+)
    \s+(?:range\s+(?P<PORT_RANGE_START>\d+)\s+(?P<PORT_RANGE_END>\d+)|eq\s+(?P<PORT_MULTI>[\w\s]+))
    \s+host\s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# IOS-XE格式 - 支持range端口（源host，目的IP+通配符）
# 示例: 3570 permit tcp host 10.65.130.233 range 6446 6447 10.66.231.8 0.0.0.7
IOSXE_RANGE_HOST_SOURCE_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    host\s+(?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)
    \s+range\s+(?P<PORT_RANGE_START>\d+)\s+(?P<PORT_RANGE_END>\d+)
    \s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<DESTINATION_WILDCARD>\d+\.\d+\.\d+\.\d+)
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# IOS-XE格式 - 支持range端口（源IP+通配符，目的IP+通配符）
# 示例: 4580 permit tcp 10.62.110.96 0.0.0.31 range 9091 9093 10.66.130.0 0.0.0.255
# 示例: 640 permit tcp 10.10.0.0 0.0.255.255 range 8000 9999 10.104.166.0 0.0.0.255
IOSXE_RANGE_WILDCARD_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<SOURCE_WILDCARD>\d+\.\d+\.\d+\.\d+)
    \s+range\s+(?P<PORT_RANGE_START>\d+)\s+(?P<PORT_RANGE_END>\d+)
    \s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<DESTINATION_WILDCARD>\d+\.\d+\.\d+\.\d+)
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# IOS-XE格式 - 支持多个eq端口（源IP+通配符，目的IP+通配符）
# 示例: 180 permit tcp 10.65.88.192 0.0.0.63 eq www 443 8400 10.62.80.0 0.0.0.7
# 示例: 260 permit tcp 10.70.130.64 0.0.0.3 eq 3366 8030 9030 10.66.130.0 0.0.0.15 log
# 示例: permit tcp 10.10.15.0 0.0.0.255 eq www 443 5480 5900 10.10.62.32 0.0.0.1 log
IOSXE_MULTI_EQ_WILDCARD_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<SOURCE_WILDCARD>\d+\.\d+\.\d+\.\d+)
    \s+eq\s+(?P<PORT_MULTI>[\w\s]+)
    \s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<DESTINATION_WILDCARD>\d+\.\d+\.\d+\.\d+)
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# IOS-XE格式 - Any关键字格式
# 示例: 10 permit tcp any any eq 80
IOSXE_ANY_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?:
        any |
        (?P<SRC_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<SRC_WC>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+.*?)?
    (?:\s+range\s+.*?)?
    \s+
    (?:
        any |
        (?P<DST_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<DST_WC>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+.*?)?
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# ASA格式 - 支持CIDR和IP地址
# 示例: permit tcp 10.10.0.0/16 10.20.0.0/16 eq www
# 示例: permit tcp any any
ASA_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SRC>any|\d+\.\d+\.\d+\.\d+(?:/\d+)?)
    (?:\s+(?:eq\s+)?(?P<PORT_A>\S+))?
    (?:\s+range\s+(?P<PORT_RANGE_START>\S+)\s+(?P<PORT_RANGE_END>\S+))?
    \s+
    (?P<DST>any|\d+\.\d+\.\d+\.\d+(?:/\d+)?)
    (?:\s+(?:eq\s+)?(?P<PORT_B>\S+))?
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# 混合格式 - 源地址wildcard + 目的地址CIDR
# 示例: 99 permit ip 10.6.26.0 0.0.0.255 10.6.26.254/32
IOSXE_WILDCARD_SOURCE_CIDR_DESTINATION_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<SOURCE_WILDCARD>\d+\.\d+\.\d+\.\d+)
    \s+(?P<DESTINATION>\d+\.\d+\.\d+\.\d+/\d+)
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# 混合格式 - 源地址CIDR + 目的地址wildcard
# 示例: 94 permit tcp 10.12.8.43/32 eq 28800 10.12.17.80 0.0.0.7
IOSXE_CIDR_SOURCE_WILDCARD_DESTINATION_REGULAR_EXPRESSION = re.compile(
    r"""
    ^\s*(?P<NUMBER>\d+)?\s*
    (?P<ACTION>permit|deny)\s+
    (?P<PROTOCOL>\S+)\s+
    (?P<SOURCE>\d+\.\d+\.\d+\.\d+/\d+)
    (?:\s+eq\s+(?P<PORT>\S+))?
    \s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)\s+(?P<DESTINATION_WILDCARD>\d+\.\d+\.\d+\.\d+)
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

# NXOS格式 - 源端口 + 多端口（源端口后跟多个端口）
# 示例: 需要进一步确认是否存在这种格式
# 暂时不添加，等待实际数据验证

# ============================================================================
# ACL规则数据类
# ============================================================================

# ACL规则数据类：存储解析后的ACL规则的所有信息，包括原始文本、动作、协议、源/目的网络、端口信息、规则格式
@dataclass
class ACLRule:
    """ACL规则数据类

    存储ACL规则的解析结果，包括源/目的网络、端口、动作等信息
    """
    raw: str
    action: str
    proto: str
    src: IPv4Network
    dst: IPv4Network
    port: Optional[int] = None  # None表示任意端口（用于兼容，通常是第一个端口）
    src_port: Optional[int] = None  # 源端口
    dst_port: Optional[int] = None  # 目标端口
    style: str = ""  # 'NXOS' / 'IOS-XE' / 'ASA'
    ports: Optional[Set[int]] = None  # 多个端口的集合（用于存储eq 22 22222这样的多个端口）

# ============================================================================
# 完整ACL解析函数
# ============================================================================

# 完整解析ACL规则行，支持所有Cisco ACL格式（包括any规则）
# 这是最底层的解析函数，支持所有格式和关键字。如果需要忽略any规则，请使用parse_acl()函数。
# Supported Formats: NXOS、IOS-XE、ASA等所有格式，支持any关键字、log、log-input、time-range等
def parse_acl_full(ACL_LINE: str) -> Tuple[Optional[ACLRule], Optional[str]]:
    """完整解析ACL规则行，支持所有Cisco ACL格式（包括any规则）

    这是最底层的解析函数，支持所有格式和关键字。如果需要忽略any规则，请使用parse_acl()函数。

    Args:
        ACL_LINE: ACL规则文本行（可包含行号前缀）

    Returns:
        元组：(ACLRule对象, None) 成功，或 (None, 错误信息) 失败
    """
    CLEANED_LINE = (ACL_LINE or "").strip()
    if not CLEANED_LINE:
        return None, "empty"


    # 去除行号前缀（如果存在）
    line_number_pattern = re.compile(r'^\s*\d+\s+(permit|deny)', re.IGNORECASE)
    if line_number_pattern.match(CLEANED_LINE):
        match = re.search(r'\b(permit|deny)\b', CLEANED_LINE, re.IGNORECASE)
        if match:
            CLEANED_LINE = CLEANED_LINE[match.start():]


    # 按优先级尝试匹配各种格式（从最具体到最通用）


    # 1. NXOS源端口格式（优先，因为更具体）
    MATCH_OUTCOME = NXOS_SOURCE_PORT_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        PORT_NUMBER = (
            service_to_port(MATCH_OUTCOME.group("PORT"))
            if MATCH_OUTCOME.group("PORT") else None
        )
        DESTINATION_PORT = (
            service_to_port(MATCH_OUTCOME.group("DST_PORT"))
            if MATCH_OUTCOME.group("DST_PORT") else None
        )
        SOURCE_NETWORK = cidr_to_network(MATCH_OUTCOME.group("SOURCE"))
        DESTINATION_NETWORK = cidr_to_network(MATCH_OUTCOME.group("DESTINATION"))
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                PORT_NUMBER,  # 源端口
                DESTINATION_PORT,  # 目标端口
                "NXOS"
            ), None
        return None, "nxos_src_port_network_parse_fail"


    # 2. NXOS端口范围格式
    MATCH_OUTCOME = NXOS_RANGE_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        SOURCE_NETWORK = cidr_to_network(MATCH_OUTCOME.group("SOURCE"))
        if MATCH_OUTCOME.group("DESTINATION_BEFORE"):
            DESTINATION_NETWORK = cidr_to_network(MATCH_OUTCOME.group("DESTINATION_BEFORE"))
        elif MATCH_OUTCOME.group("DESTINATION_AFTER"):
            DESTINATION_NETWORK = cidr_to_network(MATCH_OUTCOME.group("DESTINATION_AFTER"))
        else:
            DESTINATION_NETWORK = None
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                None,  # 端口范围，设为None
                None,
                None,
                "NXOS"
            ), None
        return None, "nxos_range_network_parse_fail"


    # 3. NXOS标准格式
    MATCH_OUTCOME = NXOS_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        PORT_NUMBER = (
            service_to_port(MATCH_OUTCOME.group("PORT"))
            if MATCH_OUTCOME.group("PORT") else None
        )
        SOURCE_NETWORK = cidr_to_network(MATCH_OUTCOME.group("SOURCE"))
        DESTINATION_NETWORK = cidr_to_network(MATCH_OUTCOME.group("DESTINATION"))
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                None,
                PORT_NUMBER,
                "NXOS"
            ), None
        return None, "nxos_network_parse_fail"


    # 4. IOS-XE多端口格式（源IP+通配符，目的IP+通配符）
    MATCH_OUTCOME = IOSXE_MULTI_EQ_WILDCARD_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        PORT_MULTI_TEXT = MATCH_OUTCOME.group("PORT_MULTI").strip()
        ALL_PORT_NUMBERS = _extract_all_ports(PORT_MULTI_TEXT)
        SOURCE_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("SOURCE_IP"),
            MATCH_OUTCOME.group("SOURCE_WILDCARD")
        )
        DESTINATION_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("DESTINATION_IP"),
            MATCH_OUTCOME.group("DESTINATION_WILDCARD")
        )
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            PORT_NUMBER = min(ALL_PORT_NUMBERS) if ALL_PORT_NUMBERS else None
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                None,
                None,
                "IOS-XE",
                ALL_PORT_NUMBERS if len(ALL_PORT_NUMBERS) > 1 else None
            ), None
        return None, "iosxe_multi_eq_wildcard_network_parse_fail"


    # 5. IOS-XE端口范围格式（源IP+通配符，目的IP+通配符）
    MATCH_OUTCOME = IOSXE_RANGE_WILDCARD_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        SOURCE_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("SOURCE_IP"),
            MATCH_OUTCOME.group("SOURCE_WILDCARD")
        )
        DESTINATION_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("DESTINATION_IP"),
            MATCH_OUTCOME.group("DESTINATION_WILDCARD")
        )
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                None,
                None,
                None,
                "IOS-XE"
            ), None
        return None, "iosxe_range_wildcard_network_parse_fail"


    # 6. IOS-XE host + 多端口 + wildcard格式
    MATCH_OUTCOME = IOSXE_HOST_MULTI_EQ_WILDCARD_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        PORT_MULTI_TEXT = MATCH_OUTCOME.group("PORT_MULTI").strip()
        ALL_PORT_NUMBERS = _extract_all_ports(PORT_MULTI_TEXT)
        SOURCE_NETWORK = host_to_network(MATCH_OUTCOME.group("SOURCE_IP"))
        DESTINATION_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("DESTINATION_IP"),
            MATCH_OUTCOME.group("DESTINATION_WILDCARD")
        )
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            PORT_NUMBER = min(ALL_PORT_NUMBERS) if ALL_PORT_NUMBERS else None
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                PORT_NUMBER,
                None,
                "IOS-XE",
                ALL_PORT_NUMBERS if len(ALL_PORT_NUMBERS) > 1 else None
            ), None
        return None, "iosxe_host_multi_eq_wildcard_network_parse_fail"


    # 7. IOS-XE range + multi eq格式（源IP+通配符，目的host）
    MATCH_OUTCOME = IOSXE_RANGE_MULTI_EQ_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        SOURCE_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("SOURCE_IP"),
            MATCH_OUTCOME.group("SOURCE_WILDCARD")
        )
        DESTINATION_NETWORK = host_to_network(MATCH_OUTCOME.group("DESTINATION_IP"))
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                None,
                None,
                None,
                "IOS-XE"
            ), None
        return None, "iosxe_range_multi_eq_network_parse_fail"


    # 8. IOS-XE range格式（源host，目的IP+通配符）
    MATCH_OUTCOME = IOSXE_RANGE_HOST_SOURCE_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        SOURCE_NETWORK = host_to_network(MATCH_OUTCOME.group("SOURCE_IP"))
        DESTINATION_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("DESTINATION_IP"),
            MATCH_OUTCOME.group("DESTINATION_WILDCARD")
        )
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                None,
                None,
                None,
                "IOS-XE"
            ), None
        return None, "iosxe_range_host_src_network_parse_fail"


    # 9. IOS-XE混合格式（host和wildcard混合）
    MATCH_OUTCOME = IOSXE_MIX_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        PORT_A_TEXT = MATCH_OUTCOME.group("PORT_A")
        PORT_B_TEXT = MATCH_OUTCOME.group("PORT_B")


        # 判断端口位置
        port_a_before_destination_position = False
        if PORT_A_TEXT and not PORT_B_TEXT:
            eq_pos = CLEANED_LINE.lower().find("eq " + PORT_A_TEXT.lower().split()[0])
            destination_host_position = CLEANED_LINE.lower().find(
                "host", eq_pos if eq_pos >= 0 else 0
            )
            destination_wildcard_position = -1
            if MATCH_OUTCOME.group("DESTINATION_IP_WILDCARD"):
                DESTINATION_IP_WILDCARD = MATCH_OUTCOME.group(
                    "DESTINATION_IP_WILDCARD"
                )
                destination_wildcard_position = CLEANED_LINE.lower().find(
                    DESTINATION_IP_WILDCARD.lower(),
                    eq_pos if eq_pos >= 0 else 0
                )
            if eq_pos >= 0:
                if destination_host_position > eq_pos or (
                    destination_wildcard_position > eq_pos
                    and destination_wildcard_position > 0
                ):
                    port_a_before_destination_position = True


        ALL_PORT_NUMBERS = set()
        SOURCE_PORT = None
        DESTINATION_PORT = None


        if port_a_before_destination_position:
            if PORT_A_TEXT:
                ALL_PORT_NUMBERS.update(_extract_all_ports(PORT_A_TEXT))
                PORT_A_ITEM_LIST = PORT_A_TEXT.strip().split()
                if PORT_A_ITEM_LIST:
                    DESTINATION_PORT = service_to_port(PORT_A_ITEM_LIST[0])
            if PORT_B_TEXT:
                ALL_PORT_NUMBERS.update(_extract_all_ports(PORT_B_TEXT))
                PORT_B_ITEM_LIST = PORT_B_TEXT.strip().split()
                if PORT_B_ITEM_LIST:
                    SOURCE_PORT = service_to_port(PORT_B_ITEM_LIST[0])
        else:
            if PORT_A_TEXT:
                ALL_PORT_NUMBERS.update(_extract_all_ports(PORT_A_TEXT))
                PORT_A_ITEM_LIST = PORT_A_TEXT.strip().split()
                if PORT_A_ITEM_LIST:
                    SOURCE_PORT = service_to_port(PORT_A_ITEM_LIST[0])
            if PORT_B_TEXT:
                ALL_PORT_NUMBERS.update(_extract_all_ports(PORT_B_TEXT))
                PORT_B_ITEM_LIST = PORT_B_TEXT.strip().split()
                if PORT_B_ITEM_LIST:
                    DESTINATION_PORT = service_to_port(PORT_B_ITEM_LIST[0])


        if (SOURCE_PORT is not None and DESTINATION_PORT is not None and
                SOURCE_PORT != DESTINATION_PORT):
            return None, "conflicting_ports"


        PORT_NUMBER = DESTINATION_PORT if DESTINATION_PORT is not None else SOURCE_PORT
        SOURCE_IP_HOST = MATCH_OUTCOME.group("SOURCE_IP_HOST")
        SOURCE_IP_WILDCARD = MATCH_OUTCOME.group("SOURCE_IP_WILDCARD")
        SOURCE_WILDCARD_WILDCARD = MATCH_OUTCOME.group("SOURCE_WILDCARD_WILDCARD")
        SOURCE_NETWORK = (
            host_to_network(SOURCE_IP_HOST) if SOURCE_IP_HOST
            else ip_and_wildcard_to_network(SOURCE_IP_WILDCARD, SOURCE_WILDCARD_WILDCARD)
        )
        DESTINATION_IP_HOST = MATCH_OUTCOME.group("DESTINATION_IP_HOST")
        DESTINATION_IP_WILDCARD = MATCH_OUTCOME.group("DESTINATION_IP_WILDCARD")
        DESTINATION_WILDCARD_WILDCARD = MATCH_OUTCOME.group("DESTINATION_WILDCARD_WILDCARD")
        DESTINATION_NETWORK = (
            host_to_network(DESTINATION_IP_HOST) if DESTINATION_IP_HOST
            else ip_and_wildcard_to_network(DESTINATION_IP_WILDCARD, DESTINATION_WILDCARD_WILDCARD)
        )
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            PORT_NUMBER_SET = ALL_PORT_NUMBERS if len(ALL_PORT_NUMBERS) > 1 else (ALL_PORT_NUMBERS if ALL_PORT_NUMBERS else None)
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                SOURCE_PORT,
                DESTINATION_PORT,
                "IOS-XE",
                PORT_NUMBER_SET
            ), None
        return None, "iosxe_mix_network_parse_fail"


    # 10. IOS-XE Wildcard格式（标准）
    MATCH_OUTCOME = IOSXE_WILDCARD_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        PORT_NUMBER = (
            service_to_port(MATCH_OUTCOME.group("PORT_A")) or
            service_to_port(MATCH_OUTCOME.group("PORT_B"))
        )
        SOURCE_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("SOURCE_IP"),
            MATCH_OUTCOME.group("SOURCE_WILDCARD")
        )
        DESTINATION_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("DESTINATION_IP"),
            MATCH_OUTCOME.group("DESTINATION_WILDCARD")
        )
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                None,
                None,
                "IOS-XE"
            ), None
        return None, "iosxe_wc_network_parse_fail"


    # 11. IOS-XE Host格式
    MATCH_OUTCOME = IOSXE_HOST_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        PORT_NUMBER = (
            service_to_port(MATCH_OUTCOME.group("PORT_A")) or
            service_to_port(MATCH_OUTCOME.group("PORT_B"))
        )
        SOURCE_NETWORK = host_to_network(MATCH_OUTCOME.group("SOURCE_IP"))
        DESTINATION_NETWORK = host_to_network(MATCH_OUTCOME.group("DESTINATION_IP"))
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                None,
                None,
                "IOS-XE"
            ), None
        return None, "iosxe_host_network_parse_fail"


    # 12. 混合格式 - 源地址wildcard + 目的地址CIDR
    MATCH_OUTCOME = IOSXE_WILDCARD_SOURCE_CIDR_DESTINATION_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        SOURCE_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("SOURCE_IP"),
            MATCH_OUTCOME.group("SOURCE_WILDCARD")
        )
        DESTINATION_NETWORK = cidr_to_network(MATCH_OUTCOME.group("DESTINATION"))
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                None,
                None,
                None,
                "IOS-XE"
            ), None
        return None, "iosxe_wc_src_cidr_dst_network_parse_fail"


    # 13. 混合格式 - 源地址CIDR + 目的地址wildcard
    MATCH_OUTCOME = IOSXE_CIDR_SOURCE_WILDCARD_DESTINATION_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        PORT_NUMBER = (
            service_to_port(MATCH_OUTCOME.group("PORT"))
            if MATCH_OUTCOME.group("PORT") else None
        )
        SOURCE_NETWORK = cidr_to_network(MATCH_OUTCOME.group("SOURCE"))
        DESTINATION_NETWORK = ip_and_wildcard_to_network(
            MATCH_OUTCOME.group("DESTINATION_IP"),
            MATCH_OUTCOME.group("DESTINATION_WILDCARD")
        )
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                PORT_NUMBER,  # 源端口
                None,
                "IOS-XE"
            ), None
        return None, "iosxe_cidr_src_wc_dst_network_parse_fail"


    # 14. IOS-XE Any关键字格式（host + any）
    # 示例: 10 permit tcp host 10.10.80.1 any eq 22 log
    IOSXE_HOST_ANY_REGULAR_EXPRESSION = re.compile(
        r"""
        ^\s*(?P<NUMBER>\d+)?\s*
        (?P<ACTION>permit|deny)\s+
        (?P<PROTOCOL>\S+)\s+
        host\s+(?P<SOURCE_IP>\d+\.\d+\.\d+\.\d+)
        (?:\s+eq\s+(?P<PORT>\S+))?
        \s+any
        (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
        """,
        re.IGNORECASE | re.VERBOSE,
    )
    MATCH_OUTCOME = IOSXE_HOST_ANY_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        SOURCE_NETWORK = host_to_network(MATCH_OUTCOME.group("SOURCE_IP"))
        DESTINATION_NETWORK = any_to_network()
        PORT_NUMBER = (
            service_to_port(MATCH_OUTCOME.group("PORT"))
            if MATCH_OUTCOME.group("PORT") else None
        )
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                None,
                PORT_NUMBER,
                "IOS-XE"
            ), None
        return None, "iosxe_host_any_network_parse_fail"


    # 15. IOS-XE Any关键字格式（any + host）
    # 示例: 10 permit tcp any host 10.10.80.1 eq 22 log
    IOSXE_ANY_HOST_REGULAR_EXPRESSION = re.compile(
        r"""
        ^\s*(?P<NUMBER>\d+)?\s*
        (?P<ACTION>permit|deny)\s+
        (?P<PROTOCOL>\S+)\s+
        any
        (?:\s+eq\s+(?P<PORT>\S+))?
        \s+host\s+(?P<DESTINATION_IP>\d+\.\d+\.\d+\.\d+)
        (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
        """,
        re.IGNORECASE | re.VERBOSE,
    )
    MATCH_OUTCOME = IOSXE_ANY_HOST_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        SOURCE_NETWORK = any_to_network()
        DESTINATION_NETWORK = host_to_network(MATCH_OUTCOME.group("DESTINATION_IP"))
        PORT_NUMBER = (
            service_to_port(MATCH_OUTCOME.group("PORT"))
            if MATCH_OUTCOME.group("PORT") else None
        )
        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                None,
                PORT_NUMBER,
                "IOS-XE"
            ), None
        return None, "iosxe_any_host_network_parse_fail"


    # 16. IOS-XE Any关键字格式（wildcard + any 或 any + wildcard）
    MATCH_OUTCOME = IOSXE_ANY_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        if MATCH_OUTCOME.group("SRC_IP") and MATCH_OUTCOME.group("SRC_WC"):
            SOURCE_NETWORK = ip_and_wildcard_to_network(
                MATCH_OUTCOME.group("SRC_IP"),
                MATCH_OUTCOME.group("SRC_WC")
            )
        else:
            SOURCE_NETWORK = any_to_network()


        if MATCH_OUTCOME.group("DST_IP") and MATCH_OUTCOME.group("DST_WC"):
            DESTINATION_NETWORK = ip_and_wildcard_to_network(
                MATCH_OUTCOME.group("DST_IP"),
                MATCH_OUTCOME.group("DST_WC")
            )
        else:
            DESTINATION_NETWORK = any_to_network()


        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                None,
                None,
                None,
                "IOS-XE"
            ), None
        return None, "iosxe_any_network_parse_fail"


    # 17. ASA格式
    MATCH_OUTCOME = ASA_REGULAR_EXPRESSION.match(CLEANED_LINE)
    if MATCH_OUTCOME:
        source_text = MATCH_OUTCOME.group("SRC")
        destination_text = MATCH_OUTCOME.group("DST")


        if source_text.lower() == "any":
            SOURCE_NETWORK = any_to_network()
        elif "/" in source_text:
            SOURCE_NETWORK = cidr_to_network(source_text)
        else:
            SOURCE_NETWORK = host_to_network(source_text)


        if destination_text.lower() == "any":
            DESTINATION_NETWORK = any_to_network()
        elif "/" in destination_text:
            DESTINATION_NETWORK = cidr_to_network(destination_text)
        else:
            DESTINATION_NETWORK = host_to_network(destination_text)


        PORT_NUMBER = (
            service_to_port(MATCH_OUTCOME.group("PORT_A")) or
            service_to_port(MATCH_OUTCOME.group("PORT_B"))
        )


        if SOURCE_NETWORK and DESTINATION_NETWORK:
            return ACLRule(
                CLEANED_LINE,
                MATCH_OUTCOME.group("ACTION").lower(),
                MATCH_OUTCOME.group("PROTOCOL").lower(),
                SOURCE_NETWORK,
                DESTINATION_NETWORK,
                PORT_NUMBER,
                None,
                None,
                "ASA"
            ), None
        return None, "asa_network_parse_fail"


    return None, "no_pattern_match"

# ============================================================================
# ACL规则判断函数
# ============================================================================

# 判断文本是否为有效的ACL规则：通过检查关键字和格式，排除明显的非ACL配置行（如as-path、route-map、prefix-list等配置命令）
def is_acl_rule(text: str) -> bool:
    """判断文本是否为有效的ACL规则

    通过检查关键字和格式，排除明显的非ACL配置行（如as-path、route-map、prefix-list等配置命令）

    Args:
        text: 待检查的文本行

    Returns:
        True表示是ACL规则，False表示不是
    """
    text = text.strip().lower()


    # 必须包含permit或deny
    if 'permit' not in text and 'deny' not in text:
        return False


    # 排除配置命令（以这些关键字开头的行，但不是ACL规则）
    exclude_patterns = [
        r'^ip\s+access-list\s+',  # ACL定义行
        r'^ip\s+as-path\s+access-list',  # as-path access-list
        r'^ip\s+prefix-list',  # prefix-list
        r'^ip\s+community-list',  # community-list
        r'^route-map\s+',  # route-map
        r'^logging\s+host',  # logging配置
        r'^certificate',  # certificate
        r'^crypto',  # crypto
        r'^interface',  # interface
        r'^router',  # router
        r'^version',  # version
        r'^hostname',  # hostname
        r'^enable',  # enable
        r'^password',  # password
        r'^username',  # username
        r'^line\s+',  # line
        r'^service\s+',  # service
        r'^logging\s+',  # logging
        r'^ntp\s+',  # ntp
        r'^snmp\s+',  # snmp
        r'^tacacs',  # tacacs
        r'^radius',  # radius
        r'^no\s+arp',  # no arp命令
    ]


    for pattern in exclude_patterns:
        if re.match(pattern, text, re.IGNORECASE):
            return False


    # 必须包含协议或IP地址
    if not (re.search(r'\b(tcp|udp|ip|icmp|ospf|eigrp|gre|esp|ah)\b', text) or

            re.search(r'\d+\.\d+\.\d+\.\d+', text)):
        return False


    return True

# ============================================================================
# 通用ACL解析函数（忽略any规则）
# ============================================================================

# 解析ACL规则行，返回完整的ACLRule对象（忽略any规则）
# 这是最常用的ACL解析函数，适用于需要完整规则信息的场景。会自动忽略包含'any'关键字的规则。
# 注意：自动去除行号前缀，自动忽略log-input、time-range等关键字，any规则会被过滤
def parse_acl(ACL_LINE: str) -> Tuple[Optional[ACLRule], Optional[str]]:
    """解析ACL规则行，自动忽略any规则

    这是最常用的ACL解析函数，适用于需要完整规则信息的场景。
    会自动忽略包含'any'关键字的规则。

    Args:
        ACL_LINE: ACL规则文本行（可包含行号前缀）

    Returns:
        元组：(ACLRule对象, None) 成功，或 (None, 错误信息) 失败，或 (None, "contains_any") any规则
    """
    CLEANED_LINE = (ACL_LINE or "").strip()
    if not CLEANED_LINE:
        return None, "empty"


    # 去除行号前缀（如果存在）：行号通常是开头的数字，后跟空格
    # 例如："464 permit ip ..." -> "permit ip ..."
    # 匹配开头的数字（可能有多位）后跟空格，然后才是permit/deny
    line_number_pattern = re.compile(r'^\s*\d+\s+(permit|deny)', re.IGNORECASE)
    if line_number_pattern.match(CLEANED_LINE):
        # 找到第一个permit或deny的位置
        match = re.search(r'\b(permit|deny)\b', CLEANED_LINE, re.IGNORECASE)
        if match:
            CLEANED_LINE = CLEANED_LINE[match.start():]


    # 忽略any规则
    if "any" in CLEANED_LINE.lower():
        return None, "contains_any"


    # 调用parse_acl_full进行完整解析
    # 注意：由于NXOS_REGULAR_EXPRESSION已统一为宽松匹配，log-input/time-range等关键字会自动被忽略
    result, error = parse_acl_full(ACL_LINE)
    if error == "contains_any":
        return None, "contains_any"
    return result, error

# ============================================================================
# 简化ACL解析函数（只提取网络信息）
# ============================================================================

# 解析ACL规则行，只提取源/目的网络信息（忽略any规则）
# 适用于只需要网络信息的场景（如网络匹配、ARP检查等），不需要完整的ACLRule对象，性能更好。
def parse_acl_network_only(
        ACL_LINE: str
) -> Tuple[Optional[Tuple[IPv4Network, IPv4Network]], Optional[str]]:
    """解析ACL规则行，只提取源/目的网络信息（忽略any规则）

    适用于只需要网络信息的场景（如网络匹配、ARP检查等），不需要完整的ACLRule对象，性能更好。

    Args:
        ACL_LINE: ACL规则文本行（可包含行号前缀）

    Returns:
        元组：((源网络, 目的网络), None) 成功，或 (None, 错误信息) 失败，或 (None, "contains_any") any规则
    """
    rule, error = parse_acl(ACL_LINE)
    if rule:
        return (rule.src, rule.dst), None
    return None, error

# ============================================================================
# Excel ACL块处理函数
# ============================================================================

# 在指定列中找到ACL块：以'ip access-list'开始，以包含'vty'和'ip'的'ip access-list'行结束（登录ACL结束标记）
def find_acl_blocks_in_column(worksheet, col: int) -> List[Tuple[int, int]]:
    """在指定列中找到ACL块

    以'ip access-list'开始，以包含'vty'和'ip'的'ip access-list'行结束（登录ACL结束标记）

    Args:
        worksheet: openpyxl Worksheet对象
        col: 列号（从1开始）

    Returns:
        ACL块列表，每个元素为(start_row, end_row)
    """
    ACL_BLOCKS = []
    CURRENT_START = None
    FOUND_VTY = False  # 标记是否遇到登录ACL结束标记


    for ROW in range(1, worksheet.max_row + 1):
        CELL_VALUE = worksheet.cell(row=ROW, column=col).value
        if CELL_VALUE and isinstance(CELL_VALUE, str):
            TEXT = str(CELL_VALUE).strip()
            TEXT_LOWER = TEXT.lower()


            # 业务ACL开始（排除登录ACL）
            if TEXT.startswith('ip access-list '):
                # 检查是否是登录ACL结束标记（包含vty和ip，忽略大小写）
                # 匹配：ip access-list VTY-ACL-IP 或 ip access-list extended vty-access-IP
                if 'vty' in TEXT_LOWER and 'ip' in TEXT_LOWER:
                    # 登录ACL标记 - 结束当前ACL块，不再处理后续ACL
                    if CURRENT_START is not None:
                        ACL_BLOCKS.append((CURRENT_START, ROW - 1))
                    FOUND_VTY = True
                    break  # 登录ACL及以下的不分析
                else:
                    # 业务ACL开始
                    if CURRENT_START is not None:
                        # 结束上一个ACL块
                        ACL_BLOCKS.append((CURRENT_START, ROW - 1))
                    CURRENT_START = ROW


    # 处理最后一个ACL块（只有在没有遇到登录ACL结束标记时才处理）
    if not FOUND_VTY and CURRENT_START is not None:
        ACL_BLOCKS.append((CURRENT_START, worksheet.max_row))


    return ACL_BLOCKS

# 从指定列的ACL块中提取ACL规则：忽略any规则，只提取有效的ACL规则
def extract_acl_rules_from_column(
        worksheet, col: int, start_row: int, end_row: int
) -> List[ACLRule]:
    """从指定列的ACL块中提取ACL规则

    忽略any规则，只提取有效的ACL规则

    Args:
        worksheet: openpyxl Worksheet对象
        col: 列号（从1开始）
        start_row: ACL块起始行号
        end_row: ACL块结束行号

    Returns:
        解析后的ACL规则列表
    """
    ACL_RULES = []
    for ROW in range(start_row, end_row + 1):
        CELL_VALUE = worksheet.cell(row=ROW, column=col).value
        if CELL_VALUE is None:
            continue
        CLEANED_TEXT = str(CELL_VALUE).strip()
        if not CLEANED_TEXT:
            continue


        # 只处理真正的ACL规则，忽略证书等数据
        if not is_acl_rule(CLEANED_TEXT):
            continue


        # 解析ACL规则（会自动忽略any规则）
        PARSED_RULE, PARSE_ERROR = parse_acl(CLEANED_TEXT)
        if PARSED_RULE:
            ACL_RULES.append(PARSED_RULE)


    return ACL_RULES

# ============================================================================
# 设备分类相关函数
# ============================================================================

# 获取设备分类规则：返回设备分类规则字典，用于识别不同类型的设备
def get_device_classification_rules() -> Dict:
    """获取设备分类规则

    返回设备分类规则字典，包含CS-N9K、LINK-AS、ASA-FW、LINK-DS、BGP、OOB-DS等设备类型的识别模式
    支持可选的 group_by_site 和 sheet_name 字段（用于DeviceBackupTask等任务）

    Returns:
        Dict: 设备分类规则字典
    """
    return {
        "cat1": {
            "name": "N9K核心交换机",
            "patterns": [
                # CS + N9K + (01|02|03|04)，统一匹配模式（参数已为小写）
                # 要求：包含n9k，且(CS+设备编号)或(CS连写模式如cs01)
                lambda filenameLower: (
                    re.search(r"\bn9k\b", filenameLower) and (
                        (re.search(r"\bcs\b", filenameLower) and
                            re.search(
                                r"(?:^|[^0-9])0?[1-4](?:[^0-9]|$)",
                                filenameLower
                            )) or
                        re.search(r"\bcs0?[1-4]", filenameLower)
                    )
                )
            ]
        },
        "cat2": {
            "name": "LINKAS接入交换机",
            "patterns": [
                # LINK + AS + (01|02)，统一匹配模式（参数已为小写）
                # 支持：LINKAS连写（link.*as01）、LINK+AS+设备编号、LINK+AS01/02连写
                lambda filenameLower: (
                    re.search(r"\blink.*as0?[12]\b", filenameLower) or
                    (re.search(r"\bas0?[12]\b", filenameLower) and
                        re.search(r"\blink\b", filenameLower)) or
                    (re.search(r"\blink\b", filenameLower) and
                        re.search(r"\bas\b", filenameLower) and
                        re.search(r"(?:^|[^0-9])0?[12](?:[^0-9]|$)", filenameLower))
                )
            ]
        },
        "cat3": {
            "name": "ASA防火墙",
            "patterns": [
                # 固定组合：fw01-frp 或 fw02-frp（参数已为小写）
                lambda filenameLower: (
                    "fw01-frp" in filenameLower or
                    "fw02-frp" in filenameLower
                )
            ]
        },
        "cat4": {
            "name": "LINK-DS交换机",
            "patterns": [
                # Link-DS + (01|02) + C9300/N9K（参数已为小写，合并为统一模式）
                lambda filenameLower: (
                    "link-ds" in filenameLower and
                    re.search(r"0?[12]", filenameLower)
                ),
                # 支持连写模式（参数已为小写）
                lambda filenameLower: re.search(
                    r"link[-_]?ds0?[12]", filenameLower
                )
            ]
        },
        "cat5": {
            "name": "BGP设备",
            "patterns": [
                # 只要包含 bgp 关键词即可（参数已为小写）
                lambda filenameLower: (
                    ("bgp" in filenameLower) or
                    re.search(r"\bbgp\b", filenameLower)
                )
            ]
        },
        "cat6": {
            "name": "OOB-DS交换机",
            "patterns": [
                # OOB-DS + (01|02)，支持连写和分隔符
                # 格式: OOB-DS01, OOB_DS01, OOB-DS02 等
                lambda filenameLower: (
                    re.search(r"\boob[-_]?ds0?[12]\b", filenameLower) or
                    (re.search(r"\boob\b", filenameLower) and
                        re.search(r"\bds\b", filenameLower) and
                        re.search(r"(?:^|[^0-9])0?[12](?:[^0-9]|$)", filenameLower))
                )
            ]
        }
    }

# 检查文本是否匹配CS-N9K设备：N9K核心交换机（CS + N9K + 01/02/03/04）
def is_cat1_device(text: str) -> bool:
    """检查文本是否匹配CS-N9K设备

    Args:
        text: 设备名称或文本

    Returns:
        bool: 如果匹配CS-N9K设备则返回True
    """
    TEXT_LOWER = text.lower()
    RULES = get_device_classification_rules()
    for PATTERN_FUNC in RULES["cat1"]["patterns"]:
        if PATTERN_FUNC(TEXT_LOWER):
            return True
    return False

# 检查文本是否匹配LINK-AS设备：LINKAS接入交换机（LINK + AS + 01/02）
def is_cat2_device(text: str) -> bool:
    """检查文本是否匹配LINK-AS设备

    Args:
        text: 设备名称或文本

    Returns:
        bool: 如果匹配LINK-AS设备则返回True
    """
    TEXT_LOWER = text.lower()
    RULES = get_device_classification_rules()
    for PATTERN_FUNC in RULES["cat2"]["patterns"]:
        if PATTERN_FUNC(TEXT_LOWER):
            return True
    return False

# 检查文本是否匹配OOB-DS设备：OOB-DS交换机（OOB-DS + 01/02）
def is_cat6_device(text: str) -> bool:
    """检查文本是否匹配OOB-DS设备

    Args:
        text: 设备名称或文本

    Returns:
        bool: 如果匹配OOB-DS设备则返回True
    """
    TEXT_LOWER = text.lower()
    RULES = get_device_classification_rules()
    for PATTERN_FUNC in RULES["cat6"]["patterns"]:
        if PATTERN_FUNC(TEXT_LOWER):
            return True
    return False

# 从设备名称中提取设备序号（01, 02, 03等）：优先匹配设备类型标识符（CS、AS、OOB-DS等）后面的数字，避免匹配站点编号（HX01等）
def extract_device_number(device_name: str) -> Optional[int]:
    """从设备名称中提取设备序号

    优先匹配设备类型标识符（CS、AS、OOB-DS等）后面的数字，避免匹配站点编号（HX01等）

    Args:
        device_name: 设备名称

    Returns:
        Optional[int]: 设备序号（1-4），如果无法提取则返回None
    """
    DEVICE_PATTERNS = [
        # CS01, AS01, Link-As01, OOB-DS01等
        r"(?:cs|as|link[-_]?as|oob[-_]?ds)(?:0?)([1-4])(?:[^0-9]|$)",
        r"(?:cs|as)(?:0?)([1-4])(?:[^0-9]|$)",  # CS01, AS01等（无连字符）
        r"(?:oob[-_]?ds)(?:0?)([12])(?:[^0-9]|$)",  # OOB-DS01, OOB-DS02等
    ]

    for PATTERN in DEVICE_PATTERNS:
        MATCH = re.search(PATTERN, device_name.lower())
        if MATCH:
            return int(MATCH.group(1))

    # 如果上述模式都不匹配，使用通用模式（但可能匹配到站点编号）
    MATCH = re.search(r"(?:^|[^0-9])(0?[1-4])(?:[^0-9]|$)", device_name.lower())
    if MATCH:
        return int(MATCH.group(1))
    return None

# 分析第一行识别设备列：识别cat1/cat2/cat6列，优先检查标识，否则回退到设备名称模式匹配，只导入指定设备编号（cat1:01/03, cat2:01, cat6:全部）
def analyze_first_row_for_cat1_cat2(worksheet) -> Tuple[List, List, List]:
    """分析Excel第一行，识别cat1/cat2/cat6列

    优先检查是否包含cat1/cat2/cat6标识，否则回退到设备名称模式匹配
    只导入指定设备编号（cat1:01/03, cat2:01, cat6:全部）

    Args:
        worksheet: openpyxl Worksheet对象

    Returns:
        Tuple[List, List, List]: (cat1_cols, cat2_cols, cat6_cols)
            cat1_cols: [(col, device_number, device_name), ...]
            cat2_cols: [(col, device_number, device_name), ...]
            cat6_cols: [(col, device_number, device_name), ...]
    """
    try:
        from openpyxl.worksheet.worksheet import Worksheet
    except ImportError:
        worksheet_type = None

    CAT1_COLS = []  # [(col, device_number, device_name), ...]
    CAT2_COLS = []  # [(col, device_number, device_name), ...]
    CAT6_COLS = []  # [(col, device_number, device_name), ...]

    # 定义允许导入的设备编号（硬编码在代码中）
    ALLOWED_CAT1_NUMBERS = {1, 3}  # cat1只导入01和03
    ALLOWED_CAT2_NUMBERS = {1}      # cat2只导入01
    # cat6不限制（导入所有匹配的设备）

    for COLUMN in range(1, worksheet.max_column + 1):
        FIRST_CELL = worksheet.cell(row=1, column=COLUMN).value
        if FIRST_CELL and isinstance(FIRST_CELL, str):
            DEVICE_NAME = FIRST_CELL.strip()
            DEVICE_NAME_LOWER = DEVICE_NAME.lower()

            # 优先检查是否包含cat1/cat2/cat6标识（不区分大小写）
            if "cat1" in DEVICE_NAME_LOWER:
                DEVICE_NUMBER = extract_device_number(DEVICE_NAME)
                # 只保留允许的设备编号（01和03）
                if DEVICE_NUMBER in ALLOWED_CAT1_NUMBERS:
                    CAT1_COLS.append((COLUMN, DEVICE_NUMBER, DEVICE_NAME))
            elif "cat2" in DEVICE_NAME_LOWER:
                DEVICE_NUMBER = extract_device_number(DEVICE_NAME)
                # 只保留允许的设备编号（01）
                if DEVICE_NUMBER in ALLOWED_CAT2_NUMBERS:
                    CAT2_COLS.append((COLUMN, DEVICE_NUMBER, DEVICE_NAME))
            elif "cat6" in DEVICE_NAME_LOWER:
                DEVICE_NUMBER = extract_device_number(DEVICE_NAME)
                # cat6不限制，直接添加
                CAT6_COLS.append((COLUMN, DEVICE_NUMBER, DEVICE_NAME))
            else:
                # 回退到设备名称模式匹配
                DEVICE_NUMBER = extract_device_number(DEVICE_NAME)
                if is_cat1_device(DEVICE_NAME):
                    # 只保留允许的设备编号（01和03）
                    if DEVICE_NUMBER in ALLOWED_CAT1_NUMBERS:
                        CAT1_COLS.append((COLUMN, DEVICE_NUMBER, DEVICE_NAME))
                elif is_cat2_device(DEVICE_NAME):
                    # 只保留允许的设备编号（01）
                    if DEVICE_NUMBER in ALLOWED_CAT2_NUMBERS:
                        CAT2_COLS.append((COLUMN, DEVICE_NUMBER, DEVICE_NAME))
                elif is_cat6_device(DEVICE_NAME):
                    # cat6不限制，直接添加
                    CAT6_COLS.append((COLUMN, DEVICE_NUMBER, DEVICE_NAME))

    return CAT1_COLS, CAT2_COLS, CAT6_COLS

