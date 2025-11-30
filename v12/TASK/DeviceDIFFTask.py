# 关键设备配置按周期DIFF审计任务
#
# 技术栈:openpyxl、正则表达式、difflib
# 目标:自动选日期做 DIFF，对比不同时期的设备配置变化
#
# 自动选日期做 DIFF:
# 日 DIFF（每天）:今天 vs 昨天；
# 周 DIFF（仅周日执行）:今天(周日) vs 本周一；若周一文件不存在，则从周一往后顺延到周六，取本周最早有文件的那天。输出名包含"周DIFF"；
# 月 DIFF（仅月末执行）:今天(当月最后一天) vs 本月 1 号；若 1 号不存在，则从 1 号往后顺延到昨日，取本月最早有文件的那天。输出名包含"月DIFF"
#
# 读取两期"基础任务 Excel"并对比:
# 输入文件目录:LOG/DeviceBackupTask/（V10新结构：从ACL/SourceACL迁移）
# 文件名模式:{YYYYMMDD}-关键设备配置备份输出EXCEL基础任务.xlsx
# 只要有改动，就生成 unified diff 行列表（含 ---/+++/@/@、加号行、减号行、上下文行）
# 着色:+ 行绿色、- 行红色，其它行（---/+++/@/@/空格开头）黑色
#
# 输出:
# 输出目录:LOG/DeviceDIFFTask/（V10新结构：从ACL/DeviceDIFFTask迁移）
# 文件名:
# - 日:{今天}-关键设备DIFF.xlsx
# - 周:{今天}-周DIFF-关键设备DIFF.xlsx
# - 月:{今天}-月DIFF-关键设备DIFF.xlsx
#
# 配置说明:支持YAML/Ignore_DIFF.yaml忽略规则

# 导入标准库
import datetime as dt
import os
import re

from difflib import unified_diff
from typing import List, Optional, Tuple

# 导入第三方库
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

# 导入本地应用
from .TaskBase import BaseTask, Level

# 关键设备差异对比任务类：对比设备配置的历史差异并生成Excel报告
class DeviceDIFFTask(BaseTask):
    """关键设备差异对比任务


    自动选日期做DIFF，对比不同时期的设备配置变化
    """


    # 初始化关键设备差异对比任务：设置任务名称和运行期缓存
    def __init__(self):
        super().__init__("关键设备DIFF")
        # 统一路径处理：在初始化时设置路径
        # V10新结构：从 LOG/DeviceBackupTask/ 读取（ACL/SourceACL已迁移）
        self.INPUT_DIR = os.path.join("LOG", "DeviceBackupTask")
        # V10新结构：直接输出到 LOG/DeviceDIFFTask/
        self.OUTPUT_DIR = os.path.join("LOG", "DeviceDIFFTask")
        self.FILENAME_PREFIX = "关键设备配置备份输出EXCEL基础任务"

        # 运行期缓存（items() 准备，run_single() 使用）
        self._PLANS = []
        self._PRIMARY = None
        self._SITES = []
        self._START_YMD = ""
        self._END_YMD = ""
        self._OUT_SUFFIX = None
        self._WB = None
        self._OVERVIEW = None
        self._SUM_ADDED = 0
        self._SUM_REMOVED = 0
        self._SUM_CHANGED = 0


        # 加载忽略规则
        self._IGNORE_RULES = self._load_ignore_rules()

    # 计算对比计划并返回Sheet列表：根据日期逻辑确定对比文件并返回Sheet名称列表
    def items(self):
        """计算对比计划并返回Sheet列表


        根据日期逻辑确定对比文件并返回Sheet名称列表


        Returns:
            list: Sheet名称列表
        """
        # 返回 Sheet 列表作为 items；items() 不做重活，仅计算计划和 sheet 名
        # 统一目录创建逻辑：在items中创建输出目录
        # 创建输出目录（如果目录已存在则不报错）
        os.makedirs(self.OUTPUT_DIR, exist_ok=True)


        TODAY_DATETIME = dt.datetime.now()
        TODAY = TODAY_DATETIME.date()
        TODAY_STRING = TODAY_DATETIME.strftime("%Y%m%d")

        # 格式化路径：根据日期格式化文件路径
        def _format_file_path(YMD: str) -> str:
            """格式化文件路径


            Args:
                YMD: 日期字符串（YYYYMMDD格式）


            Returns:
                str: 文件路径
            """
            return os.path.join(self.INPUT_DIR, f"{YMD}-{self.FILENAME_PREFIX}.xlsx")

        PLANS = []  # [(BEFORE_YMD, AFTER_YMD, OUT_SUFFIX)]

        # 日：今天 vs 昨天（最多往前找3天，找到最新的文件即可）
        # 先检查今天文件是否存在
        if os.path.isfile(_format_file_path(TODAY_STRING)):
            # 从昨天开始往前找，最多找3天
            for DAYS_BACK in range(1, 4):  # 1天前、2天前、3天前
                CHECK_DATE = TODAY - dt.timedelta(days=DAYS_BACK)
                CHECK_DATE_STRING = CHECK_DATE.strftime("%Y%m%d")
                if os.path.isfile(_format_file_path(CHECK_DATE_STRING)):
                    PLANS.append((CHECK_DATE_STRING, TODAY_STRING, None))
                    break

        # 周：仅周日，从周一顺延至周六找最早有文件
        if TODAY.weekday() == 6:  # Sunday
            MONDAY = TODAY - dt.timedelta(days=6)
            CURRENT = MONDAY
            WEEK_END_DATE = TODAY - dt.timedelta(days=1)
            FOUND = None
            while CURRENT <= WEEK_END_DATE:
                if os.path.isfile(_format_file_path(CURRENT.strftime("%Y%m%d"))):
                    FOUND = CURRENT
                    break
                CURRENT += dt.timedelta(days=1)
            if FOUND and os.path.isfile(_format_file_path(TODAY_STRING)):
                PLANS.append((FOUND.strftime("%Y%m%d"), TODAY_STRING, "周DIFF"))

        # 月：仅月末，从 1 号顺延至昨日找最早有文件
        if (TODAY + dt.timedelta(days=1)).month != TODAY.month:
            FIRST = TODAY.replace(day=1)
            CURRENT = FIRST
            FOUND = None
            while CURRENT < TODAY:
                if os.path.isfile(_format_file_path(CURRENT.strftime("%Y%m%d"))):
                    FOUND = CURRENT
                    break
                CURRENT += dt.timedelta(days=1)
            if FOUND and os.path.isfile(_format_file_path(TODAY_STRING)):
                PLANS.append((FOUND.strftime("%Y%m%d"), TODAY_STRING, "月DIFF"))

        if not PLANS:
            self._PLANS = []
            return [None]

        # 只用第一个计划驱动外层进度；其余计划最后统一生成（不额外显示进度）
        BEFORE_YMD, AFTER_YMD, OUT_SUFFIX = PLANS[0]
        self._PLANS = PLANS
        self._PRIMARY = (BEFORE_YMD, AFTER_YMD, OUT_SUFFIX)

        # 仅读取 sheet 名（很快）
        START_FILE_PATH = _format_file_path(BEFORE_YMD)
        END_FILE_PATH = _format_file_path(AFTER_YMD)
        try:
            WORKBOOK_A = load_workbook(START_FILE_PATH, read_only=True, data_only=True)
            WORKBOOK_B = load_workbook(END_FILE_PATH, read_only=True, data_only=True)
            self._SITES = sorted(set(WORKBOOK_A.sheetnames) | set(WORKBOOK_B.sheetnames))
        except Exception as ERROR:
            self._PLANS = []
            self.add_result(Level.WARN, f"读取输入Excel失败（获取 sheet 名）：{ERROR}")
            return [None]

        if not self._SITES:
            self._PLANS = []
            self.add_result(Level.WARN, "输入Excel无任何 Sheet，跳过 DIFF")
            return [None]

        # 初始化输出工作簿与概览（一次）
        self._START_YMD = BEFORE_YMD
        self._END_YMD = AFTER_YMD
        self._OUT_SUFFIX = OUT_SUFFIX
        self._WB = Workbook()
        self._OVERVIEW = self._WB.active
        self._OVERVIEW.title = "Overview"
        self._OVERVIEW.append(["起始文件", os.path.basename(START_FILE_PATH)])
        self._OVERVIEW.append(["结束文件", os.path.basename(END_FILE_PATH)])
        self._OVERVIEW.append([])
        self._OVERVIEW.append([
            "站点Sheet", "设备数(A)", "设备数(B)", "新增DIFF设备",
            "删除DIFF设备", "配置产生变更设备数量", "备注"
        ])

        self._SUM_ADDED = self._SUM_REMOVED = self._SUM_CHANGED = 0
        return self._SITES  # 关键：外层进度条按 Sheet 数显示 1/N

    # 加载忽略规则：从YAML文件加载DIFF忽略规则
    def _load_ignore_rules(self):
        """加载DIFF忽略规则


        从YAML文件加载DIFF忽略规则


        Returns:
            dict: 忽略规则字典
        """
        try:
            IGNORE_FILE_PATH = os.path.join("YAML", "Ignore_DIFF.yaml")
            if os.path.exists(IGNORE_FILE_PATH):
                with open(IGNORE_FILE_PATH, 'r', encoding='utf-8') as FILE_HANDLE:
                    return yaml.safe_load(FILE_HANDLE)
            return {}
        except Exception as ERROR:
            self.add_result(Level.WARN, f"加载忽略规则失败: {ERROR}")
            return {}

    # 检查行是否应该被忽略：根据忽略规则判断行是否应该被过滤
    def _should_ignore_line(self, line: str) -> bool:
        """检查行是否应该被忽略


        Args:
            line: 要检查的行


        Returns:
            bool: 如果应该被忽略则返回True
        """
        if not self._IGNORE_RULES or "diff_ignore" not in self._IGNORE_RULES:
            return False


        RULES = self._IGNORE_RULES["diff_ignore"]
        LINE_LOWER = line.lower()


        # 检查包含匹配
        if "message_contains" in RULES:
            for PATTERN in RULES["message_contains"]:
                if PATTERN.lower() in LINE_LOWER:
                    return True


        # 检查正则表达式匹配
        if "message_regex" in RULES:
            for PATTERN in RULES["message_regex"]:
                try:
                    if re.search(PATTERN, line, re.IGNORECASE):
                        return True
                except re.error:
                    continue


        return False

    # 过滤配置行：移除应该被忽略的行
    def _filter_configuration_lines(self, lines: list[str]) -> list[str]:
        """过滤配置行


        移除应该被忽略的行


        Args:
            lines: 配置行列表


        Returns:
            list[str]: 过滤后的配置行列表
        """
        if not self._IGNORE_RULES:
            return lines


        FILTERED_LINES = []
        for LINE in lines:
            if not self._should_ignore_line(LINE):
                FILTERED_LINES.append(LINE)


        return FILTERED_LINES

    # 标准化设备名称：从列头中提取设备标识符
    @staticmethod
    def _normalize_device(header: str) -> Tuple[Optional[str], str]:
        """标准化设备名称


        从列头中提取设备标识符


        Args:
            header: 列头字符串


        Returns:
            Tuple[Optional[str], str]: (设备标识符, 原始列头)
        """
        if header is None:
            return None, ""
        HEADER_STRING = str(header).strip()
        if not HEADER_STRING:
            return None, ""
        HEADER_NOEXT = (
            HEADER_STRING[:-4] if HEADER_STRING.lower().endswith(".log") else HEADER_STRING
        )
        if (
            len(HEADER_NOEXT) > 9
            and HEADER_NOEXT[:8].isdigit()
            and HEADER_NOEXT[8] == "-"
        ):
            DEVICE_KEY = HEADER_NOEXT[9:].strip()
        else:
            DEVICE_KEY = HEADER_NOEXT.strip()
        if not DEVICE_KEY:
            return None, HEADER_STRING
        return DEVICE_KEY, HEADER_STRING

    # 读取Excel工作表数据：解析指定工作表并返回设备配置映射
    def _read_sheet_map(self, xlsx_path: str, sheet_name: str) -> dict[str, dict]:
        """读取Excel工作表数据


        只读取指定sheet，返回设备配置映射


        Args:
            xlsx_path: Excel文件路径
            sheet_name: 工作表名称


        Returns:
            dict: {device_key: {"label": <列头>, "lines": [...]}, ...}
        """
        WORKBOOK = load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet_name not in WORKBOOK.sheetnames:
            return {}
        WORKSHEET = WORKBOOK[sheet_name]

        # 第1行：列头
        HEADER_ROW = next(WORKSHEET.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not HEADER_ROW:
            return {}
        COLS: list[tuple[int, str, str]] = []
        for COL_IDX, HEADER in enumerate(HEADER_ROW, start=1):
            KEY, LABEL = self._normalize_device(HEADER)
            if KEY:
                COLS.append((COL_IDX, KEY, LABEL))
        if not COLS:
            return {}

        # 每列：从第2行开始收集配置行，直到遇到连续空行阈值
        EMPTY_BREAK = 10000
        SHEET_MAP: dict[str, dict] = {}
        for COL, KEY, LABEL in COLS:
            LINES: list[str] = []
            EMPTY_STREAK = 0
            for CELLS in WORKSHEET.iter_rows(
                    min_row=2, min_col=COL, max_col=COL, values_only=True
            ):
                VALUE = CELLS[0]
                if VALUE is None or VALUE == "":
                    EMPTY_STREAK += 1
                    if EMPTY_STREAK >= EMPTY_BREAK:
                        break
                else:
                    EMPTY_STREAK = 0
                    LINES.append(str(VALUE))
            SHEET_MAP[KEY] = {"label": LABEL, "lines": LINES}
        return SHEET_MAP

    # 处理单个站点的差异对比：渲染指定Sheet并生成差异报告
    def run_single(self, site):
        """处理单个站点的差异对比


        渲染指定Sheet并生成差异报告
        若是最后一个Sheet，则保存主输出并顺带跑周/月计划


        Args:
            site: 站点名称
        """
        if not self._PLANS or not self._SITES or site is None:
            self.add_result(Level.WARN, "DIFF 跳过：找不到可用的基础任务 Excel")
            return

        # 仅读取当前 sheet 的数据（把重活分摊到每次调用）
        START_FILE_PATH = os.path.join(
            self.INPUT_DIR, f"{self._START_YMD}-{self.FILENAME_PREFIX}.xlsx"
        )
        END_FILE_PATH = os.path.join(
            self.INPUT_DIR, f"{self._END_YMD}-{self.FILENAME_PREFIX}.xlsx"
        )
        try:
            AMAP = self._read_sheet_map(START_FILE_PATH, site)
            BMAP = self._read_sheet_map(END_FILE_PATH, site)
        except Exception as ERROR:
            self.add_result(Level.WARN, f"读取 {site} 失败：{ERROR}")
            AMAP, BMAP = {}, {}

        ASET, BSET = set(AMAP.keys()), set(BMAP.keys())
        ADDED = sorted(list(BSET - ASET))
        REMOVED = sorted(list(ASET - BSET))
        COMMON = sorted(list(ASET & BSET))
        CHANGED = []

        WORKSHEET = self._WB.create_sheet(title=(site[:28] if len(site) > 28 else site))
        WORKSHEET.append([f"站点：{site}"])
        WORKSHEET.append([f"起始：{self._START_YMD}", f"结束：{self._END_YMD}"])
        WORKSHEET.append(["变更类型", "设备Key", "起始列头", "结束列头", "说明/DIFF"])


        # 设置列宽
        WORKSHEET.column_dimensions['A'].width = 15  # 变更类型
        WORKSHEET.column_dimensions['B'].width = 28  # 设备Key
        WORKSHEET.column_dimensions['C'].width = 45    # 起始列头
        WORKSHEET.column_dimensions['D'].width = 9    # 结束列头
        WORKSHEET.column_dimensions['E'].width = 80  # 说明/DIFF

        # ADDED / REMOVED
        for DEV in ADDED:
            WORKSHEET.append(["ADDED", DEV, "", BMAP[DEV]["label"], "设备仅存在于结束日"])
        for DEV in REMOVED:
            WORKSHEET.append(["REMOVED", DEV, AMAP[DEV]["label"], "", "设备仅存在于开始日"])

        # CHANGED：E 列按行展开 + 上色
        for DEV in COMMON:
            LINES_A = AMAP[DEV]["lines"]
            LINES_B = BMAP[DEV]["lines"]


            # 过滤忽略的行
            FILTERED_LINES_A = self._filter_configuration_lines(LINES_A)
            FILTERED_LINES_B = self._filter_configuration_lines(LINES_B)


            if FILTERED_LINES_A == FILTERED_LINES_B:
                continue
            CHANGED.append(DEV)
            DIFF_LINES = list(unified_diff(
                FILTERED_LINES_A, FILTERED_LINES_B, fromfile="before",
                tofile="after", lineterm=""
            ))

            WORKSHEET.append(["CHANGED", DEV, AMAP[DEV]["label"], BMAP[DEV]["label"], None])
            BASE_ROW_INDEX = WORKSHEET.max_row

            EXTRA = max(0, len(DIFF_LINES) - 1)
            if EXTRA > 0:
                WORKSHEET.insert_rows(idx=BASE_ROW_INDEX + 1, amount=EXTRA)

            RED, GREEN, BLACK = "FFFF0000", "FF008000", "FF000000"
            for ROW_IDX, LINE in enumerate(DIFF_LINES or ["(无差异)"]):
                CELL = WORKSHEET.cell(row=BASE_ROW_INDEX + ROW_IDX, column=5, value=LINE)
                CELL.alignment = Alignment(vertical="top", wrap_text=False)
                if LINE.startswith("+") and not LINE.startswith("+++"):
                    CELL.font = Font(color=GREEN)
                elif LINE.startswith("-") and not LINE.startswith("---"):
                    CELL.font = Font(color=RED)
                else:
                    CELL.font = Font(color=BLACK)
                WORKSHEET.row_dimensions[BASE_ROW_INDEX + ROW_IDX].height = 15

            ADD_COUNT = sum(
                1 for LINE in DIFF_LINES
                if LINE.startswith("+") and not LINE.startswith("+++")
            )
            DELETE_COUNT = sum(
                1 for LINE in DIFF_LINES
                if LINE.startswith("-") and not LINE.startswith("---")
            )
            WORKSHEET.cell(row=BASE_ROW_INDEX, column=4).value = f"-{DELETE_COUNT} / +{ADD_COUNT}"

        # 概览
        self._OVERVIEW.append([
            site, len(ASET), len(BSET), len(ADDED), len(REMOVED), len(CHANGED),
            "" if not CHANGED else "存在配置差异"
        ])
        self._SUM_ADDED += len(ADDED)
        self._SUM_REMOVED += len(REMOVED)
        self._SUM_CHANGED += len(CHANGED)

        # 输出每个站点的结果
        self.add_result(
            Level.OK,
            f"站点 {site} 处理完成，新增DIFF设备={len(ADDED)}，"
            f"删除DIFF设备={len(REMOVED)}，配置产生变更设备数量={len(CHANGED)}"
        )

        # 如果是最后一个 site：总计、保存，并顺带跑剩余计划（无额外进度条）
        if site == self._SITES[-1]:
            self._OVERVIEW.insert_rows(1)
            self._OVERVIEW["A1"] = f"DIFF 总结（{self._START_YMD} → {self._END_YMD}）"
            self._OVERVIEW.append(["总新增设备", self._SUM_ADDED])
            self._OVERVIEW.append(["总删除设备", self._SUM_REMOVED])
            self._OVERVIEW.append(["总配置产生变更设备数量", self._SUM_CHANGED])

            # 创建输出目录（如果目录已存在则不报错）
            os.makedirs(self.OUTPUT_DIR, exist_ok=True)
            OUTPUT_NAME = (
                f"{self._END_YMD}-关键设备DIFF.xlsx"
                if not self._OUT_SUFFIX
                else f"{self._END_YMD}-{self._OUT_SUFFIX}-关键设备DIFF.xlsx"
            )
            OUTPUT_PATH = os.path.join(self.OUTPUT_DIR, OUTPUT_NAME)
            self._WB.save(OUTPUT_PATH)

            # 同日若还需"周/月"DIFF：这里顺带生成（一次性写完整文件，不显示进度条）
            for (BEFORE_YMD, AFTER_YMD, TAG) in self._PLANS[1:]:
                try:
                    self._do_diff_and_save(BEFORE_YMD, AFTER_YMD, out_suffix=TAG)
                except Exception as ERROR:
                    self.add_result(Level.WARN, f"{TAG or '日'}DIFF 失败：{ERROR}")

    # —— 备用：用于顺带生成周/月 DIFF（一次性完成，不走外层进度）——
    def _do_diff_and_save(self, start_date: str, end_date: str, out_suffix: Optional[str]):
        START_FILE_PATH = os.path.join(self.INPUT_DIR, f"{start_date}-{self.FILENAME_PREFIX}.xlsx")
        END_FILE_PATH = os.path.join(self.INPUT_DIR, f"{end_date}-{self.FILENAME_PREFIX}.xlsx")
        if not os.path.isfile(START_FILE_PATH) or not os.path.isfile(END_FILE_PATH):
            raise FileNotFoundError(
                f"源文件缺失：{os.path.basename(START_FILE_PATH)} 或 "
                f"{os.path.basename(END_FILE_PATH)}"
            )

        WORKBOOK_START = load_workbook(START_FILE_PATH, read_only=True, data_only=True)
        WORKBOOK_END = load_workbook(END_FILE_PATH, read_only=True, data_only=True)
        ALL_SITES = sorted(set(WORKBOOK_START.sheetnames) | set(WORKBOOK_END.sheetnames))

        OUT_WORKBOOK = Workbook()
        OVERVIEW = OUT_WORKBOOK.active
        OVERVIEW.title = "Overview"
        OVERVIEW.append(["起始文件", os.path.basename(START_FILE_PATH)])
        OVERVIEW.append(["结束文件", os.path.basename(END_FILE_PATH)])
        OVERVIEW.append([])
        OVERVIEW.append([
            "站点Sheet", "设备数(A)", "设备数(B)", "新增DIFF设备",
            "删除DIFF设备", "配置产生变更设备数量", "备注"
        ])

        TOTAL_ADDED = TOTAL_REMOVED = TOTAL_CHANGED = 0

        for SITE in ALL_SITES:
            AMAP = self._read_sheet_map(START_FILE_PATH, SITE)
            BMAP = self._read_sheet_map(END_FILE_PATH, SITE)
            ASET, BSET = set(AMAP.keys()), set(BMAP.keys())
            ADDED = sorted(list(BSET - ASET))
            REMOVED = sorted(list(ASET - BSET))
            COMMON = sorted(list(ASET & BSET))
            CHANGED = []

            WORKSHEET = OUT_WORKBOOK.create_sheet(title=(SITE[:28] if len(SITE) > 28 else SITE))
            WORKSHEET.append([f"站点：{SITE}"])
            WORKSHEET.append([f"起始：{start_date}", f"结束：{end_date}"])
            WORKSHEET.append(["变更类型", "设备Key", "起始列头", "结束列头", "说明/DIFF"])


            # 设置列宽
            WORKSHEET.column_dimensions['A'].width = 15  # 变更类型
            WORKSHEET.column_dimensions['B'].width = 28  # 设备Key
            WORKSHEET.column_dimensions['C'].width = 45    # 起始列头
            WORKSHEET.column_dimensions['D'].width = 9    # 结束列头
            WORKSHEET.column_dimensions['E'].width = 80  # 说明/DIFF

            for DEV in ADDED:
                WORKSHEET.append(["ADDED", DEV, "", BMAP[DEV]["label"], "设备仅存在于结束日"])
            for DEV in REMOVED:
                WORKSHEET.append(["REMOVED", DEV, AMAP[DEV]["label"], "", "设备仅存在于开始日"])

            for DEV in COMMON:
                LINES_A = AMAP[DEV]["lines"]
                LINES_B = BMAP[DEV]["lines"]


                # 过滤忽略的行
                FILTERED_LINES_A = self._filter_configuration_lines(LINES_A)
                FILTERED_LINES_B = self._filter_configuration_lines(LINES_B)


                if FILTERED_LINES_A == FILTERED_LINES_B:
                    continue
                CHANGED.append(DEV)
                DIFF_LINES = list(unified_diff(
                    FILTERED_LINES_A, FILTERED_LINES_B, fromfile="before",
                    tofile="after", lineterm=""
                ))

                WORKSHEET.append(["CHANGED", DEV, AMAP[DEV]["label"], BMAP[DEV]["label"], None])
                BASE_ROW_INDEX = WORKSHEET.max_row

                EXTRA = max(0, len(DIFF_LINES) - 1)
                if EXTRA > 0:
                    WORKSHEET.insert_rows(idx=BASE_ROW_INDEX + 1, amount=EXTRA)

                RED, GREEN, BLACK = "FFFF0000", "FF008000", "FF000000"
                for ROW_IDX, LINE in enumerate(DIFF_LINES or ["(无差异)"]):
                    CELL = WORKSHEET.cell(row=BASE_ROW_INDEX + ROW_IDX, column=5, value=LINE)
                    CELL.alignment = Alignment(vertical="top", wrap_text=False)
                    if LINE.startswith("+") and not LINE.startswith("+++"):
                        CELL.font = Font(color=GREEN)
                    elif LINE.startswith("-") and not LINE.startswith("---"):
                        CELL.font = Font(color=RED)
                    else:
                        CELL.font = Font(color=BLACK)
                    WORKSHEET.row_dimensions[BASE_ROW_INDEX + ROW_IDX].height = 15

                ADD_COUNT = sum(
                    1 for LINE in DIFF_LINES
                    if LINE.startswith("+") and not LINE.startswith("+++")
                )
                DELETE_COUNT = sum(
                    1 for LINE in DIFF_LINES
                    if LINE.startswith("-") and not LINE.startswith("---")
                )
                WORKSHEET.cell(
                    row=BASE_ROW_INDEX, column=4
                ).value = f"-{DELETE_COUNT} / +{ADD_COUNT}"

            OVERVIEW.append([
                SITE, len(ASET), len(BSET), len(ADDED), len(REMOVED), len(CHANGED),
                "" if not CHANGED else "存在配置差异"
            ])
            TOTAL_ADDED += len(ADDED)
            TOTAL_REMOVED += len(REMOVED)
            TOTAL_CHANGED += len(CHANGED)

        OVERVIEW.insert_rows(1)
        OVERVIEW["A1"] = f"DIFF 总结（{start_date} → {end_date}）"
        OVERVIEW.append(["总新增设备", TOTAL_ADDED])
        OVERVIEW.append(["总删除设备", TOTAL_REMOVED])
        OVERVIEW.append(["总配置产生变更设备数量", TOTAL_CHANGED])

        # 创建输出目录（如果目录已存在则不报错）
        os.makedirs(self.OUTPUT_DIR, exist_ok=True)
        OUTPUT_FILE_NAME = f"{end_date}-关键设备DIFF.xlsx" if not out_suffix \
            else f"{end_date}-{out_suffix}-关键设备DIFF.xlsx"
        OUTPUT_FILE_PATH = os.path.join(self.OUTPUT_DIR, OUTPUT_FILE_NAME)
        OUT_WORKBOOK.save(OUTPUT_FILE_PATH)
