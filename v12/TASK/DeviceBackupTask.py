# 关键设备配置备份输出EXCEL基础任务
#
# 技术栈:openpyxl、正则表达式
# 目标:将N9K ASA设备配置备份并输出为Excel格式，为后续ACL分析提供基础数据
#
# 处理逻辑:
# 设备配置解析:从LOG目录读取设备配置文件，支持智能分组策略；
# 策略提取:解析ACL策略配置，提取关键信息；
# 配置截取优化:根据设备类别智能截取相关配置段，减少冗余信息；
# Excel输出:按设备类型和站点生成Sheet，包含设备配置和策略信息
#
# 输入文件:从settings.log_dir读取LOG目录，设备配置文件位于{log_dir}/OxidizedTask/OxidizedTaskBackup/
# 输出文件:{log_dir}/DeviceBackupTask/{日期}-关键设备配置备份输出EXCEL基础任务.xlsx（V10新结构：从ACL/SourceACL迁移）
#
# 设备分类规则:
# - CS-N9K: N9K核心交换机（按站点分组）
# - LINK-AS: LINKAS接入交换机（按站点分组）
# - ASA-FW: ASA防火墙（按站点分组）
# - LINK-DS: LINK-DS交换机（统一到LINKDS工作表）
# - BGP: BGP设备（统一到BGP工作表）
# - OOB-DS: OOB-DS交换机（按站点分组）
#
# 配置截取规则:
# - CS-N9K (NXOS N9K):从"! show running-config"开始截取
# - LINK-AS (IOS-XE):从"! Last configuration change"开始截取
# - ASA-FW:从"ASA Version"开始截取
# - LINK-DS:支持N9K格式（"! show running-config"）或IOS-XE格式（"! Last configuration change"）
# - BGP:支持N9K格式（"! show running-config"）或IOS-XE格式（"! Last configuration change"）
# - OOB-DS (IOS-XE):从"! Last configuration change"开始截取
#
# 输出:生成标准化的ACL配置Excel文件，供其他ACL分析任务使用

# 导入标准库
import os
import re
from datetime import datetime
from typing import Optional

# 导入第三方库
# (无第三方库依赖)

# 导入本地应用
from .TaskBase import (
    BaseTask, Level, CONFIG, extract_site_from_filename,
    safe_sheet_name, require_keys, get_today_str, format_datetime,
    ensure_output_dir, build_log_path, build_output_path,
    load_excel_workbook, create_excel_workbook, save_excel_workbook
)
from .CiscoBase import (
    get_device_classification_rules, is_cat1_device, is_cat2_device, is_cat6_device
)

# 关键设备配置备份任务类：将设备配置备份并输出为Excel格式
class DeviceBackupTask(BaseTask):
    """关键设备配置备份任务


    将N9K ASA设备配置备份并输出为Excel格式，为后续ACL分析提供基础数据
    """


    # 初始化关键设备配置备份任务：设置输出目录和数据结构
    def __init__(self):
        super().__init__("关键设备配置备份输出EXCEL基础任务")
        # 从配置文件读取log_dir（必须配置）
        require_keys(CONFIG, ["settings"], "root")
        require_keys(CONFIG["settings"], ["log_dir"], "settings")
        self.LOG_DIR = CONFIG["settings"]["log_dir"]
        # V10新结构：直接输出到 LOG/DeviceBackupTask/
        self.OUTPUT_DIR = os.path.join(self.LOG_DIR, "DeviceBackupTask")

        # ↓↓↓ 为 1/N 进度新增的最小状态（其余逻辑不变）
        self._TODAY = None
        self._IN_DIR = None
        self._XLSX_PATH = None
        # site -> {CS-N9K:[path...], LINK-AS:[...], ASA-FW:[...]}
        self._GROUPED_PATHS: dict[str, dict[str, list[str]]] = {}
        self._SITES_ORDER: list[str] = []
        self._WB = None
        self._TOTAL_DEVICES = 0

    # 扫描LOG目录获取站点列表：按站点分组N9K ASA设备配置文件
    def items(self):
        """扫描LOG目录获取站点列表


        按站点分组N9K ASA设备配置文件


        Returns:
            list[str]: 站点列表
        """
        from openpyxl import Workbook

        self._TODAY = get_today_str()
        # V10新结构：从 LOG/OxidizedTask/OxidizedTaskBackup/ 读取
        self._IN_DIR = os.path.join(self.LOG_DIR, "OxidizedTask", "OxidizedTaskBackup")
        if not os.path.isdir(self._IN_DIR):
            self.add_result(Level.ERROR, f"未找到当日 Oxidized 日志目录: {self._IN_DIR}")
            return []

        # 创建输出目录（如果目录已存在则不报错）
        ensure_output_dir(self.OUTPUT_DIR)
        self._XLSX_PATH = os.path.join(
            self.OUTPUT_DIR,
            f"{self._TODAY}-关键设备配置备份输出EXCEL基础任务.xlsx"
        )

        # 分组处理：根据设备分类规则进行分组
        GROUPED_PATHS: dict[str, dict[str, list[str]]] = {}
        RULES = self._get_device_classification_rules()


        # 只处理当天的文件：检查文件名是否以当天日期开头
        TODAY_DATE_TEXT = self._TODAY  # 格式：YYYYMMDD
        # 只匹配格式: YYYYMMDD-设备名.log（日期在文件名开头）
        DATE_PATTERN = re.compile(r'^' + re.escape(TODAY_DATE_TEXT) + r'-')


        for NAME in os.listdir(self._IN_DIR):
            if not NAME.lower().endswith(".log"):
                continue
            FULL_DOCUMENT_PATH = os.path.join(self._IN_DIR, NAME)
            if not os.path.isfile(FULL_DOCUMENT_PATH):
                continue


            # 检查文件名是否以当天日期开头
            if not DATE_PATTERN.match(NAME):
                continue  # 跳过不以当天日期开头的文件

            CATEGORY = self._classify(NAME)  # 保持原分类
            if not CATEGORY:
                continue


            # 获取该分类的分组策略
            CATEGORY_RULE = RULES.get(CATEGORY, {})
            GROUP_BY_SITE = CATEGORY_RULE.get("group_by_site", True)


            if GROUP_BY_SITE:
                # 按站点分组
                SITE = self._extract_site(NAME)
                SITE_DICTIONARY = GROUPED_PATHS.setdefault(SITE, {})
                SITE_DICTIONARY.setdefault(CATEGORY, []).append(FULL_DOCUMENT_PATH)
            else:
                # 不按站点分组，使用指定的工作表名称
                SHEET_NAME = CATEGORY_RULE.get("sheet_name", CATEGORY.upper())
                SHEET_DICTIONARY = GROUPED_PATHS.setdefault(SHEET_NAME, {})
                SHEET_DICTIONARY.setdefault(CATEGORY, []).append(FULL_DOCUMENT_PATH)

        if not GROUPED_PATHS:
            self.add_result(Level.OK, f"未匹配到目标设备日志，未生成 Excel（目录：{self._IN_DIR}）")
            return []

        # 站点内保证稳定顺序：按分类规则定义的顺序，再按文件名排序
        for SITE, CATS in GROUPED_PATHS.items():
            for CATEGORY in RULES.keys():
                if CATEGORY in CATS:
                    CATS[CATEGORY].sort(key=lambda PATH: os.path.basename(PATH))

        # 初始化工作簿（不立即保存）
        WORKBOOK = Workbook()
        try:
            WORKBOOK.remove(WORKBOOK.active)
        except Exception:
            pass

        self._WB = WORKBOOK
        self._GROUPED_PATHS = GROUPED_PATHS
        self._SITES_ORDER = sorted(GROUPED_PATHS.keys())  # 进度总数 N
        self._TOTAL_DEVICES = 0

        # 返回每个站点一个 item，实现进度 1/N
        return self._SITES_ORDER

    # 判断文件名是否为目标设备：检查文件名是否匹配CS+N9K或LINK+AS模式
    @staticmethod
    def _is_target(OxidizedBackup_FILENAME: str) -> bool:
        import re
        FILENAME_LOWER = OxidizedBackup_FILENAME.lower()
        # 类别1：CS + N9K + (01|02|03|04)，兼容连写
        CONTAINS_CORE_SWITCH = ("cs" in FILENAME_LOWER) or re.search(r"\bcs\b", FILENAME_LOWER)
        CONTAINS_NEXUS_9K = ("n9k" in FILENAME_LOWER) or re.search(r"\bn9k\b", FILENAME_LOWER)
        CONTAINS_DEVICE_NUMBER = re.search(r"(?:^|[^0-9])0?[1-4](?:[^0-9]|$)", FILENAME_LOWER)
        CORE_SWITCH_JOIN = re.search(r"cs0?[1-4]", FILENAME_LOWER)
        CLASS1 = (
            (CONTAINS_CORE_SWITCH and CONTAINS_NEXUS_9K and CONTAINS_DEVICE_NUMBER) or
            (CORE_SWITCH_JOIN and CONTAINS_NEXUS_9K)
        )

        # 类别2：LINK + AS + (01|02)，兼容连写
        CONTAINS_LINK = ("link" in FILENAME_LOWER) or re.search(r"\blink\b", FILENAME_LOWER)
        CONTAINS_ACCESS_SWITCH = ("as" in FILENAME_LOWER) or re.search(r"\bas\b", FILENAME_LOWER)
        CONTAINS_01_02 = re.search(r"(?:^|[^0-9])0?[12](?:[^0-9]|$)", FILENAME_LOWER)
        LINKAS_JOIN = re.search(r"link[-_]*as0?[12]", FILENAME_LOWER)
        ACCESS_SWITCH_JOIN = re.search(r"as0?[12]", FILENAME_LOWER)
        CLASS2 = (
            LINKAS_JOIN or
            (CONTAINS_LINK and CONTAINS_ACCESS_SWITCH and CONTAINS_01_02) or
            (CONTAINS_LINK and ACCESS_SWITCH_JOIN)
        )
        return bool(CLASS1 or CLASS2)

    # 获取设备分类规则（包含分组策略）：从CiscoBase获取基础规则，添加DeviceBackupTask特有的分组策略
    @staticmethod
    def _get_device_classification_rules():
        """获取设备分类规则（包含分组策略）

        从CiscoBase获取基础规则，添加DeviceBackupTask特有的分组策略（group_by_site和sheet_name）

        Returns:
            Dict: 设备分类规则字典，包含分组策略
        """
        # 从CiscoBase获取基础规则
        BASE_RULES = get_device_classification_rules()

        # 添加DeviceBackupTask特有的分组策略
        RULES = {}
        for CAT_ID, RULE_CONFIG in BASE_RULES.items():
            RULES[CAT_ID] = RULE_CONFIG.copy()
            # 添加分组策略（默认按站点分组）
            if CAT_ID in ["cat1", "cat2", "cat3", "cat6"]:
                RULES[CAT_ID]["group_by_site"] = True
            elif CAT_ID == "cat4":
                RULES[CAT_ID]["group_by_site"] = False
                RULES[CAT_ID]["sheet_name"] = "LINKDS"
            elif CAT_ID == "cat5":
                RULES[CAT_ID]["group_by_site"] = False
                RULES[CAT_ID]["sheet_name"] = "BGP"

        return RULES

    # 对目标文件进行分类：根据文件名模式将文件分为不同类别
    @staticmethod
    def _classify(FILENAME: str) -> Optional[str]:
        """对目标文件进行分类


        Args:
            FILENAME: 文件名


        Returns:
            Optional[str]: 分类ID，如果无法分类则返回None
        """
        import re
        FILENAME_LOWER = FILENAME.lower()
        # 使用CiscoBase中的统一设备分类规则
        RULES = get_device_classification_rules()

        # 遍历所有分类规则
        for CAT_ID, RULE_CONFIG in RULES.items():
            for PATTERN_FUNC in RULE_CONFIG["patterns"]:
                if PATTERN_FUNC(FILENAME_LOWER):
                    return CAT_ID
        return None

    # 从文件名提取站点名：解析文件名获取站点标识
    @staticmethod
    def _extract_site(FILENAME: str) -> str:
        """从文件名提取站点名


        Args:
            FILENAME: 文件名


        Returns:
            str: 站点名
        """
        return extract_site_from_filename(FILENAME)

    # 生成安全的Excel工作表名称：确保工作表名称符合Excel规范
    @staticmethod
    def _safe_sheet_name(NAME: str) -> str:
        """生成安全的Excel工作表名称


        Args:
            NAME: 原始名称


        Returns:
            str: 安全的Excel工作表名称
        """
        return safe_sheet_name(NAME)

    # 根据设备类别截取配置内容：从指定起始行开始截取配置
    def _extract_configuration(self, LINES: list[str], CATEGORY: str, FNAME: str) -> list[str]:
        """根据设备类别截取配置内容


        Args:
            LINES: 配置行列表
            CATEGORY: 设备类别
            FNAME: 文件名


        Returns:
            list[str]: 截取后的配置行列表
        """
        if CATEGORY == "cat1":  # CS-N9K (NXOS N9K)
            START_DELIMITER = "! show running-config"
            for LINE_INDEX, LINE in enumerate(LINES):
                if LINE.strip().startswith(START_DELIMITER):
                    return LINES[LINE_INDEX:]  # 从该行开始截取到文件末尾
        elif CATEGORY == "cat2":  # LINK-AS (LINKAS接入交换机，IOS-XE)
            # LINKAS接入交换机使用IOS-XE配置格式
            IOSXE_MARKER = "! Last configuration change"
            for LINE_INDEX, LINE in enumerate(LINES):
                if LINE.strip().startswith(IOSXE_MARKER):
                    return LINES[LINE_INDEX:]  # IOS-XE 配置
        elif CATEGORY == "cat3":  # ASA-FW (ASA防火墙)
            # ASA防火墙配置提取
            ASA_DELIMITER = "ASA Version"
            for LINE_INDEX, LINE in enumerate(LINES):
                if LINE.strip().startswith(ASA_DELIMITER):
                    return LINES[LINE_INDEX:]  # ASA 配置
        elif CATEGORY == "cat4":  # LINK-DS (LINK-DS交换机)
            # LINK-DS交换机配置提取（支持N9K和IOS-XE）
            NEXUS_9K_DELIMITER = "! show running-config"
            IOSXE_MARKER = "! Last configuration change"
            for LINE_INDEX, LINE in enumerate(LINES):
                if LINE.strip().startswith(NEXUS_9K_DELIMITER):
                    return LINES[LINE_INDEX:]  # N9K 配置
                elif LINE.strip().startswith(IOSXE_MARKER):
                    return LINES[LINE_INDEX:]  # IOS-XE 配置
        elif CATEGORY == "cat5":  # BGP (BGP设备)
            # BGP设备配置提取（支持N9K和IOS-XE）
            NEXUS_9K_DELIMITER = "! show running-config"
            IOSXE_MARKER = "! Last configuration change"
            for LINE_INDEX, LINE in enumerate(LINES):
                if LINE.strip().startswith(NEXUS_9K_DELIMITER):
                    return LINES[LINE_INDEX:]  # N9K 配置
                elif LINE.strip().startswith(IOSXE_MARKER):
                    return LINES[LINE_INDEX:]  # IOS-XE 配置
        elif CATEGORY == "cat6":  # OOB-DS (OOB-DS交换机，IOS-XE)
            # OOB-DS交换机使用IOS-XE配置格式
            IOSXE_MARKER = "! Last configuration change"
            for LINE_INDEX, LINE in enumerate(LINES):
                if LINE.strip().startswith(IOSXE_MARKER):
                    return LINES[LINE_INDEX:]  # IOS-XE 配置


        # 如果没找到标记，返回原始内容
        return LINES

    # 处理单个站点/工作表的配置备份：读取配置文件并生成Excel工作表
    def run_single(self, SITE: str):
        """处理单个站点/工作表的配置备份


        读取配置文件并生成Excel工作表
        参数SITE可能是站点名（对于CS-N9K/LINK-AS/ASA-FW）或工作表名（对于LINK-DS/BGP）


        Args:
            SITE: 站点名或工作表名
        """
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        if not self._WB or SITE not in self._GROUPED_PATHS:
            self.add_result(Level.WARN, f"站点/工作表 {SITE} 跳过（未初始化或未找到文件）")
            return

        WORKSHEET = self._WB.create_sheet(title=self._safe_sheet_name(SITE))
        WRAP = Alignment(wrap_text=True, vertical="top")

        COLUMN_INDEX = 1
        SITE_CATEGORY_DICTIONARY = self._GROUPED_PATHS[SITE]

        # 列顺序：按分类规则定义的顺序显示
        RULES = self._get_device_classification_rules()
        for CATEGORY in RULES.keys():
            for FULL_DOCUMENT_PATH in SITE_CATEGORY_DICTIONARY.get(CATEGORY, []):
                DOCUMENT_NAME = os.path.basename(FULL_DOCUMENT_PATH)
                try:
                    with open(
                        FULL_DOCUMENT_PATH, "r", encoding="utf-8", errors="ignore"
                    ) as FILE_HANDLE:
                        LINES = FILE_HANDLE.read().splitlines()


                    # 根据设备类别截取配置内容
                    CONFIGURATION_LINE_LIST = self._extract_configuration(LINES, CATEGORY, DOCUMENT_NAME)


                except Exception as EXCEPTION:
                    self.add_result(Level.ERROR, f"读取失败 {DOCUMENT_NAME}: {EXCEPTION}")
                    # 仍占一列留痕
                    CELL = WORKSHEET.cell(
                        row=1, column=COLUMN_INDEX,
                        value=f"{DOCUMENT_NAME}（读取失败：{EXCEPTION}）"
                    )
                    CELL.alignment = WRAP
                    WORKSHEET.column_dimensions[get_column_letter(COLUMN_INDEX)].width = 80
                    COLUMN_INDEX += 1
                    continue

                WORKSHEET.cell(row=1, column=COLUMN_INDEX, value=DOCUMENT_NAME).alignment = WRAP
                ROW_POSITION = 2
                for LINE in CONFIGURATION_LINE_LIST:
                    WORKSHEET.cell(row=ROW_POSITION, column=COLUMN_INDEX, value=LINE).alignment = WRAP
                    WORKSHEET.row_dimensions[ROW_POSITION].height = 15
                    ROW_POSITION += 1
                WORKSHEET.column_dimensions[get_column_letter(COLUMN_INDEX)].width = 80
                COLUMN_INDEX += 1

        # 统计：该站点的设备列数
        SITE_DEVICE_COUNT = COLUMN_INDEX - 1
        self._TOTAL_DEVICES += SITE_DEVICE_COUNT

        # 输出每个站点/工作表的结果
        self.add_result(Level.OK, f"站点/工作表 {SITE} 处理完成，设备数={SITE_DEVICE_COUNT}")

        # 检查是否是最后一个站点/工作表（通过其在列表中的位置判断）
        if SITE not in self._SITES_ORDER:
            # 如果SITE不在列表中，说明有逻辑错误，但不影响保存
            return
        CURRENT_INDEX = self._SITES_ORDER.index(SITE)
        if CURRENT_INDEX == len(self._SITES_ORDER) - 1:
            # 这是最后一个站点，保存Excel文件
            try:
                self._WB.save(self._XLSX_PATH)
            except Exception as EXCEPTION:
                self.add_result(Level.ERROR, f"Excel 保存失败: {self._XLSX_PATH} -> {EXCEPTION}")
                return

