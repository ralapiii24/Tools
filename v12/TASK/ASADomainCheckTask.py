# ASA 域名提取和检测任务
#
# 技术栈:socket、正则表达式、DNS解析、文件IO、openpyxl
# 目标:从 Excel 源文件中提取站点（仅 FW01 列），进行域名检测与标色，并支持基于手工清单的脚本生成
#
# 处理逻辑:
# - 站点来源: 读取 LOG/DeviceBackupTask/{日期}-关键设备配置备份输出EXCEL基础任务.xlsx 的各站点 sheet（V10新结构：从ACL/SourceACL迁移）
# - 域名提取: 仅扫描 FW01(-FRP) 列，支持完整域名与子串提取，跳过 URL/NAT/host-range-subnet/domain-name
# - DNS检测: 先加载本地缓存，未命中再解析；DNS不可达降级执行（仅缓存）
# - Excel标色: 保留源样式，失败标红并按场景底色区分；对象组/绑定行联动标色
# - 稳健输出: 每次任务强制删除并重建输出Excel（V10优化）
# - 输出优化: 单站点输出合并为单行汇总，静默中间OK日志（V10优化）
#
# 脚本生成（手工回收分支）:
# - 域名来源: 仅使用 CHECKRULE/ManualDomainRecovery.log（V10新结构：从DOMAIN迁移，开启 use_manual_source_first 时且非空）
# - 输出: LOG/ASADomainCheckTask/ConfigureOutput/{日期}-{SITE}-回退脚本.log / 操作脚本.log（V10新结构：从CONFIGURATION/ASADomainCheckTask/日期迁移，每次任务自动清空ConfigureOutput目录）
# - 逻辑: 先从回退脚本（按 FW01 列标色顺序）提取，再派生操作脚本（解绑→删除，保持顺序一致）
#
# DNS 缓存（统一文件）:
# - 路径: CHECKRULE/DNSLocalCache.log（V10新结构：从DOMAIN迁移），三态：S=特殊白名单、T=临时成功、D=失败
# - 行为: 月初自动清理 T；D→S 自动晋级（如果成功域名在D中，今天成功则自动升S）；S 优先且不被失败覆盖
#
# 日志优化:
# - 支持 settings.suppress_ok_logs 抑制 OK 级日志
# - 静默忽略无FW01表头的 sheet
# - 静默"已从源Excel初始化输出文件"OK日志（V10优化）
#
# 输出文件:
# - LOG/ASADomainCheckTask/{日期}-ASA域名提取及检测任务.xlsx（V10新结构：从DOMAIN/ASADomainCheckTask迁移）
# - LOG/ASADomainCheckTask/ConfigureOutput/{日期}-ManualDomainRecovery.xlsx（模板标色，V10新结构：从CONFIGURATION迁移）
#
# 日期逻辑:自动扫描LOG目录，按日期分组处理，月初第一天清空临时缓存
# 站点识别:使用与ASACompareTask相同的站点提取规则
#
# 缓存机制:
# - CHECKRULE/DNSLocalCache.log（V10新结构：从DOMAIN迁移）: 统一DNS缓存文件，三态：S=特殊白名单、T=临时成功、D=失败
# - 优先使用缓存，缓存未命中时查询DNS服务器
# - 自动去重：添加域名时检测重复，避免重复写入
#
# 失败跟踪:
# - 每天失败域名保存到LOG/日期/DNSQueryFalse.log（V10: 已废弃，不再使用）
# - 对比前一天失败域名，找出恢复的域名（V10: 已废弃，不再使用）
# - 恢复域名保存到LOG/日期/DNSQueryDontDelete.log（V10: 已废弃，不再使用）
# - 恢复域名添加到特殊缓存，REPORT中WARN提示
#
# 配置说明:所有配置硬编码在Python文件中，自动扫描LOG目录，支持fw01-frp和fw02-frp文件格式

# 导入标准库
import concurrent.futures
import os
import re
import shutil
import socket

from datetime import datetime, timedelta
from typing import Dict, List, Optional, Set, Tuple

# 导入第三方库
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# 导入本地应用
from .TaskBase import BaseTask, Level, CONFIG, require_keys

# ASA域名提取和检测任务类
# 任务1：从ASA FW配置中提取域名，进行DNS解析检测，并将结果标色到Excel
# 任务2：从手动输入的域名清单生成回收脚本（操作脚本和回退脚本）并标色到Excel
class ASADomainCheckTask(BaseTask):
    """ASA域名提取及检测任务

    从ASA防火墙配置中提取域名，并进行DNS解析检测
    """

    # 初始化ASA域名提取和检测任务：设置输入输出路径和DNS服务器
    def __init__(self):
        super().__init__("ASA域名提取及检测任务")
        # V10新结构：从 LOG/DeviceBackupTask/ 读取（ACL/SourceACL已迁移）
        self.ACL_DIR = os.path.join("LOG", "DeviceBackupTask")
        # V10新结构：直接输出到 LOG/ASADomainCheckTask/
        self.OUTPUT_DIR = os.path.join("LOG", "ASADomainCheckTask")
        self.DNS_SERVERS = ["8.8.8.8", "1.1.1.1"]
        self.DNS_TIMEOUT = 3  # DNS解析超时时间(秒)
        self._TODAY = None
        self._SITES_DATA = {}  # {site: {}} 用于记录已识别的站点


        # 统一DNS缓存文件（S/T/D 三态）：S=特殊白名单(不回收), T=临时成功缓存, D=可回收失败
        # V10新结构：从 CHECKRULE/ 读取（DOMAIN/已迁移）
        self.DNS_CACHE_FILE = os.path.join("CHECKRULE", "DNSLocalCache.log")
        self._DNS_CACHE = set()  # 内存中的成功缓存（S与T均视为成功）
        self._DNS_SPECIAL = set()  # S 集合
        self._DNS_TEMP = set()     # T 集合
        self._DNS_DELETE = set()   # D 集合（仅记录，可用于统计/参考）
        self._CACHE_LOADED = False  # 统一DNS缓存是否已加载
        # 生成操作/回退脚本参数（路径写死；是否启用与优先级由 YAML 开关控制）
        # V10新结构（调整）：脚本固定输出至 LOG/ASADomainCheckTask/ConfigureOutput/
        self.SCRIPT_OUTPUT_ROOT = os.path.join("LOG", "ASADomainCheckTask", "ConfigureOutput")
        # 域名源：仅使用手工回收清单
        # V10新结构：从 CHECKRULE/ 读取（DOMAIN/已迁移）
        self.RECOVERY_OVERRIDE = os.path.join("CHECKRULE", "ManualDomainRecovery.log")
        # YAML 开关（必须配置）
        require_keys(CONFIG, ["ASADomainCheckTask"], "root")
        require_keys(
            CONFIG["ASADomainCheckTask"],
            ["enable_domain_scripts", "use_manual_source_first"],
            "ASADomainCheckTask"
        )
        ASACFG = CONFIG["ASADomainCheckTask"]
        self.ENABLE_DOMAIN_SCRIPTS = bool(ASACFG["enable_domain_scripts"])
        self.USE_MANUAL_SOURCE_FIRST = bool(ASACFG["use_manual_source_first"])
        # 预编译域名提取正则，避免重复编译造成开销
        self._DOMAIN_FULL_RE = re.compile(
            r'(?:[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?\.)+[A-Za-z]{2,}'
        )
        self._DOMAIN_TOKEN_RE = re.compile(
            r'([A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?'
            r'(?:\.[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?)+)'
        )


    # 扫描源Excel获取站点列表：以首行识别FW01列
    def items(self):
        """返回要处理的站点列表

        从源Excel识别需处理的站点（要求存在FW01列）

        Returns:
            站点名称列表
        """
        self._TODAY = datetime.now().strftime("%Y%m%d")
        source_excel = os.path.join(self.ACL_DIR, f"{self._TODAY}-关键设备配置备份输出EXCEL基础任务.xlsx")

        if not os.path.isfile(source_excel):
            self.add_result(Level.ERROR, f"未找到源Excel: {source_excel}")
            return []

        os.makedirs(self.OUTPUT_DIR, exist_ok=True)

        # 读取源Excel，确认各sheet是否含FW01表头
        try:
            WORKBOOK = load_workbook(source_excel, read_only=True, data_only=True)
        except Exception as ERROR:
            self.add_result(Level.ERROR, f"打开源Excel失败: {ERROR}")
            return []

        candidate_sites: List[str] = []
        for SHEET_NAME in WORKBOOK.sheetnames:
            WORKSHEET = WORKBOOK[SHEET_NAME]
            fw01_present = False
            # 扫描首行表头
            for COLUMN in range(1, (WORKSHEET.max_column or 1) + 1):
                header = str(WORKSHEET.cell(row=1, column=COLUMN).value or '')
                # 与其他任务一致：fw01-frp（大小写不敏感，允许可选连字符/下划线、可选前导0）
                if re.search(r"fw\s*0*1(?:[-_]?frp)?", header, flags=re.IGNORECASE):
                    fw01_present = True
                    break
            # 仅要求存在 FW01(-FRP)
            if fw01_present:
                candidate_sites.append(SHEET_NAME)
            else:
                # 静默忽略无FW01表头的sheet（不写入日志）
                pass
        WORKBOOK.close()

        # 仅以存在FW01表头的sheet作为站点
        if not candidate_sites:
            self.add_result(Level.ERROR, "未在源Excel中找到含FW01表头的站点sheet")
        self._SITES_DATA = {SITE: {} for SITE in candidate_sites}
        return sorted(candidate_sites)


        # 加载DNS缓存（统一文件）
    # 加载统一DNS缓存文件，并在月初清理T条目：force=True为强制重载，否则若已加载过则直接返回；verbose=False为不输出OK日志
    def _load_dns_cache(self, force: bool = False, verbose: bool = True) -> None:
        if self._CACHE_LOADED and not force:
            return
        self._DNS_CACHE.clear()
        self._DNS_SPECIAL.clear()
        self._DNS_TEMP.clear()
        self._DNS_DELETE.clear()

        # 读取统一缓存
        if os.path.exists(self.DNS_CACHE_FILE):
            try:
                with open(self.DNS_CACHE_FILE, 'r', encoding='utf-8') as FILE_HANDLE:
                    for RAW_LINE in FILE_HANDLE:
                        LINE = (RAW_LINE or '').strip()
                        if not LINE:
                            continue
                        if len(LINE) < 3 or LINE[1] != ' ':
                            continue
                        TAG = LINE[0].upper()
                        DOMAIN = LINE[2:].strip()
                        if not DOMAIN:
                            continue
                        if TAG == 'S':
                            self._DNS_SPECIAL.add(DOMAIN)
                            self._DNS_CACHE.add(DOMAIN)
                        elif TAG == 'T':
                            self._DNS_TEMP.add(DOMAIN)
                            self._DNS_CACHE.add(DOMAIN)
                        elif TAG == 'D':
                            self._DNS_DELETE.add(DOMAIN)
                # 静默加载缓存：不输出统计日志
            except Exception as ERROR:
                self.add_result(Level.ERROR, f"加载DNS缓存失败: {ERROR}")

        # 月初清理：移除所有 T 项，保留 S 与 D
        today = datetime.now()
        if today.day == 1 and os.path.exists(self.DNS_CACHE_FILE):
            try:
                LINES: list[str] = []
                with open(self.DNS_CACHE_FILE, 'r', encoding='utf-8') as FILE_HANDLE:
                    for RAW_LINE in FILE_HANDLE:
                        LINE = (RAW_LINE or '').rstrip('\n')
                        if not LINE:
                            continue
                        if LINE.startswith('T '):
                            continue
                        LINES.append(LINE)
                with open(self.DNS_CACHE_FILE, 'w', encoding='utf-8') as FILE_HANDLE:
                    for LINE in LINES:
                        FILE_HANDLE.write(LINE + "\n")
                # 重新加载到内存（静默）
                self._CACHE_LOADED = False
                self._load_dns_cache(force=True, verbose=False)
                if verbose:
                    self.add_result(Level.OK, "月初第一天，已清理临时成功缓存(T)")
            except Exception as error:
                self.add_result(Level.WARN, f"月初清理DNS缓存失败: {error}")
        self._CACHE_LOADED = True


    # 保存DNS缓存
    # 将成功域名写入统一缓存(T)，自动去重；如果域名在D中（昨天失败），自动升S
    def _save_dns_cache(self, successful_domains: Set[str]) -> None:
        if not successful_domains:
            return


        try:
            # 先加载现有集合
            self._load_dns_cache()


            # 检查：如果成功域名在D中（昨天失败），今天成功则自动升S（不回收）
            recovered_domains = successful_domains & self._DNS_DELETE
            if recovered_domains:
                self._add_to_special_cache(recovered_domains)


            # 不写入已在 S 或 T 中的域名
            skip = self._DNS_SPECIAL | self._DNS_TEMP
            new_domains = sorted(DOMAIN for DOMAIN in successful_domains if DOMAIN not in skip)
            if not new_domains:
                # 静默：不输出OK日志
                return
            os.makedirs(os.path.dirname(self.DNS_CACHE_FILE), exist_ok=True)
            with open(self.DNS_CACHE_FILE, 'a', encoding='utf-8') as FILE_HANDLE:
                for DOMAIN in new_domains:
                    FILE_HANDLE.write(f"T {DOMAIN}\n")
            self._DNS_TEMP.update(new_domains)
            self._DNS_CACHE.update(new_domains)
            # 静默：不输出OK日志
        except Exception as ERROR:
            self.add_result(Level.ERROR, f"保存DNS缓存失败: {ERROR}")


    # 检查域名是否在缓存中
    # 检查域名是否在缓存中
    def _is_domain_cached(self, domain: str) -> bool:
        return domain in self._DNS_CACHE


    # 保存失败的域名到日志文件：将失败域名记录为 D 项到统一缓存（追加且去重，不影响本轮标色）
    def _save_failed_domains(self, failed_domains: Set[str]) -> None:
        if not failed_domains:
            return

        try:
            # 先加载集合
            self._load_dns_cache()
            # 不对 S(白名单) 写入 D；且避免重复
            new_domains = sorted(
                DOMAIN for DOMAIN in failed_domains
                if DOMAIN not in self._DNS_SPECIAL and DOMAIN not in self._DNS_DELETE
            )
            if not new_domains:
                # 静默：不输出OK日志
                return
            os.makedirs(os.path.dirname(self.DNS_CACHE_FILE), exist_ok=True)
            with open(self.DNS_CACHE_FILE, 'a', encoding='utf-8') as FILE_HANDLE:
                for DOMAIN in new_domains:
                    FILE_HANDLE.write(f"D {DOMAIN}\n")
            self._DNS_DELETE.update(new_domains)
            # 静默：不输出OK日志
        except Exception as ERROR:
            self.add_result(Level.ERROR, f"保存失败域名到本地缓存失败: {ERROR}")


    # 保存恢复的域名到DNSQueryDontDelete.log
    # 统一缓存下，恢复域名直接加入 S（特殊白名单）
    def _save_recovered_domains(self, recovered_domains: Set[str]) -> None:
        self._add_to_special_cache(recovered_domains)


    # 将域名添加到特殊缓存：将域名添加到统一缓存(S)，自动去重
    def _add_to_special_cache(self, domains: Set[str]) -> None:
        if not domains:
            return


        try:
            self._load_dns_cache()
            new_domains = sorted(DOMAIN for DOMAIN in domains if DOMAIN not in self._DNS_SPECIAL)
            if not new_domains:
                # 静默：不输出OK日志
                return
            os.makedirs(os.path.dirname(self.DNS_CACHE_FILE), exist_ok=True)
            with open(self.DNS_CACHE_FILE, 'a', encoding='utf-8') as FILE_HANDLE:
                for DOMAIN in new_domains:
                    FILE_HANDLE.write(f"S {DOMAIN}\n")
            self._DNS_SPECIAL.update(new_domains)
            self._DNS_CACHE.update(new_domains)
            # 静默：不输出OK日志
        except Exception as ERROR:
            self.add_result(Level.ERROR, f"添加域名到特殊白名单失败: {ERROR}")


    # 检测DNS服务器可达性
    # 检测DNS服务器是否可达
    def _check_dns_reachability(self) -> bool:
        for dns_server in self.DNS_SERVERS:
            try:
                # 尝试连接DNS服务器
                sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                sock.settimeout(self.DNS_TIMEOUT)  # 使用配置的超时时间
                sock.connect((dns_server, 53))
                sock.close()
                # 静默：DNS可达性不写入日志
                return True
            except Exception:
                continue


        self.add_result(Level.ERROR, f"所有DNS服务器不可达: {', '.join(self.DNS_SERVERS)}")
        return False


    # 解析域名
    # 解析域名，返回(是否成功, 解析结果)
    def _resolve_domain(self, domain: str) -> Tuple[bool, str]:
        try:
            # 使用系统默认DNS解析
            result = socket.gethostbyname(domain)
            return True, result
        except socket.gaierror as error:
            return False, str(error)
        except Exception as error:
            return False, str(error)


    # 重写run方法：先检查DNS可达性，再处理所有平台
    # 执行ASA域名提取和检测任务
    def run(self) -> None:
        """执行ASA域名提取及检测任务

        加载DNS缓存，处理所有站点，提取域名并进行DNS解析检测
        """
        self._load_dns_cache()
        # 确保并清空脚本固定输出目录（每次运行前清空）
        try:
            os.makedirs(self.SCRIPT_OUTPUT_ROOT, exist_ok=True)
            for _name in os.listdir(self.SCRIPT_OUTPUT_ROOT):
                _path = os.path.join(self.SCRIPT_OUTPUT_ROOT, _name)
                if os.path.isfile(_path):
                    try:
                        os.remove(_path)
                    except Exception:
                        pass
        except Exception:
            pass


        # 先检查DNS服务器可达性；不可达时不中断，降级为“仅缓存+快速失败”模式
        if not self._check_dns_reachability():
            self.add_result(Level.WARN, "DNS服务器不可达，继续执行（仅缓存命中，未命中将快速失败）")
        # 设置全局socket超时，避免解析长时间阻塞
        try:
            socket.setdefaulttimeout(self.DNS_TIMEOUT)
        except Exception:
            pass


        # 调用父类的run方法处理所有平台
        super().run()
        # 注意：脚本生成已改为在 run_single 中按站点触发，这里不再全局生成，避免重复输出


    # 处理单个站点：提取域名并进行DNS检测（任务1）
    # 任务1：处理单个站点：从Excel配置中提取域名，进行DNS解析检测，将检测结果标色到 DOMAIN/ASADomainCheckTask/{date}-ASA域名提取及检测任务.xlsx，如果启用脚本生成，触发任务2（额外分支）
    def run_single(self, site: str) -> None:
        """处理单个站点

        从源Excel的该站点sheet中读取FW01列域名，进行DNS解析检测

        Args:
            site: 站点名称
        """
        if site not in self._SITES_DATA:
            self.add_result(Level.ERROR, f"站点 {site} 数据不存在")
            return


        try:
            # 从源Excel的该站点sheet中读取FW01列域名
            source_excel_path = os.path.join(
                self.ACL_DIR,
                f"{self._TODAY}-关键设备配置备份输出EXCEL基础任务.xlsx"
            )
            WORKBOOK_SRC = load_workbook(source_excel_path, read_only=True, data_only=True)
            if site not in WORKBOOK_SRC.sheetnames:
                self.add_result(Level.ERROR, f"源Excel缺少sheet: {site}")
                return
            WORKSHEET_SRC = WORKBOOK_SRC[site]

            # 定位FW01列
            fw01_col = None
            for COLUMN in range(1, (WORKSHEET_SRC.max_column or 1) + 1):
                header = str(WORKSHEET_SRC.cell(row=1, column=COLUMN).value or '')
                if re.search(r"fw\s*0*1(?:[-_]?frp)?", header, flags=re.IGNORECASE):
                    fw01_col = COLUMN
                    break
            if not fw01_col:
                self.add_result(Level.ERROR, f"sheet {site} 未找到FW01列")
                WORKBOOK_SRC.close()
                return

            # 从指定列提取候选域名：每个单元格内支持多域名（分隔符：逗号/分号/空白/换行），也支持在长文本中用正则提取形如 a.b 的FQDN，过滤类似 *.log、*.xlsx 等文件名，去重；规则补充：若检测到"object network <name>"且紧随下一非空行不含"fqdn"，则将 <name>（若形如域名）标记为强制失败（缺少fqdn定义）
            def collect_domains(
                COL_IDX: int
            ) -> Tuple[Set[str], Dict[str, str], Set[int], Set[int], Set[int]]:
                """从指定列收集域名

                Args:
                    COL_IDX: 列索引

                Returns:
                    元组：(域名集合, 强制原因字典, 红色单元格行号集合, 跳过行号集合, 其他行号集合)
                """
                domains: Set[str] = set()
                forced_reason: Dict[str, str] = {}
                forced_cell_red: Set[int] = set()
                skip_rows: Set[int] = set()
                other_rows: Set[int] = set()
                if not COL_IDX:
                    return domains, forced_reason, forced_cell_red, skip_rows, other_rows
                empty_streak = 0
                max_empty_streak = 5  # 连续空行阈值，提前停止
                # 使用按列迭代，values_only提升性能
                rows = list(WORKSHEET_SRC.iter_rows(
                    min_row=2, max_row=WORKSHEET_SRC.max_row,
                    min_col=COL_IDX, max_col=COL_IDX, values_only=True
                ))
                total_rows = len(rows)
                for ROW_INDEX, (VAL,) in enumerate(rows):
                    if VAL is None or str(VAL).strip() == '':
                        empty_streak += 1
                        if empty_streak >= max_empty_streak:
                            break
                        continue
                    empty_streak = 0
                    TEXT = str(VAL)
                    # 若包含 URL（http/https），直接跳过此行，不参与域名提取与标色判定
                    if re.search(r"\bhttps?://", TEXT, re.IGNORECASE):
                        continue
                    # 限制单元格过长内容，避免异常数据拖慢
                    if len(TEXT) > 1024:
                        TEXT = TEXT[:1024]
                    PARTS = re.split(r"[\s,;\n\r\t]+", TEXT.strip())
                    take_count = 0
                    for PART in PARTS:
                        if take_count >= 20:  # 每格最多提取20个候选
                            break
                        PART = PART.strip()
                        if not PART:
                            continue
                        if self._DOMAIN_FULL_RE.fullmatch(PART):
                            domains.add(PART)
                            take_count += 1
                            continue
                        for MATCH in self._DOMAIN_TOKEN_RE.finditer(PART):
                            START, END = MATCH.span()
                            # 忽略邮箱场景：@紧邻域名前
                            if START > 0 and PART[START-1] == '@':
                                continue
                            candidate = MATCH.group(0).strip('.').strip()
                            if self._DOMAIN_FULL_RE.fullmatch(candidate):
                                domains.add(candidate)
                                take_count += 1
                                if take_count >= 20:
                                    break
                    # 强制失败检测：object network <name>，下一非空行不含fqdn
                    MATCH_OBJECT = re.search(r"^\s*object\s+network\s+(.+)$", TEXT, re.IGNORECASE)
                    if MATCH_OBJECT:
                        NAME = MATCH_OBJECT.group(1).strip()
                        # 查找下一非空行
                        next_text = None
                        for NEXT_ROW_INDEX in range(ROW_INDEX+1, total_rows):
                            NEXT_VALUE = rows[NEXT_ROW_INDEX][0]
                            if NEXT_VALUE is None or str(NEXT_VALUE).strip() == '':
                                continue
                            next_text = str(NEXT_VALUE)
                            break
                        # 如果下一行定义了 host/range/subnet，则视为IP对象，跳过特殊标记
                        if (next_text is not None and
                                re.search(r"\b(host|range|subnet)\b", next_text, re.IGNORECASE)):
                            # 标记当前object行与下一host/range/subnet行为需跳过标色
                            skip_rows.add(ROW_INDEX + 2)
                            if ROW_INDEX + 3 <= total_rows:
                                skip_rows.add(ROW_INDEX + 3)
                        # 如果下一行是 NAT 配置（nat (...) 或以 nat 开头），也不需要标色
                        elif (next_text is not None and
                                re.match(r"^\s*nat\b", next_text, re.IGNORECASE)):
                            skip_rows.add(ROW_INDEX + 2)
                            if ROW_INDEX + 3 <= total_rows:
                                skip_rows.add(ROW_INDEX + 3)
                        elif (next_text is None or
                                not re.search(r"\bfqdn\b", next_text, re.IGNORECASE)):
                            # 缺少fqdn：域名名合法 -> 记录为缺少定义；非法名 -> 仅在无任何有效定义时不标域名、不标色
                            if self._DOMAIN_FULL_RE.fullmatch(NAME):
                                domains.add(NAME)
                                forced_reason[NAME] = (
                                    "missing fqdn definition"
                                )
                            else:
                                # 非法域名且无host/range/subnet，也需着色（其他底色+红字）
                                other_rows.add(ROW_INDEX + 2)
                        else:
                            # 存在fqdn，若域名与对象名不一致则标红（无论对象名是否是域名）
                            MATCH_FQDN = re.search(
                                r"\bfqdn\b\s+(?:v\d+\s+)?([^\s]+)",
                                next_text, re.IGNORECASE
                            )
                            if MATCH_FQDN:
                                FQDN = MATCH_FQDN.group(1).strip()
                                # 去除可能的端口拼接（如 example.com587 或 example.com:587）
                                FQDN = re.sub(r"[:]?\d+$", "", FQDN)
                                if self._DOMAIN_FULL_RE.fullmatch(FQDN):
                                    domains.add(FQDN)
                                    if FQDN.lower() != NAME.lower():
                                        forced_reason[FQDN] = "name/fqdn mismatch"
                                        forced_cell_red.add(ROW_INDEX + 2)  # 标记object行
                                        # 同时将fqdn所在行（ROW_INDEX+3）也标记为红底（若存在该行）
                                        if ROW_INDEX + 3 <= total_rows:  # noqa: E501
                                            forced_cell_red.add(ROW_INDEX + 3)
                                else:
                                    forced_cell_red.add(ROW_INDEX + 2)
                return domains, forced_reason, forced_cell_red, skip_rows, other_rows

            (fw01_domains, fw01_forced_reason, fw01_cell_red, fw01_skip,
                    fw01_other) = collect_domains(fw01_col)
            WORKBOOK_SRC.close()


            # 解析域名
            fw01_results = self._resolve_domains(fw01_domains, forced_reasons=fw01_forced_reason)


            # 以源Excel为模板复制到DOMAIN/ASADomainCheckTask，然后在复制品上标色
            source_excel_path = os.path.join(
                self.ACL_DIR,
                f"{self._TODAY}-关键设备配置备份输出EXCEL基础任务.xlsx"
            )
            out_excel_path = os.path.join(
                self.OUTPUT_DIR,
                f"{self._TODAY}-ASA域名提取及检测任务.xlsx"
            )
            # 始终删除同名目标文件，确保与源结构完全一致（避免沿用旧文件缺少站点sheet）
            try:
                if os.path.exists(out_excel_path):
                    os.remove(out_excel_path)
            except Exception:
                # 若删除失败，依旧强制重建覆盖
                pass
            self._ensure_output_excel_from_source(
                source_excel_path, out_excel_path, force=True
            )
            success_report = self._generate_excel_report(
                site, fw01_results, out_excel_path,
                fw01_cell_red, fw01_skip, fw01_other
            )


            # 仅在报告成功生成时输出汇总，否则跳过（避免先WARN后OK的矛盾输出）
            if success_report:
                total_count = len(fw01_domains)
                success_count = sum(1 for success, _ in fw01_results.values() if success)
                fail_count = total_count - success_count
                special_count = sum(1 for DOMAIN in fw01_domains if DOMAIN in self._DNS_SPECIAL)
                self.add_result(
                    Level.OK,
                    f"站点{site}处理完成：开始解析域名，"
                    f"特殊白名单(S){special_count}个，"
                    f"本地缓存解析成功{success_count}个，"
                    f"解析失败{fail_count}个，总计{total_count}个"
                )

            # 任务2：按站点串行触发脚本生成（若有手工清单且启用），实现"边站点边输出"
            # 任务2独立于任务1，仅依赖手动输入的域名清单，不依赖任务1的DNS检测结果
            if self.ENABLE_DOMAIN_SCRIPTS:
                try:
                    domains_for_scripts = self._load_failed_domains_source()
                    if domains_for_scripts:
                        self._generate_domain_operation_scripts(site)
                except Exception as error:
                    self.add_result(
                        Level.WARN,
                        f"站点 {site} 任务2（脚本生成）触发失败: {error}"
                    )


        except Exception as error:
            self.add_result(Level.ERROR, f"处理站点 {site} 失败: {error}")


    # 批量解析域名：优先缓存，并发执行，设置超时防阻塞
    def _resolve_domains(
            self, domains: Set[str],
            forced_reasons: Dict[str, str] = None
    ) -> Dict[str, Tuple[bool, str]]:
        results: Dict[str, Tuple[bool, str]] = {}
        successful_domains: Set[str] = set()
        failed_domains: Set[str] = set()
        forced_reasons = forced_reasons or {}

        # 先标记缓存命中的
        remaining: List[str] = []
        for DOMAIN in domains:
            if self._is_domain_cached(DOMAIN):
                results[DOMAIN] = (True, "缓存命中")
                successful_domains.add(DOMAIN)
            else:
                remaining.append(DOMAIN)

        if remaining:
            max_workers = min(64, max(4, len(remaining)))
            # DNS解析工作函数：用于并发解析域名
            def worker(DOMAIN_NAME: str) -> Tuple[str, Tuple[bool, str]]:
                """DNS解析工作函数

                Args:
                    DOMAIN_NAME: 域名

                Returns:
                    元组：(域名, (解析成功标志, IP地址或错误信息))
                """
                try:
                    ok, addr = self._resolve_domain(DOMAIN_NAME)
                    return DOMAIN_NAME, (ok, addr)
                except Exception as ERROR:
                    return DOMAIN_NAME, (False, str(ERROR))
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as EXECUTOR:
                future_map = {EXECUTOR.submit(worker, DOMAIN): DOMAIN for DOMAIN in remaining}
                for future in concurrent.futures.as_completed(future_map):
                    DOMAIN_NAME = future_map[future]
                    try:
                        RESOLVED_NAME, (ok, info) = future.result(timeout=self.DNS_TIMEOUT)
                    except Exception as ERROR:
                        ok, info = False, f"timeout/{type(ERROR).__name__}"
                        RESOLVED_NAME = DOMAIN_NAME
                    results[RESOLVED_NAME] = (ok, info)
                    if ok:
                        successful_domains.add(RESOLVED_NAME)
                    else:
                        failed_domains.add(RESOLVED_NAME)


        # 应用强制失败/原因（缺少fqdn、名称不一致等）
        for DOMAIN, reason in forced_reasons.items():
            results[DOMAIN] = (False, reason)
            if DOMAIN in successful_domains:
                successful_domains.discard(DOMAIN)
            failed_domains.add(DOMAIN)

        # 保存成功的域名到临时缓存
        if successful_domains:
            self._save_dns_cache(successful_domains)


        # 保存失败的域名到统一缓存(D)
        if failed_domains:
            self._save_failed_domains(failed_domains)


        return results

    # ----------------------- 任务2：生成操作/回退脚本 -----------------------
    # 任务2：加载手动输入的域名清单，从 CHECKRULE/ManualDomainRecovery.log 读取域名，用于生成回收脚本
    def _load_failed_domains_source(self) -> List[str]:
        override_path = self.RECOVERY_OVERRIDE
        # 1) 未启用：静默跳过
        if not self.USE_MANUAL_SOURCE_FIRST:
            return []
        # 2) 路径为空：ERROR
        if not override_path:
            self.add_result(Level.ERROR, "手工域名源路径为空，跳过脚本生成")
            return []
        # 3) 文件不存在：ERROR
        if not os.path.exists(override_path):
            self.add_result(Level.ERROR, f"未找到有效的手工域名源 {override_path}，跳过脚本生成")
            return []
        # 4) 文件存在（即使为空）正常处理（空文件将返回空列表，静默不WARN）

        domains: List[str] = []
        try:
            with open(override_path, 'r', encoding='utf-8') as FILE_HANDLE:
                for LINE in FILE_HANDLE:
                    DOMAIN = (LINE or '').strip()
                    if DOMAIN:
                        domains.append(DOMAIN)
        except Exception as ERROR:
            self.add_result(Level.WARN, f"读取手工域名源失败: {override_path} -> {ERROR}")
            return []
        return sorted(set(domains))

    # 任务2：生成操作/回退脚本的入口，统一由 _generate_manual_recovery_excel 负责 Excel 标色与脚本生成
    def _generate_domain_operation_scripts(self, site: Optional[str] = None) -> None:
        domains = self._load_failed_domains_source()
        if not domains:
            # 静默：未启用/路径为空/不存在/空文件等情况均不输出WARN
            return

        date_string = datetime.now().strftime('%Y%m%d')
        out_dir = self.SCRIPT_OUTPUT_ROOT
        os.makedirs(out_dir, exist_ok=True)
        manual_excel = os.path.join(out_dir, f"{date_string}-ManualDomainRecovery.xlsx")
        source_excel = os.path.join(self.ACL_DIR, f"{date_string}-关键设备配置备份输出EXCEL基础任务.xlsx")


        # 统一由 _generate_manual_recovery_excel 负责：标色 Excel + 基于标色生成脚本
        self._generate_manual_recovery_excel(source_excel, manual_excel, domains, site)

    # 任务2：处理手动输入的域名清单：以源Excel为模板在配置目录生成一个Excel
    # （CONFIGURATION/ASADomainCheckTask/{date}/ManualDomainRecovery.xlsx），
    # 将手工回收域名在指定站点FW01列原位标色，根据标色行（匹配 object/fqdn 和
    # object-group 绑定）生成站点级回退/操作脚本
    def _generate_manual_recovery_excel(
        self, source_excel: str, excel_path: str, domains: List[str],
        site: Optional[str]
    ) -> None:
        try:
            if not site:
                return
            if not os.path.exists(source_excel):
                self.add_result(Level.WARN, f"手工回收Excel源模板不存在: {source_excel}")
                return
            # 拷贝模板到目标
            if not os.path.exists(excel_path):
                os.makedirs(os.path.dirname(excel_path), exist_ok=True)
                shutil.copy2(source_excel, excel_path)
            # 构造结果映射：手工列表全部视为失败（需要回收），以触发红色+底色标注
            fail_map = {DOMAIN: (False, "manual recovery") for DOMAIN in set(domains)}
            # 遍历模板中的所有站点sheet并标色
            try:
                workbook = load_workbook(excel_path)
                # 仅处理当前站点sheet，复用生成函数（内部会自行保存）
                if site in workbook.sheetnames:
                    workbook.close()
                    self._generate_excel_report(site, fail_map, excel_path)
                else:
                    workbook.close()
                    self.add_result(Level.WARN, f"手工回收Excel缺少sheet: {site}")
                # _generate_excel_report内部已保存
            except Exception:
                # 如果直接复用内部保存流程导致句柄冲突，则忽略该关闭
                pass
            # 基于“标色后的 FW01 列”从上到下生成回退脚本；再由回退推导操作脚本
            try:
                # 需要读取样式，不能使用 read_only 和 data_only
                workbook_for_read_styles = load_workbook(
                    excel_path, data_only=False
                )
                if site not in workbook_for_read_styles.sheetnames:
                    workbook_for_read_styles.close()
                    self.add_result(Level.WARN, f"手工回收脚本生成跳过，sheet缺失: {site}")
                    return
                worksheet = workbook_for_read_styles[site]
                # 定位FW01列
                fw01_col = None
                for COLUMN in range(1, (worksheet.max_column or 1)+1):
                    HEADER = str(worksheet.cell(1, COLUMN).value or '')
                    if re.search(r"fw\s*0*1(?:[-_]?frp)?", HEADER, re.IGNORECASE):
                        fw01_col = COLUMN
                        break
                if not fw01_col:
                    workbook_for_read_styles.close()
                    self.add_result(Level.WARN, f"手工回收脚本生成跳过，未找到FW01列: {site}")
                    return
                # 脚本路径（站点级）
                date_string = datetime.now().strftime('%Y%m%d')
                out_dir = self.SCRIPT_OUTPUT_ROOT
                os.makedirs(out_dir, exist_ok=True)
                op_path = os.path.join(
                    out_dir, f"{date_string}-{site}-操作脚本.log"
                )
                rb_path = os.path.join(
                    out_dir, f"{date_string}-{site}-回退脚本.log"
                )
                # 判定单元格是否标色：红字或填充为特定底色即视为标色
                def is_cell_colored(cell) -> bool:
                    """检查单元格是否已着色

                    Args:
                        cell: 单元格对象

                    Returns:
                        如果单元格已着色则返回True，否则返回False
                    """
                    try:
                        FONT_COLOR = (
                            cell.font and cell.font.color and
                            str(cell.font.color.rgb or '')
                        ).upper()
                        fill = (
                            cell.fill and cell.fill.start_color and
                            str(cell.fill.start_color.rgb or '')
                        ).upper()
                        return ((FONT_COLOR == 'FFFF0000') or
                                (fill in {'FFFFF2CC', 'FFFFC7CE', 'FFEEEEEE', 'FF0000FF'}))
                    except Exception:
                        return False


                # 采用追加方式，避免覆盖其他站点或前一次结果
                with (
                    open(op_path, 'a', encoding='utf-8') as operationLogFile,
                    open(rb_path, 'a', encoding='utf-8') as rollbackLogFile
                ):
                    # 1. 回退脚本：从上到下遍历 FW01 列，提取所有标色单元格的内容
                    current_group = None
                    last_group_written = None
                    empty_streak = 0
                    max_empty = 5
                    dom_to_groups: dict[str, set[str]] = {}  # 记录域名 -> 业务组集合（用于操作脚本）
                    group_to_domains_order: dict[str, list[str]] = {}  # 记录每个组中域名的出现顺序（按回退脚本顺序）
                    all_domains: set[str] = set()  # 记录所有需要删除的域名（包括单独的 object network）


                    for ROW in range(2, (worksheet.max_row or 1)+1):
                        cell = worksheet.cell(ROW, fw01_col)
                        TEXT = str(cell.value or '').strip()
                        if not TEXT:
                            empty_streak += 1
                            if empty_streak >= max_empty:
                                break
                            continue
                        empty_streak = 0


                        is_colored = is_cell_colored(cell)
                        if not is_colored:
                            # 未标色行：仅更新 current_group（用于后续绑定行判断）
                            MATCH_GROUP = re.match(
                                r"^object-group\s+network\s+(.+)$",
                                TEXT, re.IGNORECASE
                            )
                            if MATCH_GROUP:
                                current_group = MATCH_GROUP.group(1).strip()
                            continue


                        # 标色行处理
                        # object-group network：直接输出
                        MATCH_GROUP = re.match(
                            r"^object-group\s+network\s+(.+)$",
                            TEXT, re.IGNORECASE
                        )
                        if MATCH_GROUP:
                            current_group = MATCH_GROUP.group(1).strip()
                            if last_group_written != current_group:
                                if last_group_written is not None:
                                    rollbackLogFile.write("\n")
                                rollbackLogFile.write(f"object-group network {current_group}\n")
                                last_group_written = current_group
                            continue


                        # object network：输出 object + fqdn
                        MATCH_OBJECT = re.match(
                            r"^object\s+network\s+(.+)$", TEXT, re.IGNORECASE
                        )
                        if MATCH_OBJECT:
                            NAME = MATCH_OBJECT.group(1).strip()
                            # 找下一行的 fqdn
                            FQDN = None
                            # 最多向下查10行
                            for ROW_INDEX in range(
                                    ROW+1, min(ROW+10, worksheet.max_row+1)
                            ):
                                TEXT2 = str(
                                    worksheet.cell(ROW_INDEX, fw01_col).value or ''
                                ).strip()
                                if TEXT2:
                                    MATCH_FQDN = re.match(
                                        r"^fqdn\s+(?:v\d+\s+)?(.+)$",
                                        TEXT2, re.IGNORECASE
                                    )
                                    if MATCH_FQDN:
                                        FQDN = MATCH_FQDN.group(1).strip()
                                    break
                            DOMAIN = FQDN if FQDN else NAME
                            all_domains.add(DOMAIN)  # 记录需要删除的域名
                            rollbackLogFile.write(
                                f"object network {DOMAIN}\n"
                            )
                            rollbackLogFile.write(f" fqdn {DOMAIN}\n\n")
                            continue


                        # network-object object：输出绑定，并记录映射
                        MATCH_BIND = re.match(
                            r"^network-object\s+object\s+(.+)$",
                            TEXT, re.IGNORECASE
                        )
                        if MATCH_BIND and current_group:
                            DOMAIN = MATCH_BIND.group(1).strip()
                            # 记录映射（用于操作脚本）
                            if DOMAIN not in dom_to_groups:
                                dom_to_groups[DOMAIN] = set()
                            dom_to_groups[DOMAIN].add(current_group)
                            all_domains.add(DOMAIN)  # 记录需要删除的域名
                            # 记录域名在组中的出现顺序（按回退脚本顺序）
                            if current_group not in group_to_domains_order:
                                group_to_domains_order[current_group] = []
                            if DOMAIN not in group_to_domains_order[current_group]:
                                group_to_domains_order[current_group].append(DOMAIN)
                            # 输出绑定
                            if last_group_written != current_group:
                                if last_group_written is not None:
                                    rollbackLogFile.write("\n")
                                rollbackLogFile.write(f"object-group network {current_group}\n")
                                last_group_written = current_group
                            rollbackLogFile.write(f" network-object object {DOMAIN}\n")


                    # 2. 操作脚本：根据回退脚本反向生成（解绑 → 删除）
                    # 使用记录的顺序（按回退脚本顺序，保持一致性）
                    # 先输出所有组的解绑命令（按组名排序，但组内域名顺序与回退脚本一致）
                    for GROUP in sorted(group_to_domains_order.keys()):
                        # 使用回退脚本中记录的顺序，确保操作脚本和回退脚本顺序一致
                        domains_in_group = group_to_domains_order[GROUP]
                        operationLogFile.write(f"object-group network {GROUP}\n")
                        for DOMAIN in domains_in_group:
                            operationLogFile.write(f" no network-object object {DOMAIN}\n")
                        operationLogFile.write("\n")


                    # 最后输出所有 object network 的删除命令（每个域名只删除一次）
                    for DOMAIN in sorted(all_domains):
                        operationLogFile.write(f"no object network {DOMAIN}\n")

                    self.add_result(Level.OK, f"站点 {site} 基于标色生成回退/操作脚本已完成")
                workbook_for_read_styles.close()

            except Exception as ERROR2:
                self.add_result(Level.WARN, f"基于Excel生成手工回收脚本失败: {ERROR2}")
            self.add_result(Level.OK, f"生成手工回收域名Excel(模板标色): {excel_path}")
        except Exception as ERROR:
            self.add_result(Level.ERROR, f"生成手工回收域名Excel失败: {ERROR}")


    # 生成Excel报告（在源Excel的FW01列原位标色）
    # 任务1和任务2共用此方法进行Excel标色
    def _generate_excel_report(
            self, site: str, fw01_results: Dict[str, Tuple[bool, str]],
            output_path: str, fw01_cell_red: Set[int] = None,
                              fw01_skip_rows: Set[int] = None,
            fw01_other_rows: Set[int] = None
    ) -> bool:
        """任务1和任务2共用的Excel标色方法
        - 任务1：基于DNS检测结果标色（成功绿色，失败红色）
        - 任务2：基于手动输入的域名清单标色（全部标记为红色，用于回收）
        """
        if not os.path.exists(output_path):
            self.add_result(Level.ERROR, f"未找到Excel文件: {output_path}")
            return False

        try:
            workbook = load_workbook(output_path, data_only=False, keep_links=True)
        except Exception as error:
            # 目标Excel可能损坏，尝试从源模板重建后再试一次
            self.add_result(Level.WARN, f"输出Excel无法打开，尝试重建: {error}")
            source_path = os.path.join(self.ACL_DIR, f"{self._TODAY}-关键设备配置备份输出EXCEL基础任务.xlsx")
            self._ensure_output_excel_from_source(source_path, output_path)
            workbook = load_workbook(output_path, data_only=False, keep_links=True)
        sheet_title = site
        if sheet_title not in workbook.sheetnames:
            # 若目标缺少该sheet，强制用源Excel重建一次输出，再次尝试
            workbook.close()
            source_path = os.path.join(self.ACL_DIR, f"{self._TODAY}-关键设备配置备份输出EXCEL基础任务.xlsx")
            self._ensure_output_excel_from_source(source_path, output_path, force=True)
            try:
                workbook = load_workbook(output_path, data_only=False, keep_links=True)
            except Exception as error:
                self.add_result(Level.ERROR, f"重建后仍无法打开输出Excel: {error}")
                return False
            if sheet_title not in workbook.sheetnames:
                self.add_result(Level.WARN, f"源Excel缺少sheet: {sheet_title}")
                workbook.close()
                return False
        worksheet = workbook[sheet_title]

        # 定义样式（使用8位ARGB，避免Excel修复警告）
        green_font = Font(color="FF00AA00")  # 深绿
        red_font = Font(color="FFFF0000")    # 红
        recycle_fill = PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
        strict_red_fill = PatternFill(
            fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE"
        )
        # 其他：非法名且无定义
        other_fill = PatternFill(
            fill_type="solid", start_color="FFEEEEEE", end_color="FFEEEEEE"
        )

        # 定位FW01列（根据表头自动识别）
        fw01_col = None
        for COLUMN in range(1, worksheet.max_column + 1):
            header = str(worksheet.cell(row=1, column=COLUMN).value or '')
            if re.search(r"fw\s*0*1(?:[-_]?frp)?", header, flags=re.IGNORECASE):
                fw01_col = COLUMN
                break
        if not fw01_col:
            self.add_result(Level.WARN, f"sheet {sheet_title} 未找到FW01列，跳过标色")
            workbook.close()
            return

        # 构造命中映射与原因（仅FW01）
        fw01_hit = {DOMAIN: STATUS for DOMAIN, (STATUS, _) in fw01_results.items()}
        fw01_reason = {DOMAIN: MSG for DOMAIN, (_, MSG) in fw01_results.items()}

        # 遍历原有单元格进行标色（不改动单元格文本）
        # 纯蓝，用于缺少fqdn定义
        missing_fill = PatternFill(
            fill_type="solid", start_color="FF0000FF", end_color="FF0000FF"
        )
        fw01_cell_red = fw01_cell_red or set()
        fw01_skip_rows = fw01_skip_rows or set()
        fw01_other_rows = fw01_other_rows or set()
        # 第一遍：收集需要标色的组名行（如果该组下的绑定都是回收域名）
        group_rows_to_color = set()
        current_group_row = None
        for ROW in range(2, worksheet.max_row + 1):
            TEXT = str(worksheet.cell(row=ROW, column=fw01_col).value or '').strip()
            if TEXT:
                if ROW in fw01_skip_rows:
                    continue
                if re.search(r"\b(host|range|subnet)\b", TEXT, re.IGNORECASE):
                    continue
                # 检测 object-group network 行，记录行号
                MATCH_GROUP = re.match(r"^object-group\s+network\s+(.+)$", TEXT, re.IGNORECASE)
                if MATCH_GROUP:
                    current_group_row = ROW
                    continue
                # 检测 network-object object 行：如果域名在失败列表中，标记其所属的组名行也需要标色
                MATCH_BIND = re.match(r"^network-object\s+object\s+(.+)$", TEXT, re.IGNORECASE)
                if MATCH_BIND and current_group_row is not None:
                    DOMAIN = MATCH_BIND.group(1).strip()
                    if DOMAIN in fw01_hit and fw01_hit[DOMAIN] is False:
                        group_rows_to_color.add(current_group_row)


        # 第二遍：进行标色
        current_group_row = None
        for ROW in range(2, worksheet.max_row + 1):
            # FW01 列：支持"整格即域名"或"文本内包含域名子串"的两种匹配
            TEXT = str(worksheet.cell(row=ROW, column=fw01_col).value or '').strip()
            if TEXT:
                if ROW in fw01_skip_rows:
                    continue
                # 若该格为URL（http/https），跳过标色
                if re.search(r"\bhttps?://", TEXT, re.IGNORECASE):
                    continue
                # 若该格为IP对象定义（host/range/subnet），跳过一切标色
                if re.search(r"\b(host|range|subnet)\b", TEXT, re.IGNORECASE):
                    continue
                # 若该行为 NAT 配置（nat 开头），也跳过标色
                if re.match(r"^\s*nat\b", TEXT, re.IGNORECASE):
                    continue
                # 检测 object-group network 行，记录行号（用于后续 network-object 判断）
                MATCH_GROUP = re.match(r"^object-group\s+network\s+(.+)$", TEXT, re.IGNORECASE)
                if MATCH_GROUP:
                    current_group_row = ROW
                    # 如果是需要标色的组名行，立即标色
                    if ROW in group_rows_to_color:
                        worksheet.cell(row=ROW, column=fw01_col).font = red_font
                        worksheet.cell(row=ROW, column=fw01_col).fill = recycle_fill
                    continue
                verdict = None  # True=成功, False=失败, None=未知
                forced_missing = False
                # 完全等于域名
                if TEXT in fw01_hit:
                    verdict = fw01_hit[TEXT]
                    forced_missing = (fw01_reason.get(TEXT) == "missing fqdn definition")
                else:
                    # 从文本中提取域名候选，若有任一失败则按失败标色，否则有任一成功则按成功
                    tokens = set()
                    for MATCH in self._DOMAIN_TOKEN_RE.finditer(TEXT):
                        START, END = MATCH.span()
                        if START > 0 and TEXT[START-1] == '@':
                            continue
                        CANDIDATE = MATCH.group(0).strip('.').strip()
                        if self._DOMAIN_FULL_RE.fullmatch(CANDIDATE):
                            tokens.add(CANDIDATE)
                    if tokens:
                        if any(
                            fw01_hit.get(TOKEN) is False
                            for TOKEN in tokens if TOKEN in fw01_hit
                        ):
                            verdict = False
                            if any(
                                (fw01_reason.get(TOKEN) == "missing fqdn definition")
                                for TOKEN in tokens if TOKEN in fw01_reason
                            ):
                                forced_missing = True
                        elif any(
                            fw01_hit.get(TOKEN) is True
                            for TOKEN in tokens if TOKEN in fw01_hit
                        ):
                            verdict = True
                # 标色：普通行
                # 特例：全局配置项，如 "domain-name cisco.com" 不参与标色
                if re.match(r"^\s*domain-name\b", TEXT, re.IGNORECASE):
                    continue
                if ROW in fw01_other_rows:
                    worksheet.cell(row=ROW, column=fw01_col).font = red_font
                    worksheet.cell(row=ROW, column=fw01_col).fill = other_fill
                elif ROW in fw01_cell_red:
                    worksheet.cell(row=ROW, column=fw01_col).font = red_font
                    worksheet.cell(row=ROW, column=fw01_col).fill = strict_red_fill
                elif verdict is True:
                    worksheet.cell(row=ROW, column=fw01_col).font = green_font
                elif verdict is False:
                    if forced_missing:
                        worksheet.cell(row=ROW, column=fw01_col).font = red_font
                        worksheet.cell(row=ROW, column=fw01_col).fill = missing_fill
                    else:
                        worksheet.cell(row=ROW, column=fw01_col).font = red_font
                        worksheet.cell(row=ROW, column=fw01_col).fill = recycle_fill

            # 仅检查FW01

        # 原子保存，避免损坏
        tmp_path = output_path + ".tmp"
        try:
            workbook.save(tmp_path)
            workbook.close()
            os.replace(tmp_path, output_path)
        finally:
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
        # 静默：不输出Excel生成OK日志，由上层汇总输出
        return True

    # 如果目标Excel不存在，则以源Excel为模板复制数据结构到目标Excel：仅复制单元格值及sheet结构，不拷贝复杂样式，避免第三方复制限制
    def _ensure_output_excel_from_source(
            self, source_path: str, output_path: str, force: bool = False
    ) -> None:
        try:
            # 如目标不存在或损坏（0字节/不可读取），则从源复制
            need_create = True
            if not force and os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                try:
                    _ = load_workbook(output_path)
                    need_create = False
                except Exception:
                    need_create = True

            if not os.path.exists(source_path):
                self.add_result(Level.ERROR, f"未找到源Excel: {source_path}")
                return

            if need_create or force:
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                shutil.copy2(source_path, output_path)
                # 校验复制结果可打开
                _ = load_workbook(output_path)
                # 静默：不输出初始化OK日志
            else:
                # 无需重建
                return
        except Exception as error:
            self.add_result(Level.ERROR, f"复制源Excel到目标失败: {error}")
