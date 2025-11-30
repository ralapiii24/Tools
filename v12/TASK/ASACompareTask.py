# ASA防火墙主备对比检查任务
#
# 技术栈:openpyxl、正则表达式、difflib
# 目标:对比主备防火墙配置差异，生成带颜色编码的Excel报告
#
# 处理逻辑:
# - 扫描LOG目录查找设备文件，按站点分组fw01-frp和fw02-frp配置文件
# - 提取interface Port-channel到failover的配置段（包含相关的object配置）
# - 使用SequenceMatcher对比算法识别差异（内容不同的配置行、仅在FW01-FRP中存在的配置、仅在FW02-FRP中存在的配置）
# - 为缩进行显示父级配置
# - 生成带颜色编码的Excel报告（差异行用红色标记）
#
# 输出文件结构:
# - 生成以日期命名的Excel文件，每个站点对应一个Sheet
# - 每个Sheet包含：站点信息（站点名、对比时间、对比范围、文件路径）、配置差异详情、差异统计、配置对比结果
#
# 输入文件:LOG/OxidizedTask/OxidizedTaskBackup/（V10新结构：从LOG/日期/OxidizedTaskBackup迁移）
# 输出文件:LOG/ASACompareTask/{日期}-ASA防火墙主备对比检查.xlsx（V10新结构：从ACL/ASACompareTask迁移）

# 导入标准库
import os
from datetime import datetime

# 导入第三方库
from openpyxl import Workbook
from tqdm import tqdm

# 导入本地应用
from .TaskBase import (
    BaseTask, Level, extract_site_from_device, BAR_FORMAT, SHOW_PROGRESS,
    get_today_str, format_datetime, ensure_output_dir, build_log_path, build_output_path
)

# ASA防火墙主备对比检查任务类：对比ASA防火墙主备设备配置差异并生成Excel报告
class ASACompareTask(BaseTask):
    """ASA防火墙主备对比检查任务


    对比主备防火墙配置差异，生成带颜色编码的Excel报告
    """
    # 初始化ASA防火墙主备对比任务：设置任务名称和数据结构
    def __init__(self):
        super().__init__("ASA防火墙主备对比检查")
        self.LOG_DIR = "LOG"
        # V10新结构：直接输出到 LOG/ASACompareTask/
        self.OUTPUT_DIR = os.path.join("LOG", "ASACompareTask")
        self._TODAY = None
        self._SITES_DATA = {}  # {site: {fw01: content, fw02: content}}
        self._WB = None
        self._ORIGINAL_CONFIGS = {}  # 保存原始配置内容 {site: {fw01: lines, fw02: lines}}

    # 扫描LOG目录获取站点列表：按站点分组fw01-frp和fw02-frp配置文件
    def items(self):
        """扫描LOG目录获取站点列表


        按站点分组fw01-frp和fw02-frp配置文件


        Returns:
            list: 站点列表
        """
        # 返回站点列表作为items，实现1/N进度显示
        self._TODAY = get_today_str()
        # V10新结构：从 LOG/OxidizedTask/OxidizedTaskBackup/ 读取
        LOG_DIR_PATH = os.path.join(self.LOG_DIR, "OxidizedTask", "OxidizedTaskBackup")
        if not os.path.isdir(LOG_DIR_PATH):
            self.add_result(Level.ERROR, f"未找到当日日志目录: {LOG_DIR_PATH}")
            return []

        # 创建输出目录（如果目录已存在则不报错）
        ensure_output_dir(self.OUTPUT_DIR)

        # 扫描LOG目录，按站点分组fw01-frp和fw02-frp文件
        self._SITES_DATA = {}

        for filename in os.listdir(LOG_DIR_PATH):
            if not filename.lower().endswith('.log'):
                continue

            # 提取日期和设备名
            if not filename.startswith(self._TODAY + '-'):
                continue

            device_name = filename[len(self._TODAY) + 1:-4]  # 去掉日期前缀和.log后缀

            # 检查是否包含fw01-frp或fw02-frp
            device_lower = device_name.lower()
            if 'fw01-frp' in device_lower:
                site = self._extract_site_from_device(device_name)
                if site:
                    if site not in self._SITES_DATA:
                        self._SITES_DATA[site] = {}
                    self._SITES_DATA[site]['fw01'] = os.path.join(LOG_DIR_PATH, filename)
            elif 'fw02-frp' in device_lower:
                site = self._extract_site_from_device(device_name)
                if site:
                    if site not in self._SITES_DATA:
                        self._SITES_DATA[site] = {}
                    self._SITES_DATA[site]['fw02'] = os.path.join(LOG_DIR_PATH, filename)

        # 过滤出同时有fw01和fw02的站点
        valid_sites = []
        for site, data in self._SITES_DATA.items():
            if 'fw01' in data and 'fw02' in data:
                valid_sites.append(site)
            else:
                missing = []
                if 'fw01' not in data:
                    missing.append('fw01-frp')
                if 'fw02' not in data:
                    missing.append('fw02-frp')
                self.add_result(Level.WARN, f"站点 {site} 缺少设备: {', '.join(missing)}")

        if not valid_sites:
            self.add_result(Level.WARN, "未找到完整的ASA防火墙主备设备对")
            return []

        # 初始化Excel工作簿
        self._WB = Workbook()
        try:
            self._WB.remove(self._WB.active)
        except Exception:
            pass

        # 找到有效站点，无需输出日志
        return valid_sites

    # 从设备名中提取站点名：解析设备名称获取站点标识，如HX03-FW01-FRP2140-JPIDC -> HX03
    # 站点提取函数已迁移到TaskBase，直接使用extract_site_from_device
    # 保留此方法作为兼容性包装
    @staticmethod
    def _extract_site_from_device(device_name: str) -> str:
        """从设备名中提取站点名


        Args:
            device_name: 设备名称


        Returns:
            str: 站点名
        """
        return extract_site_from_device(device_name)

    # 提取配置段：从ASA配置文件中提取interface Port-channel到failover的关键配置段，包含相关的object配置
    @staticmethod
    def _extract_config_section(file_path: str) -> list:
        """从配置文件中提取interface Port-channel到failover的配置段


        Args:
            file_path: 配置文件路径


        Returns:
            list: 配置行列表
        """
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as fileHandle:
                lines = fileHandle.read().splitlines()
        except Exception:
            return []

        start_marker = "interface Port-channel"
        end_marker = "failover"

        start_index = -1
        end_index = -1

        # 查找开始标记
        for lineIndex, line in enumerate(lines):
            if start_marker in line:
                start_index = lineIndex
                break

        if start_index == -1:
            return []

        # 查找结束标记
        for lineIndex in range(start_index + 1, len(lines)):
            if end_marker in lines[lineIndex]:
                end_index = lineIndex
                break

        if end_index == -1:
            # 如果没找到failover，取到文件末尾
            end_index = len(lines)

        # 提取配置段
        config_section = lines[start_index:end_index]

        # 查找并添加相关的object配置
        object_configs = []
        for lineIndex, line in enumerate(lines):
            # 查找object network配置
            if line.strip().startswith('object network') and lineIndex < start_index:
                # 找到object network配置，提取完整的object块
                object_lines = [line]
                next_line_index = lineIndex + 1
                # 提取缩进的子配置行
                while (next_line_index < len(lines) and
                        (lines[next_line_index].startswith(' ') or
                            lines[next_line_index].startswith('\t'))):
                    object_lines.append(lines[next_line_index])
                    next_line_index += 1
                # 添加空行分隔
                object_lines.append('')
                object_configs.extend(object_lines)

        # 将object配置添加到配置段前面
        if object_configs:
            config_section = object_configs + config_section

        return config_section

    # 处理正常对比情况
    def _handle_normal_comparison(
            self, worksheet, site: str, fw01_content: list,
            fw02_content: list
    ):
        """处理正常对比情况


        Args:
            worksheet: Excel工作表对象
            site: 站点名称
            fw01_content: FW01配置内容
            fw02_content: FW02配置内容
        """
        if not fw01_content and not fw02_content:
            msg = "配置对比结果: 两台设备都无interface Port-channel到failover配置段"
            worksheet.append([msg, msg])
            self.add_result(Level.WARN, f"站点 {site} 两台设备都无interface Port-channel到failover配置段")
            return
        elif not fw01_content or not fw02_content:
            missing_device = "FW01-FRP" if not fw01_content else "FW02-FRP"
            msg = f"配置对比结果: {missing_device} 无interface Port-channel到failover配置段"
            worksheet.append([msg, msg])
            self.add_result(
                Level.WARN,
                f"站点 {site} {missing_device} "
                f"无interface Port-channel到failover配置段"
            )
            return

        # 使用简化的对比方法
        comparison_result = self._simple_comparison(fw01_content, fw02_content)

        if not comparison_result['has_differences']:
            worksheet.append(["配置对比结果: 两台设备配置完全一致", "配置对比结果: 两台设备配置完全一致"])
            self.add_result(Level.OK, f"站点{site}防火墙主备配置对比完成，ASA防火墙主备配置一致")
        else:
            fw01_only_count = len(comparison_result['fw01_only'])
            fw02_only_count = len(comparison_result['fw02_only'])
            self.add_result(
                Level.WARN,
                f"站点{site}防火墙主备配置对比完成，存在差异，"
                f"请手动对比检查(FW01独有:{fw01_only_count}, FW02独有:{fw02_only_count})"
            )
            self._fill_simple_comparison_content(worksheet, comparison_result, site)
            self._add_simple_statistics_and_result(worksheet, site, comparison_result)

    # 简单对比方法：使用SequenceMatcher进行配置差异对比，只显示真正的差异
    @staticmethod
    def _simple_comparison(fw01_content: list, fw02_content: list) -> dict:
        """简单对比方法


        优化的对比方法，只显示真正的差异


        Args:
            fw01_content: FW01配置内容
            fw02_content: FW02配置内容


        Returns:
            dict: 对比结果字典
        """
        from difflib import SequenceMatcher

        # 使用SequenceMatcher进行对比
        matcher = SequenceMatcher(None, fw01_content, fw02_content)

        # 收集真正的差异行（不包括相同的配置）
        fw01_only_lines = []
        fw02_only_lines = []
        different_lines = []

        for TAG, i1, i2, j1, j2 in matcher.get_opcodes():
            if TAG == 'delete':
                # 只在fw01中存在的行
                fw01_only_lines.extend(fw01_content[i1:i2])
            elif TAG == 'insert':
                # 只在fw02中存在的行
                fw02_only_lines.extend(fw02_content[j1:j2])
            elif TAG == 'replace':
                # 内容不同的行
                for content_index in range(i1, i2):
                    if (content_index < len(fw01_content) and
                            (j1 + content_index - i1) < len(fw02_content)):
                        different_lines.append({
                            'fw01': fw01_content[content_index],
                            'fw02': fw02_content[j1 + content_index - i1]
                        })

        # 过滤掉重新排序的差异
        fw01_set = set(fw01_only_lines)
        fw02_set = set(fw02_only_lines)

        # 找出真正只在fw01中存在的行（在fw02中不存在的）
        truly_fw01_only = [line for line in fw01_only_lines if line not in fw02_set]

        # 找出真正只在fw02中存在的行（在fw01中不存在的）
        truly_fw02_only = [line for line in fw02_only_lines if line not in fw01_set]

        # 进一步过滤：只保留有意义的差异行
        meaningful_fw01_only = []
        meaningful_fw02_only = []

        # 过滤FW01独有行：只保留非空行和有意义的内容
        for line in truly_fw01_only:
            line_stripped = line.strip()
            if line_stripped and not line_stripped.startswith('!'):  # 排除注释行
                meaningful_fw01_only.append(line)

        # 过滤FW02独有行：只保留非空行和有意义的内容
        for line in truly_fw02_only:
            line_stripped = line.strip()
            if line_stripped and not line_stripped.startswith('!'):  # 排除注释行
                meaningful_fw02_only.append(line)

        # 过滤内容不同的行：只保留真正不同的内容
        meaningful_different_lines = []
        for diff in different_lines:
            fw01_line = diff['fw01'].strip()
            fw02_line = diff['fw02'].strip()
            # 只保留非空且不同的行
            if fw01_line and fw02_line and fw01_line != fw02_line:
                if not fw01_line.startswith('!') and not fw02_line.startswith('!'):  # 排除注释行
                    meaningful_different_lines.append(diff)

        # 检查是否有真正的差异
        has_differences = (
            len(meaningful_fw01_only) > 0 or
            len(meaningful_fw02_only) > 0 or
            len(meaningful_different_lines) > 0
        )

        return {
            'has_differences': has_differences,
            'fw01_only': meaningful_fw01_only,
            'fw02_only': meaningful_fw02_only,
            'different_lines': meaningful_different_lines,
            'total_fw01': len(fw01_content),
            'total_fw02': len(fw02_content)
        }

    # 填充简化的对比内容到Excel
    def _fill_simple_comparison_content(self, worksheet, comparison_result: dict, site: str):
        from openpyxl.styles import Font

        fw01_only = comparison_result['fw01_only']
        fw02_only = comparison_result['fw02_only']
        different_lines = comparison_result['different_lines']

        # 添加差异标题
        worksheet.append(["=== 配置差异详情 ===", "=== 配置差异详情 ==="])
        worksheet.append([])

        # 处理内容不同的行
        if different_lines:
            worksheet.append(["=== 内容不同的配置行 ===", "=== 内容不同的配置行 ==="])
            for diff in different_lines:
                self._add_line_with_context(worksheet, diff['fw01'], diff['fw02'], site)
            worksheet.append([])

        # 处理只在FW01中存在的行
        if fw01_only:
            worksheet.append(["=== 仅在FW01-FRP中存在的配置 ===", "=== 仅在FW01-FRP中存在的配置 ==="])
            for line in fw01_only:
                self._add_line_with_context(worksheet, line, "", site)
            worksheet.append([])

        # 处理只在FW02中存在的行
        if fw02_only:
            worksheet.append(["=== 仅在FW02-FRP中存在的配置 ===", "=== 仅在FW02-FRP中存在的配置 ==="])
            for line in fw02_only:
                self._add_line_with_context(worksheet, "", line, site)
            worksheet.append([])

    # 添加带上下文的行，处理缩进行
    def _add_line_with_context(self, worksheet, fw01_line: str, fw02_line: str, site: str):
        from openpyxl.styles import Font

        # 检查是否有缩进（以空格开头）
        fw01_has_indent = fw01_line and fw01_line.startswith(' ')
        fw02_has_indent = fw02_line and fw02_line.startswith(' ')

        # 如果FW01有缩进，需要添加上下文
        if fw01_has_indent and fw01_line:
            context_lines = self._find_parent_context(fw01_line, 'fw01', site)
            # 只添加最近的父级配置，避免显示太多内容
            if context_lines:
                # 只显示最后一个父级配置（最近的）
                worksheet.append([context_lines[-1], ""])
                row_num = worksheet.max_row
                # 上下文行保持默认颜色（黑色）

        # 如果FW02有缩进，需要添加上下文
        if fw02_has_indent and fw02_line:
            context_lines = self._find_parent_context(fw02_line, 'fw02', site)
            # 只添加最近的父级配置，避免显示太多内容
            if context_lines:
                # 只显示最后一个父级配置（最近的）
                worksheet.append(["", context_lines[-1]])
                row_num = worksheet.max_row
                # 上下文行保持默认颜色（黑色）

        # 添加差异行
        worksheet.append([fw01_line, fw02_line])
        row_num = worksheet.max_row

        # 设置颜色：差异行用红色
        if fw01_line:
            worksheet.cell(row=row_num, column=1).font = Font(color="FF0000")
        if fw02_line:
            worksheet.cell(row=row_num, column=2).font = Font(color="FF0000")

    # 为缩进行找到真正的父级配置行
    def _find_parent_context(self, indented_line: str, device: str, site: str) -> list:
        # 检查站点是否存在原始配置
        if site not in self._ORIGINAL_CONFIGS:
            return []

        # 获取对应设备的原始配置
        original_lines = self._ORIGINAL_CONFIGS[site][device]

        # 查找缩进行在原始配置中的位置
        line_index = -1
        for lineIndex, line in enumerate(original_lines):
            if line.strip() == indented_line.strip():
                line_index = lineIndex
                break

        if line_index == -1:
            return []

        # 向上查找父级配置
        parent_lines = []
        for context_index in range(line_index - 1, -1, -1):
            line = original_lines[context_index].strip()
            if not line:  # 跳过空行
                continue

            # 如果找到不以空格开头的行，说明是父级配置
            if not line.startswith(' ') and not line.startswith('\t'):
                parent_lines.insert(0, line)
                # 继续向上查找，可能有多层父级配置
                continue
            else:
                # 如果遇到缩进行，说明已经超出了当前配置块
                break

        return parent_lines

    # 添加简化的统计信息和结果
    def _add_simple_statistics_and_result(self, worksheet, site: str, comparison_result: dict):
        """添加简化的统计信息和结果


        Args:
            worksheet: Excel工作表对象
            site: 站点名称
            comparison_result: 对比结果字典
        """
        fw01_only = comparison_result['fw01_only']
        fw02_only = comparison_result['fw02_only']
        different_lines = comparison_result['different_lines']

        # 统计差异数量
        different_count = len(different_lines)
        fw01_only_count = len(fw01_only)
        fw02_only_count = len(fw02_only)

        worksheet.append([])
        worksheet.append([f"差异统计: 内容不同 {different_count} 行", f"差异统计: 内容不同 {different_count} 行"])
        worksheet.append([f"FW01独有: {fw01_only_count} 行", f"FW02独有: {fw02_only_count} 行"])

        # 统计信息已在上面输出，这里不再重复

    # 处理单个站点的ASA防火墙主备对比：读取配置并进行差异分析
    def run_single(self, site: str):
        """处理单个站点的ASA防火墙主备对比


        读取配置并进行差异分析


        Args:
            site: 站点名称
        """
        if site not in self._SITES_DATA:
            self.add_result(Level.ERROR, f"站点 {site} 数据不存在")
            return

        data = self._SITES_DATA[site]
        fw01_path = data['fw01']
        fw02_path = data['fw02']

        try:
            # 读取原始配置内容
            with open(fw01_path, 'r', encoding='utf-8', errors='ignore') as fileHandle:
                fw01_original = fileHandle.read().splitlines()
            with open(fw02_path, 'r', encoding='utf-8', errors='ignore') as fileHandle:
                fw02_original = fileHandle.read().splitlines()

            # 保存原始配置
            self._ORIGINAL_CONFIGS[site] = {
                'fw01': fw01_original,
                'fw02': fw02_original
            }

            # 读取fw01配置并提取指定部分
            fw01_content = self._extract_config_section(fw01_path)
            if not fw01_content:
                self.add_result(
                    Level.WARN,
                    f"站点 {site} FW01-FRP "
                    f"未找到interface Port-channel到failover的配置段"
                )

            # 读取fw02配置并提取指定部分
            fw02_content = self._extract_config_section(fw02_path)
            if not fw02_content:
                self.add_result(
                    Level.WARN,
                    f"站点 {site} FW02-FRP "
                    f"未找到interface Port-channel到failover的配置段"
                )

        except Exception as error:
            self.add_result(Level.ERROR, f"站点 {site} 读取配置文件失败: {error}")
            return

        # 检查是否找到配置段
        if not fw01_content and not fw02_content:
            self.add_result(
                Level.ERROR,
                f"站点 {site} 两台设备都未找到interface Port-channel到failover的配置段"
            )
            # 即使没有配置段也要创建Sheet显示错误信息
        elif not fw01_content or not fw02_content:
            missing_device = "FW01-FRP" if not fw01_content else "FW02-FRP"
            self.add_result(
                Level.WARN,
                f"站点 {site} {missing_device} 未找到interface Port-channel到failover的配置段"
            )

        # 创建Excel Sheet并添加标题信息
        worksheet = self._WB.create_sheet(title=site)
        worksheet.append([f"ASA防火墙主备对比 - {site}"])
        worksheet.append([f"对比时间: {format_datetime(datetime.now(), '%Y-%m-%d %H:%M:%S')}"])
        worksheet.append([f"对比范围: interface Port-channel 到 failover 配置段"])
        worksheet.append([f"FW01-FRP文件: {os.path.basename(fw01_path)}"])
        worksheet.append([f"FW02-FRP文件: {os.path.basename(fw02_path)}"])
        worksheet.append([])
        worksheet.append(["FW01-FRP", "FW02-FRP"])
        worksheet.append([])

        # 处理正常对比
        if not fw01_content and not fw02_content:
            # 两台设备都没有配置段的情况
            msg = "配置对比结果: 两台设备都无interface Port-channel到failover配置段"
            worksheet.append([msg, msg])
        elif not fw01_content or not fw02_content:
            # 只有一台设备有配置段的情况
            missing_device = "FW01-FRP" if not fw01_content else "FW02-FRP"
            msg = f"配置对比结果: {missing_device} 无interface Port-channel到failover配置段"
            worksheet.append([msg, msg])
        else:
            # 正常对比
            self._handle_normal_comparison(worksheet, site, fw01_content, fw02_content)

        # 设置列宽和行高
        worksheet.column_dimensions['A'].width = 80
        worksheet.column_dimensions['B'].width = 80
        # 设置所有行的行高为15
        for rowIndex in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[rowIndex].height = 15

    # 重写run方法：在所有站点处理完成后保存Excel文件
    # 重写run方法，在所有站点处理完成后保存Excel文件
    def run(self) -> None:
        """执行ASA防火墙主备对比检查任务

        处理所有站点，对比主备防火墙配置差异并生成报告
        """
        task_items = list(self.items())
        progress = tqdm(
            total=len(task_items),
            desc=self.NAME,
            position=0,
            leave=True,
            dynamic_ncols=True,
            bar_format=BAR_FORMAT,
        ) if SHOW_PROGRESS else None

        try:
            for single_item in task_items:
                try:
                    self.run_single(single_item)
                except Exception as error:
                    self.add_result(Level.ERROR, f"{single_item} 运行异常: {error!r}")
                if progress:
                    progress.update(1)
        finally:
            if progress:
                progress.close()

            # 所有站点处理完成后保存Excel文件
            if hasattr(self, '_WB') and self._WB:
                self._save_excel_file()
            else:
                self.add_result(Level.ERROR, "工作簿不存在，无法保存Excel文件")

    # 验证工作簿是否有效
    def _validate_workbook(self):
        try:
            if not hasattr(self, '_WB') or not self._WB:
                return False

            # 检查是否有工作表
            if not self._WB.sheetnames:
                return False

            # 检查每个工作表是否有有效内容
            for sheet_name in self._WB.sheetnames:
                worksheet = self._WB[sheet_name]
                if worksheet.max_row <= 1:
                    continue  # 空工作表稍后会被清理

                # 检查是否有有效数据
                has_data = False
                max_check_row = min(worksheet.max_row, 10)
                for ROW in worksheet.iter_rows(min_row=1, max_row=max_check_row):
                    if any(cell.value for cell in ROW):
                        has_data = True
                        break

                if not has_data:
                    return False

            return True

        except Exception:
            return False

    # 清理工作簿，移除空的工作表
    def _clean_workbook(self):
        try:
            # 获取所有工作表名称
            sheet_names = self._WB.sheetnames.copy()

            for sheet_name in sheet_names:
                worksheet = self._WB[sheet_name]
                # 检查工作表是否为空（只有标题行或完全为空）
                is_empty = (
                    worksheet.max_row <= 1 or
                    (worksheet.max_row == 1 and not any(
                        worksheet.cell(1, COLUMN).value
                        for COLUMN in range(1, worksheet.max_column + 1)
                    ))
                )
                if is_empty:
                    # 删除空的工作表
                    self._WB.remove(worksheet)

        except Exception as error:
            # 清理失败不影响主流程
            pass

    # 安全保存Excel文件，确保文件完整性
    def _save_excel_file(self):
        output_file = os.path.join(
            self.OUTPUT_DIR,
            f"{self._TODAY}-ASA防火墙主备对比检查.xlsx"
        )

        try:
            # 确保目录存在
            # 创建输出目录（如果目录已存在则不报错）
            ensure_output_dir(self.OUTPUT_DIR)

            # 验证工作簿
            if not self._validate_workbook():
                self.add_result(Level.ERROR, "工作簿验证失败，无法保存Excel文件")
                return

            # 清理工作簿，移除空的工作表
            self._clean_workbook()

            # 尝试多种保存方式
            success = False

            # 方式1: 临时文件 + 原子重命名
            if not success:
                success = self._save_with_temp_file(output_file)

            # 方式2: 直接保存
            if not success:
                success = self._save_directly(output_file)

            # 方式3: 保存为CSV格式作为备选
            if not success:
                success = self._save_as_csv(output_file)

            if success:
                # Excel文件已成功生成，无需输出消息
                pass
            else:
                self.add_result(Level.ERROR, f"所有保存方式都失败: {output_file}")

        except Exception as error:
            self.add_result(Level.ERROR, f"保存Excel文件时发生异常: {error}")

    # 使用临时文件方式保存
    def _save_with_temp_file(self, output_file: str) -> bool:
        temp_file = None
        try:
            temp_file = output_file.replace('.xlsx', '_temp.xlsx')

            # 删除可能存在的临时文件
            if os.path.exists(temp_file):
                os.remove(temp_file)

            # 保存到临时文件
            self._WB.save(temp_file)

            # 验证临时文件
            if os.path.exists(temp_file) and os.path.getsize(temp_file) > 1000:  # 至少1KB
                # 尝试打开临时文件验证其完整性
                try:
                    from openpyxl import load_workbook
                    test_workbook = load_workbook(temp_file)
                    test_workbook.close()
                except Exception:
                    # 临时文件损坏，删除它
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                    return False

                # 删除旧文件
                if os.path.exists(output_file):
                    os.remove(output_file)

                # 原子重命名
                os.rename(temp_file, output_file)

                # 验证最终文件
                if os.path.exists(output_file) and os.path.getsize(output_file) > 1000:
                    # 再次验证最终文件
                    try:
                        from openpyxl import load_workbook
                        test_workbook = load_workbook(output_file)
                        test_workbook.close()
                        return True
                    except Exception:
                        return False

        except Exception as error:
            self.add_result(Level.WARN, f"临时文件保存失败: {error}")
        finally:
            # 清理临时文件
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception:
                    pass

        return False

    # 直接保存方式
    def _save_directly(self, output_file: str) -> bool:
        try:
            # 删除旧文件
            if os.path.exists(output_file):
                os.remove(output_file)

            # 直接保存
            self._WB.save(output_file)

            # 验证文件
            if os.path.exists(output_file) and os.path.getsize(output_file) > 1000:
                # 验证文件完整性
                try:
                    from openpyxl import load_workbook
                    test_workbook = load_workbook(output_file)
                    test_workbook.close()
                    return True
                except Exception:
                    return False

        except Exception as error:
            self.add_result(Level.WARN, f"直接保存失败: {error}")

        return False

    # 保存为CSV格式作为备选
    def _save_as_csv(self, output_file: str) -> bool:
        try:
            csv_file = output_file.replace('.xlsx', '.csv')

            # 获取第一个工作表
            if self._WB.sheetnames:
                worksheet = self._WB[self._WB.sheetnames[0]]

                with open(csv_file, 'w', encoding='utf-8-sig', newline='') as FILE_HANDLE:
                    import csv
                    writer = csv.writer(FILE_HANDLE)

                    for ROW in worksheet.iter_rows(values_only=True):
                        writer.writerow(ROW)

                if os.path.exists(csv_file) and os.path.getsize(csv_file) > 100:
                    self.add_result(Level.WARN, f"已保存为CSV格式: {csv_file}")
                    return True

        except Exception as error:
            self.add_result(Level.ERROR, f"CSV保存也失败: {error}")

        return False
