# ACL 无 ARP 匹配检查任务
#
# 技术栈:openpyxl、ipaddress、正则表达式、网络地址计算、xlsxwriter、CiscoBase（V11新增）
# 目标:识别源/目的两端均无 ARP（ARP 表中仅统计非 INCOMPLETE）的 ACL 规则
#
# 处理逻辑:
# 读取ARP:解析 ARP.log，提取有 MAC 的 IP 作为"有 ARP"；INCOMPLETE 视为"无 ARP"
# 解析ACL:使用CiscoBase统一解析（V11优化），支持 NX-OS CIDR、IOS-XE host/wildcard/混写、ASA any 关键字，提取源/目的网段
# ACL定界:使用CiscoBase.find_acl_blocks_in_column统一处理（V11优化）
# 多端口支持:支持eq www 443、eq 3306 3366等多端口token解析
# 特殊网段处理:支持配置特殊网段列表，特殊网段不参与 ARP 匹配检查
# 判定逻辑:根据源/目的是否为 any、特殊网段等情况，采用不同的匹配策略
# 多设备处理:按工作表逐个处理设备配置，进度显示为N/N（与ACLDupCheckTask一致）
#
# 输出:生成 {YYYYMMDD}-ACL无ARP匹配检查.xlsx，含 Sheet/Device/Row/Rule/SrcNetwork/DstNetwork 明细
# 红色标记:无 ARP 匹配且非特殊网段的网络地址标红显示
#
# 配置说明:
# - 输入文件:自动使用当天日期的ASA备份文件（LOG/DeviceBackupTask/）
# - ARP日志文件:CHECKRULE/ARP.log
# - 输出目录:LOG/ACLArpCheckTask/（V10新结构：从ACL/ACLArpCheckTask迁移）
# - 脚本输出目录:LOG/ACLArpCheckTask/ConfigureOutput/（V10新结构：从CONFIGURATION/ACLArpCheckTask/日期迁移，每次任务自动清空）
# - 脚本输出文件名格式:{日期}-{sheet_name}操作脚本.log 和 {日期}-{sheet_name}回退脚本.log
# - 任务名称:ACL无ARP匹配检查
# - 忽略第三段为86,88,108,153的IP（X.X.86.X, X.X.88.X, X.X.108.X, X.X.153.X 全部忽略）
# - 检测范围:只检测在platform_network_map（公共配置settings.platform_network_map）内的网络，不在平台网段内的网络不检测（V11优化）
# - 脚本输出:操作脚本和回退脚本，根据原始ACL定义是否包含extended关键字自动决定是否添加extended（支持cat1不带extended和cat2带extended的区分，V11优化）
# - 汇总输出控制:通过Config.yaml的enable_summary_output开关控制汇总统计输出，默认关闭
# - 依赖库:xlsxwriter
# task_switches:
#   ACLArpCheckTask: true  # 启用 ACL 无 ARP 匹配检查

# 导入标准库
import os
import re
from dataclasses import dataclass
from datetime import datetime
from ipaddress import IPv4Address, IPv4Network
from typing import Dict, List, Optional, Set, Tuple

# 导入第三方库
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import xlsxwriter

# 导入第三方库
from tqdm import tqdm

# 导入本地应用
from .TaskBase import (
    BaseTask, Level, CONFIG, safe_sheet_name, get_today_str, format_datetime,
    ensure_output_dir, build_log_path, build_output_path,
    load_excel_workbook, create_excel_workbook, save_excel_workbook
)
from .CiscoBase import (
    parse_acl, find_acl_blocks_in_column, extract_acl_rules_from_column, is_acl_rule
)

# 解析ACL正则（复用ACLDup思想，提取必要子集）：支持 NXOS/IOS-XE/ASA，any/CIDR/IP+wildcard，端口eq/range与行尾log均可忽略
NXOS_RE = re.compile(r"""
    ^\s*(?P<num>\d+)?\s*
    (?P<action>permit|deny)\s+
    (?P<proto>\S+)\s+
    (?:(?P<src_any>any)|
       (?P<src_ip>\d+\.\d+\.\d+\.\d+)\s+(?P<src_wc>\d+\.\d+\.\d+\.\d+)|
       (?P<src_cidr>\d+\.\d+\.\d+\.\d+\/\d+))
    (?:\s+(?:eq\s+\S+|range\s+\S+\s+\S+))*
    \s+
    (?:(?P<dst_any>any)|
       (?P<dst_ip>\d+\.\d+\.\d+\.\d+)\s+(?P<dst_wc>\d+\.\d+\.\d+\.\d+)|
       (?P<dst_cidr>\d+\.\d+\.\d+\.\d+\/\d+))
    (?:\s+(?:eq\s+\S+|range\s+\S+\s+\S+))*
    (?:\s+log)?
    \s*$
""", re.IGNORECASE | re.VERBOSE)

IOSXE_WC_RE = re.compile(r"""
    ^\s*(?P<num>\d+)?\s*
    (?P<action>permit|deny)\s+
    (?P<proto>\S+)\s+
    (?P<srcip>\d+\.\d+\.\d+\.\d+)\s+(?P<srcwc>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+.*?)?
    (?:\s+range\s+.*?)?
    \s+(?P<dstip>\d+\.\d+\.\d+\.\d+)\s+(?P<dstwc>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+.*?)?
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
""", re.IGNORECASE | re.VERBOSE)

IOSXE_HOST_RE = re.compile(r"""
    ^\s*(?P<num>\d+)?\s*
    (?P<action>permit|deny)\s+
    (?P<proto>\S+)\s+
    host\s+(?P<srcip>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+.*?)?
    \s+host\s+(?P<dstip>\d+\.\d+\.\d+\.\d+)
    (?:\s+eq\s+.*?)?
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
""", re.IGNORECASE | re.VERBOSE)

IOSXE_MIX_RE = re.compile(r"""
    ^\s*(?P<num>\d+)?\s*
    (?P<action>permit|deny)\s+
    (?P<proto>\S+)\s+
    (?:
        host\s+(?P<srcip_h>\d+\.\d+\.\d+\.\d+) |
        (?P<srcip_w>\d+\.\d+\.\d+\.\d+)\s+(?P<srcwc_w>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+.*?)?
    \s+
    (?:
        host\s+(?P<dstip_h>\d+\.\d+\.\d+\.\d+) |
        (?P<dstip_w>\d+\.\d+\.\d+\.\d+)\s+(?P<dstwc_w>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+.*?)?
    (?:\s+(?:log|log-input|time-range\s+\S+|\S+))*\s*$
""", re.IGNORECASE | re.VERBOSE)

# ASA 格式 ACL 规则（支持 any 关键字和端口名称，但不匹配反掩码格式）
ASA_RE = re.compile(r"""
    ^\s*(?P<num>\d+)?\s*
    (?P<action>permit|deny)\s+
    (?P<proto>\S+)\s+
    (?P<src>any|\d+\.\d+\.\d+\.\d+(?:\/\d+)?)
    (?:\s+(?:eq\s+)?(?P<port_a>\S+))?
    (?:\s+range\s+(?P<port_range_start>\S+)\s+(?P<port_range_end>\S+))?
    \s+
    (?P<dst>any|\d+\.\d+\.\d+\.\d+(?:\/\d+)?)
    (?:\s+(?:eq\s+)?(?P<port_b>\S+))?
    \s*$
""", re.IGNORECASE | re.VERBOSE)

# IOS-XE 反掩码格式（包含 any 关键字）
IOSXE_ANY_RE = re.compile(r"""
    ^\s*(?P<num>\d+)?\s*
    (?P<action>permit|deny)\s+
    (?P<proto>\S+)\s+
    (?:
        any |
        (?P<srcip>\d+\.\d+\.\d+\.\d+)\s+(?P<srcwc>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+.*?)?
    (?:\s+range\s+.*?)?
    \s+
    (?:
        any |
        (?P<dstip>\d+\.\d+\.\d+\.\d+)\s+(?P<dstwc>\d+\.\d+\.\d+\.\d+)
    )
    (?:\s+eq\s+.*?)?
    \s*$
""", re.IGNORECASE | re.VERBOSE)
# IOS-XE：ip + wildcard转网络（假设通配位连续）
def ip_and_wildcard_to_network(ip_str: str, wildcard_str: str) -> Optional[IPv4Network]:
    try:
        ip_integer = int(IPv4Address(ip_str))
        wildcard_integer = int(IPv4Address(wildcard_str))
        netmask_integer = (~wildcard_integer) & 0xFFFFFFFF
        prefix_length = 32 - bin(wildcard_integer).count("1")
        network_integer = ip_integer & netmask_integer
        return IPv4Network((IPv4Address(network_integer), prefix_length), strict=False)
    except Exception:
        return None

# 将IP地址转换为/32网络
def host_to_network(ip_str: str) -> Optional[IPv4Network]:
    try:
        return IPv4Network(f"{ip_str}/32", strict=False)
    except Exception:
        return None

# 将CIDR字符串转换为网络对象
def cidr_to_network(cidr_str: str) -> Optional[IPv4Network]:
    try:
        return IPv4Network(cidr_str, strict=False)
    except Exception:
        return None

# 将'any'关键字转换为0.0.0.0/0网络
def any_to_network() -> IPv4Network:
    return IPv4Network("0.0.0.0/0", strict=False)

# 从ACL规则中提取规则号
def _extract_rule_number(rule_text: str) -> Optional[str]:
    # 从ACL规则中提取规则号
    try:
        # 匹配规则号（数字开头）
        match = re.match(r'^\s*(\d+)', rule_text)
        if match:
            return match.group(1)
        return None
    except Exception:
        return None

# ACL规则数据类：存储解析后的ACL规则信息和ARP匹配状态
@dataclass
class ACLRule:
    sheet: str
    col: int
    row: int
    raw: str
    src: IPv4Network
    dst: IPv4Network
    DEVICE_NAME: str = ""  # 设备名称（从列标题获取）
    SOURCE_HIT: bool = False  # 源网段是否命中 ARP
    DESTINATION_HIT: bool = False  # 目的网段是否命中 ARP
    SOURCE_NO_USE: bool = False  # 源网段是否命中 NoUseIPRange
    DESTINATION_NO_USE: bool = False  # 目的网段是否命中 NoUseIPRange
# 解析ACL规则行，支持多种格式并提取源/目的网络
def parse_acl_line(text: str) -> Tuple[Optional[Tuple[IPv4Network, IPv4Network]], Optional[str]]:
    # 清理并标准化输入行
    cleaned_line = (text or "").strip()
    if not cleaned_line:
        return None, "empty"

    # 尝试匹配NXOS格式的ACL规则
    nxos_match = NXOS_RE.match(cleaned_line)
    if nxos_match:
        # 源网络
        if nxos_match.group("src_any"):
            source_network = any_to_network()
        elif nxos_match.group("src_ip") and nxos_match.group("src_wc"):
            source_network = ip_and_wildcard_to_network(nxos_match.group("src_ip"), nxos_match.group("src_wc"))
        else:
            source_network = cidr_to_network(nxos_match.group("src_cidr"))

        # 目的网络
        if nxos_match.group("dst_any"):
            destination_network = any_to_network()
        elif nxos_match.group("dst_ip") and nxos_match.group("dst_wc"):
            destination_network = ip_and_wildcard_to_network(nxos_match.group("dst_ip"), nxos_match.group("dst_wc"))
        else:
            destination_network = cidr_to_network(nxos_match.group("dst_cidr"))

        if source_network and destination_network:
            return (source_network, destination_network), None
        return None, "nxos_network_parse_fail"

    # 尝试匹配IOS-XE wildcard格式的ACL规则
    iosxe_wildcard_match = IOSXE_WC_RE.match(cleaned_line)
    if iosxe_wildcard_match:
        source_network = ip_and_wildcard_to_network(
            iosxe_wildcard_match.group("srcip"),
            iosxe_wildcard_match.group("srcwc")
        )
        destination_network = ip_and_wildcard_to_network(
            iosxe_wildcard_match.group("dstip"),
            iosxe_wildcard_match.group("dstwc")
        )
        if source_network and destination_network:
            return (source_network, destination_network), None
        return None, "iosxe_wc_network_parse_fail"

    # 尝试匹配IOS-XE host格式的ACL规则
    iosxe_host_match = IOSXE_HOST_RE.match(cleaned_line)
    if iosxe_host_match:
        source_network = host_to_network(iosxe_host_match.group("srcip"))
        destination_network = host_to_network(iosxe_host_match.group("dstip"))
        if source_network and destination_network:
            return (source_network, destination_network), None
        return None, "iosxe_host_network_parse_fail"

    # 尝试匹配IOS-XE混合格式的ACL规则（host和wildcard混合）
    iosxe_mix_match = IOSXE_MIX_RE.match(cleaned_line)
    if iosxe_mix_match:
        srcip_h = iosxe_mix_match.group("srcip_h")
        srcip_w = iosxe_mix_match.group("srcip_w")
        srcwc_w = iosxe_mix_match.group("srcwc_w")
        source_network = (
            host_to_network(srcip_h) if srcip_h
            else ip_and_wildcard_to_network(srcip_w, srcwc_w)
        )
        dstip_h = iosxe_mix_match.group("dstip_h")
        dstip_w = iosxe_mix_match.group("dstip_w")
        dstwc_w = iosxe_mix_match.group("dstwc_w")
        destination_network = (
            host_to_network(dstip_h) if dstip_h
            else ip_and_wildcard_to_network(dstip_w, dstwc_w)
        )
        if source_network and destination_network:
            return (source_network, destination_network), None
        return None, "iosxe_mix_network_parse_fail"

    # 尝试匹配IOS-XE any关键字格式的ACL规则
    iosxe_any_match = IOSXE_ANY_RE.match(cleaned_line)
    if iosxe_any_match:
        # 处理源网络
        if iosxe_any_match.group("srcip") and iosxe_any_match.group("srcwc"):
            source_network = ip_and_wildcard_to_network(iosxe_any_match.group("srcip"), iosxe_any_match.group("srcwc"))
        else:
            source_network = any_to_network()
            

        # 处理目的网络
        if iosxe_any_match.group("dstip") and iosxe_any_match.group("dstwc"):
            destination_network = ip_and_wildcard_to_network(
                iosxe_any_match.group("dstip"),
                iosxe_any_match.group("dstwc")
            )
        else:
            destination_network = any_to_network()
            

        if source_network and destination_network:
            return (source_network, destination_network), None
        return None, "iosxe_any_network_parse_fail"

    asa_match = ASA_RE.match(cleaned_line)
    if asa_match:
        src_str = asa_match.group("src")
        dst_str = asa_match.group("dst")
        

        # 处理源网络
        if src_str.lower() == "any":
            source_network = any_to_network()
        elif "/" in src_str:
            source_network = cidr_to_network(src_str)
        else:
            source_network = host_to_network(src_str)
            

        # 处理目的网络
        if dst_str.lower() == "any":
            destination_network = any_to_network()
        elif "/" in dst_str:
            destination_network = cidr_to_network(dst_str)
        else:
            destination_network = host_to_network(dst_str)
            

        if source_network and destination_network:
            return (source_network, destination_network), None
        return None, "asa_network_parse_fail"

    return None, "no_pattern_match"
# 解析ARP表，返回有MAC的IP集合（INCOMPLETE不计入）
# 解析ARP表，返回有MAC的IP集合（可按配置忽略特定段，如X.X.108.X）
def parse_arp_table(arp_path: str) -> Set[IPv4Address]:
    arp_ok_ip_addresses: Set[IPv4Address] = set()
    if not os.path.exists(arp_path):
        return arp_ok_ip_addresses

    # 从配置中读取要忽略的第三段 octet 值（例如 [108] 表示忽略 X.X.108.X）
    try:
        from .TaskBase import CONFIG as _BASE_CONFIG, require_keys
        require_keys(_BASE_CONFIG, ["ACLArpCheckTask"], "root")
        require_keys(
            _BASE_CONFIG["ACLArpCheckTask"],
            ["ignore_third_octet"],
            "ACLArpCheckTask"
        )
        ignore_octet_third_values = set(
            int(OCTET_VALUE)
            for OCTET_VALUE in _BASE_CONFIG["ACLArpCheckTask"]["ignore_third_octet"]
        )
    except Exception:
        ignore_octet_third_values = set()

    with open(arp_path, "r", encoding="utf-8", errors="ignore") as file_handle:
        for ARP_LOG_LINE in file_handle:
            ARP_LOG_LINE = ARP_LOG_LINE.strip()
            if not ARP_LOG_LINE:
                continue
            # 形如: 10.10.200.54    00:00:09  INCOMPLETE      Vlan200
            parts = ARP_LOG_LINE.split()
            if len(parts) < 3:
                continue
            ip_str, _, mac_or_state = parts[0], parts[1], parts[2]
            if mac_or_state.upper() == "INCOMPLETE":
                # 视为无 ARP
                continue
            try:
                ipAddress = IPv4Address(ip_str)
            except Exception:
                continue
            # 忽略 X.X.108.X 这类第三段等于指定值的 IP
            if ignore_octet_third_values:
                third = int(str(ipAddress).split(".")[2])
                if third in ignore_octet_third_values:
                    continue
            arp_ok_ip_addresses.add(ipAddress)
    return arp_ok_ip_addresses

# 解析NoUseIPRange.log文件，返回未使用IP网段列表
# 解析NoUseIPRange.log文件，返回未使用IP网段列表
def parse_no_use_ip_ranges(no_use_file: str) -> List[str]:
    no_use_ranges: List[str] = []
    if not os.path.exists(no_use_file):
        return no_use_ranges

    with open(no_use_file, "r", encoding="utf-8", errors="ignore") as file_handle:
        for line in file_handle:
            line = line.strip()
            if not line:
                continue
            no_use_ranges.append(line)
    return no_use_ranges

# 检查IP网段是否命中NoUseIPRange
# 检查IP网段是否命中NoUseIPRange列表（模糊匹配IP前缀）
def check_no_use_ip_range(network: IPv4Network, no_use_ranges: List[str]) -> bool:
    network_str = str(network.network_address)
    for no_use_range in no_use_ranges:
        # 检查网络地址是否以NoUseIPRange中的前缀开头（模糊匹配）
        if network_str.startswith(no_use_range):
            return True
    return False

# 优化的ARP匹配检测，找到匹配后立即返回
# 优化的ARP匹配检测，找到匹配后立即返回
def check_arp_match_optimized(network: IPv4Network, arp_ok_ip_addresses: Set[IPv4Address]) -> bool:
    for ip_address in arp_ok_ip_addresses:
        if ip_address in network:
            return True  # 找到匹配后立即退出
    return False

# 判断文本是否为ACL规则（参考ACLDupCheckTask）
# 判断文本是否为ACL规则
def _is_acl_rule(text: str) -> bool:
    text = text.strip().lower()
    # 排除证书、配置等非ACL数据
    if any(keyword in text for keyword in [
        'certificate', 'crypto', 'quit', 'exit', 'config', 'interface',
        'router', 'version', 'hostname', 'enable', 'password', 'username',
        'line', 'service', 'logging', 'ntp', 'snmp', 'tacacs', 'radius'
    ]):
        return False
    

    # 只处理包含permit/deny的规则
    return 'permit' in text or 'deny' in text

# ACL定界功能已迁移到CiscoBase
# 使用: from .CiscoBase import find_acl_blocks_in_column, extract_acl_rules_from_column


# Unicode标记方案的核心函数
# 使用Unicode标记关键字
def mark_keywords_with_unicode(text: str, keywords: list) -> str:
    marked_text = text
    for keyword in keywords:
        # 使用中间点标记关键字（不显眼且不会在配置文本中出现）
        marked_text = marked_text.replace(keyword, f'·{keyword}·')
    return marked_text

# 将Unicode标记的文本转换为富文本
def create_rich_text_from_unicode_marked(text: str, workbook, color_type: str = 'red'):
    if '·' not in text:
        normal_format = workbook.add_format({'color': 'black'})
        return [normal_format, text]
    

    # 创建格式
    if color_type == 'red':
        color_format = workbook.add_format({'color': 'red'})
    elif color_type == 'orange':
        color_format = workbook.add_format({'color': 'orange'})
    else:
        color_format = workbook.add_format({'color': 'black'})
    

    normal_format = workbook.add_format({'color': 'black'})
    

    result = []
    parts = text.split('·')
    

    for i, part in enumerate(parts):
        if part:  # 跳过空字符串
            if i % 2 == 0:  # 偶数索引：普通文本
                result.append(normal_format)
                result.append(part)
            else:  # 奇数索引：关键字（红色）
                result.append(color_format)
                result.append(part)
    

    return result

# 处理单个ACL块内的ARP检测（使用Unicode标记方案）
# 处理单个ACL块内的ARP检测，使用Unicode标记方案实现前导空格保留和部分文本标色
def process_acl_block_with_unicode_marking(
    inputWorksheet, worksheet, column_index, start_row, end_row,
    arp_ok_ip_addresses, no_use_ranges, acl_config, workbook,
    sheet_name, platform_network_map
):
    red_count = 0
    yellow_count = 0
    total_count = 0
    colored_rules_info = []
    

    # 创建格式
    red_format = workbook.add_format({'color': 'red'})
    orange_format = workbook.add_format({'color': 'orange'})
    normal_format = workbook.add_format({'color': 'black'})
    bg_format = workbook.add_format({'bg_color': '#E6F3FF'})
    

    for ROW in range(start_row, end_row + 1):
        cell_value = inputWorksheet.cell(row=ROW, column=column_index).value
        if cell_value is None:
            continue
        cleaned_text = str(cell_value).strip()
        if not cleaned_text:
            continue
        

        # 只处理真正的ACL规则，忽略证书等数据
        if not _is_acl_rule(cleaned_text):
            continue
        

        # 解析ACL规则
        parsed, err = parse_acl_line(cleaned_text)
        if not parsed:
            continue
        source_network, destination_network = parsed

        # 判定是否命中 ARP：使用优化的检测函数
        source_hit = check_arp_match_optimized(source_network, arp_ok_ip_addresses)
        destination_hit = check_arp_match_optimized(destination_network, arp_ok_ip_addresses)
        

        # 检查是否命中 NoUseIPRange
        source_no_use = check_no_use_ip_range(source_network, no_use_ranges)
        destination_no_use = check_no_use_ip_range(destination_network, no_use_ranges)
        

        # 检查源和目的是否包含 any (0.0.0.0/0)
        src_is_any = str(source_network) == "0.0.0.0/0"
        dst_is_any = str(destination_network) == "0.0.0.0/0"
        

        # 检查是否包含需要忽略的网段（如X.X.108.X）
        ignore_octet_third_values = acl_config["ignore_third_octet"]
        src_should_ignore = False
        dst_should_ignore = False
        

        if ignore_octet_third_values:
            src_str = str(source_network.network_address)
            dst_str = str(destination_network.network_address)
            

            # 检查源网络是否包含需要忽略的第三段
            if '.' in src_str:
                src_parts = src_str.split('.')
                if len(src_parts) >= 3:
                    src_third = int(src_parts[2])
                    if src_third in ignore_octet_third_values:
                        src_should_ignore = True
            

            # 检查目的网络是否包含需要忽略的第三段
            if '.' in dst_str:
                dst_parts = dst_str.split('.')
                if len(dst_parts) >= 3:
                    dst_third = int(dst_parts[2])
                    if dst_third in ignore_octet_third_values:
                        dst_should_ignore = True
        

        # 检查网络是否在平台网段内（从公共配置读取）
        platform_networks = platform_network_map.get(sheet_name, [])
        src_in_platform = False
        dst_in_platform = False
        

        if platform_networks:
            # 检查源网络是否在平台网段内（重叠或包含关系）
            for platform_network in platform_networks:
                if source_network.overlaps(platform_network) or source_network.subnet_of(platform_network):
                    src_in_platform = True
                    break
            # 检查目的网络是否在平台网段内（重叠或包含关系）
            for platform_network in platform_networks:
                if destination_network.overlaps(platform_network) or destination_network.subnet_of(platform_network):
                    dst_in_platform = True
                    break
        

        # 检测逻辑：只检测在平台网段内的网络，且忽略ignore_third_octet
        if src_should_ignore or dst_should_ignore:
            # 如果源或目的网络包含需要忽略的网段（如X.X.108.X），则不检测
            # 但如果命中NoUseIPRange，则优先级更高，仍然需要检测和标色
            if source_no_use or destination_no_use:
                should_detect = True  # NoUseIPRange优先级高于ignore_third_octet
            else:
                should_detect = False
        elif src_is_any and dst_is_any:
            # any to any 不检测
            should_detect = False
        elif src_is_any:
            # 源为any，只检测目的是否在平台网段内且命中ARP
            if dst_in_platform:
                should_detect = not destination_hit
            else:
                should_detect = False
        elif dst_is_any:
            # 目的为any，只检测源是否在平台网段内且命中ARP
            if src_in_platform:
                should_detect = not source_hit
            else:
                should_detect = False
        elif src_in_platform and dst_in_platform:
            # 源和目的都在平台网段内，检测ARP匹配
            should_detect = (not source_hit) or (not destination_hit)
        elif src_in_platform:
            # 只有源在平台网段内，只检测源是否命中ARP
            should_detect = not source_hit
        elif dst_in_platform:
            # 只有目的在平台网段内，只检测目的是否命中ARP
            should_detect = not destination_hit
        else:
            # 源和目的都不在平台网段内，不检测
            should_detect = False

        if should_detect:
            total_count += 1
            

            # 获取原始文本（保留前导空格）
            original_text = str(cell_value)
            

            # 标色逻辑：使用Unicode标记方案
            if source_no_use or destination_no_use:
                # 命中NoUseIPRange（IP前缀模糊匹配 X.X.X.），橙色标色
                keywords_to_mark = []
                if source_no_use:
                    keywords_to_mark.append(str(source_network.network_address))
                if destination_no_use:
                    keywords_to_mark.append(str(destination_network.network_address))
                

                # 使用Unicode标记
                marked_text = mark_keywords_with_unicode(original_text, keywords_to_mark)
                rich_text = create_rich_text_from_unicode_marked(marked_text, workbook, 'orange')
                if len(rich_text) >= 3:
                    worksheet.write_rich_string(ROW-1, column_index-1, *rich_text, bg_format)
                else:
                    worksheet.write(ROW-1, column_index-1, original_text, bg_format)
                yellow_count += 1
                

                # 记录标色规则信息
                colored_rules_info.append({
                    'row': ROW,
                    'column': column_index,
                    'rule_text': cleaned_text,
                    'color_type': 'orange',
                    'rule_number': _extract_rule_number(cleaned_text)
                })
            else:
                # 无ARP匹配（IP精确匹配 X.X.X.X）且非特殊网段，红色标色
                keywords_to_mark = []
                if not source_hit and not src_is_any:
                    keywords_to_mark.append(str(source_network.network_address))
                if not destination_hit and not dst_is_any:
                    keywords_to_mark.append(str(destination_network.network_address))
                

                # 使用Unicode标记
                marked_text = mark_keywords_with_unicode(original_text, keywords_to_mark)
                rich_text = create_rich_text_from_unicode_marked(marked_text, workbook, 'red')
                if len(rich_text) >= 3:
                    worksheet.write_rich_string(ROW-1, column_index-1, *rich_text, bg_format)
                else:
                    worksheet.write(ROW-1, column_index-1, original_text, bg_format)
                red_count += 1
                

                # 记录标色规则信息
                colored_rules_info.append({
                    'row': ROW,
                    'column': column_index,
                    'rule_text': cleaned_text,
                    'color_type': 'red',
                    'rule_number': _extract_rule_number(cleaned_text)
                })
        else:
            # 不需要标色的单元格，直接写入（保留前导空格）
            original_text = str(cell_value)
            worksheet.write(ROW-1, column_index-1, original_text)
    

    return red_count, yellow_count, total_count, colored_rules_info


# ACL无ARP匹配检查任务类：
# 1) 检查ACL规则在ARP表中的命中情况（无命中/命中NoUseIPRange）
# 2) 使用 xlsxwriter 输出原位富文本标色（Unicode 标记，保留前导空格）
# 3) 基于标色结果按设备/ACL分组生成操作脚本与回退脚本
class ACLArpCheckTask(BaseTask):

    # 初始化ACL无ARP匹配检查任务：设置固定配置参数
    def __init__(self):
        super().__init__("ACL无ARP匹配检查")
        # 固定配置参数
        TODAY_STR = datetime.now().strftime('%Y%m%d')
        # V10新结构：从 LOG/DeviceBackupTask/ 读取（ACL/SourceACL已迁移）
        self.INPUT_PATH = os.path.join("LOG", "DeviceBackupTask", f"{TODAY_STR}-关键设备配置备份输出EXCEL基础任务.xlsx")
        self.ARP_LOG_FILE = os.path.join("CHECKRULE", "ARP.log")
        self.NO_USE_IP_RANGE_FILE = os.path.join("CHECKRULE", "NoUseIPRange.log")
        # V10新结构：直接输出到 LOG/ACLArpCheckTask/
        self.OUTPUT_DIR = os.path.join("LOG", "ACLArpCheckTask")
        self.OUTPUT_PATH = os.path.join(self.OUTPUT_DIR, f"{TODAY_STR}-ACL无ARP匹配检查.xlsx")

    # 返回要处理的文件列表
    def items(self):
        if not os.path.exists(self.INPUT_PATH):
            return []
        try:
            inputWorkbook = load_excel_workbook(self.INPUT_PATH)
            sheet_names = [
                inputWorksheet.title
                for inputWorksheet in inputWorkbook.worksheets
                if inputWorksheet.title != 'Report'
            ]
            inputWorkbook.close()
            return sheet_names
        except Exception:
            return []

    # 处理单个工作表：解析ACL规则并检查ARP匹配
    def run_single(self, sheet_name: str) -> None:
        # 这个方法现在只用于统计，实际处理在run方法中完成
        pass
    

    # 重写run方法：直接在源文件中进行标色修改
    # 执行ACL无ARP匹配检查任务
    def run(self) -> None:
        # 获取所有工作表
        sheet_names = list(self.items())
        if not sheet_names:
            self.add_result(Level.ERROR, "未找到可处理的工作表")
            return

        # 使用父类的进度条处理

        from tqdm import tqdm

        from .TaskBase import BAR_FORMAT, SHOW_PROGRESS
        

        progress = tqdm(
            total=len(sheet_names),
            desc=self.NAME,
            position=0,
            leave=True,
            dynamic_ncols=True,
            bar_format=BAR_FORMAT,
        ) if SHOW_PROGRESS else None
        

        try:
            # 确保输出目录存在
            os.makedirs(self.OUTPUT_DIR, exist_ok=True)
            # 确保并清空脚本固定输出目录（每次运行前清空）
            config_dir = build_log_path("ACLArpCheckTask", "ConfigureOutput")
            ensure_output_dir(config_dir)
            for _name in os.listdir(config_dir):
                _path = os.path.join(config_dir, _name)
                if os.path.isfile(_path):
                    try:
                        os.remove(_path)
                    except Exception:
                        pass
            

            # 生成输出文件名（覆盖原文件）
            output_filename = f"{get_today_str()}-ACL无ARP匹配检查.xlsx"
            output_path = os.path.join(self.OUTPUT_DIR, output_filename)
            

            # 执行 ACL 无ARP匹配检查，直接在源文件中标色
            per_sheet_stats = self._process_file_with_coloring(self.INPUT_PATH, output_path, progress)
            

            # 生成操作脚本和回退脚本（仅对有标色的工作表）
            try:
                self._generate_operation_scripts(self.INPUT_PATH, per_sheet_stats)
            except Exception as error:
                self.add_result(Level.ERROR, f"生成操作脚本失败: {error}")
                import traceback
                traceback.print_exc()
            

            # 添加每个工作表的结果记录
            for sheet_name, stats in per_sheet_stats.items():
                self.add_result(
                    Level.OK, 

                    f"站点{sheet_name}处理完成："
                    f"无ARP精确匹配(红色关键字标色)={stats['red']}，"
                    f"命中NoUseIPRange模糊匹配(橙色关键字标色)={stats['yellow']}，"
                    f"总计={stats['total']}条规则"
                )
            

            # 汇总统计（通过Config.yaml的enable_summary_output开关控制，仅输出到LOG文件）
            try:
                from .TaskBase import require_keys
                require_keys(CONFIG, ["ACLArpCheckTask"], "root")
                enable_summary = CONFIG["ACLArpCheckTask"].get("enable_summary_output", False)
            except Exception:
                enable_summary = False
            

            if enable_summary:
                total_red = sum(stats.get('red', 0) for stats in per_sheet_stats.values())
                total_yellow = sum(stats.get('yellow', 0) for stats in per_sheet_stats.values())
                total_count = sum(stats.get('total', 0) for stats in per_sheet_stats.values())
                self.add_result(
                    Level.OK,
                    f"ACL无ARP匹配检查汇总：处理{len(per_sheet_stats)}个站点，"
                    f"无ARP精确匹配(红色关键字标色)={total_red}，"
                    f"命中NoUseIPRange模糊匹配(橙色关键字标色)={total_yellow}，"
                    f"总计={total_count}条规则"
                )
                

        except Exception as error:
            self.add_result(Level.ERROR, f"ACL 无ARP匹配检查失败：{str(error)}")
        finally:
            if progress:
                progress.close()
    

    # 处理文件并直接在源文件中标色（类似ACLDupCheckTask）
    # 处理Excel文件，使用Unicode标记方案实现前导空格保留和部分文本标色
    def _process_file_with_coloring(
        self, input_path: str, output_path: str, progress=None
    ) -> Dict[str, Dict[str, int]]:
        if not os.path.exists(input_path):
            self.add_result(Level.ERROR, f"未找到输入文件: {input_path}")
            return {}

        # 解析ARP表和NoUseIPRange文件
        arp_ok_ip_addresses = parse_arp_table(self.ARP_LOG_FILE)
        # self.add_result(Level.OK, f"解析 ARP 表完成，共 {len(arp_ok_ips)} 个有效 IP")

        no_use_ranges = parse_no_use_ip_ranges(self.NO_USE_IP_RANGE_FILE)
        # self.add_result(Level.OK, f"解析 NoUseIPRange 完成，共 {len(no_use_ranges)} 个未使用网段")

        # 获取配置（必须配置）
        from .TaskBase import require_keys
        require_keys(CONFIG, ["ACLArpCheckTask"], "root")
        require_keys(CONFIG["ACLArpCheckTask"], ["ignore_third_octet"], "ACLArpCheckTask")
        acl_config = CONFIG["ACLArpCheckTask"]
        

        # 加载平台网段映射（从公共配置读取）
        platform_network_map = {}
        try:
            config_map = CONFIG.get("settings", {}).get("platform_network_map", {})
            for sheet_name, network_strings in config_map.items():
                networks = [IPv4Network(net_str, strict=False) for net_str in network_strings]
                platform_network_map[sheet_name] = networks
        except Exception:
            # 如果配置读取失败，使用空字典
            pass

        # 使用xlsxwriter创建新的工作簿
        workbook = xlsxwriter.Workbook(output_path)
        

        # 创建格式
        red_format = workbook.add_format({'color': 'red'})
        orange_format = workbook.add_format({'color': 'orange'})
        normal_format = workbook.add_format({'color': 'black'})
        bg_format = workbook.add_format({'bg_color': '#E6F3FF'})
        

        per_sheet_stats: Dict[str, Dict[str, int]] = {}
        colored_rules: Dict[str, List[Dict]] = {}  # 存储标色的规则信息

        # 读取原始文件
        inputWorkbook = load_excel_workbook(input_path)
        

        for inputWorksheet in inputWorkbook.worksheets:
            sheet_name = inputWorksheet.title
            if sheet_name == 'Report':
                continue

            sheet_red_count = 0
            sheet_yellow_count = 0
            sheet_total_count = 0
            sheet_colored_rules = []
            

            # 创建新的工作表
            worksheet = workbook.add_worksheet(sheet_name)
            

            # 重新设置列宽（确保不被覆盖）
            for col_idx in range(1, inputWorksheet.max_column + 1):
                column_letter = inputWorksheet.cell(row=1, column=col_idx).column_letter
                if column_letter in inputWorksheet.column_dimensions:
                    col_dim = inputWorksheet.column_dimensions[column_letter]
                    if col_dim.width:
                        worksheet.set_column(col_idx-1, col_idx-1, col_dim.width)
            

            # 先复制所有原始数据（保留边界以外的内容）
            for row_idx in range(1, inputWorksheet.max_row + 1):
                for col_idx in range(1, inputWorksheet.max_column + 1):
                    cell = inputWorksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        # 保留原始数据的前导空格
                        cell_value = str(cell.value)
                        worksheet.write(row_idx-1, col_idx-1, cell_value)
            

            # 获取工作表的最大列数
            max_col = inputWorksheet.max_column
            

            # 遍历所有列，使用ACL块优化处理
            for col_idx in range(1, max_col + 1):
                # 找到该列中的ACL块
                acl_blocks = find_acl_blocks_in_column(inputWorksheet, col_idx)
                

                # 处理每个ACL块
                for start_row, end_row in acl_blocks:
                    red_count, yellow_count, total_count, colored_rules_info = (
                        process_acl_block_with_unicode_marking(
                            inputWorksheet, worksheet, col_idx, start_row, end_row,
                            arp_ok_ip_addresses, no_use_ranges, acl_config,
                            workbook, sheet_name, platform_network_map
                        )
                    )
                    

                    sheet_red_count += red_count
                    sheet_yellow_count += yellow_count
                    sheet_total_count += total_count
                    

                    # 收集标色的规则信息
                    for rule_info in colored_rules_info:
                        sheet_colored_rules.append(rule_info)
            

            per_sheet_stats[sheet_name] = {
                "red": sheet_red_count,
                "yellow": sheet_yellow_count,
                "total": sheet_total_count
            }
            

            colored_rules[sheet_name] = sheet_colored_rules
            

            # 更新进度条
            if progress:
                progress.update(1)

        # 关闭工作簿
        workbook.close()
        inputWorkbook.close()
        

        # 将标色规则信息存储到实例变量中
        self.colored_rules = colored_rules
        

        return per_sheet_stats

    # 从指定列中向上搜索ACL名称，返回(ACL名称, 是否包含extended)
    def _find_acl_name_in_column(
        self, worksheet, column_index: int, start_row_index: int
    ) -> Optional[Tuple[str, bool]]:
        # 从start_row向上搜索，最多搜索2000行（增加搜索范围）
        for ROW in range(start_row_index, max(1, start_row_index - 2000), -1):
            cell_value = worksheet.cell(row=ROW, column=column_index).value
            if cell_value and isinstance(cell_value, str):
                cell_str = str(cell_value).strip()
                # 查找ACL定义，检查是否包含extended关键字
                match = re.search(r'ip access-list (?:extended )?(\S+)', cell_str, re.IGNORECASE)
                if match:
                    acl_name = match.group(1)
                    # 检查原始定义中是否包含extended关键字（忽略大小写）
                    has_extended = 'extended' in cell_str.lower()
                    return (acl_name, has_extended)
        

        return None

    # 生成操作脚本和回退脚本（仅对有标色的工作表）
    def _generate_operation_scripts(self, input_path: str, per_sheet_stats: Dict[str, Dict[str, int]]) -> None:
        try:
            # 生成当日日期前缀，文件名保留日期
            date_str = get_today_str()
            # V10新结构（调整）：脚本固定输出至 LOG/ACLArpCheckTask/ConfigureOutput（不再使用日期目录）
            config_dir = build_log_path("ACLArpCheckTask", "ConfigureOutput")
            ensure_output_dir(config_dir)
            

            # 读取原始Excel文件
            inputWorkbook = load_excel_workbook(input_path)
            

            # 遍历所有工作表，只处理有标色的工作表
            for inputWorksheet in inputWorkbook.worksheets:
                sheet_name = inputWorksheet.title
                if sheet_name == 'Report':
                        continue
                

                # 获取统计信息
                stats = per_sheet_stats.get(sheet_name, {"red": 0, "yellow": 0, "total": 0})
                

                # 只处理有标色的工作表
                if stats["total"] == 0:
                        continue
                

                # 获取该工作表的标色规则信息
                colored_rules = getattr(self, 'colored_rules', {}).get(sheet_name, [])
                

                # 生成文件名（带日期前缀）
                operation_file = os.path.join(config_dir, f"{date_str}-{sheet_name}操作脚本.log")
                rollback_file = os.path.join(config_dir, f"{date_str}-{sheet_name}回退脚本.log")
                

                with open(operation_file, 'w', encoding='utf-8') as operationLogFile, \
                     open(rollback_file, 'w', encoding='utf-8') as rollbackLogFile:
                    

                    # 写入文件头
                    timestamp = format_datetime(datetime.now(), '%Y-%m-%d %H:%M:%S')
                    operationLogFile.write(
                        f"# {sheet_name}操作脚本 - {timestamp}\n"
                    )
                    operationLogFile.write(
                        f"# 统计: 红色标记={stats['red']}, "
                        f"橙色标记={stats['yellow']}, 总计={stats['total']}\n\n"
                    )
                    

                    rollbackLogFile.write(f"# {sheet_name}回退脚本 - {format_datetime(datetime.now(), '%Y-%m-%d %H:%M:%S')}\n")
                    rollbackLogFile.write(f"# 统计: 红色标记={stats['red']}, 橙色标记={stats['yellow']}, 总计={stats['total']}\n\n")
                    

                    # 按设备（列）分组处理标色规则
                    device_groups = {}
                    for rule_info in colored_rules:
                        column_index_value = rule_info['column']
                        row_index_value = rule_info['row']
                        

                        # 获取设备信息（从第1行获取设备名称）
                        device_name = None
                        if inputWorksheet.cell(row=1, column=column_index_value).value:
                            device_name = str(inputWorksheet.cell(row=1, column=column_index_value).value).strip()
                        

                        if device_name:
                            if column_index_value not in device_groups:
                                device_groups[column_index_value] = {
                                    'device_name': device_name,
                                    'rules': []
                                }
                            device_groups[column_index_value]['rules'].append(rule_info)
                    

                    # 遍历设备组
                    for column_index_key, device_info in device_groups.items():
                        device_name = device_info['device_name']
                        rules = device_info['rules']
                        

                        # 写入设备信息头
                        operationLogFile.write(f"#####{device_name}#####\n")
                        rollbackLogFile.write(f"#####{device_name}#####\n")
                        

                        # 按ACL名称分组该设备的规则（保留extended信息）
                        acl_groups = {}
                        for rule_info in rules:
                            row_index_value = rule_info['row']
                            acl_info = self._find_acl_name_in_column(inputWorksheet, column_index_key, row_index_value)
                            

                            if acl_info:
                                acl_name, has_extended = acl_info
                                # 使用(acl_name, has_extended)作为键，确保同一ACL的extended信息一致
                                acl_key = (acl_name, has_extended)
                                if acl_key not in acl_groups:
                                    acl_groups[acl_key] = []
                                acl_groups[acl_key].append(rule_info)
                        

                        # 遍历该设备的ACL组
                        for (acl_name, has_extended), acl_rules in acl_groups.items():
                            # 根据原始ACL定义是否包含extended来决定是否添加extended关键字
                            if has_extended:
                                operationLogFile.write(f"ip access-list extended {acl_name}\n")
                                rollbackLogFile.write(f"ip access-list extended {acl_name}\n")
                            else:
                                operationLogFile.write(f"ip access-list {acl_name}\n")
                                rollbackLogFile.write(f"ip access-list {acl_name}\n")
                            

                            # 写入操作脚本（删除规则）
                            for rule_info in acl_rules:
                                rule_number = rule_info['rule_number']
                                if rule_number:
                                    operationLogFile.write(f" no {rule_number}\n")
                            

                            operationLogFile.write("\n")
                            

                            # 写入回退脚本（恢复规则）
                            for rule_info in acl_rules:
                                rule_text = rule_info['rule_text']
                                rollbackLogFile.write(f" {rule_text}\n")
                            

                            rollbackLogFile.write("\n")
            

            inputWorkbook.close()
            

            # 记录生成的文件
            for sheet_name, stats in per_sheet_stats.items():
                if stats["total"] > 0:
                    op_script = (
                        f"LOG/ACLArpCheckTask/ConfigureOutput/"
                        f"{date_str}-{sheet_name}操作脚本.log"
                    )
                    self.add_result(Level.OK, f"生成{sheet_name}操作脚本: {op_script}")
                    rb_script = (
                        f"LOG/ACLArpCheckTask/ConfigureOutput/"
                        f"{date_str}-{sheet_name}回退脚本.log"
                    )
                    self.add_result(Level.OK, f"生成{sheet_name}回退脚本: {rb_script}")
            

        except Exception as error:
            self.add_result(Level.ERROR, f"生成操作脚本失败: {error}")

